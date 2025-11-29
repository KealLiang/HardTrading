import json
import os
import re
from collections import Counter

import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# 指数顺序配置
INDEX_ORDER = [
    "上证指数",
    "深证成指",
    "创业板指",
    "科创50",
    "北证50",
    "中证500",
    "大盘高贝",
    "小盘高贝"
]

# 选择一个同义词分组规则
from data.reasons.origin_synonym_groups import synonym_groups

# from data.reasons.updated_synonym_groups import synonym_groups

# 概念正统度分层配置（可选）
try:
    from analysis.group_config.concept_tiers import (
        GROUP_TIER,
        GROUP_ROLE,
        TIER_WEIGHT,
        ROLE_WEIGHT,
        CORE_KEYWORDS,
        CORE_KEYWORD_BONUS,
        FREQ_WEIGHT,
        HOT_WEIGHT,
        REASON_OVERRIDES,
    )

    CONCEPT_TIERING_AVAILABLE = True
except Exception:
    # 安全降级：无配置时，打分功能禁用
    CONCEPT_TIERING_AVAILABLE = False
    GROUP_TIER = {}
    GROUP_ROLE = {}
    TIER_WEIGHT = {1: 1.0, 2: 0.6, 3: 0.3}
    ROLE_WEIGHT = {"brand": 0.2, "event": 0.2, "chain": 0.4, "industry": 0.3, "generic": 0.1}
    CORE_KEYWORDS = {}
    CORE_KEYWORD_BONUS = 0.5
    FREQ_WEIGHT = 0.2
    HOT_WEIGHT = 0.1
    REASON_OVERRIDES = {}

# 排除列表 - 这些原因不会被选为热门原因
EXCLUDED_REASONS = [
    "预期改善",
    "国企",
]

# 优先列表 - 这些原因必定会被选为热门原因，并且按照列表顺序排序
PRIORITY_REASONS = [
    # "旅游",
    # "房地产",
]

# 选取top n的原因着色
TOP_N = 9

# 颜色列表 - 彩虹色系(深色)
COLORS = [
    "FF5A5A",  # 红色
    "FF8C42",  # 橙色
    "FFCE30",  # 黄色
    "6AD15A",  # 绿色
    "45B5FF",  # 蓝色
    "9966FF",  # 紫色
    "FF66B3",  # 粉色
    "5ACDCD",  # 青色
    "DDA0DD",  # 浅紫红
    "FFB366",  # 浅橙色
    "FFF066",  # 浅黄色
    "90EE90",  # 浅绿色
    "87CEFA",  # 浅蓝色
    "B19CD9",  # 浅紫色
    "FFB6C1",  # 浅粉色
    "E1FFFF",  # 浅青色
]

# 多次上榜但无热门原因的颜色
MULTI_COLOR = "1abc9c"

# 表头颜色
HEADER_COLOR = "E0E0E0"  # 浅灰色

# 涨停板颜色映射（根据连板数）
BOARD_COLORS = {
    1: "FFB3B3",  # 首板 (淡红色)
    2: "FF9999",  # 2板 (红色)
    3: "FF8080",  # 3板 (深红色)
    4: "FF6666",  # 4板 (大红色)
    5: "FF4D4D",  # 5板 (深红色)
    6: "FF3333",  # 6板 (暗红色)
    7: "FF1A1A",  # 7板 (更暗红色)
    8: "FF0000",  # 8板 (非常暗红色)
    9: "E60000",  # 9板 (极暗红色)
    10: "CC0000",  # 10板及以上 (近黑红色)
}

# 定义蓝色系梯度颜色映射（4板及以上，每2板一个档次）
HIGH_BOARD_COLORS = {
    4: "E6F3FF",  # 4板 - 最浅蓝色
    6: "CCE8FF",  # 6板 - 浅蓝色
    8: "99D1FF",  # 8板 - 中蓝色
    10: "66BAFF",  # 10板 - 深蓝色
    12: "3394FF",  # 12板 - 更深蓝色
    14: "0073E6",  # 14板及以上 - 最深蓝色
}

# 重复入选梯度颜色映射
REENTRY_COLORS = {
    2: "D3D3D3",  # 2次入选 - 浅灰色
    3: "A9A9A9",  # 3次入选 - 中灰色
    4: "696969",  # 4次及以上 - 深灰色
}

# 红绿色系定义，用于涨跌幅颜色表示
RED_COLORS = [
    "FFCCCC",  # 浅红色 (0-1%)
    "FF9999",  # 淡红色 (1-2%)
    "FF6666",  # 红色 (2-3%)
    "FF3333",  # 深红色 (3-5%)
    "FF0000",  # 大红色 (>5%)
]

GREEN_COLORS = [
    "CCFFCC",  # 浅绿色 (0-1%)
    "99FF99",  # 淡绿色 (1-2%)
    "66FF66",  # 绿色 (2-3%)
    "33FF33",  # 深绿色 (3-5%)
    "00CC00",  # 大绿色 (>5%)
]

# 涨跌幅颜色阈值
PCT_CHANGE_THRESHOLDS = [1.0, 2.0, 3.0, 5.0]

# 未分类原因的最小打印阈值
UNCLASSIFIED_PRINT_THRESHOLD = 5


def normalize_reason(reason):
    """
    将原因标准化，处理同一类型的不同表述，优先匹配最具体的模式
    """
    # 移除所有空格
    reason = re.sub(r'\s+', '', reason)
    original_reason = reason

    # 存储所有匹配结果及其匹配信息
    matches = []

    # 检查原因属于哪个组
    for main_reason, synonyms in synonym_groups.items():
        for synonym in synonyms:
            # 处理通配符匹配
            if '%' in synonym:
                # 转换SQL风格通配符为正则表达式
                pattern = synonym.replace('%', '(.*)')
                regex = re.compile(f"^{pattern}$")
                match = regex.search(reason)

                if match:
                    # 计算匹配的具体部分
                    matched_text = reason
                    for group in match.groups():
                        matched_text = matched_text.replace(group, '')

                    # 匹配强度 = 匹配文本长度 / 原文长度
                    match_strength = len(matched_text) / len(reason)
                    matches.append((main_reason, match_strength, synonym))

            # 保留原有的包含匹配
            elif synonym in reason:
                match_strength = len(synonym) / len(reason)
                matches.append((main_reason, match_strength, synonym))

    # 如果有匹配，选择匹配强度最高的
    if matches:
        matches.sort(key=lambda x: x[1], reverse=True)
        return matches[0][0]

    return f"未分类_{original_reason}"


def get_reason_match_type(original_reason, normalized_reason):
    """
    判断原因的匹配类型：精确匹配(exact)或模糊匹配(fuzzy)
    
    Args:
        original_reason: 原始原因文本
        normalized_reason: 规范化后的原因（即所属的主分组名）
    
    Returns:
        str: 'exact' 表示精确匹配（不带%），'fuzzy' 表示模糊匹配（带%），'unmatched' 表示未匹配
    """
    # 移除空格
    original_reason = re.sub(r'\s+', '', original_reason)

    # 如果是未分类，返回unmatched
    if normalized_reason.startswith("未分类_"):
        return 'unmatched'

    # 找到normalized_reason对应的同义词列表
    if normalized_reason not in synonym_groups:
        return 'unmatched'

    synonyms = synonym_groups[normalized_reason]

    # 检查是否有精确匹配（不带%的同义词）
    for synonym in synonyms:
        if '%' not in synonym and synonym in original_reason:
            # 精确匹配
            return 'exact'

    # 检查是否有模糊匹配（带%的同义词）
    for synonym in synonyms:
        if '%' in synonym:
            pattern = synonym.replace('%', '(.*)')
            regex = re.compile(f"^{pattern}$")
            if regex.search(original_reason):
                # 模糊匹配
                return 'fuzzy'

    return 'unmatched'


def extract_reasons_with_match_type(reason_text):
    """
    从原因文本中提取所有原因，并返回匹配类型信息
    
    Args:
        reason_text: 原因文本，如"英伟达液冷+人形机器人"
    
    Returns:
        list: [(normalized_reason, match_type, original_reason), ...]
              match_type为'exact'或'fuzzy'
    """
    if not reason_text or isinstance(reason_text, float):
        return []

    # 以"+"分割不同原因
    reasons = reason_text.split('+')
    result = []

    for r in reasons:
        r = r.strip()
        if not r:
            continue

        # 规范化原因
        normalized = normalize_reason(r)

        # 获取匹配类型
        match_type = get_reason_match_type(r, normalized)

        result.append((normalized, match_type, r))

    return result


def extract_reasons_with_original(reason_text):
    """
    与 extract_reasons 相同，但保留原始短语与归并后的主类。
    Returns: list of tuples (original_reason, grouped_reason)
    """
    if not reason_text or isinstance(reason_text, float):
        return []

    parts = [p.strip() for p in reason_text.split('+') if p and p.strip()]
    result = []
    for p in parts:
        grouped = normalize_reason(p)
        result.append((p, grouped))
    return result


def _get_group_meta(group_name):
    """读取分组的 tier/role 元数据，若无配置返回默认。"""
    tier = GROUP_TIER.get(group_name)
    role = GROUP_ROLE.get(group_name)
    # 默认：未知组视为 industry+tier2
    if tier is None:
        tier = 2
    if role is None:
        role = "industry"
    return tier, role


def _has_core_keyword(group_name, original_reason):
    kws = CORE_KEYWORDS.get(group_name, [])
    return any(k in original_reason for k in kws)


def _top_reason_rank(reason, top_reasons):
    if not top_reasons or reason not in top_reasons:
        return None
    # top_reasons[0] 最热门，我们希望越热门加分越多
    idx = top_reasons.index(reason)
    # 映射为分值：前排更大
    return max(0, len(top_reasons) - idx)


def score_reason(original_reason, grouped_reason, group_count_in_stock=1, top_reasons=None):
    """
    计算单个原因的正统分。
    参数:
      - original_reason: 原始短语（如 "AI服务器电源"）
      - grouped_reason: 归并后的主类（如 "算力半导体"）
      - group_count_in_stock: 该股票内该主类出现次数
      - top_reasons: 全局热门主类顺序
    返回:
      - float 分数
    """
    tier, role = _get_group_meta(grouped_reason)
    base = TIER_WEIGHT.get(tier, 0.0) + ROLE_WEIGHT.get(role, 0.0)

    bonus_core = CORE_KEYWORD_BONUS if _has_core_keyword(grouped_reason, original_reason) else 0.0
    bonus_freq = FREQ_WEIGHT * max(0, group_count_in_stock)

    hot_rank = _top_reason_rank(grouped_reason, top_reasons)
    bonus_hot = HOT_WEIGHT * hot_rank if hot_rank else 0.0

    # overrides（按原始短语或分组名进行匹配）
    override_bonus = 0.0
    if REASON_OVERRIDES:
        # 精确匹配原始短语或包含匹配
        for key, cfg in REASON_OVERRIDES.items():
            if key in original_reason or key == grouped_reason:
                override_bonus += float(cfg.get("bonus", 0.0))

    return base + bonus_core + bonus_freq + bonus_hot + override_bonus


def get_stock_reason_labels(all_stocks, top_reasons, k=2):
    """
    为每只股票计算主+次标签（默认最多2个标签）。
    返回: { stock_key: {"primary": str, "secondaries": [str], "details": [(original, grouped, score)]} }
    说明: details 为“按主类去重后”的明细（每个分组只保留最高分的一条），便于用于批注展示，避免重复。
    """
    labels = {}
    if not CONCEPT_TIERING_AVAILABLE:
        return labels

    for stock_key, data in all_stocks.items():
        reasons = data.get("reason_details") or []  # [(original, grouped)]
        if not reasons:
            # 若上游未提供细节，则尝试从简化列表构造
            grouped_only = data.get("reasons", [])
            reasons = [(g, g) for g in grouped_only]

        if not reasons:
            continue

        # 统计组内频次
        group_counts = Counter([g for _, g in reasons])

        # 逐条打分
        scored = []
        for original, grouped in reasons:
            cnt = group_counts.get(grouped, 1)
            s = score_reason(original, grouped, cnt, top_reasons)
            scored.append((original, grouped, s))

        # 汇总到主类维度，取每组最高分
        group_best = {}
        for original, grouped, s in scored:
            if grouped not in group_best or s > group_best[grouped][1]:
                group_best[grouped] = (original, s)

        # 排序（分数降序，tier升序，热度优先）
        def sort_key(item):
            grouped, (original, s) = item
            tier, _role = _get_group_meta(grouped)
            hot_rank = _top_reason_rank(grouped, top_reasons) or 0
            return (-s, tier, -hot_rank, -group_counts.get(grouped, 0))

        sorted_groups = sorted(group_best.items(), key=sort_key)
        if not sorted_groups:
            continue

        primary = sorted_groups[0][0]
        secondaries = [g for g, _ in sorted_groups[1:k]] if k > 1 else []

        # 构造“按组去重”的明细，顺序与排序一致
        unique_details = [(group_best[g][0], g, group_best[g][1]) for g, _ in sorted_groups]

        labels[stock_key] = {
            "primary": primary,
            "secondaries": secondaries,
            "details": unique_details,
        }

    return labels


def extract_reasons(reason_text):
    """
    从原因文本中提取所有原因
    """
    if not reason_text or isinstance(reason_text, float):  # 处理pd.isna()的情况
        return []

    # 以"+"分割不同原因
    reasons = reason_text.split('+')
    return [normalize_reason(r.strip()) for r in reasons if r.strip()]


def get_reason_colors(all_reasons, top_n=TOP_N, priority_reasons=None):
    """
    根据原因出现频率，为热门原因分配颜色

    Args:
        all_reasons: 所有原因的列表
        top_n: 选取的热门原因数量
        priority_reasons: 优先选择的原因列表，默认为None时使用全局PRIORITY_REASONS

    Returns:
        tuple: (reason_colors, top_reasons) - 原因到颜色的映射字典和热门原因列表
    """
    # 使用传入的优先原因列表或全局定义的列表
    priority_list = priority_reasons if priority_reasons is not None else PRIORITY_REASONS

    # 统计所有原因出现次数
    reason_counter = Counter(all_reasons)

    # 过滤掉未分类的原因和排除列表中的原因
    classified_reasons = [reason for reason in reason_counter.keys()
                          if not reason.startswith('未分类_') and reason not in EXCLUDED_REASONS]

    # 确保所有原因的出现次数都被正确计算
    # 首先创建包含所有可能原因的计数字典
    all_reason_counts = {reason: 0 for reason in synonym_groups.keys()}

    # 然后用实际统计的次数更新
    for reason, count in reason_counter.items():
        if not reason.startswith('未分类_'):
            all_reason_counts[reason] = count

    # 首先添加优先列表中的原因（如果它们在数据中出现过）
    top_reasons = []
    for reason in priority_list:
        if reason in all_reason_counts and all_reason_counts[reason] > 0 and reason not in EXCLUDED_REASONS:
            top_reasons.append(reason)

    # 然后按出现次数倒序添加其他热门原因，直到达到TOP_N个
    remaining_slots = top_n - len(top_reasons)
    if remaining_slots > 0:
        # 排除已经在优先列表中的原因
        other_reasons = [reason for reason, count in sorted(all_reason_counts.items(),
                                                            key=lambda x: x[1],
                                                            reverse=True)
                         if count > 0 and reason not in EXCLUDED_REASONS and reason not in top_reasons][
                        :remaining_slots]
        top_reasons.extend(other_reasons)

    # 为每个原因分配颜色
    reason_colors = {reason: COLORS[i % len(COLORS)] for i, reason in enumerate(top_reasons)}

    return reason_colors, top_reasons


def get_stock_reason_group(all_stocks, top_reasons):
    """
    确定每支股票主要属于哪个原因组
    
    优先级规则：
    1. 匹配类型：精确匹配 > 模糊匹配（最高优先级）
    2. 出现次数：在该股票内出现次数多的优先
    3. 热门度：全市场热门度高的优先（top_reasons排名）

    Args:
        all_stocks: 股票信息字典，包含每只股票的原因列表
                   如果包含'reason_details'字段，则为[(normalized, match_type, original), ...]格式
                   否则使用'reasons'字段（向后兼容）
        top_reasons: 热门原因列表

    Returns:
        dict: 股票到主要原因的映射字典
    """
    stock_reason_group = {}

    for stock_key, data in all_stocks.items():
        # 尝试获取详细的原因信息（包含匹配类型）
        reason_details = data.get('reason_details')

        if reason_details:
            # 新格式：包含匹配类型信息
            # reason_details: [(normalized_reason, match_type, original_reason), ...]

            # 构建评分列表：(normalized_reason, match_type, count, top_rank)
            reason_info = {}
            for normalized, match_type, original in reason_details:
                if normalized not in reason_info:
                    reason_info[normalized] = {
                        'match_type': match_type,
                        'count': 0,
                        'original_reasons': []
                    }
                reason_info[normalized]['count'] += 1
                reason_info[normalized]['original_reasons'].append(original)

                # 如果有更好的匹配类型，更新
                if match_type == 'exact' and reason_info[normalized]['match_type'] != 'exact':
                    reason_info[normalized]['match_type'] = 'exact'

            # 评分和排序
            scored_reasons = []
            for normalized, info in reason_info.items():
                # 计算热门度排名（数字越小越靠前）
                top_rank = top_reasons.index(normalized) if normalized in top_reasons else 9999

                # 匹配类型分数：exact=2, fuzzy=1, unmatched=0
                match_score = 2 if info['match_type'] == 'exact' else (1 if info['match_type'] == 'fuzzy' else 0)

                scored_reasons.append({
                    'reason': normalized,
                    'match_score': match_score,
                    'count': info['count'],
                    'top_rank': top_rank,
                    'match_type': info['match_type']
                })

            if scored_reasons:
                # 排序：匹配类型(降序) > 出现次数(降序) > 热门度排名(升序)
                scored_reasons.sort(key=lambda x: (-x['match_score'], -x['count'], x['top_rank']))
                stock_reason_group[stock_key] = scored_reasons[0]['reason']

        else:
            # 旧格式：只有规范化后的原因列表（向后兼容）
            reasons = data.get('reasons', [])
            if not reasons:
                continue

            # 统计该股票的原因
            stock_reason_counter = Counter(reasons)

            # 先检查哪些原因是热门原因
            top_reasons_found = [reason for reason in top_reasons if reason in stock_reason_counter]

            if top_reasons_found:
                # 如果有多个热门原因，选择出现次数最多的
                top_reason_counts = [(reason, stock_reason_counter[reason]) for reason in top_reasons_found]
                top_reason_counts.sort(key=lambda x: x[1], reverse=True)
                stock_reason_group[stock_key] = top_reason_counts[0][0]
            elif stock_reason_counter:
                # 如果没有热门原因，使用该股票最常见的原因
                most_common_reason = stock_reason_counter.most_common(1)[0][0]
                stock_reason_group[stock_key] = most_common_reason

    return stock_reason_group


def get_color_by_pct_change(pct_change):
    """
    根据涨跌幅返回对应的颜色代码

    :param pct_change: 涨跌幅百分比
    :return: 颜色代码
    """
    if pd.isna(pct_change):
        return "FFFFFF"  # 白色

    # 初始化颜色索引为0 (最浅)
    color_idx = 0
    abs_change = abs(pct_change)

    # 根据涨跌幅绝对值确定颜色深浅
    for i, threshold in enumerate(PCT_CHANGE_THRESHOLDS):
        if abs_change >= threshold:
            color_idx = i + 1

    # 如果超过最后一个阈值，使用最深的颜色
    color_idx = min(color_idx, len(RED_COLORS) - 1)

    # 根据正负选择红色或绿色
    if pct_change >= 0:
        return RED_COLORS[color_idx]
    else:
        return GREEN_COLORS[color_idx]


# 涨跌幅颜色映射函数
def get_color_for_pct_change(pct_change):
    """
    根据涨跌幅返回颜色代码，区别连板的红
    小于7%的涨跌幅不着色，7%以上才进行梯度着色

    Args:
        pct_change: 涨跌幅百分比

    Returns:
        str: 16进制颜色代码，或None表示不着色
    """
    if pct_change is None:
        return None  # 数据缺失不着色

    # 将涨跌幅转换为浮点数
    try:
        pct = float(pct_change)
    except:
        return None  # 无法转换时不着色

    # 涨跌停板情况
    if pct >= 9.5:
        return "FFD700"  # 涨停 - 金黄色
    if pct <= -9.5:
        return "00CC00"  # 跌停 - 深绿色

    # 绝对值小于7%不着色
    if abs(pct) < 7:
        return None

    # 普通涨跌幅，仅对7%以上的进行着色
    if pct >= 7:
        # 上涨 - 黄色系
        # 将7%-10%映射到颜色梯度
        intensity = min(255, int(200 * (pct - 7) / 3) + 55)  # 根据涨幅计算黄色强度
        return f"FFFF{hex(255 - intensity)[2:].zfill(2).upper()}"  # 黄色系，涨幅越大越深
    elif pct <= -7:
        # 下跌 - 绿色系
        # 将-7%至-10%映射到颜色梯度
        intensity = min(255, int(200 * (abs(pct) - 7) / 3) + 55)  # 根据跌幅计算绿色强度
        green_value = 255 - intensity + 55  # 保证最小值
        green = hex(max(0, min(255, green_value)))[2:].zfill(2).upper()
        return f"{green}FF{green}"
    else:
        # 平盘或小幅波动 - 不着色
        return None


def create_legend_sheet(wb, reason_counter, reason_colors, top_reasons, high_board_colors=None,
                        reentry_colors=None, source_sheet_name=None, concept_analysis_data=None):
    """
    创建颜色图例工作表

    Args:
        wb: openpyxl工作簿对象
        reason_counter: 原因计数器(Counter对象)
        reason_colors: 原因到颜色的映射字典
        top_reasons: 热门原因列表
        high_board_colors: 高板数颜色映射字典，默认为None
        reentry_colors: 重复入选颜色映射字典，默认为None
        source_sheet_name: 源数据工作表名称，用于生成图例工作表名称，默认为None
        concept_analysis_data: 概念分析数据，格式为(top_concepts, new_concepts)，默认为None

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: 创建的图例工作表对象
    """
    # 确定图例工作表名称
    if source_sheet_name:
        legend_sheet_name = f"图例_{source_sheet_name}"
    else:
        legend_sheet_name = "题材颜色图例"

    # 检查图例工作表是否已存在，如果存在则删除
    if legend_sheet_name in wb.sheetnames:
        wb.remove(wb[legend_sheet_name])

    # 创建图例工作表
    legend_ws = wb.create_sheet(title=legend_sheet_name)

    # 设置标题
    title_cell = legend_ws.cell(row=1, column=1, value="热门概念图例")
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    legend_ws.column_dimensions['A'].width = 25
    legend_ws.column_dimensions['B'].width = 15

    # 添加各个热门原因图例（带次数）
    current_row = 2

    # 添加表头
    legend_ws.cell(row=current_row, column=1, value="原因").font = Font(bold=True)
    legend_ws.cell(row=current_row, column=2, value="出现次数").font = Font(bold=True)
    current_row += 1

    # 添加热门原因（带颜色）
    for i, reason in enumerate(top_reasons, start=0):
        count = reason_counter.get(reason, 0)

        # 添加原因名称
        name_cell = legend_ws.cell(row=current_row, column=1, value=reason)
        name_cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")

        # 如果背景色较深，使用白色字体
        if reason_colors[reason] in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
            name_cell.font = Font(color="FFFFFF")

        # 添加出现次数
        count_cell = legend_ws.cell(row=current_row, column=2, value=count)

        current_row += 1

    # 添加分隔行
    separator_cell = legend_ws.cell(row=current_row, column=1, value="其他概念")
    separator_cell.border = Border(top=Side(style='thin', color='000000'))
    separator_cell.font = Font(bold=True)
    current_row += 1

    # 添加未入选着色的原因（按出现次数倒序）
    non_top_reasons = [r for r in synonym_groups.keys()
                       if r not in top_reasons and r not in EXCLUDED_REASONS]
    sorted_non_top = sorted(non_top_reasons, key=lambda x: reason_counter.get(x, 0), reverse=True)

    for reason in sorted_non_top:
        count = reason_counter.get(reason, 0)
        if count > 0:  # 只显示出现过的概念
            legend_ws.cell(row=current_row, column=1, value=reason)
            legend_ws.cell(row=current_row, column=2, value=count)
            current_row += 1

    # 添加排除概念分隔行
    if EXCLUDED_REASONS:
        separator_cell = legend_ws.cell(row=current_row, column=1, value="排除的概念")
        separator_cell.border = Border(top=Side(style='thin', color='000000'))
        separator_cell.font = Font(bold=True)
        current_row += 1

        # 添加EXCLUDED_REASONS原因
        for reason in EXCLUDED_REASONS:
            count = reason_counter.get(reason, 0)
            if count > 0:  # 只显示出现过的概念
                legend_ws.cell(row=current_row, column=1, value=reason)
                legend_ws.cell(row=current_row, column=2, value=count)
                current_row += 1

    # 添加多次上榜图例
    multi_cell = legend_ws.cell(row=current_row, column=1, value="多次上榜")
    multi_cell.fill = PatternFill(start_color=MULTI_COLOR, fill_type="solid")
    current_row += 1

    # 添加板数颜色图例（如果提供）
    if high_board_colors:
        # 添加分隔行
        separator_cell = legend_ws.cell(row=current_row, column=1, value="高板数颜色")
        separator_cell.border = Border(top=Side(style='thin', color='000000'))
        separator_cell.font = Font(bold=True)
        current_row += 1

        # 添加高板数颜色图例
        for board_level, color in sorted(high_board_colors.items()):
            # 如果是最后一个，显示为"X板及以上"
            if board_level == max(high_board_colors.keys()):
                label = f"{board_level}板及以上"
            else:
                label = f"{board_level}板"

            name_cell = legend_ws.cell(row=current_row, column=1, value=label)
            name_cell.fill = PatternFill(start_color=color, fill_type="solid")

            # 对于深色背景，使用白色字体
            if board_level >= 12:
                name_cell.font = Font(color="FFFFFF")

            current_row += 1

    # 添加重复入选颜色图例（如果提供）
    if reentry_colors:
        # 添加分隔行
        separator_cell = legend_ws.cell(row=current_row, column=1, value="重复入选颜色")
        separator_cell.border = Border(top=Side(style='thin', color='000000'))
        separator_cell.font = Font(bold=True)
        current_row += 1

        # 添加重复入选颜色图例
        for entry_count, color in sorted(reentry_colors.items()):
            # 如果是最后一个，显示为"X次及以上"
            if entry_count == max(reentry_colors.keys()):
                label = f"{entry_count}次及以上入选"
            else:
                label = f"{entry_count}次入选"

            name_cell = legend_ws.cell(row=current_row, column=1, value=label)
            name_cell.fill = PatternFill(start_color=color, fill_type="solid")

            # 对于深色背景，使用白色字体
            if entry_count >= 4:
                name_cell.font = Font(color="FFFFFF")

            # 添加边框
            name_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 添加第二列的边框
            legend_ws.cell(row=current_row, column=2).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            current_row += 1

    # 设置所有单元格的边框
    for row in legend_ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # 添加概念分析数据（如果提供）
    if concept_analysis_data:
        add_concept_analysis_to_legend_sheet(legend_ws, concept_analysis_data)

    return legend_ws


def add_concept_analysis_to_legend_sheet(legend_ws, concept_analysis_data):
    """
    在图例工作表中添加原因分析数据（并列显示）

    Args:
        legend_ws: 图例工作表对象
        concept_analysis_data: 原因分析数据，格式为(reason_stats, new_reasons)
    """
    if not concept_analysis_data:
        return

    reason_stats, new_reasons = concept_analysis_data

    if not reason_stats:
        return

    from analysis.concept_analyzer import TOP_CONCEPTS_COUNT, NEW_CONCEPT_DAYS, NEW_REASON_MIN_COUNT

    # 设置起始位置
    start_col = 4  # D列开始

    # 设置列宽 - 并列显示
    legend_ws.column_dimensions['D'].width = 20  # 热门原因名称
    legend_ws.column_dimensions['E'].width = 15  # 热门原因首次出现
    legend_ws.column_dimensions['F'].width = 12  # 热门原因出现次数
    legend_ws.column_dimensions['G'].width = 25  # 新原因名称（调宽）
    legend_ws.column_dimensions['H'].width = 15  # 新原因首次出现
    legend_ws.column_dimensions['I'].width = 12  # 新原因出现次数

    current_row = 1

    # 并列显示标题
    # 热门原因标题
    hot_title_cell = legend_ws.cell(row=current_row, column=start_col, value=f"热门原因统计 Top {TOP_CONCEPTS_COUNT}")
    hot_title_cell.font = Font(bold=True, size=12)
    hot_title_cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")  # 浅蓝色背景
    hot_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # 合并热门原因标题单元格
    legend_ws.merge_cells(start_row=current_row, start_column=start_col,
                          end_row=current_row, end_column=start_col + 2)

    # 新原因标题（如果有新原因数据）
    if new_reasons:
        new_title_cell = legend_ws.cell(row=current_row, column=start_col + 3,
                                        value=f"新原因统计 (最近{NEW_CONCEPT_DAYS}天，≥{NEW_REASON_MIN_COUNT}次)")
        new_title_cell.font = Font(bold=True, size=12)
        new_title_cell.fill = PatternFill(start_color="FFE6F0", fill_type="solid")  # 浅粉色背景
        new_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # 合并新原因标题单元格
        legend_ws.merge_cells(start_row=current_row, start_column=start_col + 3,
                              end_row=current_row, end_column=start_col + 5)
    else:
        # 如果没有新原因，显示无新原因
        no_new_title_cell = legend_ws.cell(row=current_row, column=start_col + 3,
                                           value=f"无新原因 (最近{NEW_CONCEPT_DAYS}天)")
        no_new_title_cell.font = Font(bold=True, size=12, italic=True)
        no_new_title_cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")  # 灰色背景
        no_new_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # 合并无新原因标题单元格
        legend_ws.merge_cells(start_row=current_row, start_column=start_col + 3,
                              end_row=current_row, end_column=start_col + 5)

    current_row += 1

    # 添加表头
    # 热门原因表头
    hot_headers = ["原因", "首次出现", "出现次数"]
    for i, header in enumerate(hot_headers):
        header_cell = legend_ws.cell(row=current_row, column=start_col + i, value=header)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="F0F8FF", fill_type="solid")  # 更浅的蓝色
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 新原因表头（如果有新原因数据）
    if new_reasons:
        new_headers = ["原因", "首次出现", "出现次数"]
        for i, header in enumerate(new_headers):
            header_cell = legend_ws.cell(row=current_row, column=start_col + 3 + i, value=header)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="FFF0F5", fill_type="solid")  # 更浅的粉色
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    current_row += 1

    # 准备数据
    # reason_stats 已经是排序后的列表 [(reason, stats), ...]
    if isinstance(reason_stats, list):
        sorted_reasons = reason_stats[:TOP_CONCEPTS_COUNT]
    else:
        sorted_reasons = sorted(reason_stats.items(), key=lambda x: x[1]['count'], reverse=True)[:TOP_CONCEPTS_COUNT]

    # 新原因排序：出现次数（倒序），首次出现（倒序）
    # 注意：first_date是字符串，需要用reverse=True来实现倒序
    sorted_new_reasons = sorted(new_reasons.items(), key=lambda x: (x[1]['count'], x[1]['first_date']),
                                reverse=True) if new_reasons else []

    # 并列填充数据
    max_rows = max(len(sorted_reasons), len(sorted_new_reasons))

    for i in range(max_rows):
        # 填充热门原因数据
        if i < len(sorted_reasons):
            reason, stats = sorted_reasons[i]

            # 原因名称
            reason_cell = legend_ws.cell(row=current_row, column=start_col, value=reason)
            reason_cell.alignment = Alignment(horizontal="left", vertical="center")
            reason_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为前5名添加不同的背景色
            if i < 5:
                colors = ["FFE6E6", "FFF0E6", "FFFBE6", "E6FFE6", "E6F0FF"]  # 红、橙、黄、绿、蓝的浅色版本
                reason_cell.fill = PatternFill(start_color=colors[i], fill_type="solid")

            # 首次出现日期
            first_date = stats.get('historical_first_date', stats.get('first_date', ''))
            first_date_cell = legend_ws.cell(row=current_row, column=start_col + 1, value=first_date)
            first_date_cell.alignment = Alignment(horizontal="center", vertical="center")
            first_date_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 出现次数
            count_cell = legend_ws.cell(row=current_row, column=start_col + 2, value=stats['count'])
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 填充新原因数据
        if new_reasons and i < len(sorted_new_reasons):
            reason, stats = sorted_new_reasons[i]

            # 原因名称
            reason_cell = legend_ws.cell(row=current_row, column=start_col + 3, value=reason)
            reason_cell.alignment = Alignment(horizontal="left", vertical="center")
            reason_cell.fill = PatternFill(start_color="E6FFE6", fill_type="solid")  # 浅绿色背景
            reason_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 首次出现日期
            first_date_cell = legend_ws.cell(row=current_row, column=start_col + 4, value=stats['first_date'])
            first_date_cell.alignment = Alignment(horizontal="center", vertical="center")
            first_date_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 出现次数
            count_cell = legend_ws.cell(row=current_row, column=start_col + 5, value=stats['count'])
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        current_row += 1


def load_index_data(index_file="./data/indexes/sz399006_创业板指.csv"):
    """
    加载创业板指数数据

    Args:
        index_file: 指数数据文件路径，默认为创业板指数

    Returns:
        dict: 以日期为键的字典，包含涨跌幅和成交量信息
    """
    try:
        # 检查指数数据文件是否存在
        if not os.path.exists(index_file):
            print(f"创业板指数文件不存在: {index_file}")
            return {}

        # 读取CSV文件 (无表头)
        df = pd.read_csv(index_file, header=None,
                         names=['date', 'open', 'high', 'low', 'close', 'volume'])

        # 计算每日涨跌幅
        df['pct_change'] = df['close'].pct_change() * 100

        # 将成交量转换为亿元 (原始数据单位为成交量，而非手数)
        # 注意：volume可能已经是以元为单位，需根据实际数据调整
        df['volume_100m'] = df['volume'] / 100000000  # 转换为亿元

        # 以日期为键创建字典
        index_data = {}
        for _, row in df.iterrows():
            # 确保日期格式与date_columns中的键完全一致
            date_obj = pd.to_datetime(row['date'])
            date_str = date_obj.strftime('%Y年%m月%d日')
            index_data[date_str] = {
                'pct_change': row['pct_change'],
                'volume_100m': row['volume_100m']
            }

        return index_data
    except Exception as e:
        print(f"读取创业板指数数据失败: {e}")
        return {}


def load_all_index_data(index_dir="./data/indexes"):
    """
    加载所有指数数据

    Args:
        index_dir: 指数数据文件目录

    Returns:
        dict: 以指数名称为键的字典，每个值包含该指数的日期数据
    """
    all_index_data = {}

    try:
        # 检查目录是否存在
        if not os.path.exists(index_dir):
            print(f"指数数据目录不存在: {index_dir}")
            return {}

        # 遍历目录中的所有CSV文件
        for filename in os.listdir(index_dir):
            if filename.endswith('.csv'):
                # 从文件名提取指数名称
                # 文件名格式: sh000001_上证指数.csv
                index_name = filename.replace('.csv', '').split('_', 1)
                if len(index_name) >= 2:
                    index_display_name = index_name[1]  # 使用中文名称
                else:
                    index_display_name = filename.replace('.csv', '')

                # 加载单个指数数据
                index_file = os.path.join(index_dir, filename)
                index_data = load_single_index_data(index_file)

                if index_data:
                    all_index_data[index_display_name] = index_data
                    print(f"成功加载指数数据: {index_display_name}")
                else:
                    print(f"加载指数数据失败: {filename}")

        print(f"总共加载了 {len(all_index_data)} 个指数的数据")
        return all_index_data

    except Exception as e:
        print(f"加载所有指数数据失败: {e}")
        return {}


def load_single_index_data(index_file):
    """
    加载单个指数数据文件

    Args:
        index_file: 指数数据文件路径

    Returns:
        dict: 以日期为键的字典，包含涨跌幅和成交量信息
    """
    try:
        # 读取CSV文件 (无表头)
        df = pd.read_csv(index_file, header=None,
                         names=['date', 'open', 'high', 'low', 'close', 'volume'])

        # 计算每日涨跌幅
        df['pct_change'] = df['close'].pct_change() * 100

        # 将成交量转换为亿元
        df['volume_100m'] = df['volume'] / 100000000  # 转换为亿元

        # 以日期为键创建字典
        index_data = {}
        for _, row in df.iterrows():
            # 确保日期格式与date_columns中的键完全一致
            date_obj = pd.to_datetime(row['date'])
            date_str = date_obj.strftime('%Y年%m月%d日')
            index_data[date_str] = {
                'pct_change': row['pct_change'],
                'volume_100m': row['volume_100m']
            }

        return index_data
    except Exception as e:
        print(f"读取指数数据失败 {index_file}: {e}")
        return {}


def add_market_indicators(ws, date_columns, index_data=None, index_file="./data/indexes/sz399006_创业板指.csv",
                          label_col=1):
    """
    在Excel表格中添加大盘指标行（创业指和成交量）

    Args:
        ws: openpyxl工作表对象
        date_columns: 日期列映射字典，键为日期字符串，值为列索引
        index_data: 已加载的指数数据，如果为None则会尝试加载
        index_file: 指数数据文件路径，默认为创业板指数
        label_col: 标签列索引，默认为1

    Returns:
        bool: 是否成功添加指标
    """
    # 如果没有提供指数数据，尝试加载
    if index_data is None:
        index_data = load_index_data(index_file)
        if not index_data:
            print("未能加载创业板指数数据，将不显示大盘指标")
            return False

    # 在第一列添加涨跌幅和成交量的标签，并设置样式
    label_cell_1 = ws.cell(row=2, column=label_col, value="创业指")
    label_cell_1.alignment = Alignment(horizontal="center", vertical="center")
    label_cell_1.font = Font(bold=True, size=9)  # 设置小一号字体
    label_cell_1.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

    label_cell_2 = ws.cell(row=3, column=label_col, value="成交量(亿)")
    label_cell_2.alignment = Alignment(horizontal="center", vertical="center")
    label_cell_2.font = Font(bold=True, size=9)  # 设置小一号字体
    label_cell_2.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

    # 为每个日期列添加指数数据
    for date_str, col_idx in date_columns.items():
        if date_str in index_data:
            # 涨跌幅数据 (第二行)
            pct_change = index_data[date_str]['pct_change']
            pct_cell = ws.cell(row=2, column=col_idx, value=f"{pct_change:.2f}%")
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.fill = PatternFill(start_color=get_color_by_pct_change(pct_change), fill_type="solid")
            pct_cell.font = Font(size=9)  # 设置小一号字体

            # 成交量数据 (第三行)
            volume = index_data[date_str]['volume_100m']
            volume_cell = ws.cell(row=3, column=col_idx, value=f"{volume:.2f}")
            volume_cell.alignment = Alignment(horizontal="center", vertical="center")
            volume_cell.font = Font(size=9)  # 设置小一号字体
        else:
            # 如果没有指数数据，添加空单元格
            ws.cell(row=2, column=col_idx, value="--").font = Font(size=9)
            ws.cell(row=3, column=col_idx, value="--").font = Font(size=9)

    # 添加指数行和个股行之间的分隔线
    for col_idx in range(1, max(date_columns.values()) + 2):
        cell = ws.cell(row=3, column=col_idx)
        cell.border = Border(
            bottom=Side(style='double', color='000000')
        )

    return True


def format_date_header_for_index(date_str):
    """
    为指数工作表格式化日期表头，与其他sheet保持一致

    Args:
        date_str: 日期字符串（格式：YYYY年MM月DD日）

    Returns:
        str: 格式化后的日期表头（格式：YYYY-MM-DD\n星期X）
    """
    try:
        if '年' in date_str:
            date_obj = pd.to_datetime(date_str, format='%Y年%m月%d日')
        else:
            date_obj = pd.to_datetime(date_str)

        # 获取星期几的中文名称
        weekday_map = {0: "星期一", 1: "星期二", 2: "星期三", 3: "星期四", 4: "星期五", 5: "星期六", 6: "星期日"}
        weekday = weekday_map[date_obj.weekday()]

        # 格式化为：YYYY-MM-DD\n星期X
        return f"{date_obj.strftime('%Y-%m-%d')}\n{weekday}"
    except:
        return date_str


def create_index_sheet_optimized(wb, date_columns, all_index_data=None, index_dir="./data/indexes",
                                 sheet_name="指数数据"):
    """
    创建指数专用的工作表（优化版本，支持增量更新）

    Args:
        wb: openpyxl工作簿对象
        date_columns: 日期列映射字典，键为日期字符串，值为列索引
        all_index_data: 已加载的所有指数数据，如果为None则会尝试加载
        index_dir: 指数数据文件目录
        sheet_name: 工作表名称

    Returns:
        bool: 是否成功创建指数工作表
    """
    try:
        # 如果没有提供指数数据，尝试加载
        if all_index_data is None:
            all_index_data = load_all_index_data(index_dir)
            if not all_index_data:
                print("未能加载指数数据，将不创建指数工作表")
                return False

        # 检查指数工作表是否已存在
        index_ws = None
        if sheet_name in wb.sheetnames:
            index_ws = wb[sheet_name]
            print(f"指数工作表 {sheet_name} 已存在，将进行增量更新")
        else:
            # 创建新的指数工作表
            index_ws = wb.create_sheet(title=sheet_name)
            print(f"创建新的指数工作表: {sheet_name}")

        # 创建日期列映射，从第2列开始（第1列是指数名称）
        compact_date_columns = {}
        col_idx = 2
        for date_str in sorted(date_columns.keys()):
            compact_date_columns[date_str] = col_idx
            col_idx += 1

        # 设置表头
        # 第一行：指数名称列
        header_cell = index_ws.cell(row=1, column=1, value="指数名称")
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加日期列表头（与其他sheet格式一致）
        for date_str, col_idx in compact_date_columns.items():
            formatted_date = format_date_header_for_index(date_str)

            header_cell = index_ws.cell(row=1, column=col_idx, value=formatted_date)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 设置列宽
        index_ws.column_dimensions['A'].width = 15  # 指数名称列

        # 按照配置的顺序排列指数
        ordered_indexes = []
        for index_name in INDEX_ORDER:
            if index_name in all_index_data:
                ordered_indexes.append(index_name)

        # 添加不在配置中的指数（按字母顺序）
        remaining_indexes = sorted([name for name in all_index_data.keys() if name not in INDEX_ORDER])
        ordered_indexes.extend(remaining_indexes)

        # 为每个指数创建两行：涨跌幅行和成交量行，指数之间空一行
        current_row = 2
        for i, index_name in enumerate(ordered_indexes):
            index_data = all_index_data[index_name]

            # 如果不是第一个指数，先空一行
            if i > 0:
                current_row += 1

            # 涨跌幅行
            pct_change_cell = index_ws.cell(row=current_row, column=1, value=f"{index_name}")
            pct_change_cell.alignment = Alignment(horizontal="left", vertical="center")
            pct_change_cell.font = Font(bold=True, size=9)
            pct_change_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

            # 成交量行
            volume_cell = index_ws.cell(row=current_row + 1, column=1, value=f"{index_name}(亿)")
            volume_cell.alignment = Alignment(horizontal="left", vertical="center")
            volume_cell.font = Font(bold=True, size=9)
            volume_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

            # 为每个日期列添加指数数据
            for date_str, col_idx in compact_date_columns.items():
                if date_str in index_data:
                    # 涨跌幅数据
                    pct_change = index_data[date_str]['pct_change']
                    if pd.notna(pct_change):
                        pct_cell = index_ws.cell(row=current_row, column=col_idx, value=f"{pct_change:.2f}%")
                        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
                        pct_cell.fill = PatternFill(start_color=get_color_by_pct_change(pct_change), fill_type="solid")
                        pct_cell.font = Font(size=9)
                    else:
                        index_ws.cell(row=current_row, column=col_idx, value="--").font = Font(size=9)

                    # 成交量数据
                    volume = index_data[date_str]['volume_100m']
                    if pd.notna(volume):
                        volume_cell = index_ws.cell(row=current_row + 1, column=col_idx, value=f"{volume:.2f}")
                        volume_cell.alignment = Alignment(horizontal="center", vertical="center")
                        volume_cell.font = Font(size=9)
                    else:
                        index_ws.cell(row=current_row + 1, column=col_idx, value="--").font = Font(size=9)
                else:
                    # 如果没有指数数据，添加空单元格
                    index_ws.cell(row=current_row, column=col_idx, value="--").font = Font(size=9)
                    index_ws.cell(row=current_row + 1, column=col_idx, value="--").font = Font(size=9)

            current_row += 2  # 每个指数占用两行

        # 设置所有单元格的边框
        max_row = current_row - 1
        max_col = max(compact_date_columns.values()) if compact_date_columns else 1

        for row in index_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # 冻结第一列和第一行
        index_ws.freeze_panes = index_ws.cell(row=2, column=2)

        print(f"成功创建指数工作表: {sheet_name}，包含 {len(ordered_indexes)} 个指数")
        return True

    except Exception as e:
        print(f"创建指数工作表失败: {e}")
        return False


# 为了向后兼容，保留原函数名
def create_index_sheet(wb, date_columns, all_index_data=None, index_dir="./data/indexes", sheet_name="指数数据"):
    """
    创建指数专用的工作表（兼容性包装函数）
    """
    return create_index_sheet_optimized(wb, date_columns, all_index_data, index_dir, sheet_name)


def save_unique_reasons(all_reasons, output_file="./data/reasons/unique_reasons.json"):
    """
    将所有涨停原因去重并保存为JSON文件

    Args:
        all_reasons: 所有原因的列表
        output_file: 输出文件路径

    Returns:
        tuple: (success, message) - 是否成功保存和相关信息
    """
    try:
        # 确保目录存在
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # 统计原因出现次数
        reason_counter = Counter(all_reasons)

        # 分离已分类和未分类的原因
        classified_reasons = {}
        unclassified_reasons = {}

        for reason, count in reason_counter.items():
            if reason.startswith('未分类_'):
                # 移除"未分类_"前缀
                original_reason = reason.replace('未分类_', '')
                unclassified_reasons[original_reason] = count
            else:
                classified_reasons[reason] = count

        # 创建包含所有信息的字典
        reason_data = {
            "classified": {k: v for k, v in sorted(classified_reasons.items(), key=lambda item: item[1], reverse=True)},
            "unclassified": {k: v for k, v in
                             sorted(unclassified_reasons.items(), key=lambda item: item[1], reverse=True)},
            "total_classified": len(classified_reasons),
            "total_unclassified": len(unclassified_reasons),
            "total_reasons": len(reason_counter)
        }

        # 保存为JSON文件
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(reason_data, f, ensure_ascii=False, indent=2)

        return True, f"成功保存{len(reason_counter)}个涨停原因到 {output_file}"

    except Exception as e:
        return False, f"保存涨停原因失败: {str(e)}"
