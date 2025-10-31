"""
从涨停梯队Excel文件中提取股票代码并导出到txt文件

功能：
1. 从【涨停梯队】sheet提取所有股票代码
2. 从【涨停梯队_概念分组】sheet提取"默默上涨"分组的股票代码
3. 合并去重后写入到候选股票txt文件
"""

import os

from openpyxl import load_workbook


def extract_stock_codes_from_excel(excel_file, output_txt):
    """
    从涨停梯队Excel文件中提取股票代码并写入txt文件
    
    Args:
        excel_file: Excel文件路径
        output_txt: 输出txt文件路径
    """
    if not os.path.exists(excel_file):
        print(f"错误: Excel文件不存在: {excel_file}")
        return False

    try:
        print(f"开始从Excel文件提取股票代码: {excel_file}")
        wb = load_workbook(excel_file, data_only=True)

        all_stock_codes = set()  # 使用集合自动去重

        # 1. 从【涨停梯队】sheet提取所有股票代码
        extracted_from_main = extract_codes_from_main_sheet(wb, all_stock_codes)

        # 2. 从【涨停梯队_概念分组】sheet提取"默默上涨"分组的股票代码
        extracted_from_momo = extract_codes_from_momo_group(wb, all_stock_codes)

        if not all_stock_codes:
            print("警告: 未提取到任何股票代码")
            return False

        # 3. 写入txt文件
        write_codes_to_file(all_stock_codes, output_txt)

        # 计算重复数量
        total_extracted = extracted_from_main + extracted_from_momo
        duplicates = total_extracted - len(all_stock_codes)

        print(f"✓ 成功提取并写入 {len(all_stock_codes)} 只股票代码到: {output_txt}")
        print(f"  - 从【涨停梯队】提取: {extracted_from_main} 只")
        print(f"  - 从【默默上涨】提取: {extracted_from_momo} 只")
        if duplicates > 0:
            print(f"  - 去重: {duplicates} 只重复股票（重复入选或同时出现在多个分组）")
        print(f"  → 最终不重复股票数量: {len(all_stock_codes)} 只")

        return True

    except Exception as e:
        print(f"提取股票代码时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def extract_codes_from_main_sheet(wb, stock_codes_set):
    """
    从【涨停梯队】sheet提取股票代码
    
    Args:
        wb: Excel工作簿对象
        stock_codes_set: 股票代码集合（用于收集结果）
    
    Returns:
        int: 提取的股票数量
    """
    count = 0

    # 查找【涨停梯队】开头的sheet
    main_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("涨停梯队") and "概念分组" not in sheet_name and "成交量" not in sheet_name:
            main_sheet = wb[sheet_name]
            print(f"  找到主sheet: {sheet_name}")
            break

    if main_sheet is None:
        print("  警告: 未找到【涨停梯队】主sheet")
        return count

    # 从第4行开始读取（前3行是表头和指标行）
    for row in range(4, main_sheet.max_row + 1):
        stock_code = main_sheet.cell(row=row, column=1).value

        # 跳过空行
        if stock_code is None or str(stock_code).strip() == '':
            continue

        # 清理股票代码（去除空格、前导零等）
        clean_code = str(stock_code).strip()
        if clean_code and clean_code.isdigit():
            stock_codes_set.add(clean_code)
            count += 1

    return count


def extract_codes_from_momo_group(wb, stock_codes_set):
    """
    从【涨停梯队_概念分组】sheet的"默默上涨"分组提取股票代码
    
    Args:
        wb: Excel工作簿对象
        stock_codes_set: 股票代码集合（用于收集结果）
    
    Returns:
        int: 提取的股票数量
    """
    count = 0

    # 查找【涨停梯队_概念分组】sheet
    concept_sheet = None
    for sheet_name in wb.sheetnames:
        if "概念分组" in sheet_name and "成交量" not in sheet_name:
            concept_sheet = wb[sheet_name]
            print(f"  找到概念分组sheet: {sheet_name}")
            break

    if concept_sheet is None:
        print("  警告: 未找到【涨停梯队_概念分组】sheet")
        return count

    # 查找【默默上涨】分组
    in_momo_group = False

    for row in range(4, concept_sheet.max_row + 1):
        # 读取第一列的值（可能是分组标题或股票代码）
        cell_value = concept_sheet.cell(row=row, column=1).value

        if cell_value is None:
            continue

        cell_str = str(cell_value).strip()

        # 检查是否是分组标题行
        if cell_str.startswith("【") and cell_str.endswith("】"):
            # 判断是否进入或离开"默默上涨"分组
            if "默默上涨" in cell_str:
                in_momo_group = True
                print(f"  找到【默默上涨】分组，开始提取...")
            else:
                # 遇到其他分组标题，说明默默上涨分组已结束
                if in_momo_group:
                    print(f"  【默默上涨】分组结束")
                in_momo_group = False
            continue

        # 如果在"默默上涨"分组内，提取股票代码
        if in_momo_group and cell_str.isdigit():
            stock_codes_set.add(cell_str)
            count += 1

    return count


def write_codes_to_file(stock_codes_set, output_txt):
    """
    将股票代码写入txt文件
    
    Args:
        stock_codes_set: 股票代码集合
        output_txt: 输出文件路径
    """
    # 确保输出目录存在
    output_dir = os.path.dirname(output_txt)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"  创建输出目录: {output_dir}")

    # 排序后写入（方便查看）
    sorted_codes = sorted(stock_codes_set)

    with open(output_txt, 'w', encoding='utf-8') as f:
        for code in sorted_codes:
            f.write(code + '\n')


if __name__ == "__main__":
    # 测试用
    from analysis.loader.fupan_data_loader import OUTPUT_FILE

    excel_file = OUTPUT_FILE
    output_txt = "bin/candidate_temp/candidate_stocks.txt"

    extract_stock_codes_from_excel(excel_file, output_txt)
