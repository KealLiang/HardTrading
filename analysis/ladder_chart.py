import os
import re
from collections import Counter
from datetime import datetime
from functools import lru_cache

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from analysis.concept_analyzer import (
    analyze_concepts_from_ladder_data, format_concept_analysis_summary
)
from analysis.helper.ladder_chart_helpers import (
    # 常量
    VOLUME_DAYS, VOLUME_RATIO_THRESHOLD, VOLUME_RATIO_LOW_THRESHOLD,
    NEW_HIGH_MARKER,
    MA_SLOPE_DAYS, HIGH_GAIN_TRACKING_THRESHOLD, COLLAPSE_DAYS_AFTER_BREAK,
    # 缓存管理
    clear_helper_caches, cache_zaban_format,  # 股票数据
    get_stock_data_df, get_stock_data, get_stock_daily_pct_change,
    # 成交量
    add_volume_ratio_to_text,
    # 新高标记
    get_new_high_markers_cached,
    # MA斜率
    get_ma_slope_indicator, clear_ma_slope_cache, print_slope_statistics,
    # 跟踪判断
    clear_high_gain_cache,
    should_track_after_break as _should_track_after_break,
    should_track_before_entry, calculate_last_board_date,
    should_collapse_row as _should_collapse_row,
    # 炸板检查
    check_stock_in_zaban, check_stock_in_shouban,
)
from analysis.loader.fupan_data_loader import (
    OUTPUT_FILE, load_stock_data
)
from analysis.momo_shangzhang_processor import (
    identify_momo_shangzhang_stocks
)
from decorators.practical import timer
from utils.date_util import get_trading_days, count_trading_days_between, get_n_trading_days_before, \
    get_valid_trading_date_pair
from utils.stock_util import get_stock_market
from utils.theme_color_util import (
    extract_reasons, get_reason_colors, get_stock_reason_group, normalize_reason,
    create_legend_sheet, get_color_for_pct_change, add_market_indicators,
    create_index_sheet, HIGH_BOARD_COLORS, REENTRY_COLORS, BOARD_COLORS,
    extract_reasons_with_match_type
)

# 断板后跟踪的最大天数，超过这个天数后不再显示涨跌幅
# 例如设置为5，会显示断板后的第1、2、3、4、5个交易日，从第6个交易日开始不再显示
# 设置为None表示一直跟踪到分析周期结束
MAX_TRACKING_DAYS_AFTER_BREAK = 11

# 断板后折叠行的天数阈值，超过这个天数的股票会在Excel中自动折叠（隐藏）
# 例如设置为7，断板7天后该股票所在行会被折叠，减少显示数据量
# 设置为None表示不自动折叠任何行
COLLAPSE_DAYS_AFTER_BREAK = 12

# 入选前跟踪的最大天数，显示入选前的第1、2、3、...个交易日的涨跌幅
# 例如设置为3，会显示入选前的第1、2、3个交易日的涨跌幅
# 设置为0表示不显示入选前的走势
MAX_TRACKING_DAYS_BEFORE_ENTRY = 7

# 断板后再次达到入选的交易日间隔阈值
# 例如设置为4，当股票断板后第5个交易日或之后再次达到入选时，会作为新的一行记录
# 如果一只股票断板后超过这个交易日天数又再次达到入选，则视为新的一行记录
REENTRY_DAYS_THRESHOLD = 4

# 计算入选日与之前X个交易日的涨跌幅
# 例如设置为20，会计算入选日与20个交易日之前的涨跌幅
PERIOD_DAYS_CHANGE = 10

# 计算最近30天的涨跌幅
PERIOD_DAYS_LONG = 30

# 注意：以下常量已移动到 ladder_chart_helpers.py：
# - HIGH_GAIN_TRACKING_THRESHOLD, VOLUME_DAYS, VOLUME_RATIO_THRESHOLD, VOLUME_RATIO_LOW_THRESHOLD
# - NEW_HIGH_DAYS, NEW_HIGH_MARKER, MA_SLOPE_DAYS, MA_SLOPE_THRESHOLD_PCT, COLLAPSE_DAYS_AFTER_BREAK

# 关注度榜加粗相关参数
# 关注度榜取前N名
ATTENTION_TOP_N = 15
# 统计最近N个交易日
ATTENTION_DAYS_WINDOW = 3

# 调试模式开关（控制详细分析日志的输出）
DEBUG_MODE = False

# ==================== 龙头股筛选相关参数 ====================
# 【筛选门槛 - 主板股】
MIN_BOARD_LEVEL_FOR_LEADER = 2  # 主板股最低连板数门槛
MIN_SHORT_PERIOD_CHANGE_FOR_LEADER = 20.0  # 主板股最低短周期涨幅门槛（%）
MIN_LONG_PERIOD_CHANGE_FOR_LEADER = 80.0  # 主板股最低长周期涨幅门槛（%），min和max设成一样表示不要求长周期涨多少
MAX_LONG_PERIOD_CHANGE_FOR_LEADER = 120.0  # 主板股最高长周期涨幅门槛（%，避免涨幅过高）

# 【筛选门槛 - 非主板股（创业板/科创板/北交所）】
MIN_BOARD_LEVEL_FOR_LEADER_NON_MAIN = 1  # 非主板股最低连板数门槛
MIN_SHORT_PERIOD_CHANGE_FOR_LEADER_NON_MAIN = 30.0  # 非主板股最低短周期涨幅门槛（%）
MIN_LONG_PERIOD_CHANGE_FOR_LEADER_NON_MAIN = 90.0  # 非主板股最低长周期涨幅门槛（%）
MAX_LONG_PERIOD_CHANGE_FOR_LEADER_NON_MAIN = 130.0  # 非主板股最高长周期涨幅门槛（%，避免涨幅过高）

# 【名额分配规则】按板块活跃度排名动态分配龙头数量
LEADER_QUOTA_TOP1 = 4  # 最热板块（排名第1）
LEADER_QUOTA_TOP2 = 4  # 次热板块（排名第2）
LEADER_QUOTA_TOP3 = 3  # 第三热板块（排名第3）
LEADER_QUOTA_TOP4 = 3  # 第四热板块（排名第4）
LEADER_QUOTA_DEFAULT = 2  # 默认板块（排名第5到默认阈值之间）
LEADER_QUOTA_COLD = 1  # 非热门板块（排名在默认阈值之后）
LEADER_QUOTA_DEFAULT_THRESHOLD = 0.2  # 默认/冷门分界线（例如0.5表示前50%为默认，后50%为冷门）

# 【筛选策略】
SELECT_LEADERS_FROM_ACTIVE_ONLY = True  # 是否只从活跃股中选择（True=只从未被折叠的股票中选，False=从全部符合条件的股票中选）
LEADER_EXCLUDE_CONCEPTS = ['默默上涨']  # 排除在龙头股筛选之外的特殊概念组（列表形式，方便扩展）

# 【工作表管理】
MAX_LEADER_SHEETS = 3  # 最大龙头股工作表保留数量（超过此数量会自动归档旧的sheet）

# 关注度榜前N名股票缓存
_top_attention_stocks_cache = None

# 单元格边框样式
BORDER_STYLE = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 周期涨跌幅颜色映射
PERIOD_CHANGE_COLORS = {
    "EXTREME_POSITIVE": "9933FF",  # 深紫色 - 极强势 (≥100%)
    "STRONG_POSITIVE": "AA66FF",  # 中深紫色 - 强势 (≥70%)
    "MODERATE_POSITIVE": "BB99FF",  # 中紫色 - 较强势 (≥40%)
    "MILD_POSITIVE": "CCCCFF",  # 浅紫色 - 偏强势 (≥20%)
    "SLIGHT_POSITIVE": "E6E6FF",  # 极浅紫色 - 微强势 (≥0%)
    "SLIGHT_NEGATIVE": "E6EBC9",  # 极浅橄榄绿 - 微弱势 (≥-20%)
    "MILD_NEGATIVE": "C9D58C",  # 浅橄榄绿 - 偏弱势 (≥-40%)
    "MODERATE_NEGATIVE": "A3B86C",  # 中橄榄绿 - 较弱势 (≥-60%)
    "STRONG_NEGATIVE": "7D994D",  # 深橄榄绿 - 弱势 (<-60%)
}


def clear_caches():
    """清理所有缓存"""
    global _top_attention_stocks_cache
    # 清理 helpers 模块的缓存
    clear_helper_caches()
    _top_attention_stocks_cache = None
    print("已清理所有缓存")


def load_top_attention_stocks(end_date_yyyymmdd, days_window=ATTENTION_DAYS_WINDOW, top_n=ATTENTION_TOP_N):
    """
    加载最近N个交易日内进入关注度榜前N名的股票
    
    Args:
        end_date_yyyymmdd: 结束日期 (YYYYMMDD格式)
        days_window: 向前查找的交易日天数
        top_n: 取前N名
        
    Returns:
        set: 股票代码集合（已标准化，不含市场前缀）
    """
    try:
        from analysis.loader.fupan_data_loader import FUPAN_FILE
        from openpyxl import load_workbook

        # 读取关注度榜数据（从复盘数据源文件读取）
        wb = load_workbook(FUPAN_FILE, data_only=True)

        # 获取最近N个交易日
        recent_dates = get_n_recent_trading_dates(end_date_yyyymmdd, days_window)

        attention_stocks = set()

        # 处理两个sheet：【关注度榜】和【非主关注度榜】
        for sheet_name in ['关注度榜', '非主关注度榜']:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # 遍历所有列，查找最近N日的数据
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                if not header_cell.value:
                    continue

                # 解析日期（格式：2025年11月18日）
                col_date = parse_date_from_header(header_cell.value)
                if not col_date or col_date not in recent_dates:
                    continue

                # 读取该列的前top_n行数据（从第2行开始）
                for row_idx in range(2, min(2 + top_n, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if not cell_value:
                        continue

                    # 解析数据：600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1
                    stock_code = extract_stock_code_from_attention_data(cell_value)
                    if stock_code:
                        attention_stocks.add(stock_code)

        print(f"✓ 加载关注度榜数据：最近{days_window}日前{top_n}名，共{len(attention_stocks)}只股票")
        return attention_stocks

    except Exception as e:
        print(f"✗ 加载关注度榜数据失败: {e}")
        import traceback
        traceback.print_exc()
        return set()


def get_n_recent_trading_dates(end_date_yyyymmdd, n):
    """
    获取最近N个交易日的日期集合（YYYYMMDD格式）
    
    Args:
        end_date_yyyymmdd: 结束日期 (YYYYMMDD格式)
        n: 交易日天数
        
    Returns:
        set: 日期集合
    """
    dates = set()
    current_date = end_date_yyyymmdd
    dates.add(current_date)

    for i in range(1, n):
        prev_date = get_n_trading_days_before(end_date_yyyymmdd, i)
        if '-' in prev_date:
            prev_date = prev_date.replace('-', '')
        dates.add(prev_date)

    return dates


def parse_date_from_header(header_value):
    """
    从表头解析日期：2025年11月18日 -> 20251118
    
    Args:
        header_value: 表头值
        
    Returns:
        str: YYYYMMDD格式的日期，解析失败返回None
    """
    try:
        if isinstance(header_value, str) and '年' in header_value:
            date_obj = datetime.strptime(header_value, '%Y年%m月%d日')
            return date_obj.strftime('%Y%m%d')
    except:
        pass
    return None


def extract_stock_code_from_attention_data(cell_value):
    """
    从关注度榜数据中提取股票代码
    
    输入: "600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1"
    输出: "600340"（标准化后的纯代码）
    
    Args:
        cell_value: 单元格值
        
    Returns:
        str: 标准化后的股票代码，解析失败返回None
    """
    try:
        parts = str(cell_value).split(';')
        if len(parts) >= 1:
            stock_code = parts[0].strip()  # "600340.SH"
            # 去除市场后缀 .SH/.SZ
            if '.' in stock_code:
                stock_code = stock_code.split('.')[0]
            return stock_code
    except:
        pass
    return None


def get_top_attention_stocks_cached(end_date_yyyymmdd):
    """
    获取缓存的关注度榜前N名股票，避免重复加载
    
    Args:
        end_date_yyyymmdd: 结束日期 (YYYYMMDD格式)
        
    Returns:
        set: 股票代码集合
    """
    global _top_attention_stocks_cache

    # 如果缓存为空，则加载数据
    if _top_attention_stocks_cache is None:
        _top_attention_stocks_cache = load_top_attention_stocks(end_date_yyyymmdd)

    return _top_attention_stocks_cache


def get_loose_board_level(board_level):
    """
    获取宽松的连板数
    """
    return 1 if board_level == 1 else board_level - 1


def check_attention_criteria(stock_code, board_days, market, min_board_level, non_main_board_level,
                             current_date_col, attention_data_main, attention_data_non_main):
    """
    检查股票是否满足关注度榜入选条件

    Args:
        stock_code: 股票代码
        board_days: 连板天数
        market: 市场类型
        min_board_level: 主板股票最小显著连板天数
        non_main_board_level: 非主板股票最小显著连板天数
        current_date_col: 当前日期列名
        attention_data_main: 主板关注度榜数据
        attention_data_non_main: 非主板关注度榜数据

    Returns:
        tuple: (是否满足条件, 入选类型)
    """
    # 默认不满足条件，入选类型为normal
    is_significant = False
    entry_type = 'normal'

    # 清理股票代码，去除可能的市场前缀
    clean_stock_code = stock_code.split('.')[0] if '.' in stock_code else stock_code

    # 解析当前日期
    current_date = datetime.strptime(current_date_col, '%Y年%m月%d日') if '年' in current_date_col else pd.to_datetime(
        current_date_col)

    # 根据市场类型选择不同的检查逻辑
    if market == 'main' and board_days == get_loose_board_level(min_board_level):
        # 检查主板股票
        if attention_data_main is not None and not attention_data_main.empty:
            is_significant, entry_type = check_stock_attention(
                clean_stock_code, current_date, attention_data_main,
                f"主板股票 {stock_code} 在(board_level-1)={get_loose_board_level(min_board_level)}板"
            )
    elif market in ['gem', 'star', 'bse'] and board_days == get_loose_board_level(non_main_board_level):
        # 检查非主板股票
        if attention_data_non_main is not None and not attention_data_non_main.empty:
            is_significant, entry_type = check_stock_attention(
                clean_stock_code, current_date, attention_data_non_main,
                f"非主板股票 {stock_code} 在(board_level-1)={get_loose_board_level(non_main_board_level)}板"
            )

    return is_significant, entry_type


def check_stock_attention(stock_code, current_date, attention_data, log_prefix):
    """
    检查股票在关注度榜中的出现情况

    Args:
        stock_code: 股票代码
        current_date: 当前日期
        attention_data: 关注度榜数据
        log_prefix: 日志前缀

    Returns:
        tuple: (是否满足条件, 入选类型)
    """
    # 过滤出当前股票的关注度数据
    stock_attention = attention_data[attention_data['股票代码'].str.contains(stock_code)]

    if stock_attention.empty:
        return False, 'normal'

    # 计算每个关注度日期与当前日期的交易日差
    attention_dates = []
    for _, att_row in stock_attention.iterrows():
        att_date_str = att_row['日期']
        att_date = datetime.strptime(att_date_str, '%Y年%m月%d日') if '年' in att_date_str else pd.to_datetime(
            att_date_str)
        days_diff = count_trading_days_between(current_date, att_date)
        # 当天之前2个或之后的3个交易日
        if -2 <= days_diff <= 3:  # 负值表示之前的交易日，正值表示之后的交易日
            attention_dates.append(att_date)

    # 如果在指定交易日范围内出现了至少两次，则认为符合条件
    if len(attention_dates) >= 2:
        if DEBUG_MODE:
            print(f"    {log_prefix}指定交易日范围内两次入选关注度榜前20，符合额外入选条件")
        return True, 'attention'

    return False, 'normal'


def is_stock_significant(board_days, market, min_board_level, non_main_board_level):
    """
    判断股票是否达到显著连板条件

    Args:
        board_days: 连板天数
        market: 市场类型
        min_board_level: 主板股票最小显著连板天数
        non_main_board_level: 非主板股票最小显著连板天数

    Returns:
        bool: 是否达到显著连板条件
    """
    if not board_days:
        return False

    if market == 'main':
        return board_days >= min_board_level
    elif market in ['gem', 'star', 'bse']:
        return board_days >= non_main_board_level
    else:
        # 其他情况，使用主板标准
        return board_days >= min_board_level


def select_concept_group_for_stock(concept_str, entry_type, top_reasons, low_priority_reasons=None):
    """
    为股票选择最合适的概念组
    
    优先级规则：
    1. 排除低优先级概念（最高优先级）
    2. 匹配类型：精确匹配 > 模糊匹配
    3. 热门度：top_reasons中的排名
    
    Args:
        concept_str: 概念字符串
        entry_type: 入选类型（如'momo_shangzhang'）
        top_reasons: 热门原因列表
        low_priority_reasons: 低优先级原因列表
        
    Returns:
        str: 选择的概念组名称
    """
    # 特殊类型：【默默上涨】
    if entry_type == 'momo_shangzhang':
        return "默默上涨"

    # 空概念
    if pd.isna(concept_str) or not concept_str:
        return "其他"

    # 提取概念及其匹配类型
    concepts_with_type = extract_reasons_with_match_type(concept_str)
    if not concepts_with_type:
        return "其他"

    # 性能优化：转换为set加速查找
    top_reasons_set = set(top_reasons) if top_reasons else set()
    low_priority_set = set(low_priority_reasons) if low_priority_reasons else set()

    # 分离精确匹配和模糊匹配的概念
    exact_concepts = [c[0] for c in concepts_with_type if c[1] == 'exact']
    fuzzy_concepts = [c[0] for c in concepts_with_type if c[1] == 'fuzzy']

    # ========== 优先级规则：排除低优先级 > 匹配类型 > 热门度 ==========

    # 优先级1：精确匹配的非低优先级概念
    if exact_concepts:
        # 1a. 优先选择热门的、非低优先级的精确匹配概念
        for top_reason in top_reasons:
            if top_reason in exact_concepts and top_reason not in low_priority_set:
                return top_reason

        # 1b. 如果有非低优先级的精确匹配（即使不是热门），也优先返回
        for concept in exact_concepts:
            if concept not in low_priority_set:
                return concept

    # 优先级2：模糊匹配的非低优先级概念
    if fuzzy_concepts:
        # 2a. 优先选择热门的、非低优先级的模糊匹配概念
        for top_reason in top_reasons:
            if top_reason in fuzzy_concepts and top_reason not in low_priority_set:
                return top_reason

        # 2b. 如果有非低优先级的模糊匹配（即使不是热门），也优先返回
        for concept in fuzzy_concepts:
            if concept not in low_priority_set:
                return concept

    # 优先级3：如果所有概念都是低优先级，则按原逻辑选择
    # 3a. 精确匹配的低优先级概念
    if exact_concepts:
        for top_reason in top_reasons:
            if top_reason in exact_concepts:
                return top_reason
        return exact_concepts[0]

    # 3b. 模糊匹配的低优先级概念
    if fuzzy_concepts:
        for top_reason in top_reasons:
            if top_reason in fuzzy_concepts:
                return top_reason
        return fuzzy_concepts[0]

    # 兜底：返回第一个概念（如果有的话）
    all_concepts = [c[0] for c in concepts_with_type]
    return all_concepts[0] if all_concepts else "其他"


def check_reentry_condition(current_date, continuous_board_dates, reentry_days_threshold,
                            board_days, market, min_board_level, non_main_board_level,
                            enable_attention_criteria=False, attention_data_main=None,
                            attention_data_non_main=None, stock_code=None):
    """
    检查是否满足断板后再次入选条件

    Args:
        current_date: 当前日期
        continuous_board_dates: 连续连板日期列表
        reentry_days_threshold: 断板后再次上榜的天数阈值
        board_days: 连板天数
        market: 市场类型
        min_board_level: 主板股票最小显著连板天数
        non_main_board_level: 非主板股票最小显著连板天数
        enable_attention_criteria: 是否启用关注度榜入选条件
        attention_data_main: 主板关注度榜数据
        attention_data_non_main: 非主板关注度榜数据
        stock_code: 股票代码

    Returns:
        tuple: (是否满足再次入选条件, 是否需要清空连续连板日期)
    """
    # 获取之前的连板日期
    previous_board_dates = [d for d in continuous_board_dates if d < current_date]

    if not previous_board_dates:
        return False, False

    # 获取上一个连板区间的最后日期
    last_board_date = max(previous_board_dates)

    # 计算间隔交易日天数
    days_since_last_board = count_trading_days_between(last_board_date, current_date)

    if DEBUG_MODE:
        print(f"    上一次连板日期: {last_board_date.strftime('%Y-%m-%d')}, "
              f"当前日期: {current_date.strftime('%Y-%m-%d')}, "
              f"交易日间隔: {days_since_last_board}天")

    # 判断是否满足再次入选条件
    if days_since_last_board > reentry_days_threshold:
        # 首先检查是否达到基本的显著连板条件
        is_significant_reentry = is_stock_significant(board_days, market, min_board_level, non_main_board_level)

        # 如果不满足基本条件，但启用了关注度榜入选条件，则检查是否满足关注度榜条件
        if not is_significant_reentry and enable_attention_criteria and stock_code:
            current_date_str = current_date.strftime('%Y年%m月%d日')
            is_significant_reentry, _ = check_attention_criteria(
                stock_code, board_days, market, min_board_level, non_main_board_level,
                current_date_str, attention_data_main, attention_data_non_main
            )

        if is_significant_reentry:
            if DEBUG_MODE:
                print(f"    断板后{days_since_last_board}个交易日再次达到入选条件，作为新记录")
            return True, True

    return False, False


@lru_cache(maxsize=1000)
def get_cached_concept_group(stock_code, stock_name, concept_str, priority_reasons_str=None,
                             low_priority_reasons_str=None):
    """缓存股票的概念组信息，用于排序"""
    # 将列表类型的priority_reasons转换为元组，使其可哈希
    priority_reasons = None if priority_reasons_str is None else tuple(priority_reasons_str.split(','))

    # 将列表类型的low_priority_reasons转换为元组，使其可哈希
    low_priority_reasons = None if low_priority_reasons_str is None else tuple(low_priority_reasons_str.split(','))

    # 提取概念
    if pd.isna(concept_str) or not concept_str:
        return "其他"

    # 提取所有概念（包含匹配类型信息）
    from utils.theme_color_util import extract_reasons_with_match_type
    reason_details = extract_reasons_with_match_type(concept_str)

    if not reason_details:
        return "其他"

    # 提取规范化后的概念列表（用于颜色映射）
    reasons = [normalized for normalized, _, _ in reason_details]

    # 创建简化版的股票概念数据结构
    stock_key = f"{stock_code}_{stock_name}"
    all_concepts = reasons.copy()

    # 获取热门概念的颜色映射
    _, top_reasons = get_reason_colors(all_concepts, priority_reasons=priority_reasons,
                                       low_priority_reasons=low_priority_reasons)

    # 创建简化版的股票数据（新格式：包含匹配类型信息）
    all_stocks = {
        stock_key: {
            'name': stock_name,
            'reasons': reasons,
            'reason_details': reason_details,  # 新增：包含匹配类型
            'appearances': [1]
        }
    }

    # 获取股票的主要概念组
    stock_reason_groups = get_stock_reason_group(all_stocks, top_reasons, low_priority_reasons=low_priority_reasons)

    # 返回股票的概念组
    return stock_reason_groups.get(stock_key, "其他")


def identify_first_significant_board(df, shouban_df=None, min_board_level=2,
                                     reentry_days_threshold=REENTRY_DAYS_THRESHOLD, non_main_board_level=1,
                                     enable_attention_criteria=False, attention_data_main=None,
                                     attention_data_non_main=None, priority_reasons=None, low_priority_reasons=None):
    """
    识别每只股票首次达到显著连板（例如2板或以上）的日期，以及断板后再次达到的情况

    Args:
        df: 连板数据DataFrame，已透视处理，每行一只股票，每列一个日期
        shouban_df: 首板数据DataFrame，已透视处理，每行一只股票，每列一个日期
        min_board_level: 主板股票最小显著连板天数，默认为2
        reentry_days_threshold: 断板后再次上榜的天数阈值，超过这个天数再次达到入选条件会作为新记录
        non_main_board_level: 非主板股票最小显著连板天数，默认为1
        enable_attention_criteria: 是否启用关注度榜入选条件，默认为False
        attention_data_main: 主板关注度榜数据，当enable_attention_criteria=True时需要提供
        attention_data_non_main: 非主板关注度榜数据，当enable_attention_criteria=True时需要提供
        priority_reasons: 优先选择的原因列表，默认为None
        low_priority_reasons: 低优先级原因列表，默认为None

    Returns:
        pandas.DataFrame: 添加了连板信息的DataFrame
    """
    # 创建结果列表
    result = []

    # 找出日期列
    date_columns = [col for col in df.columns if col not in ['纯代码', '股票名称', '股票代码', '概念']]
    if not date_columns:
        print("无法找到日期列")
        return pd.DataFrame()

    # 将日期列按时间排序
    date_columns.sort()

    # 获取所有股票数据
    all_stocks = get_all_stocks(df, date_columns, shouban_df)

    # 确定每只股票的市场类型
    determine_stock_markets(all_stocks)

    # 分析每只股票，确定是否入选显著连板
    analyze_stocks_for_significant_boards(
        all_stocks, date_columns, result, min_board_level,
        reentry_days_threshold, non_main_board_level,
        enable_attention_criteria, attention_data_main, attention_data_non_main
    )

    # 转换为DataFrame
    result_df = pd.DataFrame(result)

    # 如果没有符合条件的数据，返回空DataFrame
    if result_df.empty:
        print("未找到符合条件的连板数据")
        return result_df

    print(f"找到{len(result_df)}只达到显著连板的股票")

    # 计算全局热门概念用于分组
    all_concepts = []
    for _, row in result_df.iterrows():
        concept_str = row.get('concept', '')
        if pd.notna(concept_str) and concept_str:
            concepts = extract_reasons(concept_str)
            all_concepts.extend(concepts)

    # 获取全局热门概念的顺序
    global_reason_colors, global_top_reasons = get_reason_colors(all_concepts, priority_reasons=priority_reasons,
                                                                 low_priority_reasons=low_priority_reasons)

    # 创建概念优先级映射（热门概念排在前面）
    concept_priority = {}
    for i, concept in enumerate(global_top_reasons):
        concept_priority[concept] = i
    concept_priority["其他"] = len(global_top_reasons)  # "其他"排在最后

    # 添加概念组信息用于排序，使用全局热门概念和低优先级过滤
    result_df['concept_group'] = result_df.apply(
        lambda row: select_concept_group_for_stock(
            row.get('concept', ''),
            row.get('entry_type'),
            global_top_reasons,
            low_priority_reasons
        ),
        axis=1
    )

    # 计算每个概念组的股票数量，用于非热门概念的排序
    concept_counts = result_df['concept_group'].value_counts().to_dict()

    # 为非热门概念（不在global_top_reasons中的概念）按股票数量重新分配优先级
    non_hot_concepts = [concept for concept in concept_counts.keys()
                        if concept not in global_top_reasons and concept != "其他"]
    # 按股票数量倒序排列非热门概念
    non_hot_concepts_sorted = sorted(non_hot_concepts, key=lambda x: concept_counts[x], reverse=True)

    # 重新构建概念优先级映射
    concept_priority = {}
    # 热门概念保持原有优先级
    for i, concept in enumerate(global_top_reasons):
        concept_priority[concept] = i

    # 非热门概念按股票数量倒序排列，优先级在热门概念之后
    start_priority = len(global_top_reasons)
    for i, concept in enumerate(non_hot_concepts_sorted):
        concept_priority[concept] = start_priority + i

    # 【默默上涨】排在非热门概念之后，"其他"之前
    concept_priority["默默上涨"] = start_priority + len(non_hot_concepts_sorted)

    # "其他"排在最后
    concept_priority["其他"] = start_priority + len(non_hot_concepts_sorted) + 1

    # 添加概念优先级列用于排序
    result_df['concept_priority'] = result_df['concept_group'].map(lambda x: concept_priority.get(x, 999))

    # 按首次显著连板日期、概念优先级、概念组和连板天数排序
    result_df = result_df.sort_values(
        by=['first_significant_date', 'concept_priority', 'concept_group', 'board_level_at_first'],
        ascending=[True, True, True, False]
    )

    return result_df


def get_all_stocks(df, date_columns, shouban_df=None):
    """
    合并连板数据和首板数据，创建完整的股票池

    Args:
        df: 连板数据DataFrame
        date_columns: 日期列列表
        shouban_df: 首板数据DataFrame

    Returns:
        dict: 股票池字典，键为股票代码，值为股票数据
    """
    all_stocks = {}

    # 1. 处理连板数据
    print(f"处理连板数据，共有{len(df)}只股票")
    for idx, row in df.iterrows():
        stock_code = row['纯代码']
        stock_name = row['股票名称']

        # 跳过代码或名称为空的行
        if not stock_code or not stock_name:
            continue

        # 创建股票数据字典
        stock_data = {
            'stock_code': stock_code,
            'stock_name': stock_name,
            'concept': row.get('概念', '其他'),
            'board_data': {},
            'market': None  # 将在后面填充
        }

        # 填充连板数据
        for col in date_columns:
            if pd.notna(row[col]):
                stock_data['board_data'][col] = row[col]

        all_stocks[stock_code] = stock_data

    # 2. 处理首板数据，合并到股票池中
    if shouban_df is not None and not shouban_df.empty:
        print(f"处理首板数据，共有{len(shouban_df)}只股票")
        shouban_date_columns = [col for col in shouban_df.columns if
                                col not in ['纯代码', '股票名称', '股票代码', '概念']]
        shouban_date_columns.sort()

        for idx, row in shouban_df.iterrows():
            stock_code = row['纯代码']
            stock_name = row['股票名称']

            # 跳过代码或名称为空的行
            if not stock_code or not stock_name:
                continue

            # 如果股票已在连板数据中，则合并首板数据
            if stock_code in all_stocks:
                for col in shouban_date_columns:
                    if pd.notna(row[col]) and col not in all_stocks[stock_code]['board_data']:
                        all_stocks[stock_code]['board_data'][col] = row[col]
            else:
                # 创建新的股票数据
                stock_data = {
                    'stock_code': stock_code,
                    'stock_name': stock_name,
                    'concept': row.get('概念', '其他'),
                    'board_data': {},
                    'market': None  # 将在后面填充
                }

                # 填充首板数据
                for col in shouban_date_columns:
                    if pd.notna(row[col]):
                        stock_data['board_data'][col] = row[col]

                all_stocks[stock_code] = stock_data

    return all_stocks


def determine_stock_markets(all_stocks):
    """
    确定每只股票的市场类型

    Args:
        all_stocks: 股票池字典
    """
    for stock_code, stock_data in all_stocks.items():
        try:
            market = get_stock_market(stock_code)
            stock_data['market'] = market
        except Exception as e:
            print(f"判断股票 {stock_code} 市场类型时出错: {e}")
            stock_data['market'] = 'unknown'


def analyze_stocks_for_significant_boards(all_stocks, date_columns, result, min_board_level,
                                          reentry_days_threshold, non_main_board_level,
                                          enable_attention_criteria, attention_data_main, attention_data_non_main):
    """
    分析每只股票，确定是否入选显著连板

    Args:
        all_stocks: 股票池字典
        date_columns: 日期列列表
        result: 结果列表，用于存储入选的股票记录
        min_board_level: 主板股票最小显著连板天数
        reentry_days_threshold: 断板后再次上榜的天数阈值
        non_main_board_level: 非主板股票最小显著连板天数
        enable_attention_criteria: 是否启用关注度榜入选条件
        attention_data_main: 主板关注度榜数据
        attention_data_non_main: 非主板关注度榜数据
    """
    if DEBUG_MODE:
        print(f"分析股票池中的所有股票，共有{len(all_stocks)}只股票")
    for stock_code, stock_data in all_stocks.items():
        stock_name = stock_data['stock_name']
        market = stock_data['market']
        board_data = stock_data['board_data']

        if DEBUG_MODE:
            print(f"分析股票: {stock_code}_{stock_name} (市场: {market})")

        # 按日期排序的板块数据
        sorted_dates = sorted(board_data.keys())

        # 存储所有板块出现的时间点
        significant_board_dates = []  # 存储显著连板日期
        continuous_board_dates = []  # 存储连续的连板日期，用于判断断板

        # 检查每一个日期
        for col in sorted_dates:
            board_days = board_data[col]
            if DEBUG_MODE:
                print(f"  日期 {col}: {board_days}")

            # 判断是否为显著连板
            is_significant = False
            entry_type = 'normal'  # 默认入选类型

            # 首先检查是否满足基本的显著连板条件
            is_significant = is_stock_significant(board_days, market, min_board_level, non_main_board_level)

            # 如果不满足基本条件，但启用了关注度榜入选条件，则检查是否满足关注度榜条件
            if not is_significant and enable_attention_criteria:
                is_significant, entry_type = check_attention_criteria(
                    stock_code, board_days, market, min_board_level, non_main_board_level,
                    col, attention_data_main, attention_data_non_main
                )

            if is_significant:
                if DEBUG_MODE:
                    print(f"    找到显著连板: {board_days}板")

                # 记录为显著连板日期
                current_date = datetime.strptime(col, '%Y年%m月%d日') if '年' in col else pd.to_datetime(col)

                # 记录连续连板日期，用于确定最后一次连板的日期
                continuous_board_dates.append(current_date)

                # 判断是否为新的入选记录
                is_new_entry = False

                if not significant_board_dates:
                    # 第一次显著连板
                    is_new_entry = True
                elif continuous_board_dates:
                    # 检查是否是断板后间隔足够长再次达到入选条件
                    is_reentry, clear_dates = check_reentry_condition(
                        current_date, continuous_board_dates, reentry_days_threshold,
                        board_days, market, min_board_level, non_main_board_level,
                        enable_attention_criteria, attention_data_main, attention_data_non_main, stock_code
                    )

                    if is_reentry:
                        is_new_entry = True
                        # 清空连续连板日期列表，开始新的连板区间记录
                        if clear_dates:
                            continuous_board_dates = [current_date]

                # 添加到显著连板日期列表
                significant_board_dates.append(current_date)

                if is_new_entry:
                    # 构建完整的板块数据字典
                    all_board_data = {}
                    for date_col in date_columns:
                        all_board_data[date_col] = board_data.get(date_col)

                    # 确定入选类型
                    if entry_type == 'attention':
                        final_entry_type = 'attention'
                    elif len(significant_board_dates) > 1:
                        final_entry_type = 'reentry'
                    elif market in ['gem', 'star', 'bse'] and board_days == 1:
                        final_entry_type = 'non_main_first_board'
                    else:
                        final_entry_type = 'first'

                    # 添加到结果列表
                    entry = {
                        'stock_code': stock_code,
                        'stock_name': stock_name,
                        'first_significant_date': current_date,
                        'board_level_at_first': board_days,
                        'all_board_data': all_board_data,
                        'concept': stock_data['concept'],
                        'entry_type': final_entry_type
                    }

                    result.append(entry)
                    if DEBUG_MODE:
                        print(f"    记录显著连板: {stock_name} 在 {col} 达到 {board_days}板")


def get_concept_from_board_text(board_text):
    """
    从连板文本中提取概念信息

    Args:
        board_text: 连板文本，可能包含概念信息

    Returns:
        str: 提取出的概念，如果没有则返回"其他"
    """
    if not board_text or pd.isna(board_text):
        return "其他"

    # 将连板文本转换为字符串并去除空格
    board_text = str(board_text).strip()

    # 尝试提取概念（假设概念位于连板天数之后的括号中）
    match = re.search(r'\(([^)]+)\)', board_text)
    if match:
        concept = match.group(1).strip()
        # 使用theme_color_util中的normalize_reason对概念进行规范化
        try:
            normalized_concept = normalize_reason(concept)
            return normalized_concept if not normalized_concept.startswith("未分类") else concept
        except:
            return concept

    # 如果没有找到括号内的概念，返回默认值
    return "其他"


def get_market_marker(stock_code):
    """
    根据股票代码获取市场标记

    Args:
        stock_code: 股票代码

    Returns:
        str: 市场标记，创业板或科创板返回"*"，北交所返回"**"，其他返回空字符串
    """
    try:
        market = get_stock_market(stock_code)
        if market == 'gem' or market == 'star':  # 创业板或科创板
            return "*"
        elif market == 'bse':  # 北交所
            return "**"
        return ""
    except Exception as e:
        print(f"获取股票 {stock_code} 市场类型出错: {e}")
        return ""


def get_color_for_period_change(pct_change):
    """
    根据周期涨跌幅获取背景色（紫色系为正，橄榄绿系为负）

    Args:
        pct_change: 涨跌幅百分比

    Returns:
        str: 颜色代码
    """
    if pct_change is None:
        return None

    # 正值使用紫色系配色
    if pct_change >= 100:
        return PERIOD_CHANGE_COLORS["EXTREME_POSITIVE"]
    elif pct_change >= 70:
        return PERIOD_CHANGE_COLORS["STRONG_POSITIVE"]
    elif pct_change >= 40:
        return PERIOD_CHANGE_COLORS["MODERATE_POSITIVE"]
    elif pct_change >= 20:
        return PERIOD_CHANGE_COLORS["MILD_POSITIVE"]
    elif pct_change >= 0:
        return PERIOD_CHANGE_COLORS["SLIGHT_POSITIVE"]
    # 负值使用橄榄绿系配色
    elif pct_change >= -20:
        return PERIOD_CHANGE_COLORS["SLIGHT_NEGATIVE"]
    elif pct_change >= -40:
        return PERIOD_CHANGE_COLORS["MILD_NEGATIVE"]
    elif pct_change >= -60:
        return PERIOD_CHANGE_COLORS["MODERATE_NEGATIVE"]
    else:
        return PERIOD_CHANGE_COLORS["STRONG_NEGATIVE"]


@lru_cache(maxsize=1000)
def calculate_stock_period_change(stock_code, start_date_yyyymmdd, end_date_yyyymmdd, stock_name=None):
    """
    计算股票在两个日期之间的涨跌幅

    Args:
        stock_code: 股票代码
        start_date_yyyymmdd: 开始日期 (YYYYMMDD)
        end_date_yyyymmdd: 结束日期 (YYYYMMDD)
        stock_name: 股票名称 (不再使用)

    Returns:
        float: 涨跌幅百分比，如果数据不存在则返回None
    """
    try:
        if not stock_code:
            return None

        # 获取股票数据
        df = get_stock_data_df(stock_code)

        if df is None or df.empty:
            return None

        # 计算原始的交易日间隔
        original_trading_days = count_trading_days_between(start_date_yyyymmdd, end_date_yyyymmdd)
        if original_trading_days <= 0:
            return None

        # 使用新的函数获取有效的日期对
        valid_start_date, valid_end_date = get_valid_trading_date_pair(
            end_date_yyyymmdd, original_trading_days, df
        )

        if valid_start_date is None or valid_end_date is None:
            return None

        # 格式化日期进行查询
        start_date_fmt = f"{valid_start_date[:4]}-{valid_start_date[4:6]}-{valid_start_date[6:8]}"
        end_date_fmt = f"{valid_end_date[:4]}-{valid_end_date[4:6]}-{valid_end_date[6:8]}"

        # 查找数据
        start_row = df[df['日期'] == start_date_fmt]
        end_row = df[df['日期'] == end_date_fmt]

        # 如果还是找不到，使用最接近的日期
        if start_row.empty or end_row.empty:
            all_dates = pd.to_datetime(df['日期'])
            start_date_dt = pd.to_datetime(start_date_fmt)
            end_date_dt = pd.to_datetime(end_date_fmt)

            if start_row.empty:
                valid_dates = all_dates[all_dates <= start_date_dt]
                if not valid_dates.empty:
                    closest_start_date = valid_dates.max()
                    start_row = df[df['日期'] == closest_start_date.strftime('%Y-%m-%d')]

            if end_row.empty:
                valid_dates = all_dates[all_dates >= end_date_dt]
                if not valid_dates.empty:
                    closest_end_date = valid_dates.min()
                    end_row = df[df['日期'] == closest_end_date.strftime('%Y-%m-%d')]

        if start_row.empty or end_row.empty:
            return None

        # 获取价格数据
        start_price = start_row['开盘'].values[0]
        end_price = end_row['收盘'].values[0]

        # 检查价格数据是否有效
        if pd.isna(start_price) or pd.isna(end_price) or start_price <= 0 or end_price <= 0:
            return None

        # 计算涨跌幅
        period_change = ((end_price / start_price) - 1) * 100

        # 如果使用了调整后的日期，输出调试信息
        if valid_start_date != start_date_yyyymmdd or valid_end_date != end_date_yyyymmdd:
            print(
                f"股票 {stock_code} 日期范围调整: {start_date_yyyymmdd}->{valid_start_date}, {end_date_yyyymmdd}->{valid_end_date}")

        return period_change

    except Exception as e:
        print(
            f"计算股票 {stock_code} ({stock_name}) 在 {start_date_yyyymmdd} 至 {end_date_yyyymmdd} 的涨跌幅时出错: {e}")
        return None


def build_comment_text(details):
    """
    构建备注文本

    Args:
        details: 详细信息字典

    Returns:
        str: 备注文本
    """
    comment_text = ""

    if '连板信息' in details:
        comment_text += f"{details['连板信息']} "
    if '首次涨停时间' in details:
        comment_text += f"\n首次涨停: {details['首次涨停时间']} "
    if '最终涨停时间' in details:
        comment_text += f"\n最终涨停: {details['最终涨停时间']} "
    if '涨停开板次数' in details:
        comment_text += f"\n开板次数: {details['涨停开板次数']}"

    return comment_text


def format_header_cell(ws, row, col, value, is_center=True, is_bold=True, font_size=None):
    """
    格式化表头单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        value: 单元格值
        is_center: 是否居中对齐
        is_bold: 是否加粗
        font_size: 字体大小，None表示使用默认大小
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BORDER_STYLE

    # 设置对齐方式
    if is_center:
        cell.alignment = Alignment(horizontal='center')

    # 设置字体
    font_args = {'bold': is_bold}
    if font_size:
        font_args['size'] = font_size
    cell.font = Font(**font_args)

    return cell


def format_stock_code_cell(ws, row, col, stock_code):
    """
    格式化股票代码单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        stock_code: 股票代码
    """
    # 使用Excel单元格格式设置为文本，而不是添加单引号前缀
    code_cell = ws.cell(row=row, column=col, value=f'{stock_code.split(".")[0]}')
    code_cell.alignment = Alignment(horizontal='center')
    code_cell.border = BORDER_STYLE
    code_cell.font = Font(size=8)  # 设置比正常小的字体
    code_cell.number_format = '@'  # 设置单元格格式为文本，保留前导零

    return code_cell


def format_concept_cell(ws, row, col, concept, stock_key, stock_reason_group, reason_colors):
    """
    格式化概念单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        concept: 概念文本
        stock_key: 股票键名
        stock_reason_group: 股票概念组映射
        reason_colors: 概念颜色映射
    """
    # 设置概念列
    concept_cell = ws.cell(row=row, column=col, value=f"[{concept}]")
    concept_cell.alignment = Alignment(horizontal='left')
    concept_cell.border = BORDER_STYLE
    concept_cell.font = Font(size=9)  # 设置小一号字体

    # 根据股票所属概念组设置颜色
    if stock_key in stock_reason_group:
        reason = stock_reason_group[stock_key]
        if reason in reason_colors:
            concept_cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")
            # 如果背景色较深，使用白色字体
            if reason_colors[reason] in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
                concept_cell.font = Font(color="FFFFFF", size=9)  # 保持小一号字体

    return concept_cell


def format_stock_name_cell(ws, row, col, stock_name, market_type, max_board_level, apply_high_board_color,
                           stock_code, stock_entry_count, end_date_yyyymmdd=None):
    """
    格式化股票名称单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        stock_name: 股票名称
        market_type: 市场类型标记
        max_board_level: 最高板数
        apply_high_board_color: 是否应用高板颜色
        stock_code: 股票代码
        stock_entry_count: 股票入选次数映射
        end_date_yyyymmdd: 结束日期，用于计算均线斜率标记
    """
    global _top_attention_stocks_cache

    # 添加均线斜率标记
    ma_slope_indicator = ''
    if end_date_yyyymmdd:
        ma_slope_indicator = get_ma_slope_indicator(stock_code, end_date_yyyymmdd)

    # 设置股票简称列，添加市场标记和均线斜率标记
    stock_display_name = f"{stock_name}{market_type}{ma_slope_indicator}"
    name_cell = ws.cell(row=row, column=col, value=stock_display_name)
    name_cell.alignment = Alignment(horizontal='left')
    name_cell.border = BORDER_STYLE

    # 判断是否需要加粗（在关注度榜前十）
    should_bold = False
    # 提取纯股票代码（去除市场后缀）用于匹配
    pure_stock_code = stock_code.split('.')[0] if '.' in stock_code else stock_code
    if _top_attention_stocks_cache and pure_stock_code in _top_attention_stocks_cache:
        should_bold = True

    # 为曾经到过4板及以上的个股，设置蓝色背景
    if max_board_level >= 4:
        # 找出最接近的颜色档位
        color_level = 4
        for level in sorted(HIGH_BOARD_COLORS.keys()):
            if max_board_level >= level:
                color_level = level

        # 设置背景色
        bg_color = HIGH_BOARD_COLORS[color_level]
        name_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
        apply_high_board_color = True

        # 对于深色背景，使用白色字体
        if color_level >= 12:
            name_cell.font = Font(color="FFFFFF", bold=should_bold)
        else:
            name_cell.font = Font(bold=should_bold)
    # 如果没有应用高板数颜色，且该股票是重复入选，则应用灰色背景
    elif stock_code in stock_entry_count and stock_entry_count[stock_code] > 1:
        # 获取入选次数
        entry_count = stock_entry_count[stock_code]
        # 确定颜色深度，超过4次使用最深的灰色
        color_level = min(entry_count, 4)
        bg_color = REENTRY_COLORS.get(color_level, REENTRY_COLORS[4])
        name_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")

        # 对于深色灰色背景，使用白色字体
        if color_level >= 4:
            name_cell.font = Font(color="FFFFFF", bold=should_bold)
        else:
            name_cell.font = Font(bold=should_bold)
    else:
        # 普通情况，只在需要时设置加粗
        if should_bold:
            name_cell.font = Font(bold=True)

    return name_cell, apply_high_board_color


def format_period_change_cell(ws, row, col, stock_code, stock_name, entry_date, period_days,
                              period_days_long=PERIOD_DAYS_LONG, end_date_str=None):
    """
    格式化周期涨跌幅单元格
    
    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        stock_code: 股票代码
        stock_name: 股票名称
        entry_date: 入选日期
        period_days: 短周期天数
        period_days_long: 长周期天数，默认为PERIOD_DAYS_LONG
        end_date_str: 结束日期字符串，如果为None则使用当前日期
    """
    try:
        # 获取结束日期字符串
        if end_date_str is None:
            # 如果未提供结束日期，则使用当前日期（影响后面计算）
            end_date_str = datetime.now().strftime('%Y%m%d')

        # 获取结束日期前短周期交易日的日期
        prev_date_short = get_n_trading_days_before(end_date_str, period_days)
        if '-' in prev_date_short:
            prev_date_short = prev_date_short.replace('-', '')

        # 获取结束日期前长周期交易日的日期
        prev_date_long = get_n_trading_days_before(end_date_str, period_days_long)
        if '-' in prev_date_long:
            prev_date_long = prev_date_long.replace('-', '')

            # 计算短周期涨跌幅
        period_change_short = calculate_stock_period_change(stock_code, prev_date_short, end_date_str, stock_name)

        # 计算长周期涨跌幅
        period_change_long = calculate_stock_period_change(stock_code, prev_date_long, end_date_str, stock_name)

        # 计算入选日前period_days交易日的涨跌幅（用于备注）
        entry_date_str = entry_date.strftime('%Y%m%d')
        prev_entry_date = get_n_trading_days_before(entry_date_str, period_days)
        if '-' in prev_entry_date:
            prev_entry_date = prev_entry_date.replace('-', '')
        entry_period_change = calculate_stock_period_change(stock_code, prev_entry_date, entry_date_str, stock_name)

        # 设置单元格值和格式
        if period_change_short is not None and period_change_long is not None:
            # 显示格式：短周期涨幅/长周期涨幅，保留1位小数
            period_cell = ws.cell(row=row, column=col, value=f"{period_change_short:.1f}/{period_change_long:.1f}")

            # 设置背景色 - 根据长周期涨跌幅
            color = get_color_for_period_change(period_change_long)
            if color:
                period_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 添加入选日涨跌幅信息作为备注
            if entry_period_change is not None:
                comment_text = f"入选日前{period_days}日涨幅: {entry_period_change:.2f}%"
                period_cell.comment = Comment(comment_text, "入选日涨跌幅")
        else:
            period_cell = ws.cell(row=row, column=col, value="--/--")
    except Exception as e:
        print(f"计算周期涨跌幅时出错: {e}, 股票: {stock_name}")
        period_cell = ws.cell(row=row, column=col, value="--/--")

    # 设置周期涨跌幅单元格格式
    period_cell.alignment = Alignment(horizontal='center')
    period_cell.border = BORDER_STYLE
    period_cell.font = Font(size=9)  # 设置小一号字体

    return period_cell


def format_board_cell(ws, row, col, board_days, pure_stock_code, stock_detail_key, stock_details, current_date_obj):
    """
    格式化连板单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        board_days: 连板天数
        pure_stock_code: 纯股票代码
        stock_detail_key: 股票详细信息键名
        stock_details: 股票详细信息映射
        current_date_obj: 当前日期对象

    Returns:
        tuple: (单元格对象, 最后连板日期)
    """
    cell = ws.cell(row=row, column=col)

    # 对于首板（board_days=1），显示为"首板"，否则显示为"N板"
    if board_days == 1:
        # 获取市场标记
        market_marker = get_market_marker(pure_stock_code)
        board_text = f"首板{market_marker}"
    else:
        board_text = f"{int(board_days)}板"

    # 获取当前日期的YYYYMMDD格式
    date_yyyymmdd = current_date_obj.strftime('%Y%m%d')

    # 添加成交量比信息
    board_text = add_volume_ratio_to_text(board_text, pure_stock_code, date_yyyymmdd)

    cell.value = board_text
    # 设置连板颜色
    board_level = int(board_days)
    color = BOARD_COLORS.get(board_level, BOARD_COLORS[10] if board_level > 10 else BOARD_COLORS[1])
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    # 文字颜色设为白色以增强可读性
    cell.font = Font(color="FFFFFF", bold=True)
    # 添加边框样式
    cell.border = BORDER_STYLE

    # 添加备注信息（仅对连板股票）
    if stock_detail_key in stock_details:
        details = stock_details[stock_detail_key]
        comment_text = build_comment_text(details)
        if comment_text:
            cell.comment = Comment(comment_text.strip(), "涨停信息")

    # 记录当前日期为最后一次连板的日期
    last_board_date = current_date_obj

    return cell, last_board_date


def format_shouban_cell(ws, row, col, pure_stock_code, current_date_obj):
    """
    格式化首板单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        pure_stock_code: 纯股票代码
        current_date_obj: 当前日期对象

    Returns:
        tuple: (单元格对象, 最后连板日期)
    """
    cell = ws.cell(row=row, column=col)

    # 获取市场标记
    market_marker = get_market_marker(pure_stock_code)
    # 显示为"首板"并使用特殊颜色
    board_text = f"首板{market_marker}"

    # 获取当前日期的YYYYMMDD格式
    date_yyyymmdd = current_date_obj.strftime('%Y%m%d')

    # 添加成交量比信息
    board_text = add_volume_ratio_to_text(board_text, pure_stock_code, date_yyyymmdd)

    cell.value = board_text
    cell.fill = PatternFill(start_color=BOARD_COLORS[1], fill_type="solid")
    # 文字颜色设为白色以增强可读性
    cell.font = Font(color="FFFFFF", bold=True)
    # 添加边框样式
    cell.border = BORDER_STYLE

    # 更新最后连板日期以便可以继续跟踪
    last_board_date = current_date_obj

    return cell, last_board_date


def format_momo_entry_cell(ws, row, col, pure_stock_code, current_date_obj, stock):
    """
    格式化【默默上涨】入选日期单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        pure_stock_code: 纯股票代码
        current_date_obj: 当前日期对象
        stock: 股票数据

    Returns:
        单元格对象
    """
    cell = ws.cell(row=row, column=col)

    # 获取当日涨跌幅
    date_yyyymmdd = current_date_obj.strftime('%Y%m%d')
    pct_change = get_stock_daily_pct_change(pure_stock_code, date_yyyymmdd)

    if pd.notna(pct_change):
        # 显示正常的涨跌幅
        cell_value = f"{pct_change:.2f}%"

        # 添加成交量比信息
        cell_value = add_volume_ratio_to_text(cell_value, pure_stock_code, date_yyyymmdd)

        cell.value = cell_value

        # 设置背景色 - 根据涨跌幅正常上色
        color = get_color_for_pct_change(pct_change)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # 设置深紫色字体并加粗
        cell.font = Font(color="663399", bold=True)  # 深紫色字体
    else:
        cell.value = "停牌"
        # 停牌时也使用深紫色字体并加粗
        cell.font = Font(color="663399", bold=True)

    # 设置单元格格式
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER_STYLE

    return cell


def format_daily_pct_change_cell(ws, row, col, current_date_obj, stock_code):
    """
    格式化日涨跌幅单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        current_date_obj: 当前日期对象
        stock_code: 股票代码

    Returns:
        单元格对象
    """
    cell = ws.cell(row=row, column=col)

    # 获取当日涨跌幅
    date_yyyymmdd = current_date_obj.strftime('%Y%m%d')
    pct_change = get_stock_daily_pct_change(stock_code, date_yyyymmdd)
    if pd.notna(pct_change):
        # 基本涨跌幅显示
        cell_value = f"{pct_change:.2f}%"

        # 添加成交量比信息
        cell_value = add_volume_ratio_to_text(cell_value, stock_code, date_yyyymmdd)

        cell.value = cell_value

        # 设置背景色 - 根据涨跌幅
        color = get_color_for_pct_change(pct_change)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    else:
        cell.value = "停牌"

    # 设置单元格格式
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER_STYLE

    return cell


def process_daily_cell(ws, row_idx, col_idx, formatted_day, board_days, found_in_shouban,
                       pure_stock_code, stock_details, stock, date_mapping, max_tracking_days,
                       max_tracking_days_before, zaban_df, period_days=PERIOD_DAYS_CHANGE, new_high_markers=None):
    """
    处理每日单元格

    Args:
        ws: Excel工作表
        row_idx: 行索引
        col_idx: 列索引
        formatted_day: 格式化的日期
        board_days: 连板天数
        found_in_shouban: 是否在首板数据中找到
        pure_stock_code: 纯股票代码
        stock_details: 股票详细信息映射
        stock: 股票数据
        date_mapping: 日期映射
        max_tracking_days: 断板后跟踪的最大天数
        max_tracking_days_before: 入选前跟踪的最大天数
        zaban_df: 炸板数据DataFrame
        period_days: 计算涨跌幅的周期天数
        new_high_markers: 新高标记映射

    Returns:
        更新后的股票数据
    """
    cell = ws.cell(row=row_idx, column=col_idx)

    # 解析当前日期对象
    current_date_obj = datetime.strptime(formatted_day, '%Y年%m月%d日') if '年' in formatted_day else datetime.strptime(
        formatted_day, '%Y/%m/%d')

    # 检查是否为炸板股票
    is_zaban = check_stock_in_zaban(zaban_df, pure_stock_code, formatted_day)

    # 缓存炸板格式信息，供B sheet使用
    cache_zaban_format(pure_stock_code, formatted_day, is_zaban)

    # 检查是否为【默默上涨】数据
    is_momo_shangzhang = stock.get('entry_type') == 'momo_shangzhang'

    # 处理【默默上涨】数据 - 显示涨跌幅而不是连板信息
    if is_momo_shangzhang:
        # 【默默上涨】数据在入选日期显示特殊标记，其他日期显示涨跌幅
        if current_date_obj == stock['first_significant_date']:
            # 在入选日期显示【默默上涨】标记
            format_momo_entry_cell(ws, row_idx, col_idx, pure_stock_code, current_date_obj, stock)
        else:
            # 其他日期显示涨跌幅
            format_daily_pct_change_cell(ws, row_idx, col_idx, current_date_obj, stock['stock_code'])
    # 处理连板数据
    elif pd.notna(board_days) and board_days:
        stock_detail_key = f"{pure_stock_code}_{formatted_day}"
        _, last_board_date = format_board_cell(ws, row_idx, col_idx, board_days, pure_stock_code,
                                               stock_detail_key, stock_details, current_date_obj)
        stock['last_board_date'] = last_board_date
    # 处理首板数据
    elif found_in_shouban:
        _, last_board_date = format_shouban_cell(ws, row_idx, col_idx, pure_stock_code, current_date_obj)
        stock['last_board_date'] = last_board_date
    # 处理其他情况
    else:
        try:
            # 检查当日日期是否在首次显著连板日期之后
            if current_date_obj >= stock['first_significant_date']:
                # 处理断板后的跟踪
                if should_track_after_break(stock, current_date_obj, max_tracking_days, period_days):
                    format_daily_pct_change_cell(ws, row_idx, col_idx, current_date_obj, stock['stock_code'])
            # 检查当日日期是否在首次显著连板日期之前，且在跟踪天数范围内
            elif should_track_before_entry(current_date_obj, stock['first_significant_date'], max_tracking_days_before):
                format_daily_pct_change_cell(ws, row_idx, col_idx, current_date_obj, stock['stock_code'])
        except Exception as e:
            print(f"处理日期时出错: {e}, 日期: {formatted_day}")

    # 如果是炸板股票，添加下划线标记
    if is_zaban:
        add_zaban_underline(cell)

    # 检查是否需要添加新高标记
    if new_high_markers and stock.get('stock_code') in new_high_markers:
        marked_date = new_high_markers[stock['stock_code']]
        if formatted_day == marked_date:
            # 在当前单元格内容后添加新高标记
            current_value = cell.value if cell.value else ""
            cell.value = f"{current_value}{NEW_HIGH_MARKER}"

    return stock


def should_track_after_break(stock, current_date_obj, max_tracking_days, period_days=PERIOD_DAYS_CHANGE):
    """
    判断是否应该跟踪断板后的股票（包装函数）
    """
    return _should_track_after_break(
        stock, current_date_obj, max_tracking_days, period_days,
        calculate_stock_period_change
    )


def should_collapse_row(stock, formatted_trading_days, date_mapping):
    """
    判断是否应该折叠此行（包装函数）
    """
    return _should_collapse_row(stock, formatted_trading_days, date_mapping, COLLAPSE_DAYS_AFTER_BREAK)

    # 注意：以下函数已移动到 ladder_chart_helpers.py：
    # calculate_ma_slope, get_ma_slope_indicator, clear_ma_slope_cache, print_slope_statistics
    # should_track_before_entry, calculate_last_board_date, check_stock_in_zaban


def add_zaban_underline(cell):
    """
    为炸板股票的单元格添加下划线（更显眼）：
    - 保持原有字体样式，添加单下划线
    - 在单元格底部添加深灰色双线边框，增强可见性
    """
    # 保持原有的字体设置，只添加下划线
    current_font = cell.font
    cell.font = Font(
        color=current_font.color,
        bold=current_font.bold,
        size=current_font.size,
        underline='single'
    )

    # 在不破坏原有边框的前提下，增强底边样式
    existing_border = cell.border
    new_border = Border(
        left=existing_border.left,
        right=existing_border.right,
        top=existing_border.top,
        bottom=Side(style='double', color="000000")
    )
    cell.border = new_border


def setup_excel_header(ws, formatted_trading_days, show_period_change, period_days, date_column_start,
                       show_warning_column=True):
    """
    设置Excel表头

    Args:
        ws: Excel工作表
        formatted_trading_days: 格式化的交易日列表
        show_period_change: 是否显示周期涨跌幅
        period_days: 周期天数
        date_column_start: 日期列开始位置
        show_warning_column: 是否显示异动预警列

    Returns:
        dict: 日期到列索引的映射
    """
    # 设置日期表头（第1行）
    format_header_cell(ws, 1, 1, "股票代码")
    format_header_cell(ws, 1, 2, "题材概念")
    format_header_cell(ws, 1, 3, "股票简称")

    # 添加周期涨跌幅列（第4列）
    if show_period_change:
        period_header = f"近{period_days}日\n近{PERIOD_DAYS_LONG}日"
        header_cell = format_header_cell(ws, 1, 4, period_header, font_size=9)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 设置日期列标题
    date_columns = {}  # 用于保存日期到列索引的映射
    for i, formatted_day in enumerate(formatted_trading_days):
        date_obj = datetime.strptime(formatted_day, '%Y年%m月%d日')
        col = i + date_column_start  # 日期列的起始位置根据是否显示周期涨跌幅决定
        date_columns[formatted_day] = col

        # 添加日期标题和星期几
        weekday = date_obj.weekday()
        weekday_map = {0: "星期一", 1: "星期二", 2: "星期三", 3: "星期四", 4: "星期五", 5: "星期六", 6: "星期日"}
        date_with_weekday = f"{date_obj.strftime('%Y-%m-%d')}\n{weekday_map[weekday]}"

        # 设置日期单元格样式：居中、自动换行、边框、字体加粗
        date_cell = ws.cell(row=1, column=col, value=date_with_weekday)
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        date_cell.border = BORDER_STYLE
        date_cell.font = Font(bold=True)

    # 添加异动预警列（在最后一个日期列之后）
    if show_warning_column and formatted_trading_days:
        warning_col = len(formatted_trading_days) + date_column_start
        warning_header = "异动预警"
        warning_cell = format_header_cell(ws, 1, warning_col, warning_header, font_size=10)
        warning_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 设置预警列的背景色为浅黄色
        warning_cell.fill = PatternFill(start_color="FFFACD", fill_type="solid")

    return date_columns


def prepare_workbook(output_file, sheet_name, start_date):
    """
    准备Excel工作簿，根据参数决定是创建新工作簿还是使用现有工作簿

    Args:
        output_file: 输出文件路径
        sheet_name: 工作表名称，如果为None则使用默认名称
        start_date: 开始日期，用于生成默认工作表名称

    Returns:
        tuple: (工作簿, 工作表, 工作表名称)
    """
    # 确定工作表名称和类型
    if sheet_name is None:
        sheet_name = f"涨停梯队{start_date[:6]}"
        is_default_pattern = True
    else:
        is_default_pattern = "涨停梯队" in sheet_name

    # 创建新的工作簿（如果不存在）或加载现有工作簿
    if not os.path.exists(output_file):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        print(f"创建新的工作簿和工作表: {sheet_name}")
        return wb, ws, sheet_name

    # 尝试加载现有工作簿
    wb = load_workbook(output_file)

    # 处理已存在的工作表
    sheet_exists = sheet_name in wb.sheetnames

    # 情况1: 用户自定义工作表已存在 - 保留原样
    if sheet_exists and not is_default_pattern:
        print(f"保留用户自定义工作表: {sheet_name}")
        return wb, None, sheet_name

    # 情况2: 默认模式工作表已存在 - 更新它
    if sheet_exists and is_default_pattern:
        wb.remove(wb[sheet_name])
        ws = wb.create_sheet(title=sheet_name)
        print(f"已更新默认模式工作表: {sheet_name}")
        return wb, ws, sheet_name

    # 情况3: 工作表不存在 - 创建新的
    ws = wb.create_sheet(title=sheet_name)
    print(f"在现有工作簿中创建新工作表: {sheet_name}")
    return wb, ws, sheet_name


def prepare_stock_data(result_df, priority_reasons, low_priority_reasons=None):
    """
    准备股票数据，包括概念收集、颜色映射等

    Args:
        result_df: 显著连板股票DataFrame
        priority_reasons: 优先选择的原因列表
        low_priority_reasons: 低优先级原因列表，默认为None

    Returns:
        dict: 股票数据字典
    """
    # 收集所有概念信息，用于生成颜色映射
    all_concepts, stock_concepts = collect_stock_concepts(result_df)

    # 获取热门概念的颜色映射
    reason_colors, top_reasons = get_reason_colors(all_concepts, priority_reasons=priority_reasons,
                                                   low_priority_reasons=low_priority_reasons)

    # 为每只股票确定主要概念组
    all_stocks = prepare_stock_concept_data(stock_concepts)

    # 获取每只股票的主要概念组
    stock_reason_group = get_stock_reason_group(all_stocks, top_reasons, low_priority_reasons=low_priority_reasons)

    # 创建理由计数器
    reason_counter = Counter(all_concepts)

    # 返回股票数据
    return {
        'reason_colors': reason_colors,
        'top_reasons': top_reasons,
        'stock_reason_group': stock_reason_group,
        'reason_counter': reason_counter
    }


def collect_stock_concepts(result_df):
    """
    收集股票概念信息

    Args:
        result_df: 显著连板股票DataFrame

    Returns:
        tuple: (所有概念列表, 股票概念详情映射)
               股票概念详情映射格式：{stock_key: (reasons, reason_details)}
               reason_details: [(normalized, match_type, original), ...]
    """
    all_concepts = []
    stock_concepts = {}

    # 从result_df中收集所有概念
    for _, row in result_df.iterrows():
        concept = row.get('concept', '其他')
        if pd.isna(concept) or not concept:
            concept = "其他"

        # 提取概念中的原因（包含匹配类型信息）
        from utils.theme_color_util import extract_reasons_with_match_type
        reason_details = extract_reasons_with_match_type(concept)

        if reason_details:
            # 提取规范化后的概念列表
            reasons = [normalized for normalized, _, _ in reason_details]
            all_concepts.extend(reasons)

            # 存储完整信息：(reasons, reason_details)
            stock_concepts[f"{row['stock_code']}_{row['stock_name']}"] = (reasons, reason_details)

    return all_concepts, stock_concepts


def prepare_stock_concept_data(stock_concepts):
    """
    准备股票概念数据

    Args:
        stock_concepts: 股票概念详情映射
                       格式：{stock_key: (reasons, reason_details)} 或 {stock_key: reasons}（向后兼容）

    Returns:
        dict: 股票概念数据
    """
    all_stocks = {}
    for stock_key, concept_data in stock_concepts.items():
        # 兼容新旧格式
        if isinstance(concept_data, tuple):
            # 新格式：(reasons, reason_details)
            reasons, reason_details = concept_data
            all_stocks[stock_key] = {
                'name': stock_key.split('_')[1],
                'reasons': reasons,
                'reason_details': reason_details,  # 新增：包含匹配类型
                'appearances': [1]
            }
        else:
            # 旧格式：reasons（向后兼容）
            reasons = concept_data
            all_stocks[stock_key] = {
                'name': stock_key.split('_')[1],
                'reasons': reasons,
                'appearances': [1]
            }

    return all_stocks


def _get_leader_sheet_info(wb, last_trading_day_in_run):
    """
    获取所有龙头工作表及其对应的日期
    """
    leader_sheets_info = []
    current_year = last_trading_day_in_run.year
    # 用于处理跨年时的情况
    current_month = last_trading_day_in_run.month

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("龙头"):
            match = re.search(r'龙头(\d{4})', sheet_name)
            if not match:
                continue

            sheet_mmdd = match.group(1)
            sheet_month = int(sheet_mmdd[:2])

            # 跨年处理：如果当前月份是早期（如1-3月），而sheet的月份是晚期（如10-12月），
            # 则认为该sheet属于上一年
            year_to_use = current_year
            if current_month <= 3 and sheet_month >= 10:
                year_to_use = current_year - 1

            try:
                sheet_date = datetime.strptime(f"{year_to_use}{sheet_mmdd}", "%Y%m%d")
                leader_sheets_info.append({'name': sheet_name, 'date': sheet_date})
            except ValueError:
                print(f"警告: 无法从工作表名 '{sheet_name}' 解析有效日期")
    return leader_sheets_info


def copy_worksheet(source_ws, target_ws):
    """
    复制工作表的内容和样式
    
    Args:
        source_ws: 源工作表
        target_ws: 目标工作表
    """
    # 复制列宽
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    # 复制行高
    for row_idx, dim in source_ws.row_dimensions.items():
        if dim.height:
            target_ws.row_dimensions[row_idx].height = dim.height

    # 复制单元格内容和样式
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]

            # 复制值
            if cell.value is not None:
                target_cell.value = cell.value

            # 复制样式
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection.copy()
                target_cell.alignment = cell.alignment.copy()

            # 复制备注
            if cell.comment:
                target_cell.comment = cell.comment

    # 复制合并单元格
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # 不复制冻结窗格 - 避免因回填数据导致的冻结位置错乱
    # 归档文件主要用于查看历史数据，不需要冻结窗格
    # if source_ws.freeze_panes:
    #     target_ws.freeze_panes = source_ws.freeze_panes


def archive_leader_sheets(wb, sheets_to_archive, output_file):
    """
    将要删除的龙头工作表归档到单独的Excel文件中
    
    Args:
        wb: 当前工作簿
        sheets_to_archive: 要归档的工作表信息列表 [{'name': 'sheet_name', 'date': datetime_obj}, ...]
        output_file: 当前输出文件路径
    """
    if not sheets_to_archive:
        return

    # 生成归档文件路径
    output_dir = os.path.dirname(output_file)
    base_name = os.path.splitext(os.path.basename(output_file))[0]
    archive_file = os.path.join(output_dir, f"{base_name}_龙头归档.xlsx")

    print(f"开始归档 {len(sheets_to_archive)} 个龙头工作表到: {archive_file}")

    # 加载或创建归档工作簿
    if os.path.exists(archive_file):
        try:
            archive_wb = load_workbook(archive_file)
            print(f"加载已有归档文件: {archive_file}")
        except Exception as e:
            print(f"加载归档文件失败: {e}，创建新的归档文件")
            archive_wb = Workbook()
            # 删除默认的Sheet
            if 'Sheet' in archive_wb.sheetnames:
                archive_wb.remove(archive_wb['Sheet'])
    else:
        archive_wb = Workbook()
        # 删除默认的Sheet
        if 'Sheet' in archive_wb.sheetnames:
            archive_wb.remove(archive_wb['Sheet'])
        print(f"创建新的归档文件: {archive_file}")

    # 复制要归档的工作表
    archived_count = 0
    for sheet_info in sheets_to_archive:
        sheet_name = sheet_info['name']

        # 如果归档文件中已存在同名sheet，跳过（避免重复归档）
        if sheet_name in archive_wb.sheetnames:
            print(f"  归档文件中已存在工作表 {sheet_name}，跳过")
            continue

        try:
            # 复制工作表
            source_ws = wb[sheet_name]
            target_ws = archive_wb.create_sheet(title=sheet_name)

            # 复制内容和样式
            copy_worksheet(source_ws, target_ws)

            archived_count += 1
            print(f"  已归档工作表: {sheet_name} (日期: {sheet_info['date'].strftime('%Y-%m-%d')})")
        except Exception as e:
            print(f"  归档工作表 {sheet_name} 时出错: {e}")

    # 保存归档文件
    if archived_count > 0:
        try:
            archive_wb.save(archive_file)
            print(f"成功归档 {archived_count} 个工作表到: {archive_file}")
        except Exception as e:
            print(f"保存归档文件时出错: {e}")
    else:
        print("没有新的工作表需要归档")


def manage_leader_sheets(wb, last_trading_day_in_run, output_file=OUTPUT_FILE):
    """
    管理龙头工作表数量，删除超过数量限制的最旧的工作表
    
    Args:
        wb: 工作簿对象
        last_trading_day_in_run: 本次运行的最后交易日
        output_file: 输出文件路径，用于生成归档文件名
    """
    leader_sheets_info = _get_leader_sheet_info(wb, last_trading_day_in_run)

    if len(leader_sheets_info) > MAX_LEADER_SHEETS:
        # 按日期排序，旧的在前
        leader_sheets_info.sort(key=lambda x: x['date'])
        # 确定要删除的工作表
        sheets_to_delete = leader_sheets_info[:len(leader_sheets_info) - MAX_LEADER_SHEETS]

        # 在删除之前，先归档这些工作表
        archive_leader_sheets(wb, sheets_to_delete, output_file)

        # 删除工作表
        for sheet_info in sheets_to_delete:
            if sheet_info['name'] in wb.sheetnames:
                wb.remove(wb[sheet_info['name']])
                print(f"已删除旧的龙头工作表: {sheet_info['name']}")


def backfill_historical_leader_sheets(wb, last_trading_day_in_run, all_formatted_trading_days, date_mapping):
    """
    回填历史龙头股工作表的最新数据
    """
    leader_sheets_info = _get_leader_sheet_info(wb, last_trading_day_in_run)

    if not leader_sheets_info:
        return

    # 按日期倒序，最新的在前
    leader_sheets_info.sort(key=lambda x: x['date'], reverse=True)

    # 当前运行的所有交易日 (datetime 对象)
    run_dates = [datetime.strptime(date_mapping[d], '%Y%m%d') for d in all_formatted_trading_days]

    # 遍历历史龙头工作表进行回填 (跳过最新的一个，即当前运行生成的)
    for sheet_info in leader_sheets_info[1:MAX_LEADER_SHEETS]:
        ws = wb[sheet_info['name']]
        sheet_date = sheet_info['date']
        print(f"检查历史龙头工作表: {sheet_info['name']} (日期: {sheet_date.strftime('%Y-%m-%d')})")

        # 找出该工作表现有的所有日期列
        existing_dates_in_sheet = set()
        for col_idx in range(4, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val and isinstance(header_val, str) and '\n' in header_val:
                date_part = header_val.split('\n')[0]
                try:
                    existing_dates_in_sheet.add(datetime.strptime(date_part, '%Y-%m-%d').date())
                except ValueError:
                    continue

        # 找出当前运行中、比该工作表最后日期新、且尚未存在于该工作表中的日期
        dates_to_add = [d for d in run_dates if d > sheet_date and d.date() not in existing_dates_in_sheet]
        dates_to_add.sort()  # 按时间顺序添加

        if not dates_to_add:
            print(f"工作表 {sheet_info['name']} 数据已是最新，无需回填。")
            continue

        print(f"为 {sheet_info['name']} 回填 {len(dates_to_add)} 天的数据...")

        # 在表头添加新日期列
        start_col = ws.max_column + 1
        for i, date_obj in enumerate(dates_to_add):
            col_idx = start_col + i
            weekday = date_obj.weekday()
            weekday_map = {0: "星期一", 1: "星期二", 2: "星期三", 3: "星期四", 4: "星期五", 5: "星期六", 6: "星期日"}
            date_with_weekday = f"{date_obj.strftime('%Y-%m-%d')}\n{weekday_map[weekday]}"

            date_cell = ws.cell(row=1, column=col_idx, value=date_with_weekday)
            date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            date_cell.border = BORDER_STYLE
            date_cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        # 填充新日期列的数据
        for row_idx in range(4, ws.max_row + 1):  # 数据从第4行开始
            stock_code = ws.cell(row=row_idx, column=1).value
            if not stock_code or not isinstance(stock_code, (str, int)):
                continue

            for i, date_obj in enumerate(dates_to_add):
                col_idx = start_col + i
                format_daily_pct_change_cell(ws, row_idx, col_idx, date_obj, str(stock_code))


@timer
def build_ladder_chart(start_date, end_date, output_file=OUTPUT_FILE, min_board_level=2,
                       max_tracking_days=MAX_TRACKING_DAYS_AFTER_BREAK, reentry_days=REENTRY_DAYS_THRESHOLD,
                       non_main_board_level=1, max_tracking_days_before=MAX_TRACKING_DAYS_BEFORE_ENTRY,
                       period_days=PERIOD_DAYS_CHANGE, period_days_long=PERIOD_DAYS_LONG, show_period_change=False,
                       priority_reasons=None, low_priority_reasons=None, enable_attention_criteria=False,
                       sheet_name=None,
                       create_leader_sheet=False, enable_momo_shangzhang=True, create_volume_sheet=False,
                       enable_reason_analysis=None):
    """
    构建梯队形态的涨停复盘图

    Args:
        start_date: 开始日期 (YYYYMMDD)
        end_date: 结束日期 (YYYYMMDD)，如果为None则使用当前日期
        output_file: 输出文件路径
        min_board_level: 主板股票最小显著连板天数，默认为2
        max_tracking_days: 断板后跟踪的最大天数，默认取全局配置
        reentry_days: 断板后再次上榜的天数阈值，默认取全局配置
        non_main_board_level: 非主板股票最小显著连板天数，默认为1
        max_tracking_days_before: 入选前跟踪的最大天数，默认取全局配置
        period_days: 计算最近X个交易日的短周期涨跌幅，默认取全局配置
        period_days_long: 计算最近X个交易日的长周期涨跌幅，默认取全局配置
        show_period_change: 是否显示周期涨跌幅列，默认取全局配置
        priority_reasons: 优先选择的原因列表，默认为None
        low_priority_reasons: 低优先级原因列表，默认为None
        enable_attention_criteria: 是否启用关注度榜入选条件，默认为False
        sheet_name: 工作表名称，默认为None，表示使用"涨停梯队{start_date[:6]}"；如果指定，则使用指定的名称
        create_leader_sheet: 是否创建龙头股工作表，默认为False
        enable_momo_shangzhang: 是否启用【默默上涨】数据，默认为False
        create_volume_sheet: 是否创建成交量涨跌幅分析工作表，默认为False
        enable_reason_analysis: 是否启用原因分析，None表示使用默认配置，True强制启用，False强制禁用
    """
    # 清除缓存
    get_stock_data.cache_clear()
    # 清理高涨幅计算缓存
    clear_high_gain_cache()
    # 清理均线斜率计算缓存
    clear_ma_slope_cache()

    # 设置结束日期，如果未指定则使用当前日期
    if end_date is None:
        end_date = datetime.now().strftime('%Y%m%d')
        print(f"未指定结束日期，使用当前日期: {end_date}")

    print(f"开始构建梯队形态涨停复盘图 ({start_date} 至 {end_date})...")

    # 清理缓存，确保每次运行都是全新计算
    clear_caches()

    # 获取并处理交易日数据
    trading_days = get_trading_days(start_date, end_date)
    if trading_days is None or len(trading_days) == 0:
        print("未获取到交易日列表")
        return

    # 格式化交易日列表和映射
    formatted_trading_days, date_mapping = format_trading_days(trading_days)

    # 加载股票数据
    lianban_df, shouban_df, attention_data, zaban_df, momo_df = load_stock_data(start_date, end_date,
                                                                                enable_attention_criteria,
                                                                                enable_momo_shangzhang)
    if lianban_df.empty:
        print("未获取到有效的连板数据")
        return

    # 获取股票详细信息映射
    stock_details = lianban_df.attrs.get('stock_details', {})

    # 加载最近n日关注度前十股票
    get_top_attention_stocks_cached(end_date)

    # 识别显著连板股票（不包含【默默上涨】数据）
    result_df = identify_significant_boards(lianban_df, shouban_df, min_board_level, reentry_days,
                                            non_main_board_level, enable_attention_criteria,
                                            attention_data, priority_reasons, low_priority_reasons, None, start_date,
                                            end_date)
    if result_df.empty:
        print(f"未找到在{start_date}至{end_date}期间有符合条件的显著连板股票")
        return

    # 准备工作簿和工作表
    wb, ws, sheet_name_used = prepare_workbook(output_file, sheet_name, start_date)

    # 如果ws为None，表示用户自定义工作表已存在且应保留
    if ws is None:
        print(f"工作表 {sheet_name_used} 已存在且为用户自定义工作表，保留现有内容且不更新图例")
        return True

    # 准备股票数据（仅在需要更新工作表时）
    stock_data = prepare_stock_data(result_df, priority_reasons, low_priority_reasons)

    # 设置Excel表头和日期列
    period_column = 4
    date_column_start = 4 if not show_period_change else 5
    show_warning_column = True  # 默认显示异动预警列
    date_columns = setup_excel_header(ws, formatted_trading_days, show_period_change, period_days, date_column_start,
                                      show_warning_column)

    # 添加大盘指标行
    add_market_indicators(ws, date_columns, label_col=2)

    # 统计股票入选情况
    stock_entry_count = count_stock_entries(result_df)

    # 填充数据行（【涨停梯队】sheet不启用折叠功能）
    fill_data_rows(ws, result_df, shouban_df, stock_data['stock_reason_group'], stock_data['reason_colors'],
                   stock_entry_count, formatted_trading_days, date_column_start, show_period_change,
                   period_column, period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                   max_tracking_days_before, zaban_df, show_warning_column, enable_collapse=False)

    # 调整列宽
    adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, show_warning_column)

    # 冻结前三列和前三行
    ws.freeze_panes = ws.cell(row=4, column=date_column_start)

    # 创建按概念分组的工作表
    concept_grouped_sheet_name = f"{sheet_name_used}_概念分组"

    # 检查是否需要创建或更新按概念分组的工作表
    is_default_pattern = "涨停梯队" in concept_grouped_sheet_name
    sheet_exists = concept_grouped_sheet_name in wb.sheetnames

    # 决定是否需要创建/更新工作表
    should_create_sheet = False

    if not sheet_exists:
        # 工作表不存在，需要创建
        should_create_sheet = True
        print(f"在现有工作簿中创建新工作表: {concept_grouped_sheet_name}")
    elif is_default_pattern:
        # 默认模式工作表已存在，需要更新
        wb.remove(wb[concept_grouped_sheet_name])
        should_create_sheet = True
        print(f"已更新默认模式工作表: {concept_grouped_sheet_name}")
    else:
        # 用户自定义工作表已存在，保留原样
        print(f"保留用户自定义工作表: {concept_grouped_sheet_name}")

    # 用于存储【概念分组】sheet需要折叠的行索引
    concept_grouped_rows_to_collapse = []

    # 只有当需要创建工作表时才创建内容
    if should_create_sheet:
        concept_ws = wb.create_sheet(title=concept_grouped_sheet_name)

        # 为"涨停梯队 按概念分组"sheet去重，但保留重入标记色，并添加【默默上涨】数据
        grouped_df = result_df.copy() if not result_df.empty else pd.DataFrame()

        # 添加【默默上涨】数据到按概念分组工作表
        if enable_momo_shangzhang and momo_df is not None and not momo_df.empty:
            try:
                from analysis.momo_shangzhang_processor import identify_momo_shangzhang_stocks
                momo_result_df = identify_momo_shangzhang_stocks(momo_df, start_date, end_date)
                if not momo_result_df.empty:
                    grouped_df = pd.concat([grouped_df, momo_result_df], ignore_index=True)
                    print(f"按概念分组工作表合并【默默上涨】数据后总股票数量: {len(grouped_df)}")
            except Exception as e:
                print(f"处理【默默上涨】数据时出错，跳过: {e}")
                # 继续执行，不影响概念分析

        # 去重处理
        if not grouped_df.empty:
            # 检查 is_reentry 列是否存在
            if 'is_reentry' in grouped_df.columns:
                is_reentry_map = grouped_df.groupby('stock_code')['is_reentry'].any()
                grouped_df = grouped_df.drop_duplicates(subset='stock_code', keep='first').copy()
                grouped_df['is_reentry'] = grouped_df['stock_code'].map(is_reentry_map)
            else:
                # 如果 is_reentry 列不存在，直接去重
                grouped_df = grouped_df.drop_duplicates(subset='stock_code', keep='first').copy()

        # 创建【概念分组】sheet并收集需要折叠的行索引
        concept_grouped_rows_to_collapse = create_concept_grouped_sheet_content(
            concept_ws, grouped_df, shouban_df, stock_data,
            stock_entry_count, formatted_trading_days, date_column_start,
            show_period_change, period_column, period_days, period_days_long,
            stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
            zaban_df, enable_collapse=True
        )

    # 创建成交量涨跌幅分析工作表（如果启用）
    if create_volume_sheet:
        volume_sheet_name = f"{concept_grouped_sheet_name}_成交量"

        # 检查是否需要创建成交量工作表
        volume_sheet_exists = volume_sheet_name in wb.sheetnames
        should_create_volume_sheet = False

        if not volume_sheet_exists:
            # 工作表不存在，需要创建
            should_create_volume_sheet = True
            print(f"创建成交量涨跌幅分析工作表: {volume_sheet_name}")
        elif is_default_pattern:
            # 默认模式下，覆盖更新
            wb.remove(wb[volume_sheet_name])
            should_create_volume_sheet = True
            print(f"已更新成交量涨跌幅分析工作表: {volume_sheet_name}")
        else:
            # 用户自定义工作表已存在，保留原样
            print(f"保留用户自定义成交量工作表: {volume_sheet_name}")

        # 创建成交量工作表
        if should_create_volume_sheet:
            from analysis.helper.volume_ladder_chart import create_volume_concept_grouped_sheet

            # 使用与概念分组相同的数据
            volume_grouped_df = grouped_df if 'grouped_df' in locals() else result_df.copy()

            create_volume_concept_grouped_sheet(wb, volume_sheet_name, volume_grouped_df, shouban_df, stock_data,
                                                stock_entry_count, formatted_trading_days, date_column_start,
                                                show_period_change, period_column, period_days, period_days_long,
                                                stock_details, date_mapping, max_tracking_days,
                                                max_tracking_days_before,
                                                zaban_df)

            # 强制同步单元格格式，确保与概念分组表一致
            if 'concept_ws' in locals() and volume_sheet_name in wb.sheetnames:
                volume_ws = wb[volume_sheet_name]
                print(f"同步工作表 '{concept_ws.title}' 和 '{volume_ws.title}' 的单元格格式...")

                # 1. 同步列宽
                for col_letter, dim in concept_ws.column_dimensions.items():
                    volume_ws.column_dimensions[col_letter].width = dim.width

                # 2. 同步行高
                for row_idx, dim in concept_ws.row_dimensions.items():
                    volume_ws.row_dimensions[row_idx].height = dim.height

    # 创建龙头股工作表（如果启用）
    if create_leader_sheet:
        # 从最后一个交易日提取MMDD格式的日期作为sheet名称
        last_trading_day = formatted_trading_days[-1] if formatted_trading_days else ""
        if last_trading_day:
            # 解析日期并格式化为MMDD
            try:
                if '年' in last_trading_day:
                    # 中文格式: YYYY年MM月DD日
                    date_obj = datetime.strptime(last_trading_day, '%Y年%m月%d日')
                else:
                    # 标准格式: YYYY/MM/DD
                    date_obj = datetime.strptime(last_trading_day, '%Y/%m/%d')

                date_suffix = date_obj.strftime('%m%d')  # MMDD格式
                leader_sheet_name = f"龙头{date_suffix}"
            except Exception as e:
                print(f"解析日期时出错: {e}，使用默认命名")
                leader_sheet_name = f"{sheet_name_used}_龙头股"
        else:
            leader_sheet_name = f"{sheet_name_used}_龙头股"

        # 检查是否需要创建或更新龙头股工作表
        sheet_exists = leader_sheet_name in wb.sheetnames

        # 决定是否需要创建/更新工作表
        should_create_leader_sheet = False

        if not sheet_exists:
            # 工作表不存在，需要创建
            should_create_leader_sheet = True
            print(f"在现有工作簿中创建新的龙头股工作表: {leader_sheet_name}")
        else:
            # 龙头股工作表已存在，直接覆盖更新（每日独立保存）
            wb.remove(wb[leader_sheet_name])
            should_create_leader_sheet = True
            print(f"已更新龙头股工作表: {leader_sheet_name}")

        # 只有当需要创建工作表时才创建内容
        if should_create_leader_sheet:
            leader_ws = wb.create_sheet(title=leader_sheet_name)

            # 使用按概念分组的数据作为输入，这样可以复用已经计算好的long_period_change
            if 'concept_ws' in locals() and 'grouped_df' in locals():
                # 如果前面创建了概念分组工作表，则复用其数据
                concept_data_for_leader = grouped_df
            else:
                # 否则基于result_df重新计算概念分组数据
                concept_data_for_leader = result_df.copy()

                # 计算长周期涨跌幅
                def calculate_long_period_change(row):
                    try:
                        stock_code = row['stock_code']
                        end_date = date_mapping.get(formatted_trading_days[-1])
                        if not end_date:
                            return 0.0
                        start_date = get_n_trading_days_before(end_date, period_days_long)
                        if '-' in start_date:
                            start_date = start_date.replace('-', '')
                        long_change = calculate_stock_period_change(stock_code, start_date, end_date)
                        return long_change if long_change is not None else 0.0
                    except:
                        return 0.0

                concept_data_for_leader['long_period_change'] = concept_data_for_leader.apply(
                    calculate_long_period_change, axis=1)

            create_leader_stocks_sheet_content(leader_ws, concept_data_for_leader, shouban_df, stock_data,
                                               stock_entry_count, formatted_trading_days, date_column_start,
                                               show_period_change, period_column, period_days, period_days_long,
                                               stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                                               zaban_df)

        # 管理龙头sheet并回填历史数据
        last_day_str = date_mapping[formatted_trading_days[-1]]
        last_trading_day_obj = datetime.strptime(last_day_str, '%Y%m%d')
        manage_leader_sheets(wb, last_trading_day_obj, output_file)
        backfill_historical_leader_sheets(wb, last_trading_day_obj, formatted_trading_days, date_mapping)

    # 创建指数数据工作表
    print("开始创建指数数据工作表...")
    create_index_sheet(wb, date_columns, sheet_name="指数数据")

    # 进行原因分析（根据开关控制）
    from analysis.concept_analyzer import ENABLE_REASON_ANALYSIS

    # 确定是否启用原因分析
    should_enable_reason_analysis = enable_reason_analysis
    if should_enable_reason_analysis is None:
        should_enable_reason_analysis = ENABLE_REASON_ANALYSIS

    concept_analysis_data = None
    if should_enable_reason_analysis:
        print("开始进行原因分析...")
        try:
            # 从原始连板数据中分析原因
            print("从原始连板数据中分析原因...")
            concept_analysis_data = analyze_concepts_from_ladder_data(lianban_df, date_columns, start_date, end_date)

            if concept_analysis_data:
                reason_stats, new_reasons = concept_analysis_data
                # 打印原因分析摘要
                summary = format_concept_analysis_summary(reason_stats, new_reasons)
                print(summary)
            else:
                print("原因分析未返回数据")

        except Exception as e:
            print(f"原因分析过程中出现错误: {e}")
            import traceback
            traceback.print_exc()
            concept_analysis_data = None
    else:
        print("原因分析已禁用，跳过分析步骤")

    # 创建图例工作表，传入对应的sheet名和概念分析数据
    create_legend_sheet(wb, stock_data['reason_counter'], stock_data['reason_colors'],
                        stock_data['top_reasons'], HIGH_BOARD_COLORS, REENTRY_COLORS,
                        source_sheet_name=sheet_name_used, concept_analysis_data=concept_analysis_data)

    # 在所有计算完成后，统一应用行折叠（最后一步）
    if concept_grouped_rows_to_collapse:
        print(f"开始应用行折叠：【{concept_grouped_sheet_name}】需要折叠 {len(concept_grouped_rows_to_collapse)} 行")

        # 对【概念分组】sheet应用折叠
        if concept_grouped_sheet_name in wb.sheetnames:
            ws_concept = wb[concept_grouped_sheet_name]
            for row_idx in concept_grouped_rows_to_collapse:
                ws_concept.row_dimensions[row_idx].hidden = True
            print(f"  已折叠【{concept_grouped_sheet_name}】的 {len(concept_grouped_rows_to_collapse)} 行")

        # 对【成交量】sheet复用相同的行索引进行折叠
        if create_volume_sheet and f"{concept_grouped_sheet_name}_成交量" in wb.sheetnames:
            volume_sheet_name = f"{concept_grouped_sheet_name}_成交量"
            ws_volume = wb[volume_sheet_name]
            for row_idx in concept_grouped_rows_to_collapse:
                ws_volume.row_dimensions[row_idx].hidden = True
            print(f"  已折叠【{volume_sheet_name}】的 {len(concept_grouped_rows_to_collapse)} 行（复用概念分组的折叠索引）")

    # 保存Excel文件
    try:
        save_excel_file(wb, output_file)

        # 显示均线斜率统计信息
        print_slope_statistics()

        return True
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")
        return False


def format_trading_days(trading_days):
    """
    格式化交易日列表

    Args:
        trading_days: 交易日列表

    Returns:
        tuple: (格式化的交易日列表, 日期映射)
    """
    formatted_trading_days = []
    date_mapping = {}  # 用于将格式化日期映射回YYYYMMDD格式

    for day in trading_days:
        # 处理不同类型的日期输入
        if isinstance(day, str):
            date_obj = datetime.strptime(day, '%Y%m%d')
            day_str = day
        else:
            # 如果是Timestamp对象，直接使用
            date_obj = day.to_pydatetime() if hasattr(day, 'to_pydatetime') else day
            day_str = date_obj.strftime('%Y%m%d')

        # 标准格式 YYYY/MM/DD
        formatted_day_slash = date_obj.strftime('%Y/%m/%d')
        # 中文格式 YYYY年MM月DD日
        formatted_day_cn = date_obj.strftime('%Y年%m月%d日')

        # 使用中文格式作为主要格式
        formatted_trading_days.append(formatted_day_cn)
        date_mapping[formatted_day_cn] = day_str
        date_mapping[formatted_day_slash] = day_str  # 同时保存标准格式的映射

    return formatted_trading_days, date_mapping


def identify_significant_boards(lianban_df, shouban_df, min_board_level, reentry_days,
                                non_main_board_level, enable_attention_criteria, attention_data,
                                priority_reasons, low_priority_reasons=None, momo_df=None, start_date=None,
                                end_date=None):
    """
    识别显著连板股票

    Args:
        lianban_df: 连板数据DataFrame
        shouban_df: 首板数据DataFrame
        min_board_level: 主板股票最小显著连板天数
        reentry_days: 断板后再次上榜的天数阈值
        non_main_board_level: 非主板股票最小显著连板天数
        enable_attention_criteria: 是否启用关注度榜入选条件
        attention_data: 关注度榜数据
        priority_reasons: 优先选择的原因列表
        low_priority_reasons: 低优先级原因列表
        momo_df: 【默默上涨】数据DataFrame
        start_date: 分析开始日期
        end_date: 分析结束日期

    Returns:
        pandas.DataFrame: 显著连板股票DataFrame（包含【默默上涨】数据）
    """
    # 识别连板股票
    result_df = identify_first_significant_board(
        lianban_df, shouban_df, min_board_level, reentry_days, non_main_board_level,
        enable_attention_criteria, attention_data['main'], attention_data['non_main'],
        priority_reasons, low_priority_reasons
    )

    # 如果有【默默上涨】数据，则处理并合并
    if momo_df is not None and not momo_df.empty and start_date and end_date:
        print("处理【默默上涨】数据...")
        momo_result_df = identify_momo_shangzhang_stocks(momo_df, start_date, end_date)

        if not momo_result_df.empty:
            print(f"【默默上涨】处理完成，共{len(momo_result_df)}只股票")
            # 合并【默默上涨】数据到结果中
            result_df = pd.concat([result_df, momo_result_df], ignore_index=True)
            print(f"合并后总股票数量: {len(result_df)}")
        else:
            print("【默默上涨】数据处理后为空")

    return result_df


def count_stock_entries(result_df):
    """
    统计股票入选次数

    Args:
        result_df: 显著连板股票DataFrame

    Returns:
        dict: 股票入选次数映射
    """
    # 收集所有已入选连板梯队的股票代码
    lianban_stock_codes = set()
    # 统计每只股票入选次数
    stock_entry_count = {}

    for _, row in result_df.iterrows():
        # 纯代码作为标识
        pure_code = row['stock_code'].split('_')[0] if '_' in row['stock_code'] else row['stock_code']
        lianban_stock_codes.add(pure_code)

        # 统计入选次数
        if pure_code in stock_entry_count:
            stock_entry_count[pure_code] += 1
        else:
            stock_entry_count[pure_code] = 1

    return stock_entry_count


def fill_data_rows(ws, result_df, shouban_df, stock_reason_group, reason_colors, stock_entry_count,
                   formatted_trading_days, date_column_start, show_period_change, period_column,
                   period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                   max_tracking_days_before, zaban_df, show_warning_column=True, enable_collapse=False):
    """
    填充数据行

    Args:
        ws: Excel工作表
        result_df: 显著连板股票DataFrame
        shouban_df: 首板数据DataFrame
        stock_reason_group: 股票概念组映射
        reason_colors: 概念颜色映射
        stock_entry_count: 股票入选次数映射
        formatted_trading_days: 格式化的交易日列表
        date_column_start: 日期列开始位置
        show_period_change: 是否显示周期涨跌幅
        period_column: 周期涨跌幅列索引
        period_days: 周期天数
        stock_details: 股票详细信息映射
        date_mapping: 日期映射
        max_tracking_days: 断板后跟踪的最大天数
        max_tracking_days_before: 入选前跟踪的最大天数
        zaban_df: 炸板数据DataFrame
        show_warning_column: 是否显示异动预警列
        enable_collapse: 是否启用行折叠功能（True=收集折叠行索引，False=不折叠）
    
    Returns:
        list: 需要折叠的行索引列表（仅当enable_collapse=True时返回）
    """
    # 计算所有股票的新高标记（使用缓存版本）
    new_high_markers = get_new_high_markers_cached(result_df, formatted_trading_days, date_mapping)

    # 获取结束日期用于均线斜率计算
    end_date_for_ma = date_mapping.get(formatted_trading_days[-1])

    # 初始化异动检测器（如果需要显示预警列）
    abnormal_detector = None
    if show_warning_column:
        try:
            from analysis.abnormal_movement_detector import AbnormalMovementDetector
            abnormal_detector = AbnormalMovementDetector()
            print("异动检测器初始化成功")
        except Exception as e:
            print(f"异动检测器初始化失败: {e}")
            abnormal_detector = None

    # 用于收集需要折叠的行索引
    rows_to_collapse = []

    for i, (_, stock) in enumerate(result_df.iterrows()):
        row_idx = i + 4  # 行索引，从第4行开始（第1行是日期标题，第2-3行是大盘指标）

        # 提取基本股票信息
        stock_code = stock['stock_code']
        stock_name = stock['stock_name']
        all_board_data = stock['all_board_data']

        # 提取纯代码
        pure_stock_code = extract_pure_stock_code(stock_code)

        # 设置股票代码列（第一列）
        format_stock_code_cell(ws, row_idx, 1, pure_stock_code)

        # 获取概念
        concept = get_stock_concept(stock)

        # 设置概念列（第二列）
        stock_key = f"{stock_code}_{stock_name}"
        format_concept_cell(ws, row_idx, 2, concept, stock_key, stock_reason_group, reason_colors)

        # 计算股票的最高板数
        max_board_level = calculate_max_board_level(all_board_data)

        # 根据股票代码确定市场类型
        market_type = get_market_marker(pure_stock_code)

        # 设置股票简称列（第三列）
        apply_high_board_color = False
        _, apply_high_board_color = format_stock_name_cell(ws, row_idx, 3, stock_name, market_type,
                                                           max_board_level, apply_high_board_color,
                                                           pure_stock_code, stock_entry_count, end_date_for_ma)

        # 填充周期涨跌幅列
        if show_period_change:
            # 使用最后一个交易日作为结束日期
            end_date = date_mapping.get(formatted_trading_days[-1])
            format_period_change_cell(ws, row_idx, period_column, stock_code, stock_name,
                                      stock['first_significant_date'], period_days, period_days_long, end_date)

        # 填充每个交易日的数据
        fill_daily_data(ws, row_idx, formatted_trading_days, date_column_start, all_board_data,
                        shouban_df, pure_stock_code, stock_details, stock, date_mapping,
                        max_tracking_days, max_tracking_days_before, zaban_df, period_days, new_high_markers)

        # 填充异动预警列
        if show_warning_column and abnormal_detector and formatted_trading_days:
            warning_col = len(formatted_trading_days) + date_column_start
            warning_message = ""

            try:
                # 使用最后一个交易日作为检查日期
                check_date = date_mapping.get(formatted_trading_days[-1])
                if check_date:
                    # 准备周期数据用于性能优化
                    period_data = None
                    try:
                        from analysis.ladder_chart import calculate_stock_period_change
                        from utils.date_util import get_n_trading_days_before

                        end_date_str = check_date.strftime('%Y%m%d')

                        # 计算5日、10日、30日涨幅
                        prev_5d = get_n_trading_days_before(end_date_str, 5).replace('-', '')
                        prev_10d = get_n_trading_days_before(end_date_str, 10).replace('-', '')
                        prev_30d = get_n_trading_days_before(end_date_str, 30).replace('-', '')

                        change_5d = calculate_stock_period_change(stock_code, prev_5d, end_date_str, stock_name)
                        change_10d = calculate_stock_period_change(stock_code, prev_10d, end_date_str, stock_name)
                        change_30d = calculate_stock_period_change(stock_code, prev_30d, end_date_str, stock_name)

                        if change_5d is not None and change_10d is not None and change_30d is not None:
                            period_data = {
                                '5d': change_5d / 100.0,  # 转换为小数
                                '10d': change_10d / 100.0,
                                '30d': change_30d / 100.0
                            }
                    except Exception:
                        period_data = None

                    warning_message = abnormal_detector.get_warning_message(pure_stock_code, check_date, period_data)
            except Exception as e:
                print(f"获取股票{pure_stock_code}预警信息时出错: {e}")
                warning_message = ""

            # 设置预警单元格
            warning_cell = ws.cell(row=row_idx, column=warning_col, value=warning_message if warning_message else "")
            warning_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            warning_cell.border = BORDER_STYLE
            warning_cell.font = Font(size=9)

            # 设置固定行高（与默认行高一致）
            ws.row_dimensions[row_idx].height = 14

            # 根据预警类型设置背景色（使用浅色方案）
            if warning_message:
                if "已触发严重异常波动" in warning_message:
                    warning_cell.fill = PatternFill(start_color="FFCCCB", fill_type="solid")  # 浅红色
                    warning_cell.font = Font(color="8B0000", size=9, bold=True)  # 深红色字体
                elif "已触发异常波动" in warning_message:
                    warning_cell.fill = PatternFill(start_color="FFE4B5", fill_type="solid")  # 浅橙色
                    warning_cell.font = Font(color="FF4500", size=9, bold=True)  # 橙红色字体
                elif "将触发严重异动" in warning_message:
                    warning_cell.fill = PatternFill(start_color="FFB6C1", fill_type="solid")  # 浅粉红色
                    warning_cell.font = Font(color="8B008B", size=9)  # 深紫色字体
                elif "将触发异常波动" in warning_message:
                    # 异动预警和正常状态都不设置背景色
                    pass

        # 判断是否需要折叠此行（仅在启用折叠功能时）
        if enable_collapse and should_collapse_row(stock, formatted_trading_days, date_mapping):
            # 收集需要折叠的行索引，稍后统一处理
            rows_to_collapse.append(row_idx)

    # 返回需要折叠的行索引（如果启用了折叠功能）
    if enable_collapse:
        return rows_to_collapse


def extract_pure_stock_code(stock_code):
    """
    提取纯股票代码

    Args:
        stock_code: 股票代码

    Returns:
        str: 纯股票代码
    """
    pure_stock_code = stock_code.split('_')[0] if '_' in stock_code else stock_code
    if pure_stock_code.startswith(('sh', 'sz', 'bj')):
        pure_stock_code = pure_stock_code[2:]

    return pure_stock_code


def get_stock_concept(stock):
    """
    获取股票概念

    Args:
        stock: 股票数据

    Returns:
        str: 概念文本
    """
    # 检查是否为【默默上涨】股票
    if stock.get('entry_type') == 'momo_shangzhang':
        # 【默默上涨】股票使用特殊的概念格式
        momo_data = stock.get('momo_data', {})
        period_volume = momo_data.get('区间成交额', '')
        period_change = momo_data.get('区间涨跌幅', '')
        return f"成交额:{period_volume} 涨幅:{period_change}"

    # 普通股票的概念处理
    concept = stock.get('concept', '其他')
    if pd.isna(concept) or not concept:
        concept = "其他"

    return concept


def calculate_max_board_level(all_board_data):
    """
    计算股票的最高板数

    Args:
        all_board_data: 所有板块数据

    Returns:
        int: 最高板数
    """
    max_board_level = 0
    for day_data in all_board_data.values():
        if pd.notna(day_data) and day_data and day_data > max_board_level:
            max_board_level = int(day_data)

    return max_board_level


def fill_daily_data(ws, row_idx, formatted_trading_days, date_column_start, all_board_data,
                    shouban_df, pure_stock_code, stock_details, stock, date_mapping,
                    max_tracking_days, max_tracking_days_before, zaban_df, period_days=PERIOD_DAYS_CHANGE,
                    new_high_markers=None):
    """
    填充每日数据

    Args:
        ws: Excel工作表
        row_idx: 行索引
        formatted_trading_days: 格式化的交易日列表
        date_column_start: 日期列开始位置
        all_board_data: 所有板块数据
        shouban_df: 首板数据DataFrame
        pure_stock_code: 纯股票代码
        stock_details: 股票详细信息映射
        stock: 股票数据
        date_mapping: 日期映射
        max_tracking_days: 断板后跟踪的最大天数
        max_tracking_days_before: 入选前跟踪的最大天数
        zaban_df: 炸板数据DataFrame
        period_days: 计算涨跌幅的周期天数
        new_high_markers: 新高标记映射
    """
    for j, formatted_day in enumerate(formatted_trading_days):
        col_idx = j + date_column_start  # 列索引，从date_column_start开始

        # 获取当日的连板信息
        board_days = all_board_data.get(formatted_day)

        # 标记是否在首板数据中找到该股票
        found_in_shouban = check_stock_in_shouban(shouban_df, pure_stock_code, formatted_day)

        # 处理单元格
        stock = process_daily_cell(ws, row_idx, col_idx, formatted_day, board_days, found_in_shouban,
                                   pure_stock_code, stock_details, stock, date_mapping,
                                   max_tracking_days, max_tracking_days_before, zaban_df, period_days, new_high_markers)


def adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, show_warning_column=True):
    """
    调整列宽

    Args:
        ws: Excel工作表
        formatted_trading_days: 格式化的交易日列表
        date_column_start: 日期列开始位置
        show_period_change: 是否显示周期涨跌幅
        show_warning_column: 是否显示异动预警列
    """
    # 调整列宽
    ws.column_dimensions['A'].width = 8  # 股票代码列宽度设置窄一些
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 15

    if show_period_change:
        ws.column_dimensions['D'].width = 9  # 周期涨跌幅列宽度

    for i in range(len(formatted_trading_days)):
        col_letter = get_column_letter(i + date_column_start)
        ws.column_dimensions[col_letter].width = 12

    # 调整异动预警列宽度
    if show_warning_column and formatted_trading_days:
        warning_col = len(formatted_trading_days) + date_column_start
        warning_col_letter = get_column_letter(warning_col)
        ws.column_dimensions[warning_col_letter].width = 20  # 预警列需要更宽一些

    # 调整行高，确保日期和星期能完整显示
    ws.row_dimensions[1].height = 30


def create_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                         formatted_trading_days, date_column_start, show_period_change, period_column,
                                         period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                         max_tracking_days_before, zaban_df, enable_collapse=False):
    """
    创建按概念分组的工作表内容

    Args:
        ws: Excel工作表
        result_df: 显著连板股票DataFrame
        shouban_df: 首板数据DataFrame
        stock_data: 股票数据字典
        stock_entry_count: 股票入选次数映射
        formatted_trading_days: 格式化的交易日列表
        date_column_start: 日期列开始位置
        show_period_change: 是否显示周期涨跌幅
        period_column: 周期涨跌幅列索引
        period_days: 周期天数
        period_days_long: 长周期天数
        stock_details: 股票详细信息映射
        date_mapping: 日期映射
        max_tracking_days: 断板后跟踪的最大天数
        max_tracking_days_before: 入选前跟踪的最大天数
        zaban_df: 炸板数据DataFrame
        enable_collapse: 是否启用行折叠功能（True=收集折叠行索引，False=不折叠）
    
    Returns:
        list: 需要折叠的行索引列表（仅当enable_collapse=True时返回）
    """
    print(f"填充按概念分组的工作表内容")

    # 按概念分组重新排序数据，添加长周期涨跌幅计算
    concept_grouped_df = result_df.copy()

    # 计算长周期涨跌幅用于排序
    def calculate_long_period_change(row):
        try:
            stock_code = row['stock_code']

            # 使用最后一个交易日作为结束日期
            end_date = date_mapping.get(formatted_trading_days[-1])
            if not end_date:
                return 0.0

            # 计算长周期前的开始日期
            start_date = get_n_trading_days_before(end_date, period_days_long)
            if '-' in start_date:
                start_date = start_date.replace('-', '')

            # 使用已有的函数计算涨跌幅
            long_change = calculate_stock_period_change(stock_code, start_date, end_date)
            return long_change if long_change is not None else 0.0
        except:
            return 0.0

    concept_grouped_df['long_period_change'] = concept_grouped_df.apply(calculate_long_period_change, axis=1)

    # 为【默默上涨】分组添加特殊排序字段
    def get_momo_sort_keys(row):
        if row.get('concept_group') == '默默上涨':
            momo_data = row.get('momo_data', {})
            # 提取成交额数值（去掉"亿"字符）
            volume_str = momo_data.get('区间成交额', '0')
            try:
                volume = float(volume_str.replace('亿', '')) if '亿' in str(volume_str) else 0.0
            except:
                volume = 0.0

            # 提取涨幅数值（去掉"%"字符）
            change_str = momo_data.get('区间涨跌幅', '0')
            try:
                change = float(change_str.replace('%', '')) if '%' in str(change_str) else 0.0
            except:
                change = 0.0

            return volume, change
        else:
            return 0.0, 0.0

    # 添加【默默上涨】排序字段
    concept_grouped_df[['momo_volume', 'momo_change']] = concept_grouped_df.apply(
        lambda row: pd.Series(get_momo_sort_keys(row)), axis=1
    )

    # 分别处理【默默上涨】分组和其他分组的排序
    momo_mask = concept_grouped_df['concept_group'] == '默默上涨'
    momo_df = concept_grouped_df[momo_mask].copy()
    other_df = concept_grouped_df[~momo_mask].copy()

    # 【默默上涨】分组：按概念优先级、成交额、涨幅倒序排列
    if not momo_df.empty:
        momo_df = momo_df.sort_values(
            by=['concept_priority', 'momo_volume', 'momo_change'],
            ascending=[True, False, False]
        )

    # 其他分组：按原有逻辑排序
    if not other_df.empty:
        other_df = other_df.sort_values(
            by=['concept_priority', 'concept_group', 'first_significant_date', 'long_period_change',
                'board_level_at_first'],
            ascending=[True, True, True, False, False]
        )

    # 合并排序结果
    concept_grouped_df = pd.concat([other_df, momo_df], ignore_index=True)

    # 删除临时排序字段
    concept_grouped_df = concept_grouped_df.drop(columns=['momo_volume', 'momo_change'])

    # 设置Excel表头和日期列
    show_warning_column = True  # 默认显示异动预警列
    date_columns = setup_excel_header(ws, formatted_trading_days, show_period_change, period_days, date_column_start,
                                      show_warning_column)

    # 添加大盘指标行
    add_market_indicators(ws, date_columns, label_col=2)

    # 填充数据行，使用重新排序的数据
    rows_to_collapse = fill_data_rows_with_concept_groups(ws, concept_grouped_df, shouban_df,
                                                          stock_data['stock_reason_group'],
                                                          stock_data['reason_colors'], stock_entry_count,
                                                          formatted_trading_days,
                                                          date_column_start, show_period_change, period_column,
                                                          period_days,
                                                          period_days_long, stock_details, date_mapping,
                                                          max_tracking_days,
                                                          max_tracking_days_before, zaban_df, show_warning_column,
                                                          enable_collapse)

    # 调整列宽
    adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, show_warning_column)

    # 冻结前三列和前三行
    from openpyxl.utils import get_column_letter
    freeze_cell = f"{get_column_letter(date_column_start)}4"
    ws.freeze_panes = freeze_cell

    # 返回需要折叠的行索引（如果启用了折叠）
    if enable_collapse:
        return rows_to_collapse


def create_concept_grouped_sheet(wb, sheet_name, result_df, shouban_df, stock_data, stock_entry_count,
                                 formatted_trading_days, date_column_start, show_period_change, period_column,
                                 period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                 max_tracking_days_before, zaban_df, collapse_days=None):
    """
    创建按概念分组的工作表（保持向后兼容）

    Args:
        wb: Excel工作簿
        sheet_name: 工作表名称
        collapse_days: 断板后折叠行的天数阈值
        其他参数与create_concept_grouped_sheet_content相同
    """
    print(f"创建按概念分组的工作表: {sheet_name}")

    # 创建新的工作表
    ws = wb.create_sheet(title=sheet_name)

    # 填充内容
    create_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                         formatted_trading_days, date_column_start, show_period_change, period_column,
                                         period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                         max_tracking_days_before, zaban_df, collapse_days)


def fill_data_rows_with_concept_groups(ws, result_df, shouban_df, stock_reason_group, reason_colors,
                                       stock_entry_count, formatted_trading_days, date_column_start,
                                       show_period_change, period_column, period_days, period_days_long,
                                       stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                                       zaban_df, show_warning_column=True, enable_collapse=False):
    """
    填充数据行，按概念分组并在组间添加分隔行

    Args:
        ws: Excel工作表
        result_df: 按概念分组排序的显著连板股票DataFrame
        show_warning_column: 是否显示异动预警列
        enable_collapse: 是否启用行折叠功能（True=收集折叠行索引，False=不折叠）
        其他参数与fill_data_rows相同
    
    Returns:
        list: 需要折叠的行索引列表（仅当enable_collapse=True时返回）
    """
    # 计算所有股票的新高标记（使用缓存版本）
    new_high_markers = get_new_high_markers_cached(result_df, formatted_trading_days, date_mapping)

    # 初始化异动检测器（如果需要显示预警列）
    abnormal_detector = None
    if show_warning_column:
        try:
            from analysis.abnormal_movement_detector import AbnormalMovementDetector
            abnormal_detector = AbnormalMovementDetector()
            print("按概念分组工作表：异动检测器初始化成功")
        except Exception as e:
            print(f"按概念分组工作表：异动检测器初始化失败: {e}")
            abnormal_detector = None

    # 用于收集需要折叠的行索引
    rows_to_collapse = []

    current_row = 4  # 从第4行开始（第1行是日期标题，第2-3行是大盘指标）
    current_concept_group = None

    for _, stock in result_df.iterrows():
        stock_concept_group = stock.get('concept_group', '其他')

        # 如果概念组发生变化，添加分隔行
        if current_concept_group is not None and current_concept_group != stock_concept_group:
            # 添加空行作为分隔
            current_row += 1

            # 添加概念组标题行，合并所有列
            concept_title_cell = ws.cell(row=current_row, column=1, value=f"【{stock_concept_group}】")
            concept_title_cell.font = Font(bold=True, size=12)
            concept_title_cell.alignment = Alignment(horizontal='left')

            # 为概念组标题行设置背景色
            if stock_concept_group in reason_colors:
                concept_title_cell.fill = PatternFill(start_color=reason_colors[stock_concept_group], fill_type="solid")
                # 如果背景色较深，使用白色字体
                if reason_colors[stock_concept_group] in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
                    concept_title_cell.font = Font(bold=True, size=12, color="FFFFFF")

            # 计算合并的结束列
            end_col = date_column_start + len(formatted_trading_days) - 1
            if show_period_change:
                end_col += 1
            if show_warning_column:
                end_col += 1

            # 合并概念组标题行的所有列
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=end_col)

            current_row += 1
        elif current_concept_group is None:
            # 第一个概念组，添加标题行，合并所有列
            concept_title_cell = ws.cell(row=current_row, column=1, value=f"【{stock_concept_group}】")
            concept_title_cell.font = Font(bold=True, size=12)
            concept_title_cell.alignment = Alignment(horizontal='left')

            # 为概念组标题行设置背景色
            if stock_concept_group in reason_colors:
                concept_title_cell.fill = PatternFill(start_color=reason_colors[stock_concept_group], fill_type="solid")
                # 如果背景色较深，使用白色字体
                if reason_colors[stock_concept_group] in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
                    concept_title_cell.font = Font(bold=True, size=12, color="FFFFFF")

            # 计算合并的结束列
            end_col = date_column_start + len(formatted_trading_days) - 1
            if show_period_change:
                end_col += 1
            if show_warning_column:
                end_col += 1

            # 合并概念组标题行的所有列
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=end_col)

            current_row += 1

        current_concept_group = stock_concept_group

        # 填充股票数据行（复用原有逻辑）
        should_collapse = fill_single_stock_row(ws, current_row, stock, shouban_df, stock_reason_group, reason_colors,
                                                stock_entry_count, formatted_trading_days, date_column_start,
                                                show_period_change,
                                                period_column, period_days, period_days_long, stock_details,
                                                date_mapping,
                                                max_tracking_days, max_tracking_days_before, zaban_df, new_high_markers,
                                                show_warning_column, abnormal_detector, enable_collapse)

        # 如果启用折叠并且此行需要折叠
        if enable_collapse and should_collapse:
            rows_to_collapse.append(current_row)

        current_row += 1

    # 返回需要折叠的行索引（如果启用了折叠）
    if enable_collapse:
        return rows_to_collapse


def fill_single_stock_row(ws, row_idx, stock, shouban_df, stock_reason_group, reason_colors, stock_entry_count,
                          formatted_trading_days, date_column_start, show_period_change, period_column,
                          period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                          max_tracking_days_before, zaban_df, new_high_markers=None,
                          show_warning_column=True, abnormal_detector=None, enable_collapse=False):
    """
    填充单个股票的数据行

    Args:
        ws: Excel工作表
        row_idx: 行索引
        stock: 股票数据
        new_high_markers: 新高标记映射
        show_warning_column: 是否显示异动预警列
        abnormal_detector: 异动检测器实例
        enable_collapse: 是否启用行折叠功能（True=返回是否需要折叠，False=不判断）
        其他参数与fill_data_rows相同
    
    Returns:
        bool: 是否需要折叠此行（仅当enable_collapse=True时返回）
    """
    # 提取基本股票信息
    stock_code = stock['stock_code']
    stock_name = stock['stock_name']
    all_board_data = stock['all_board_data']

    # 获取结束日期用于均线斜率计算
    end_date_for_ma = date_mapping.get(formatted_trading_days[-1])

    # 提取纯代码
    pure_stock_code = extract_pure_stock_code(stock_code)

    # 设置股票代码列（第一列）
    format_stock_code_cell(ws, row_idx, 1, pure_stock_code)

    # 获取概念
    concept = get_stock_concept(stock)

    # 设置概念列（第二列）
    stock_key = f"{stock_code}_{stock_name}"
    format_concept_cell(ws, row_idx, 2, concept, stock_key, stock_reason_group, reason_colors)

    # 计算股票的最高板数
    max_board_level = calculate_max_board_level(all_board_data)

    # 根据股票代码确定市场类型
    market_type = get_market_marker(pure_stock_code)

    # 设置股票简称列（第三列）
    apply_high_board_color = False
    _, apply_high_board_color = format_stock_name_cell(ws, row_idx, 3, stock_name, market_type,
                                                       max_board_level, apply_high_board_color,
                                                       pure_stock_code, stock_entry_count, end_date_for_ma)

    # 填充周期涨跌幅列
    if show_period_change:
        # 使用最后一个交易日作为结束日期
        end_date = date_mapping.get(formatted_trading_days[-1])
        format_period_change_cell(ws, row_idx, period_column, stock_code, stock_name,
                                  stock['first_significant_date'], period_days, period_days_long, end_date)

    # 填充每个交易日的数据
    fill_daily_data(ws, row_idx, formatted_trading_days, date_column_start, all_board_data,
                    shouban_df, pure_stock_code, stock_details, stock, date_mapping,
                    max_tracking_days, max_tracking_days_before, zaban_df, period_days, new_high_markers)

    # 填充异动预警列
    if show_warning_column and abnormal_detector and formatted_trading_days:
        warning_col = len(formatted_trading_days) + date_column_start
        warning_message = ""

        try:
            # 使用最后一个交易日作为检查日期
            check_date = date_mapping.get(formatted_trading_days[-1])
            if check_date:
                # 准备周期数据用于性能优化
                period_data = None
                try:
                    from analysis.ladder_chart import calculate_stock_period_change
                    from utils.date_util import get_n_trading_days_before

                    end_date_str = check_date.strftime('%Y%m%d')

                    # 计算5日、10日、30日涨幅
                    prev_5d = get_n_trading_days_before(end_date_str, 5).replace('-', '')
                    prev_10d = get_n_trading_days_before(end_date_str, 10).replace('-', '')
                    prev_30d = get_n_trading_days_before(end_date_str, 30).replace('-', '')

                    change_5d = calculate_stock_period_change(stock_code, prev_5d, end_date_str, stock_name)
                    change_10d = calculate_stock_period_change(stock_code, prev_10d, end_date_str, stock_name)
                    change_30d = calculate_stock_period_change(stock_code, prev_30d, end_date_str, stock_name)

                    if change_5d is not None and change_10d is not None and change_30d is not None:
                        period_data = {
                            '5d': change_5d / 100.0,  # 转换为小数
                            '10d': change_10d / 100.0,
                            '30d': change_30d / 100.0
                        }
                except Exception:
                    period_data = None

                warning_message = abnormal_detector.get_warning_message(pure_stock_code, check_date, period_data)
        except Exception as e:
            print(f"获取股票{pure_stock_code}预警信息时出错: {e}")
            warning_message = ""

        # 设置预警单元格
        warning_cell = ws.cell(row=row_idx, column=warning_col, value=warning_message if warning_message else "")
        warning_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        warning_cell.border = BORDER_STYLE
        warning_cell.font = Font(size=9)

        # 设置固定行高（与默认行高一致）
        ws.row_dimensions[row_idx].height = 14

        # 根据预警类型设置背景色（使用浅色方案）
        if warning_message:
            if "已触发严重异常波动" in warning_message:
                warning_cell.fill = PatternFill(start_color="FFCCCB", fill_type="solid")  # 浅红色
                warning_cell.font = Font(color="8B0000", size=9, bold=True)  # 深红色字体
            elif "已触发异常波动" in warning_message:
                warning_cell.fill = PatternFill(start_color="FFE4B5", fill_type="solid")  # 浅橙色
                warning_cell.font = Font(color="FF4500", size=9, bold=True)  # 橙红色字体
            elif "将触发严重异动" in warning_message:
                warning_cell.fill = PatternFill(start_color="FFB6C1", fill_type="solid")  # 浅粉红色
                warning_cell.font = Font(color="8B008B", size=9)  # 深紫色字体
            elif "将触发异常波动" in warning_message:
                # 异动预警和正常状态都不设置背景色
                pass

    # 判断是否需要折叠此行（仅在启用折叠功能时）
    if enable_collapse:
        should_collapse = should_collapse_row(stock, formatted_trading_days, date_mapping)
        return should_collapse

    # 如果未启用折叠，返回False
    return False


def save_excel_file(wb, output_file):
    """
    保存Excel文件

    Args:
        wb: Excel工作簿
        output_file: 输出文件路径
    """
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    wb.save(output_file)
    print(f"梯队形态涨停复盘图已生成: {output_file}")


def select_leader_stocks_from_concept_groups(concept_grouped_df, date_mapping, formatted_trading_days,
                                             period_days, period_days_long):
    """
    从按概念分组的数据中选出龙头股
    
    动态名额分配策略（由全局参数控制）：
    - 最热板块（第1名）：LEADER_QUOTA_TOP1 只
    - 次热板块（第2名）：LEADER_QUOTA_TOP2 只
    - 第三热板块（第3名）：LEADER_QUOTA_TOP3 只
    - 第四热板块（第4名）：LEADER_QUOTA_TOP4 只
    - 默认板块（第5名到 LEADER_QUOTA_DEFAULT_THRESHOLD 之间）：LEADER_QUOTA_DEFAULT 只
    - 非热门板块（LEADER_QUOTA_DEFAULT_THRESHOLD 之后）：LEADER_QUOTA_COLD 只
    
    筛选条件（由全局参数控制，区分主板和非主板）：
    - 主板股：
      * 最低连板数：MIN_BOARD_LEVEL_FOR_LEADER (默认2)
      * 最低短周期涨幅：MIN_SHORT_PERIOD_CHANGE_FOR_LEADER (默认20%)
      * 最低长周期涨幅：MIN_LONG_PERIOD_CHANGE_FOR_LEADER (默认60%)
    - 非主板股（创业板/科创板/北交所）：
      * 最低连板数：MIN_BOARD_LEVEL_FOR_LEADER_NON_MAIN (默认1)
      * 最低短周期涨幅：MIN_SHORT_PERIOD_CHANGE_FOR_LEADER_NON_MAIN (默认30%)
      * 最低长周期涨幅：MIN_LONG_PERIOD_CHANGE_FOR_LEADER_NON_MAIN (默认70%)
    - 是否只从活跃股选择：SELECT_LEADERS_FROM_ACTIVE_ONLY
    - 排除概念组：LEADER_EXCLUDE_CONCEPTS
    
    Args:
        concept_grouped_df: 按概念分组且已计算长周期涨跌幅的DataFrame
        date_mapping: 日期映射
        formatted_trading_days: 格式化的交易日列表
        period_days: 短周期天数
        period_days_long: 长周期天数
        
    Returns:
        pandas.DataFrame: 筛选出的龙头股DataFrame
    """
    print(f"开始从概念分组中筛选龙头股...")

    # 补充计算短周期涨跌幅和最高连板数（如果还没有的话）
    def calculate_additional_metrics(row):
        """计算额外的指标用于筛选"""
        try:
            stock_code = row['stock_code']

            # 计算最高连板数
            max_board_level = calculate_max_board_level(row['all_board_data'])

            # 计算短周期涨跌幅
            end_date = date_mapping.get(formatted_trading_days[-1])
            if end_date:
                start_date = get_n_trading_days_before(end_date, period_days)
                if '-' in start_date:
                    start_date = start_date.replace('-', '')
                short_change = calculate_stock_period_change(stock_code, start_date, end_date)
            else:
                short_change = 0.0

            return max_board_level, short_change if short_change is not None else 0.0
        except:
            return 0, 0.0

    # 为DataFrame添加必要的指标
    temp_df = concept_grouped_df.copy()

    # 确保long_period_change列存在，如果不存在则重新计算（calculate_stock_period_change已有lru_cache缓存）
    if 'long_period_change' not in temp_df.columns:
        end_date_str = date_mapping.get(formatted_trading_days[-1])
        if end_date_str:
            temp_df['long_period_change'] = temp_df.apply(
                lambda row: calculate_stock_period_change(
                    row['stock_code'],
                    get_n_trading_days_before(end_date_str, period_days_long).replace('-', ''),
                    end_date_str
                ) or 0.0, axis=1
            )
        else:
            temp_df['long_period_change'] = 0.0

    # 计算额外的指标
    metrics = temp_df.apply(calculate_additional_metrics, axis=1, result_type='expand')
    temp_df['max_board_level'] = metrics[0]
    temp_df['short_period_change'] = metrics[1]

    # 第一步：为所有股票计算last_board_date（用于判断是否应该折叠）
    print("\n第一步：计算所有股票的最后连板日期...")
    temp_df['last_board_date'] = temp_df.apply(
        lambda row: calculate_last_board_date(row, formatted_trading_days), axis=1
    )

    # 第二步：统计每个概念组的活跃股票数（未被折叠的股票数）
    print("\n第二步：统计各概念组活跃度...")
    concept_activity = {}

    for concept_group, group_df in temp_df.groupby('concept_group'):
        # 跳过特殊概念组
        if concept_group in LEADER_EXCLUDE_CONCEPTS:
            print(f"  跳过特殊概念组: {concept_group}")
            continue

        # 统计未被折叠的股票数量
        active_count = 0
        for _, stock in group_df.iterrows():
            if not should_collapse_row(stock, formatted_trading_days, date_mapping):
                active_count += 1

        concept_activity[concept_group] = active_count
        print(f"  概念组 {concept_group}: 总计{len(group_df)}只股票, 活跃{active_count}只")

    # 第三步：根据活跃度排名确定每个概念组的龙头名额
    print("\n第三步：根据活跃度分配龙头名额...")
    sorted_concepts = sorted(concept_activity.items(), key=lambda x: x[1], reverse=True)
    total_concept_count = len(sorted_concepts)

    concept_quota = {}  # 概念组 -> 龙头名额

    for rank, (concept_group, active_count) in enumerate(sorted_concepts, start=1):
        if rank == 1:
            quota = LEADER_QUOTA_TOP1  # 最热门
        elif rank == 2:
            quota = LEADER_QUOTA_TOP2  # 次热门
        elif rank == 3:
            quota = LEADER_QUOTA_TOP3  # 第三热门
        elif rank == 4:
            quota = LEADER_QUOTA_TOP4  # 第四热门
        elif rank >= 5 and rank <= total_concept_count * LEADER_QUOTA_DEFAULT_THRESHOLD:
            quota = LEADER_QUOTA_DEFAULT  # 默认（第5名到前N%）
        else:
            quota = LEADER_QUOTA_COLD  # 非热门（后N%）

        concept_quota[concept_group] = quota
        print(f"  排名第{rank}: {concept_group} (活跃{active_count}只) -> 分配{quota}个龙头名额")

    # 第四步：按概念分组筛选龙头股
    print(f"\n第四步：筛选龙头股 (候选范围: {'仅活跃股' if SELECT_LEADERS_FROM_ACTIVE_ONLY else '全部股票'})...")
    leader_stocks = []
    total_concepts = 0
    qualified_concepts = 0

    for concept_group, group_df in temp_df.groupby('concept_group'):
        # 跳过没有分配名额的概念组（如【默默上涨】）
        if concept_group not in concept_quota:
            continue

        total_concepts += 1
        quota = concept_quota[concept_group]

        # 根据开关决定候选范围
        if SELECT_LEADERS_FROM_ACTIVE_ONLY:
            # 只从活跃股（未被折叠）中选择，使用.copy()确保数据完整性
            candidate_df = group_df[~group_df.apply(
                lambda row: should_collapse_row(row, formatted_trading_days, date_mapping), axis=1
            )].copy()
            print(f"  处理概念组: {concept_group} (总计{len(group_df)}只, 活跃{len(candidate_df)}只, 名额{quota})")
        else:
            # 从全部股票中选择，使用.copy()确保数据完整性
            candidate_df = group_df.copy()
            print(f"  处理概念组: {concept_group} (总计{len(group_df)}只股票, 名额{quota})")

        # 筛选符合龙头条件的股票（区分主板和非主板）
        # 首先确定每只股票的市场类型
        def get_stock_market_type(stock_code):
            """获取股票市场类型"""
            try:
                pure_code = stock_code.split('_')[0] if '_' in stock_code else stock_code
                if pure_code.startswith(('sh', 'sz', 'bj')):
                    pure_code = pure_code[2:]
                return get_stock_market(pure_code)
            except:
                return 'main'  # 默认按主板处理

        candidate_df['market_type'] = candidate_df['stock_code'].apply(get_stock_market_type)

        # 1. 连板数门槛筛选（根据市场类型）
        def check_board_level(row):
            if row['market_type'] == 'main':
                return row['max_board_level'] >= MIN_BOARD_LEVEL_FOR_LEADER
            else:  # 非主板（创业板/科创板/北交所）
                return row['max_board_level'] >= MIN_BOARD_LEVEL_FOR_LEADER_NON_MAIN

        board_mask = candidate_df.apply(check_board_level, axis=1)

        # 2. 涨幅门槛筛选（根据市场类型，短周期或长周期满足一个即可，且长周期涨幅不能超过上限）
        def check_change_threshold(row):
            if row['market_type'] == 'main':
                short_threshold = MIN_SHORT_PERIOD_CHANGE_FOR_LEADER
                long_threshold = MIN_LONG_PERIOD_CHANGE_FOR_LEADER
                long_max_threshold = MAX_LONG_PERIOD_CHANGE_FOR_LEADER
            else:  # 非主板
                short_threshold = MIN_SHORT_PERIOD_CHANGE_FOR_LEADER_NON_MAIN
                long_threshold = MIN_LONG_PERIOD_CHANGE_FOR_LEADER_NON_MAIN
                long_max_threshold = MAX_LONG_PERIOD_CHANGE_FOR_LEADER_NON_MAIN

            short_ok = row['short_period_change'] >= short_threshold
            long_ok = row['long_period_change'] >= long_threshold
            long_not_too_high = row['long_period_change'] < long_max_threshold

            # 短周期或长周期满足一个即可，且长周期涨幅不能超过上限
            return (short_ok | long_ok) & long_not_too_high

        change_mask = candidate_df.apply(check_change_threshold, axis=1)

        # 3. 组合筛选条件
        qualified_df = candidate_df[board_mask & change_mask].copy()

        # 删除临时添加的market_type列
        if 'market_type' in qualified_df.columns:
            qualified_df = qualified_df.drop(columns=['market_type'])

        if qualified_df.empty:
            print(f"    概念组 {concept_group} 无符合条件的股票")
            continue

        qualified_concepts += 1
        print(f"    概念组 {concept_group} 有{len(qualified_df)}只符合条件的股票")

        # 选领涨股的排序逻辑，排序后跳着选
        qualified_df = qualified_df.sort_values(
            by=['long_period_change', 'short_period_change', 'max_board_level'],
            ascending=[False, False, False]
        )

        # 根据动态名额选出龙头
        leaders = qualified_df.head(quota)
        print(f"    从概念组 {concept_group} 选出{len(leaders)}只龙头股")

        # 添加到龙头股列表
        for _, leader in leaders.iterrows():
            leader_stocks.append(leader.to_dict())

    print(
        f"\n龙头股筛选完成: 处理了{total_concepts}个概念组，{qualified_concepts}个概念组有符合条件的股票，共选出{len(leader_stocks)}只龙头股")

    if not leader_stocks:
        return pd.DataFrame()

    # 转换为DataFrame
    leader_df = pd.DataFrame(leader_stocks)

    # 按概念优先级排序后，再按入选日期和涨幅排序（与概念分组sheet保持一致的分组效果）
    leader_df = leader_df.sort_values(
        by=['concept_priority', 'concept_group', 'first_significant_date', 'long_period_change', 'board_level_at_first'],
        ascending=[True, True, True, False, False]
    )

    return leader_df


def create_leader_stocks_sheet_content(ws, concept_grouped_df, shouban_df, stock_data, stock_entry_count,
                                       formatted_trading_days, date_column_start, show_period_change, period_column,
                                       period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                       max_tracking_days_before, zaban_df):
    """
    创建龙头股工作表内容
    
    Args:
        ws: Excel工作表
        concept_grouped_df: 按概念分组且已计算长周期涨跌幅的DataFrame
        其他参数与create_concept_grouped_sheet_content相同
    """
    print(f"开始创建龙头股工作表内容")

    # 选出龙头股
    leader_df = select_leader_stocks_from_concept_groups(
        concept_grouped_df, date_mapping, formatted_trading_days, period_days, period_days_long
    )

    if leader_df.empty:
        print("未找到符合条件的龙头股，跳过龙头股工作表创建")
        return

    print(f"成功筛选出{len(leader_df)}只龙头股，开始填充工作表内容")

    # 复用现有的表头和数据填充逻辑
    date_columns = setup_excel_header(ws, formatted_trading_days, show_period_change, period_days, date_column_start)

    # 添加大盘指标行
    add_market_indicators(ws, date_columns, label_col=2)

    # 填充龙头股数据行
    fill_data_rows(ws, leader_df, shouban_df, stock_data['stock_reason_group'], stock_data['reason_colors'],
                   stock_entry_count, formatted_trading_days, date_column_start, show_period_change,
                   period_column, period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                   max_tracking_days_before, zaban_df)

    # 调整列宽
    adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, False)  # 龙头股工作表不显示预警列

    # 冻结前三列和前三行
    ws.freeze_panes = ws.cell(row=4, column=date_column_start)

    print(f"龙头股工作表内容创建完成")


if __name__ == "__main__":
    import argparse

    # 命令行参数解析
    parser = argparse.ArgumentParser(description='生成梯队形态的涨停复盘图')
    parser.add_argument('--start_date', type=str, default="20230501",
                        help='开始日期 (格式: YYYYMMDD)')
    parser.add_argument('--end_date', type=str, default="20230531",
                        help='结束日期 (格式: YYYYMMDD)')
    parser.add_argument('--output', type=str, default=OUTPUT_FILE,
                        help=f'输出文件路径 (默认: {OUTPUT_FILE})')
    parser.add_argument('--min_board', type=int, default=2,
                        help='主板股票最小显著连板天数 (默认: 2)')
    parser.add_argument('--max_tracking', type=int, default=MAX_TRACKING_DAYS_AFTER_BREAK,
                        help=f'断板后跟踪的最大天数 (默认: {MAX_TRACKING_DAYS_AFTER_BREAK}，设为-1表示一直跟踪)')
    parser.add_argument('--reentry_days', type=int, default=REENTRY_DAYS_THRESHOLD,
                        help=f'断板后再次达到2板作为新行的天数阈值 (默认: {REENTRY_DAYS_THRESHOLD})')
    parser.add_argument('--non_main_board', type=int, default=1,
                        help='非主板股票最小显著连板天数 (默认: 1)')
    parser.add_argument('--max_tracking_before', type=int, default=MAX_TRACKING_DAYS_BEFORE_ENTRY,
                        help=f'入选前跟踪的最大天数 (默认: {MAX_TRACKING_DAYS_BEFORE_ENTRY}，设为0表示不显示入选前走势)')
    parser.add_argument('--period_days', type=int, default=PERIOD_DAYS_CHANGE,
                        help=f'计算最近X个交易日的短周期涨跌幅 (默认: {PERIOD_DAYS_CHANGE})')
    parser.add_argument('--period_days_long', type=int, default=PERIOD_DAYS_LONG,
                        help=f'计算最近X个交易日的长周期涨跌幅 (默认: {PERIOD_DAYS_LONG})')
    parser.add_argument('--show_period_change', action='store_true',
                        help='是否显示周期涨跌幅列 (默认: 不显示)')
    parser.add_argument('--priority_reasons', type=str, default="",
                        help='优先选择的原因列表，使用逗号分隔 (例如: "旅游,房地产,AI")')
    parser.add_argument('--low_priority_reasons', type=str, default="",
                        help='低优先级原因列表，使用逗号分隔 (例如: "预期改善")，只有在没有其他分组可匹配时才使用')
    parser.add_argument('--enable_attention_criteria', action='store_true',
                        help='是否启用关注度榜入选条件：(board_level-1)连板后5天内两次入选关注度榜前20 (默认: 不启用)')
    parser.add_argument('--sheet_name', type=str, default=None,
                        help='工作表名称，如果不指定则使用默认名称；如果指定，则保留现有Excel文件内容，并在指定名称的工作表中添加数据')
    parser.add_argument('--volume_days', type=int, default=VOLUME_DAYS,
                        help=f'计算成交量比的天数，当天成交量与前X天平均成交量的比值 (默认: {VOLUME_DAYS})')
    parser.add_argument('--volume_ratio', type=float, default=VOLUME_RATIO_THRESHOLD,
                        help=f'成交量比高阈值，超过该值则在单元格中显示成交量比 (默认: {VOLUME_RATIO_THRESHOLD})')
    parser.add_argument('--volume_ratio_low', type=float, default=VOLUME_RATIO_LOW_THRESHOLD,
                        help=f'成交量比低阈值，低于该值则在单元格中显示成交量比 (默认: {VOLUME_RATIO_LOW_THRESHOLD})')
    parser.add_argument('--high_gain_threshold', type=float, default=HIGH_GAIN_TRACKING_THRESHOLD,
                        help=f'持续跟踪的涨幅阈值，如果股票在period_days天内涨幅超过此值，即便没有涨停也会继续跟踪 (默认: {HIGH_GAIN_TRACKING_THRESHOLD}%)')
    parser.add_argument('--create_leader_sheet', action='store_true',
                        help='是否创建龙头股工作表，从每个概念分组中筛选出最强的股票 (默认: 不创建)')
    parser.add_argument('--ma_slope_days', type=int, default=MA_SLOPE_DAYS,
                        help=f'计算均线斜率的天数，用于在股票简称后显示趋势标记 (默认: {MA_SLOPE_DAYS})')

    args = parser.parse_args()

    # 处理max_tracking参数
    max_tracking = None if args.max_tracking == -1 else args.max_tracking

    # 处理优先原因列表
    priority_reasons = [reason.strip() for reason in
                        args.priority_reasons.split(',')] if args.priority_reasons else None

    # 处理低优先级原因列表
    low_priority_reasons = [reason.strip() for reason in
                            args.low_priority_reasons.split(',')] if args.low_priority_reasons else None

    # 更新全局成交量参数
    VOLUME_DAYS = args.volume_days
    VOLUME_RATIO_THRESHOLD = args.volume_ratio
    VOLUME_RATIO_LOW_THRESHOLD = args.volume_ratio_low

    # 更新全局高涨幅跟踪阈值
    import analysis.ladder_chart as ladder_chart_module

    ladder_chart_module.HIGH_GAIN_TRACKING_THRESHOLD = args.high_gain_threshold

    # 更新全局均线斜率天数
    ladder_chart_module.MA_SLOPE_DAYS = args.ma_slope_days

    # 构建梯队图
    build_ladder_chart(args.start_date, args.end_date, args.output, args.min_board,
                       max_tracking, args.reentry_days, args.non_main_board,
                       args.max_tracking_before, args.period_days, args.period_days_long, args.show_period_change,
                       priority_reasons=priority_reasons, low_priority_reasons=low_priority_reasons,
                       enable_attention_criteria=args.enable_attention_criteria,
                       sheet_name=args.sheet_name, create_leader_sheet=args.create_leader_sheet)
