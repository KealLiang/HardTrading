from collections import Counter
from datetime import datetime

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def analyze_zt_reasons(excel_file='./excel/fupan_stocks.xlsx', start_date=None, end_date=None, top_n=20, plot=False):
    """
    分析涨停原因类别，对数据进行聚合统计
    
    参数:
    excel_file: Excel文件路径
    start_date: 开始日期，格式为 "YYYYMMDD"，为None时分析所有日期数据
    end_date: 结束日期，格式为 "YYYYMMDD"，为None时默认等于start_date（分析单日）
    top_n: 显示前多少个最常见的原因类别
    plot: 是否生成可视化图表
    
    返回:
    分析结果字典列表，每个字典包含单日分析数据
    """
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    # 读取连板数据和首板数据
    lianban_data = pd.read_excel(excel_file, sheet_name="连板数据", index_col=0)
    shouban_data = pd.read_excel(excel_file, sheet_name="首板数据", index_col=0)

    # 转换日期格式函数
    def convert_date_format(date_str):
        """将YYYYMMDD格式转换为YYYY年MM月DD日格式"""
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, "%Y%m%d").strftime("%Y年%m月%d日")
        except ValueError:
            # 如果已经是YYYY年MM月DD日格式，直接返回
            return date_str

    # 选择要分析的日期列
    all_dates = lianban_data.columns
    start_date_formatted = convert_date_format(start_date)
    end_date_formatted = convert_date_format(end_date) if end_date else start_date_formatted

    if start_date_formatted:
        if start_date_formatted not in all_dates:
            print(f"错误: 未找到开始日期 {start_date_formatted}")
            return []

        if end_date_formatted and end_date_formatted not in all_dates:
            print(f"错误: 未找到结束日期 {end_date_formatted}")
            return []

        # 获取日期范围内的所有日期
        date_indices = [list(all_dates).index(d) for d in all_dates if
                        d in [start_date_formatted, end_date_formatted] or
                        (d > start_date_formatted and d < end_date_formatted)]
        date_indices.sort()
        analysis_dates = [all_dates[i] for i in date_indices]
    else:
        analysis_dates = all_dates

    # 统计所有类别
    all_reasons = Counter()
    daily_results = []

    # 处理每日数据
    for date in analysis_dates:
        # 提取当日连板数据
        lianban_col = lianban_data[date].dropna()
        lianban_stocks = lianban_col.str.split(';').apply(lambda x: [item.strip() for item in x])

        lianban_df = pd.DataFrame(lianban_stocks.tolist(), columns=[
            '股票代码', '股票简称', '涨停开板次数', '最终涨停时间',
            '几天几板', '最新价', '首次涨停时间', '最新涨跌幅',
            '连续涨停天数', '涨停原因类别'
        ])

        # 提取当日首板数据
        shouban_col = shouban_data[date].dropna()
        shouban_stocks = shouban_col.str.split(';').apply(lambda x: [item.strip() for item in x])

        # 首板数据可能有不同的列结构，这里假设最后一列是涨停原因类别
        shouban_df = pd.DataFrame(shouban_stocks.tolist())
        shouban_reasons_col = shouban_df.iloc[:, -1] if not shouban_df.empty else pd.Series([])

        # 分析连板数据中的涨停原因
        lianban_reasons = Counter()
        for reason_str in lianban_df['涨停原因类别']:
            if pd.notna(reason_str):
                # 拆分"+"连接的原因
                reasons = [r.strip() for r in reason_str.split('+')]
                for r in reasons:
                    if r:  # 确保不是空字符串
                        lianban_reasons[r] += 1
                        all_reasons[r] += 1

        # 分析首板数据中的涨停原因
        shouban_reasons = Counter()
        for reason_str in shouban_reasons_col:
            if pd.notna(reason_str):
                # 拆分"+"连接的原因
                reasons = [r.strip() for r in reason_str.split('+')]
                for r in reasons:
                    if r:  # 确保不是空字符串
                        shouban_reasons[r] += 1
                        all_reasons[r] += 1

        # 合并两种数据的统计结果
        day_reasons = lianban_reasons + shouban_reasons

        # 存储当日结果
        daily_results.append({
            'date': date,
            'total_stocks': len(lianban_df) + len(shouban_df),
            'lianban_count': len(lianban_df),
            'shouban_count': len(shouban_df),
            'reasons': day_reasons,
            'lianban_df': lianban_df,
            'shouban_df': shouban_df
        })

    # 标准输出模式
    for day_data in daily_results:
        print(f"\n=== {day_data['date']} 涨停热点 ===")
        print(
            f"📊 涨停: {day_data['total_stocks']}只 (连板: {day_data['lianban_count']}, 首板: {day_data['shouban_count']})")
        print("\n📈 热点类别 (Top {}):\n".format(top_n))

        for i, (reason, count) in enumerate(day_data['reasons'].most_common(top_n), 1):
            print(f"{i}. {reason}: {count}次")

    # 打印所有日期统计结果（如果分析了多个日期）
    if len(daily_results) > 1:
        print("\n=== 所有日期涨停原因类别汇总 (Top {}) ===".format(top_n))

        for i, (reason, count) in enumerate(all_reasons.most_common(top_n), 1):
            print(f"{i}. {reason}: {count}次")

    # 如果需要绘图
    if plot and daily_results:
        # 如果仅有一天数据，则显示该天的图
        if len(daily_results) == 1:
            day_data = daily_results[0]
            plot_reason_distribution(day_data['date'], day_data['reasons'], top_n)
        else:
            # 如果有多天数据，显示汇总图
            plot_reason_distribution("汇总", all_reasons, top_n)

    # 返回分析结果，便于其他函数使用
    return daily_results


def plot_reason_distribution(date, reasons_counter, top_n=15):
    """
    绘制涨停原因分布图
    
    参数:
    date: 日期字符串
    reasons_counter: Counter对象，包含原因和频次
    top_n: 显示前多少个最常见的原因
    """
    # 获取前N个最常见的原因
    top_reasons = reasons_counter.most_common(top_n)

    # 提取原因和频次
    reasons = [reason for reason, _ in top_reasons]
    counts = [count for _, count in top_reasons]

    # 创建水平条形图
    plt.figure(figsize=(12, 8))

    # 绘制水平条形图
    bars = plt.barh(range(len(reasons)), counts, align='center', color='cornflowerblue')

    # 设置y轴标签
    plt.yticks(range(len(reasons)), reasons)

    # 添加数据标签
    for bar in bars:
        width = bar.get_width()
        plt.text(width + 0.3, bar.get_y() + bar.get_height() / 2,
                 f'{width:.0f}', ha='left', va='center')

    # 设置标题和标签
    plt.title(f'{date} 涨停原因类别分布 (Top {top_n})', fontsize=14)
    plt.xlabel('出现频次', fontsize=12)
    plt.ylabel('涨停原因', fontsize=12)
    plt.tight_layout()

    # 显示图表
    plt.show()


def get_latest_date_data(excel_file):
    """
    获取Excel文件中最新的日期
    
    参数:
    excel_file: Excel文件路径
    
    返回:
    最新的日期字符串，格式为 YYYYMMDD
    """
    try:
        lianban_data = pd.read_excel(excel_file, sheet_name="连板数据", index_col=0)
        latest_date = lianban_data.columns[-1]

        # 将日期从 "YYYY年MM月DD日" 转换为 "YYYYMMDD"
        dt = datetime.strptime(latest_date, "%Y年%m月%d日")
        return dt.strftime("%Y%m%d")
    except Exception as e:
        print(f"获取最新日期时出错: {e}")
        return None


def format_excel_sheet(worksheet, columns):
    """
    设置Excel工作表的格式：调整列宽并为不同日期设置交替背景色
    
    参数:
        worksheet: openpyxl的worksheet对象
        columns: 列名列表
    """
    # 定义背景颜色
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # 设置列宽 - 特殊处理宽列
    for i, column in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        # 特殊处理宽列
        if column == '涨停原因类别':
            worksheet.column_dimensions[col_letter].width = 45
        elif column == '覆盖热点':
            worksheet.column_dimensions[col_letter].width = 30
        else:
            worksheet.column_dimensions[col_letter].width = 15

    # 设置表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, column in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 应用交替背景色（按日期分组）
    current_date = None
    use_gray = True

    # 从第2行开始（跳过标题行）
    for row_idx in range(2, worksheet.max_row + 1):
        date_value = worksheet.cell(row=row_idx, column=1).value
        if date_value != current_date:
            current_date = date_value
            use_gray = not use_gray  # 切换颜色

        fill = light_gray_fill if use_gray else white_fill

        # 为该行的所有单元格设置背景色
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = fill


def save_to_excel(stock_scores, result_file='./excel/limit_up_history.xlsx', sheet_name='每日热门',
                  skip_existing_dates=True):
    """
    将分析结果保存到Excel文件
    
    参数:
    stock_scores: 股票得分列表
    result_file: 保存结果的Excel文件路径
    sheet_name: 工作表名称
    skip_existing_dates: 是否跳过已存在的日期数据
    
    返回:
    保存是否成功
    """
    if not stock_scores:
        print("没有数据需要保存")
        return False

    try:
        # 将结果转换为DataFrame
        results_df = pd.DataFrame(stock_scores)

        # 将列表类型转为字符串格式 "[item1, item2, item3]"
        if '覆盖热点' in results_df.columns:
            results_df['覆盖热点'] = results_df['覆盖热点'].apply(lambda x: str(x) if isinstance(x, list) else x)

        # 调整列顺序，确保日期在最前面
        if '日期' in results_df.columns:
            cols = ['日期'] + [col for col in results_df.columns if col != '日期']
            results_df = results_df[cols]

        # 将日期转换为YYYYMMDD格式方便排序
        # 创建临时列用于排序
        results_df['日期排序'] = results_df['日期'].apply(lambda x:
                                                          datetime.strptime(x, "%Y年%m月%d日").strftime(
                                                              "%Y%m%d") if isinstance(x, str) else "")

        # 排序：先按日期降序，再按总得分降序
        results_df = results_df.sort_values(['日期排序', '总得分'], ascending=[False, False])

        # 删除临时排序列
        results_df = results_df.drop('日期排序', axis=1)

        # 获取要处理的日期列表
        processing_dates = set(results_df['日期'].unique())

        # 尝试读取现有数据
        try:
            existing_df = pd.read_excel(result_file, sheet_name=sheet_name)

            # 如果需要跳过已存在的日期
            if skip_existing_dates and '日期' in existing_df.columns:
                existing_dates = set(existing_df['日期'].unique())

                # 找出已存在的日期
                dates_to_skip = processing_dates.intersection(existing_dates)
                if dates_to_skip:
                    print(f"跳过已存在的日期: {', '.join(sorted(dates_to_skip))}")

                    # 只保留不存在的日期数据
                    results_df = results_df[~results_df['日期'].isin(dates_to_skip)]

                    # 如果过滤后没有数据，则直接返回
                    if results_df.empty:
                        print("所有日期数据已存在，无需更新")
                        return True

            # 合并数据（新数据在前）
            combined_df = pd.concat([results_df, existing_df], ignore_index=True)

            # 去重（基于股票代码、日期和总得分的组合）
            combined_df = combined_df.drop_duplicates(subset=['股票代码', '日期', '总得分'])

            # 重新排序
            combined_df['日期排序'] = combined_df['日期'].apply(lambda x:
                                                                datetime.strptime(x, "%Y年%m月%d日").strftime(
                                                                    "%Y%m%d") if isinstance(x, str) else "")
            combined_df = combined_df.sort_values(['日期排序', '总得分'], ascending=[False, False])
            combined_df = combined_df.drop('日期排序', axis=1)

            results_df = combined_df
        except:
            # 如果没有现有数据或读取出错，就使用新数据
            pass

        # 在最终输出前进行一次四舍五入处理
        decimal_columns = ['关注度加分', '原始得分', '总得分']
        for col in decimal_columns:
            if col in results_df.columns:
                results_df[col] = results_df[col].round(3)

        # 准备写入Excel
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows

        # 检查文件是否存在，如果不存在则创建新的工作簿
        try:
            wb = openpyxl.load_workbook(result_file)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # 检查sheet是否存在，如果存在则删除
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        # 创建新的sheet
        ws = wb.create_sheet(title=sheet_name)

        # 将DataFrame写入到工作表
        for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 设置格式
        format_excel_sheet(ws, results_df.columns.tolist())

        # 保存工作簿
        wb.save(result_file)

        print(f"\n✅ 分析结果已保存到 {result_file} 的 '{sheet_name}' 工作表")
        return True
    except Exception as e:
        print(f"\n❌ 保存结果时出错: {e}")
        import traceback
        traceback.print_exc()  # 打印详细错误堆栈
        return False


def find_stocks_by_hot_themes(start_date=None, end_date=None, top_n=5, weight_factor=2,
                              attention_weight_factor=3, excel_file='./excel/fupan_stocks.xlsx',
                              save_result=True, result_file='./excel/limit_up_history.xlsx', skip_existing_dates=True):
    """
    根据热点类别找出覆盖多个热点的股票
    
    参数:
    start_date: 开始日期，格式为 "YYYYMMDD"，None时使用最新日期
    end_date: 结束日期，格式为 "YYYYMMDD"，None时等于start_date（单日）
    top_n: 获取排名前几的热点类别
    weight_factor: 权重因子，决定第一名热点与最后一名热点的权重比例
    attention_weight_factor: 关注度榜权重因子，决定第一名关注度与最后一名的权重比例
    excel_file: Excel文件路径
    save_result: 是否保存结果到Excel
    result_file: 保存结果的Excel文件路径
    skip_existing_dates: 是否跳过已存在的日期数据
    
    返回:
    无，直接打印结果
    """
    # 如果没有指定开始日期，获取最新日期
    if start_date is None:
        start_date = get_latest_date_data(excel_file)
        if start_date is None:
            print("无法获取有效日期")
            return

    # 分析涨停数据
    daily_results = analyze_zt_reasons(excel_file, start_date, end_date, top_n=top_n, plot=False)

    if not daily_results:
        print("未获取到有效数据")
        return

    # 读取关注度榜数据
    try:
        attention_data = pd.read_excel(excel_file, sheet_name="关注度榜", index_col=0)
    except Exception as e:
        print(f"读取关注度榜数据时出错: {e}")
        attention_data = None

    # 单独处理每一天的数据，不合并
    all_stock_scores = []

    # 处理每一天的数据
    for day_data in daily_results:
        current_date = day_data['date']
        day_reasons = day_data['reasons']

        # 获取当天热点排名
        if not day_reasons:
            print(f"未找到 {current_date} 的热点类别")
            continue

        # 获取第N个热点的频次（考虑并列情况）
        hot_reasons = day_reasons.most_common()
        nth_count = hot_reasons[min(top_n - 1, len(hot_reasons) - 1)][1]

        # 找出所有频次≥第N个热点频次的热点
        top_hot_reasons = [(reason, count) for reason, count in hot_reasons if count >= nth_count]

        # 打印热点排名
        print(f"\n🔥 {current_date} 热点类别 Top {len(top_hot_reasons)}:")
        for i, (reason, count) in enumerate(top_hot_reasons, 1):
            print(f"{i}. {reason}: {count}次")

        # 处理当天的股票数据
        lianban_df = day_data['lianban_df']
        shouban_df = day_data['shouban_df']

        # 合并当天的连板和首板数据
        if shouban_df.empty:
            day_stocks_df = lianban_df.copy()
        else:
            # 确保首板DataFrame的列与连板DataFrame匹配
            shouban_df_adjusted = pd.DataFrame()

            # 复制连板数据的列结构
            for col in lianban_df.columns:
                if col in shouban_df.columns:
                    shouban_df_adjusted[col] = shouban_df[col]
                else:
                    # 如果首板数据缺少某列，用空值填充
                    shouban_df_adjusted[col] = pd.NA

            # 合并当日数据
            day_stocks_df = pd.concat([lianban_df, shouban_df_adjusted], ignore_index=True)

        # 获取当天的关注度榜数据（如果存在）
        day_attention = None
        if attention_data is not None and current_date in attention_data.columns:
            day_attention = attention_data[current_date].dropna()
            day_attention_stocks = day_attention.str.split(';').apply(lambda x: [item.strip() for item in x])

            # 创建关注度榜DataFrame
            attention_df = pd.DataFrame()
            if not day_attention_stocks.empty:
                attention_df = pd.DataFrame(day_attention_stocks.tolist())
                # 假设第一列是股票代码，第二列是股票名称
                if attention_df.shape[1] >= 2:
                    attention_df.columns = ['股票代码', '股票简称'] + [f'列{i + 3}' for i in
                                                                       range(attention_df.shape[1] - 2)]

        # 计算每只股票的热点覆盖得分
        stock_scores = []

        # 计算权重系数
        # 使用线性插值计算权重: 从weight_factor到1的线性变化
        num_hot_reasons = len(top_hot_reasons)

        # 遍历当天的所有股票
        for _, stock_row in day_stocks_df.iterrows():
            # 提取股票信息
            stock_code = stock_row['股票代码']
            stock_name = stock_row['股票简称']
            stock_board = stock_row['几天几板'] if pd.notna(stock_row['几天几板']) else ''
            stock_reasons_str = stock_row['涨停原因类别'] if pd.notna(stock_row['涨停原因类别']) else ''

            # 拆分股票的涨停原因
            stock_reasons = [r.strip() for r in stock_reasons_str.split('+')] if stock_reasons_str else []

            # 计算得分：覆盖的热点越靠前，分值越高
            score = 0
            covered_hot_reasons = []

            for hot_reason_idx, (hot_reason, _) in enumerate(top_hot_reasons):
                # 计算权重: 从weight_factor递减到1
                # 使用线性插值
                weight = 1 + (weight_factor - 1) * (num_hot_reasons - 1 - hot_reason_idx) / max(1, num_hot_reasons - 1)

                if any(hot_reason in r for r in stock_reasons):
                    score += weight
                    covered_hot_reasons.append(hot_reason)

            # 检查股票是否在关注度榜中，如果在则加分
            attention_rank = -1
            attention_bonus = 0

            if day_attention is not None and not attention_df.empty:
                # 在关注度榜中查找该股票
                for idx, att_row in enumerate(day_attention_stocks):
                    if len(att_row) >= 2 and (att_row[0] == stock_code or att_row[1] == stock_name):
                        attention_rank = idx + 1  # 排名从1开始

                        # 计算关注度加分，排名越靠前加分越高
                        # 使用和热点相似的线性插值方法
                        total_attention_stocks = len(day_attention_stocks)
                        attention_bonus = 1 + (attention_weight_factor - 1) * (total_attention_stocks - idx - 1) / max(
                            1, total_attention_stocks - 1)
                        score += attention_bonus
                        break

            # 记录股票得分和覆盖的热点
            if score > 0:  # 只关注有覆盖热点的股票
                stock_scores.append({
                    '日期': current_date,  # 将日期放在最前面
                    '股票代码': stock_code,
                    '股票简称': stock_name,
                    '几天几板': stock_board,
                    '涨停原因类别': stock_reasons_str,
                    '覆盖热点': covered_hot_reasons,
                    '热点数量': len(covered_hot_reasons),
                    '关注度排名': attention_rank if attention_rank > 0 else '未上榜',
                    '关注度加分': attention_bonus,
                    '原始得分': score - attention_bonus,
                    '总得分': score
                })

        # 根据得分对股票排序
        stock_scores.sort(key=lambda x: x['总得分'], reverse=True)

        # 输出当天结果
        print(f"\n🏆 {current_date} 覆盖热点最多的股票:")
        for i, stock in enumerate(stock_scores[:15], 1):  # 只显示前15只
            covered_hot_str = ', '.join(stock['覆盖热点'])
            attention_info = ""
            if stock['关注度排名'] != '未上榜':
                attention_info = f" | 关注度排名: {stock['关注度排名']} (+{stock['关注度加分']:.2f}分)"

            print(
                f"{i}. {stock['股票代码']} {stock['股票简称']} | {stock['几天几板']} | 得分: {stock['总得分']:.2f}{attention_info} | 覆盖热点: {covered_hot_str}")
            print(f"   原始涨停原因: {stock['涨停原因类别']}")
            print()

        # 保存当天的分析结果
        all_stock_scores.extend(stock_scores)

    # 保存结果到Excel
    if save_result and all_stock_scores:
        save_to_excel(all_stock_scores, result_file, '每日热门', skip_existing_dates)

    return all_stock_scores


if __name__ == '__main__':
    # 文件路径
    excel_file = "E:/demo/MachineLearning/HardTrading/excel/fupan_stocks.xlsx"
    result_file = "E:/demo/MachineLearning/HardTrading/excel/limit_up_history.xlsx"

    # 获取最新日期
    latest_date = get_latest_date_data(excel_file)

    # 找出覆盖热点最多的股票，权重因子为2（即第一名热点权重是最后一名的2倍）
    # 单日分析，关注度榜权重为3
    find_stocks_by_hot_themes(start_date="20250425", top_n=5, weight_factor=2, attention_weight_factor=3,
                              excel_file=excel_file, save_result=True, result_file=result_file)

    # 多日分析示例
    # find_stocks_by_hot_themes(start_date="20250420", end_date="20250425", top_n=5, weight_factor=2, 
    #                         attention_weight_factor=3, excel_file=excel_file, save_result=True, result_file=result_file)

    # 其他使用示例:
    # 单日分析并生成图表
    # analyze_zt_reasons(excel_file, start_date="20250425", top_n=10, plot=True)

    # 多日分析并生成图表
    # analyze_zt_reasons(excel_file, start_date="20250420", end_date="20250425", top_n=10, plot=True)
