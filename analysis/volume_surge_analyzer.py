"""
爆量分歧转一致形态分析器

形态定义：
- 强势连板股在某日出现爆量（当日量能较近期明显放大）
- 但当日仍然上涨（今收 > 昨收，不要求涨停）
- 这种形态代表分歧后资金选择继续做多

核心目的：
- 寻找这类形态的规律
- 观察后续走势
- 分析什么时段什么形态的股票资金最愿意买入

作者：AI Assistant
版本：v1.1
日期：2025-12-23
"""

import logging
import re
from dataclasses import dataclass
from datetime import datetime
from typing import List, Dict, Optional, Any, Tuple, Union

import mplfinance as mpf
import pandas as pd
from tqdm import tqdm

from analysis.pattern_analyzer_base import (
    PatternAnalyzerBase, PatternAnalysisConfig, PatternInfo
)

# 关注度榜检查相关全局配置
# 统计最近N个交易日内的关注度榜数据
ATTENTION_DAYS_WINDOW = 1
# 关注度榜取前N名
ATTENTION_TOP_N = 15

# 调试跟踪配置（用于定位特定股票未入选原因）
# 格式：{股票代码: [日期列表]}，例如：{'002347': ['20251230']}
# 只有配置了待跟踪的股票代码和日期才会打印调试日志
# 示例：
#   DEBUG_TRACK_STOCKS = {'002347': ['20251230']}  # 跟踪002347在20251230的检测过程
#   DEBUG_TRACK_STOCKS = {'002347': []}  # 跟踪002347在所有日期的检测过程
#   DEBUG_TRACK_STOCKS = {}  # 不跟踪任何股票（默认）
DEBUG_TRACK_STOCKS: Dict[str, List[str]] = {}


@dataclass
class VolumeSurgeConfig(PatternAnalysisConfig):
    """
    爆量分歧转一致分析配置
    
    Attributes:
        volume_surge_ratio: 爆量阈值（当日量/前N日均量），可以是单个值或元组。
            如果是元组，数量必须等于continuous_surge_days，不同连续天数对应不同阈值：
            - 连续N天使用第1个阈值
            - 连续N-1天使用第2个阈值
            - ...
            - 单日使用第N个阈值
            例如：continuous_surge_days=2时，volume_surge_ratio=(2.0, 3.0)表示
            连续2天需要>=2.0，单日需要>=3.0
        volume_avg_days: 计算均量的天数，默认5天
        min_lianban_count: 最小连板数，只分析达到此连板数的股票，默认2
        min_pct_change: 信号日最小涨幅(%)，默认3.0，用于过滤大阴线
        continuous_surge_days: 连续爆量检测天数，默认2。如果单日爆量检测失败，
            会检查最近N日是否连续爆量上涨，使用连续爆量开始之前的均量作为基准
        enable_attention_criteria: 是否启用关注度榜入选条件，默认为False。
            启用时，对于在关注度榜中的股票，连板数要求减1（例如min_lianban_count=2时，关注度榜股票只需1板即可）
    """
    volume_surge_ratio: Union[float, Tuple[float, ...]] = 2.0
    volume_avg_days: int = 5
    min_lianban_count: int = 2
    min_pct_change: float = 3.0  # 信号日最小涨幅%
    continuous_surge_days: int = 2  # 连续爆量检测天数
    enable_attention_criteria: bool = False  # 是否启用关注度榜入选条件


class VolumeSurgeAnalyzer(PatternAnalyzerBase):
    """
    爆量分歧转一致形态分析器
    
    筛选逻辑：
    1. 从复盘数据中筛选达到min_lianban_count连板的股票
    2. 检测这些股票在指定日期范围内是否出现爆量日
    3. 爆量日需满足：当日量能 >= 前N日均量 * volume_surge_ratio
    4. 爆量日需满足：当日上涨（收盘价 > 前一日收盘价）
    """

    def __init__(self, config: VolumeSurgeConfig):
        """
        初始化分析器
        
        Args:
            config: 爆量分歧分析配置
        """
        super().__init__(config, "爆量分歧转一致")
        self.config: VolumeSurgeConfig = config

        # 验证 volume_surge_ratio 和 continuous_surge_days 的匹配关系
        if isinstance(config.volume_surge_ratio, (tuple, list)):
            if len(config.volume_surge_ratio) != config.continuous_surge_days:
                raise ValueError(
                    f"volume_surge_ratio 的数量({len(config.volume_surge_ratio)}) "
                    f"必须等于 continuous_surge_days({config.continuous_surge_days})"
                )
            # 转换为元组并验证所有值都是正数
            config.volume_surge_ratio = tuple(config.volume_surge_ratio)
            if any(ratio <= 0 for ratio in config.volume_surge_ratio):
                raise ValueError("volume_surge_ratio 的所有值必须大于0")
        else:
            # 单个值的情况，转换为元组以便统一处理
            config.volume_surge_ratio = (config.volume_surge_ratio,) * config.continuous_surge_days
        self.lianban_data = None
        self.shouban_data = None
        self.date_columns = []
        # 存储连板股候选池：{stock_code: {name, first_board_date, max_board}}
        self.lianban_candidates: Dict[str, Dict] = {}
        # 存储每个信号的详细数据，用于多信号场景
        self._signal_details: Dict[str, Dict[str, Dict]] = {}  # {code: {date: surge_info}}
        # 存储关注度榜股票集合（仅在启用关注度榜条件时加载）
        self.attention_stocks: set = set()
        # 存储按日期索引的关注度榜数据：{date_str: set(stock_codes)}
        # 用于回测时按日期快速获取关注度榜股票集合，避免重复加载
        self.attention_stocks_by_date: Dict[str, set] = {}

    def _should_debug(self, code: str, date_str: str) -> bool:
        """
        判断是否需要打印调试日志
        
        Args:
            code: 股票代码（6位或带后缀）
            date_str: 日期字符串 YYYYMMDD
            
        Returns:
            是否需要打印调试日志
        """
        if not DEBUG_TRACK_STOCKS:
            return False

        # 提取纯代码（去除市场后缀）
        pure_code = code.split('.')[0] if '.' in code else code

        # 检查是否在跟踪列表中
        if pure_code in DEBUG_TRACK_STOCKS:
            track_dates = DEBUG_TRACK_STOCKS[pure_code]
            # 如果日期列表为空，表示跟踪该股票的所有日期
            if not track_dates:
                return True
            # 检查是否在指定日期列表中
            return date_str in track_dates

        return False

    def _debug_log(self, code: str, date_str: str, message: str):
        """
        打印调试日志（仅在配置了跟踪目标时）
        
        Args:
            code: 股票代码
            date_str: 日期字符串
            message: 日志消息
        """
        if self._should_debug(code, date_str):
            pure_code = code.split('.')[0] if '.' in code else code
            logging.info(f"[调试跟踪] {pure_code} @ {date_str}: {message}")

    def load_data(self):
        """加载连板数据和首板数据"""
        logging.info(f"正在加载连板数据: {self.config.fupan_file}")

        try:
            # 读取连板数据
            self.lianban_data = pd.read_excel(
                self.config.fupan_file,
                sheet_name="连板数据",
                index_col=0
            )

            # 读取首板数据
            self.shouban_data = pd.read_excel(
                self.config.fupan_file,
                sheet_name="首板数据",
                index_col=0
            )

            logging.info(f"连板数据加载成功，共 {len(self.lianban_data)} 行")
            logging.info(f"首板数据加载成功，共 {len(self.shouban_data)} 行")

            # 提取并过滤日期列
            self._filter_date_columns()

            # 如果启用关注度榜条件，加载关注度榜数据
            if self.config.enable_attention_criteria:
                self._load_attention_stocks()

        except FileNotFoundError:
            raise FileNotFoundError(f"未找到复盘数据文件: {self.config.fupan_file}")
        except Exception as e:
            raise Exception(f"加载连板数据失败: {e}")

    def _filter_date_columns(self):
        """过滤出指定日期范围内的列"""
        all_columns = self.lianban_data.columns.tolist()

        start_dt = datetime.strptime(self.config.start_date, '%Y%m%d')
        end_dt = datetime.strptime(self.config.end_date, '%Y%m%d')

        self.date_columns = []
        for col in all_columns:
            try:
                col_date = self._parse_column_date(col)
                if start_dt <= col_date <= end_dt:
                    self.date_columns.append(col)
            except:
                continue

        self.date_columns.sort(key=lambda x: self._parse_column_date(x))
        logging.info(f"筛选出 {len(self.date_columns)} 个日期列")

    def _parse_column_date(self, column_date: str) -> datetime:
        """将Excel列名日期转为datetime"""
        match = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', column_date)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            day = int(match.group(3))
            return datetime(year, month, day)
        else:
            raise ValueError(f"无法解析日期列: {column_date}")

    def _load_attention_stocks(self):
        """
        一次性加载所有关注度榜数据，按日期索引存储
        参考 ladder_chart.py 中的 load_top_attention_stocks 实现
        
        优化：一次性加载所有数据，避免重复读取Excel文件
        """
        try:
            from analysis.loader.fupan_data_loader import FUPAN_FILE
            from openpyxl import load_workbook

            # 读取关注度榜数据（从复盘数据源文件读取）
            wb = load_workbook(FUPAN_FILE, data_only=True)

            # 清空按日期索引的数据
            self.attention_stocks_by_date = {}

            # 处理两个sheet：【关注度榜】和【非主关注度榜】
            for sheet_name in ['关注度榜', '非主关注度榜']:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # 遍历所有列，加载所有日期的数据
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=1, column=col_idx)
                    if not header_cell.value:
                        continue

                    # 解析日期（格式：2025年11月18日）
                    col_date = self._parse_attention_date_from_header(header_cell.value)
                    if not col_date:
                        continue

                    # 初始化该日期的关注度榜集合
                    if col_date not in self.attention_stocks_by_date:
                        self.attention_stocks_by_date[col_date] = set()

                    # 读取该列的前top_n行数据（从第2行开始）
                    for row_idx in range(2, min(2 + ATTENTION_TOP_N, ws.max_row + 1)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if not cell_value:
                            continue

                        # 解析数据：600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1
                        stock_code = self._extract_stock_code_from_attention_data(cell_value)
                        if stock_code:
                            self.attention_stocks_by_date[col_date].add(stock_code)

            # 加载默认的关注度榜集合（基于end_date）
            self._get_attention_stocks_for_date(self.config.end_date)

            logging.info(
                f"✓ 加载关注度榜数据：共{len(self.attention_stocks_by_date)}个日期，"
                f"默认集合（end_date={self.config.end_date}）共{len(self.attention_stocks)}只股票")

        except Exception as e:
            logging.warning(f"✗ 加载关注度榜数据失败: {e}")
            self.attention_stocks = set()
            self.attention_stocks_by_date = {}

    def _get_attention_stocks_for_date(self, max_date: str) -> set:
        """
        获取指定日期及之前N个交易日内的关注度榜股票集合
        
        Args:
            max_date: 最大日期（YYYYMMDD格式）
            
        Returns:
            关注度榜股票集合
        """
        from utils.date_util import get_n_trading_days_before

        # 获取最近N个交易日
        recent_dates = set()
        recent_dates.add(max_date)
        for i in range(1, ATTENTION_DAYS_WINDOW):
            prev_date = get_n_trading_days_before(max_date, i)
            if '-' in prev_date:
                prev_date = prev_date.replace('-', '')
            recent_dates.add(prev_date)

        # 合并这些日期的关注度榜股票
        attention_stocks = set()
        for date_str in recent_dates:
            if date_str in self.attention_stocks_by_date:
                attention_stocks.update(self.attention_stocks_by_date[date_str])

        return attention_stocks

    def _parse_attention_date_from_header(self, header_value) -> Optional[str]:
        """
        从表头解析日期：2025年11月18日 -> 20251118
        
        Args:
            header_value: 表头值
            
        Returns:
            str: YYYYMMDD格式的日期，解析失败返回None
        """
        try:
            if isinstance(header_value, str) and '年' in header_value:
                date_obj = datetime.strptime(header_value, '%Y年%m月%d日')
                return date_obj.strftime('%Y%m%d')
        except:
            pass
        return None

    def _extract_stock_code_from_attention_data(self, cell_value) -> Optional[str]:
        """
        从关注度榜数据中提取股票代码
        
        输入: "600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1"
        输出: "600340"（标准化后的纯代码）
        
        Args:
            cell_value: 单元格值
            
        Returns:
            str: 标准化后的股票代码，解析失败返回None
        """
        try:
            parts = str(cell_value).split(';')
            if len(parts) >= 1:
                stock_code = parts[0].strip()  # "600340.SH"
                # 去除市场后缀 .SH/.SZ
                if '.' in stock_code:
                    stock_code = stock_code.split('.')[0]
                return stock_code
        except:
            pass
        return None

    def _parse_lianban_cell(self, cell_value: str) -> Optional[Dict]:
        """解析连板数据单元格"""
        if pd.isna(cell_value) or not cell_value:
            return None

        try:
            parts = [p.strip() for p in str(cell_value).split(';')]
            if len(parts) < 10:
                return None

            return {
                'code': parts[0],
                'name': parts[1],
                'kai_ban': parts[2],
                'final_zt_time': parts[3],
                'jitian_jiban': parts[4],
                'price': parts[5],
                'first_zt_time': parts[6],
                'pct_chg': parts[7],
                'continuous_days': int(parts[8]) if parts[8] and parts[8].isdigit() else 0,
                'reason': parts[9]
            }
        except Exception as e:
            logging.warning(f"解析单元格失败: {cell_value}, 错误: {e}")
            return None

    def _extract_board_count(self, jitian_jiban: str) -> int:
        """从"几天几板"中提取板数"""
        if not jitian_jiban:
            return 0

        match = re.search(r'(\d+)天(\d+)板', str(jitian_jiban))
        if match:
            return int(match.group(2))

        return 0

    def _build_lianban_candidates(self, max_date=None):
        """
        构建连板股候选池
        筛选在日期范围内达到min_lianban_count连板的股票
        
        注意：
        - min_lianban_count=1 时，需要同时读取首板数据
        - 启用关注度榜条件且 min_lianban_count-1=1 时，也需要读取首板数据（因为关注度榜股票可能只需1板）
        - min_lianban_count>=2 且未启用关注度榜条件时，只读取连板数据
        
        Args:
            max_date: 最大日期限制（YYYYMMDD格式），只收集该日期及之前的连板股。
                     如果为None，则使用config.end_date。回测时应该传入信号日，避免使用未来数据。
        """
        logging.info("构建连板股候选池...")

        # 判断是否需要读取首板数据
        # 1. 如果 min_lianban_count=1，需要读取首板数据
        # 2. 如果启用关注度榜条件且降低后需要1板（min_lianban_count-1=1），也需要读取首板数据
        need_shouban = (self.config.min_lianban_count == 1) or \
                       (self.config.enable_attention_criteria and
                        max(1, self.config.min_lianban_count - 1) == 1)

        if need_shouban:
            self._collect_from_sheet(self.shouban_data, board_count_default=1, max_date=max_date)

        # 读取连板数据
        self._collect_from_sheet(self.lianban_data, board_count_default=None, max_date=max_date)

        logging.info(f"候选池中共有 {len(self.lianban_candidates)} 只连板股")

    def _collect_from_sheet(self, sheet_data, board_count_default=None, max_date=None):
        """
        从指定sheet收集候选股票
        
        Args:
            sheet_data: 数据sheet（连板数据或首板数据）
            board_count_default: 默认板数（首板时为1，连板时为None从数据中解析）
            max_date: 最大日期限制（YYYYMMDD格式），只收集该日期及之前的连板股。
                     如果为None，则使用config.end_date。回测时应该传入信号日，避免使用未来数据。
        """
        # 确定最大日期限制
        if max_date is None:
            max_date = self.config.end_date
        max_date_dt = datetime.strptime(max_date, '%Y%m%d')

        # 关键修复：获取该日期及之前的关注度榜股票集合
        attention_stocks_for_date = set()
        if self.config.enable_attention_criteria:
            attention_stocks_for_date = self._get_attention_stocks_for_date(max_date)

        for date_col in self.date_columns:
            if date_col not in sheet_data.columns:
                continue

            # 关键修复：只收集max_date及之前的连板股，避免使用未来数据
            col_date = self._parse_column_date(date_col)
            if col_date > max_date_dt:
                continue

            for idx, cell_value in sheet_data[date_col].items():
                parsed = self._parse_lianban_cell(cell_value)
                if not parsed:
                    continue

                code = parsed['code']
                # 提取纯代码（去除市场后缀）用于关注度榜匹配
                pure_code = code.split('.')[0] if '.' in code else code
                date_str = col_date.strftime('%Y%m%d')

                # 解析板数
                if board_count_default is not None:
                    board_count = board_count_default  # 首板固定为1
                else:
                    board_count = self._extract_board_count(parsed['jitian_jiban'])

                # 确定实际需要的最小连板数
                # 如果启用关注度榜条件且该股票在关注度榜中，则连板数要求减1
                required_board_count = self.config.min_lianban_count
                if self.config.enable_attention_criteria and pure_code in attention_stocks_for_date:
                    required_board_count = max(1, self.config.min_lianban_count - 1)

                # 只收集达到最小连板数的股票
                if board_count >= required_board_count:
                    if code not in self.lianban_candidates:
                        self.lianban_candidates[code] = {
                            'name': parsed['name'],
                            'max_board': board_count,
                            'lianban_dates': [],
                            'reason': parsed.get('reason', '')  # 涨停原因
                        }
                    else:
                        self.lianban_candidates[code]['max_board'] = max(
                            self.lianban_candidates[code]['max_board'],
                            board_count
                        )
                        # 更新为最新的涨停原因
                        if parsed.get('reason'):
                            self.lianban_candidates[code]['reason'] = parsed['reason']

                    # 记录涨停日期（用于后续检测爆量）
                    date_str = self._parse_column_date(date_col).strftime('%Y%m%d')
                    if date_str not in self.lianban_candidates[code]['lianban_dates']:
                        self.lianban_candidates[code]['lianban_dates'].append(date_str)
                else:
                    # 关键失败点1：板数不足，未进入候选池
                    if self._should_debug(pure_code, date_str):
                        self._debug_log(pure_code, date_str,
                                        f"✗ 未进入候选池: 板数{board_count} < 要求{required_board_count}")

    def _get_volume_surge_info(self, code_6digit: str, date_str: str) -> Optional[Dict]:
        """
        检测指定日期是否为爆量日（支持单日爆量和连续爆量检测）
        
        Args:
            code_6digit: 6位股票代码
            date_str: 日期字符串 YYYYMMDD
            
        Returns:
            爆量信息字典，如果不是爆量日返回None
        """
        from utils.backtrade.visualizer import read_stock_data

        stock_data = read_stock_data(code_6digit, self.config.data_dir)
        if stock_data is None or stock_data.empty:
            if self._should_debug(code_6digit, date_str):
                self._debug_log(code_6digit, date_str, "✗ 无法读取股票数据或数据为空")
            return None

        try:
            target_date = datetime.strptime(date_str, '%Y%m%d')

            # 关键修复：回测时只能使用目标日期及之前的数据，避免使用未来数据
            # 过滤掉目标日期之后的数据
            stock_data = stock_data[stock_data.index <= target_date].copy()
            if stock_data.empty:
                if self._should_debug(code_6digit, date_str):
                    self._debug_log(code_6digit, date_str, f"✗ 过滤后数据为空（可能停牌）")
                return None

            # 确保目标日期在数据范围内
            if target_date not in stock_data.index:
                # 尝试查找最接近的日期（用于处理日期格式微小差异）
                idx = stock_data.index.searchsorted(target_date, side='right') - 1
                if idx < 0 or idx >= len(stock_data):
                    if self._should_debug(code_6digit, date_str):
                        self._debug_log(code_6digit, date_str, "✗ 无法找到目标日期（可能停牌）")
                    return None
                found_date = stock_data.index[idx]

                # 关键修复：如果找到的日期早于目标日期，说明目标日期后无数据（停牌）
                if found_date < target_date:
                    if self._should_debug(code_6digit, date_str):
                        self._debug_log(code_6digit, date_str, f"✗ 目标日期无数据（停牌）")
                    return None

                target_date = found_date
                # 如果日期差距太大，放弃
                if abs((target_date - datetime.strptime(date_str, '%Y%m%d')).days) > 3:
                    if self._should_debug(code_6digit, date_str):
                        self._debug_log(code_6digit, date_str, f"✗ 日期差距过大")
                    return None

            idx = stock_data.index.get_loc(target_date)

            # 需要足够的历史数据计算均量
            if idx < self.config.volume_avg_days:
                if self._should_debug(code_6digit, date_str):
                    self._debug_log(code_6digit, date_str,
                                    f"✗ 历史数据不足（需要{self.config.volume_avg_days}天）")
                return None

            # 获取当日数据
            current_row = stock_data.iloc[idx]
            current_volume = current_row['Volume']
            current_close = current_row['Close']

            # 检查数据有效性（停牌日数据为空或成交量为0）
            if pd.isna(current_volume) or pd.isna(current_close) or current_volume <= 0:
                if self._should_debug(code_6digit, date_str):
                    self._debug_log(code_6digit, date_str, "✗ 当日数据无效（停牌）")
                return None

            # 获取前一日收盘价
            prev_close = stock_data.iloc[idx - 1]['Close']

            # 关键失败点3：未上涨
            if current_close <= prev_close:
                if self._should_debug(code_6digit, date_str):
                    self._debug_log(code_6digit, date_str,
                                    f"✗ 未上涨: 今收={current_close:.2f} <= 昨收={prev_close:.2f}")
                return None

            # 使用连续爆量检测（支持1到continuous_surge_days天，整合了单日爆量检测）
            result = self._check_continuous_surge(stock_data, idx, current_close, prev_close, code_6digit, date_str)
            if result:
                return result

            # 关键失败点4：爆量检测失败（具体原因在_check_continuous_surge中已记录）
            return None

        except Exception as e:
            logging.error(f"检测股票 {code_6digit} 在 {date_str} 的爆量信息时出错: {e}")
            return None

    def _check_continuous_surge(self, stock_data: pd.DataFrame, idx: int,
                                current_close: float, prev_close: float,
                                code_6digit: str = '', date_str: str = '') -> Optional[Dict]:
        """
        检测连续爆量（支持1到continuous_surge_days天，整合了单日爆量检测）
        
        逻辑：
        1. 从1天开始，尝试检测连续1天、2天、...、continuous_surge_days天的爆量
        2. 找到最长的满足条件的连续爆量期间
        3. 对于每个候选期间，计算连续爆量开始之前的均量作为基准
        4. 检查连续爆量期间的每一天是否都：
           - 上涨
           - 满足最小涨幅阈值
           - 单日量能 >= 基准均量 * volume_surge_ratio（严格每日都爆量）
        
        Returns:
            爆量信息字典，如果不是连续爆量日返回None
        """
        max_continuous_days = self.config.continuous_surge_days
        failed_reasons = []  # 汇总所有失败原因

        # 从最长的连续天数开始尝试，找到最长的满足条件的期间
        for continuous_days in range(max_continuous_days, 0, -1):
            # 需要足够的历史数据
            if idx < continuous_days + self.config.volume_avg_days:
                continue

            # 回溯检查最近 continuous_days 日
            surge_period_start = idx - continuous_days + 1
            if surge_period_start < 0:
                continue

            # 先计算连续爆量开始之前的均量作为基准（跳过连续爆量期间）
            # 需要往前找足够数量的有效交易日（跳过停牌日）
            base_avg_start = surge_period_start - self.config.volume_avg_days
            if base_avg_start < 0:
                continue

            # 获取基准期间的数据，过滤掉停牌日（开盘价为空或为0，或成交量为空或为0）
            base_period_data = stock_data.iloc[base_avg_start:surge_period_start]
            # 过滤有效交易日：开盘价和成交量都不为空且大于0
            valid_base_data = base_period_data[
                base_period_data['Open'].notna() &
                (base_period_data['Open'] > 0) &
                base_period_data['Volume'].notna() &
                (base_period_data['Volume'] > 0)
                ]

            # 如果有效交易日不足，继续往前找
            if len(valid_base_data) < self.config.volume_avg_days:
                # 往前扩展查找范围，直到找到足够的有效交易日
                extended_start = base_avg_start
                while extended_start > 0 and len(valid_base_data) < self.config.volume_avg_days:
                    extended_start -= 1
                    extended_data = stock_data.iloc[extended_start:surge_period_start]
                    valid_base_data = extended_data[
                        extended_data['Open'].notna() &
                        (extended_data['Open'] > 0) &
                        extended_data['Volume'].notna() &
                        (extended_data['Volume'] > 0)
                        ]

                # 如果还是不够，跳过这个连续天数
                if len(valid_base_data) < self.config.volume_avg_days:
                    continue

            base_volumes = valid_base_data['Volume']
            base_avg_volume = base_volumes.mean()

            if pd.isna(base_avg_volume) or base_avg_volume <= 0:
                continue

            # 检查连续爆量期间的每一天是否都满足条件
            surge_volumes = []
            surge_pct_changes = []
            volume_ratios = []  # 记录每一天的量比
            all_days_valid = True
            failed_reason = None  # 记录失败原因

            for i in range(surge_period_start, idx + 1):
                row = stock_data.iloc[i]
                check_date = stock_data.index[i].strftime('%Y%m%d')

                volume = row['Volume']
                close = row['Close']

                # 检查当前日数据有效性（停牌日）
                if pd.isna(volume) or pd.isna(close) or volume <= 0:
                    all_days_valid = False
                    failed_reason = f"{check_date}数据无效"
                    break

                # 往前找最近的有效交易日作为前一日（跳过停牌日）
                prev_close_val = None
                for j in range(i - 1, -1, -1):
                    prev_row = stock_data.iloc[j]
                    prev_open = prev_row['Open']
                    prev_close = prev_row['Close']
                    # 找到有效交易日（开盘价和收盘价都不为空且大于0）
                    if pd.notna(prev_open) and prev_open > 0 and pd.notna(prev_close) and prev_close > 0:
                        prev_close_val = prev_close
                        break

                # 如果找不到有效的前一日，跳过这个连续天数
                if prev_close_val is None:
                    all_days_valid = False
                    failed_reason = f"{check_date}找不到有效的前一日"
                    break

                # 必须上涨
                if close <= prev_close_val:
                    all_days_valid = False
                    failed_reason = f"{check_date}未上涨(今收{close:.2f}<=昨收{prev_close_val:.2f})"
                    break

                # 计算涨幅
                pct_change = (close - prev_close_val) / prev_close_val * 100

                # 每日涨幅都要满足最小涨幅阈值
                if pct_change < self.config.min_pct_change:
                    all_days_valid = False
                    failed_reason = f"{check_date}涨幅不足({pct_change:.2f}%<{self.config.min_pct_change}%)"
                    break

                # 关键：每一天都要单独与基准均量比较，必须满足爆量条件
                # 根据连续天数获取对应的阈值：连续N天使用第1个阈值，连续N-1天使用第2个阈值，以此类推
                threshold_index = max_continuous_days - continuous_days
                required_ratio = self.config.volume_surge_ratio[threshold_index]

                daily_volume_ratio = volume / base_avg_volume
                if daily_volume_ratio < required_ratio:
                    all_days_valid = False
                    failed_reason = f"{check_date}量比不足({daily_volume_ratio:.2f}<{required_ratio}, 成交量{volume:.0f}, 基准均量{base_avg_volume:.0f})"
                    break  # 如果某一天不满足爆量条件，尝试更短的期间

                surge_volumes.append(volume)
                surge_pct_changes.append(pct_change)
                volume_ratios.append(daily_volume_ratio)

            # 如果所有天都满足条件，返回结果
            if all_days_valid and surge_volumes:
                target_date = stock_data.index[idx]
                final_pct_change = surge_pct_changes[-1]
                final_volume = surge_volumes[-1]
                # 使用最小量比，表示连续爆量期间最保守的爆量倍数
                min_volume_ratio = min(volume_ratios)

                return {
                    'date': target_date.strftime('%Y%m%d'),
                    'volume_ratio': round(min_volume_ratio, 2),  # 使用最小量比
                    'pct_change': round(final_pct_change, 2),
                    'current_volume': final_volume,
                    'avg_volume': round(base_avg_volume, 0),
                    'current_close': current_close,
                    'prev_close': prev_close,
                    'surge_type': 'continuous' if continuous_days > 1 else 'single',  # 标记类型
                    'continuous_days': continuous_days  # 实际连续天数
                }
            else:
                # 记录失败原因（只记录第一次尝试的失败原因，通常是最有意义的）
                if failed_reason and not failed_reasons:
                    failed_reasons.append(failed_reason)

        # 关键失败点4：所有连续天数检测均失败，汇总失败原因
        if self._should_debug(code_6digit, date_str):
            if failed_reasons:
                self._debug_log(code_6digit, date_str, f"✗ 爆量检测失败: {failed_reasons[0]}")
            else:
                self._debug_log(code_6digit, date_str,
                                f"✗ 爆量检测失败: 历史数据不足或无法计算基准均量")
        return None

    def filter_stocks(self):
        """
        筛选符合爆量分歧转一致形态的股票
        
        形态定义：
        - t-1日：至少 min_lianban_count 连板（涨停日）
        - t日（信号日）：上涨 + 爆量（不要求涨停）
        
        信号日是涨停后的下一个交易日
        """
        logging.info("开始筛选爆量分歧转一致形态...")

        from utils.date_util import get_next_trading_day

        # 按日期逐个检测，确保回测时只使用该日期及之前的数据
        # 生成所有需要检测的信号日期
        signal_dates = set()
        for date_col in self.date_columns:
            lianban_date = self._parse_column_date(date_col).strftime('%Y%m%d')
            signal_date = get_next_trading_day(lianban_date)
            if signal_date and self.config.start_date <= signal_date <= self.config.end_date:
                signal_dates.add(signal_date)

        signal_dates = sorted(list(signal_dates))

        # 对每个信号日，只使用该日期及之前的数据来收集候选股和检测
        for signal_date in tqdm(signal_dates, desc="按日期检测爆量日"):
            # 清空候选池，重新收集该日期及之前的候选股
            # 关键修复：使用该信号日及之前的关注度榜数据
            self.lianban_candidates = {}
            self._build_lianban_candidates(max_date=signal_date)

            if not self.lianban_candidates:
                continue

            # 关键失败点2：检查配置了跟踪的股票是否在候选池中
            if DEBUG_TRACK_STOCKS:
                for track_code, track_dates in DEBUG_TRACK_STOCKS.items():
                    if not track_dates or signal_date in track_dates:
                        # 检查该股票是否在候选池中
                        found_in_candidates = False
                        for code in self.lianban_candidates.keys():
                            pure_code = code.split('.')[0] if '.' in code else code
                            if pure_code == track_code:
                                found_in_candidates = True
                                break
                        if not found_in_candidates:
                            self._debug_log(track_code, signal_date,
                                            f"✗ 不在候选池中（未达到最小连板数要求: {self.config.min_lianban_count}板）")

            # 遍历候选池，检测该信号日的爆量
            for code, info in self.lianban_candidates.items():
                code_6digit = code.split('.')[0] if '.' in code else code

                # 初始化信号详情存储
                if code not in self._signal_details:
                    self._signal_details[code] = {}

                # 遍历该股票的所有涨停日期
                for lianban_date in info['lianban_dates']:
                    # 信号日是涨停日的下一个交易日
                    expected_signal_date = get_next_trading_day(lianban_date)
                    if not expected_signal_date or expected_signal_date != signal_date:
                        continue

                    # 检测信号日是否满足：上涨 + 爆量
                    surge_info = self._get_volume_surge_info(code_6digit, signal_date)

                    if surge_info:
                        # 存储信号详情
                        self._signal_details[code][surge_info['date']] = surge_info

                        pattern_info = PatternInfo(
                            code=code,
                            name=info['name'],
                            pattern_date=surge_info['date'],
                            pattern_type="爆量分歧转一致",
                            extra_data={
                                'max_board': info['max_board'],
                                'volume_ratio': surge_info['volume_ratio'],
                                'pct_change': surge_info['pct_change'],
                                'current_volume': surge_info['current_volume'],
                                'avg_volume': surge_info['avg_volume'],
                                'reason': info.get('reason', ''),  # 涨停原因
                                'prev_lianban_date': lianban_date  # 前一日涨停日期
                            }
                        )
                        self.filtered_stocks.append(pattern_info)

        logging.info(f"筛选完成，符合条件的形态: {len(self.filtered_stocks)} 个")

    def get_chart_markers(self, pattern_info: PatternInfo, chart_df: pd.DataFrame) -> List:
        """获取图表标记配置（支持多个信号日期）"""
        addplots = []

        # 为每个信号日期添加标记
        surge_markers = [float('nan')] * len(chart_df)
        volume_markers = [float('nan')] * len(chart_df)

        for date_str in pattern_info.pattern_dates:
            pattern_date_dt = datetime.strptime(date_str, '%Y%m%d')

            try:
                idx = chart_df.index.searchsorted(pattern_date_dt, side='right') - 1
                if 0 <= idx < len(chart_df):
                    surge_markers[idx] = chart_df.iloc[idx]['Low'] * 0.97
                    volume_markers[idx] = chart_df.iloc[idx]['Volume']
            except:
                pass

        # 爆量日标记（蓝色向上三角）
        addplots.append(mpf.make_addplot(
            surge_markers,
            type='scatter',
            marker='^',
            color='blue',
            markersize=150,
            label=f'爆量日({len(pattern_info.pattern_dates)}个)'
        ))

        # 成交量面板标记
        addplots.append(mpf.make_addplot(
            volume_markers,
            type='scatter',
            marker='*',
            color='magenta',
            markersize=100,
            panel=1,
            label='爆量'
        ))

        return addplots

    def get_chart_title(self, pattern_info: PatternInfo) -> str:
        """获取图表标题"""
        extra = pattern_info.extra_data
        header = self._format_stock_header(pattern_info)
        signal_count = extra.get('signal_count', 1)

        if signal_count > 1:
            # 多信号情况
            all_signals = extra.get('all_signals', [])
            avg_volume_ratio = sum(s.get('volume_ratio', 0) for s in all_signals) / len(all_signals)
            dates_str = ', '.join(pattern_info.pattern_dates)
            title = (
                f"{header}\n"
                f"信号日期: {dates_str} ({signal_count}次)\n"
                f"平均量比: {avg_volume_ratio:.1f}倍 | 最高连板: {extra.get('max_board', 'N/A')}板"
            )
        else:
            # 单信号情况
            title = (
                f"{header}\n"
                f"形态日期: {pattern_info.pattern_date} | "
                f"量比: {extra.get('volume_ratio', 'N/A')}倍 | "
                f"涨幅: {extra.get('pct_change', 'N/A')}% | "
                f"最高连板: {extra.get('max_board', 'N/A')}板"
            )
        return title

    def get_summary_columns(self) -> List[str]:
        """获取汇总报告的列名列表"""
        return [
            '股票代码', '股票名称', '涨停原因', '信号次数', '形态日期', '量比',
            '当日涨幅%', '最高连板数', '图表路径'
        ]

    def build_summary_row(self, pattern_info: PatternInfo) -> Dict[str, Any]:
        """构建汇总报告的单行数据"""
        extra = pattern_info.extra_data
        signal_count = extra.get('signal_count', 1)

        # 文件名：日期在前便于排序
        filename = f"{pattern_info.pattern_date}_{pattern_info.name}.png"

        # 量比和涨幅（多信号时显示范围）
        if signal_count > 1:
            all_signals = extra.get('all_signals', [])
            volume_ratios = [s.get('volume_ratio', 0) for s in all_signals]
            pct_changes = [s.get('pct_change', 0) for s in all_signals]
            volume_ratio_str = f"{min(volume_ratios):.1f}-{max(volume_ratios):.1f}"
            pct_change_str = f"{min(pct_changes):.1f}-{max(pct_changes):.1f}"
        else:
            volume_ratio_str = str(extra.get('volume_ratio', ''))
            pct_change_str = str(extra.get('pct_change', ''))

        return {
            '股票代码': pattern_info.code,
            '股票名称': pattern_info.name,
            '涨停原因': extra.get('reason', ''),
            '信号次数': signal_count,
            '形态日期': ', '.join(pattern_info.pattern_dates),
            '量比': volume_ratio_str,
            '当日涨幅%': pct_change_str,
            '最高连板数': extra.get('max_board', ''),
            '图表路径': f"./{filename}"
        }

    def generate_html_charts(self, columns: int = 2) -> str:
        """
        生成HTML交互式图表（单个HTML文件）
        
        Args:
            columns: 横向并排显示的列数（1、2或3），默认2
        
        Returns:
            str: 生成的HTML文件路径
        """
        from analysis.volume_surge_html_chart import generate_html_charts_for_analyzer
        return generate_html_charts_for_analyzer(self, columns=columns)


if __name__ == '__main__':
    # 测试代码
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    config = VolumeSurgeConfig(
        start_date='20251101',
        end_date='20251220',
        volume_surge_ratio=2.0,
        volume_avg_days=5,
        min_lianban_count=2,
        before_days=30,
        after_days=10
    )

    analyzer = VolumeSurgeAnalyzer(config)
    analyzer.run()
