"""
龙头候选股后续表现统计（胜率 / 盈亏比 / 期望值）

数据来源：复盘 Excel 中「龙头*」工作表（可选合并龙头归档）。
T-1 为选出日（sheet 快照日），T 为买入日（T-1 次一交易日开盘价），统计至 T+x 收盘价（x 默认 3）。
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union

from openpyxl import Workbook, load_workbook

from analysis.helper.ladder_chart_helpers import get_stock_data
from analysis.html_gen.leader_sheet_html_chart import (
    _leader_archive_paths,
    _load_leader_snapshots_main_and_all,
)
from utils.date_util import get_latest_trade_date, get_next_trading_day

# 买入日 T 后第 x 个交易日收盘卖出，默认 x=3
DEFAULT_HOLD_DAYS_AFTER_BUY = 3

HoldDaysArg = Union[int, Sequence[int]]

# 名额日志行：存储芯片: 3; 柏诚股份, 博敏电子, (和远气体)
_QUOTA_LOG_LINE_RE = re.compile(r'^(.+?)[:：]\s*(\d+)\s*[;；]\s*(.+)$')


LEADER_TYPE_NORMAL = 'normal'
LEADER_TYPE_EXTRA = 'extra'


@dataclass
class LeaderSample:
    code: str
    name: str
    sheet_concept: str  # 题材概念列原文（标签串）
    concept_group: str  # 龙头筛选名额口径的概念组
    leader_type: str  # normal=普通龙头, extra=大龙股
    signal_date: str  # YYYY-MM-DD
    sheet_name: str


@dataclass
class TradeOutcome:
    sample: LeaderSample
    buy_date: str  # YYYYMMDD
    sell_date: str  # YYYYMMDD
    buy_price: float
    sell_price: float
    return_pct: float


@dataclass
class HorizonStats:
    label: str
    hold_days: int  # 买入日 T 后第 x 个交易日
    total_samples: int
    valid_samples: int
    skipped_samples: int
    win_rate: float
    avg_win: float
    avg_loss: float
    profit_loss_ratio: Optional[float]
    expectancy: float
    median_return: float
    avg_return: float


def normalize_hold_days(hold_days: HoldDaysArg) -> Tuple[int, ...]:
    """
    将 hold_days 规范为观测窗口元组。

    - 传 int x：统计 T+1 … T+x（含），例如 3 → (1, 2, 3)
    - 传序列：仅统计指定 x，例如 (1, 3, 5)
    """
    if isinstance(hold_days, int):
        if hold_days <= 0:
            raise ValueError('hold_days 须 >= 1')
        return tuple(range(1, hold_days + 1))
    days = tuple(int(d) for d in hold_days)
    if not days or any(d <= 0 for d in days):
        raise ValueError('hold_days 须均为正整数')
    return days


def _yyyymmdd_from_iso(date_iso: str) -> str:
    return date_iso.replace('-', '')


def _iso_from_yyyymmdd(date_yyyymmdd: str) -> str:
    return f"{date_yyyymmdd[:4]}-{date_yyyymmdd[4:6]}-{date_yyyymmdd[6:8]}"


def _normalize_stock_display_name(name: str) -> str:
    """去掉 Excel 中常见的 ↑ 等后缀，便于与名额日志中的股票名匹配。"""
    return re.sub(r'\s*↑.*$', '', str(name or '')).strip()


def parse_leader_quota_log(log_text: str) -> Dict[str, str]:
    """
    解析龙头 sheet 名额汇总日志，返回 {股票简称: 概念组}。

    日志格式与 ladder_chart 写入一致，例如::
        存储芯片: 3; 柏诚股份, 博敏电子, (和远气体)
    """
    mapping: Dict[str, str] = {}
    for raw_line in str(log_text or '').splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = _QUOTA_LOG_LINE_RE.match(line)
        if not m:
            continue
        group = m.group(1).strip()
        names_part = m.group(3).strip()
        if ', (' in names_part:
            names_part = names_part.split(', (', 1)[0]
        elif '，(' in names_part:
            names_part = names_part.split('，(', 1)[0]
        for part in re.split(r'[,，]', names_part):
            name = part.strip().strip('()（）')
            if name:
                mapping[name] = group
    return mapping


def _is_valid_stock_code(code_val) -> bool:
    if code_val is None:
        return False
    return bool(re.match(r'^\d{6}$', str(code_val).strip()))


def _is_quota_log_row(ws, row_idx: int) -> bool:
    b_val = ws.cell(row=row_idx, column=2).value
    if not b_val or not isinstance(b_val, str):
        return False
    if not re.search(r'[:：]\s*\d+\s*[;；]', b_val):
        return False
    return not _is_valid_stock_code(ws.cell(row=row_idx, column=1).value)


def _parse_stock_row(ws, row_idx: int) -> Optional[tuple[str, str, str]]:
    code_val = ws.cell(row=row_idx, column=1).value
    if not _is_valid_stock_code(code_val):
        return None
    code = str(code_val).strip()
    name = str(ws.cell(row=row_idx, column=3).value or '').strip()
    sheet_concept = str(ws.cell(row=row_idx, column=2).value or '').strip()
    return code, name, sheet_concept


def _split_entries_by_row_gap(entries: Sequence[tuple[int, str, str, str]]) -> List[List[tuple[int, str, str, str]]]:
    """按行号不连续（中间有空行）拆成多块。"""
    if not entries:
        return []
    groups: List[List[tuple[int, str, str, str]]] = [[entries[0]]]  # type: ignore[list-item]
    for item in entries[1:]:
        if item[0] - groups[-1][-1][0] > 1:
            groups.append([item])
        else:
            groups[-1].append(item)
    return groups


def _parse_leader_sheet_stock_rows(ws) -> List[tuple[str, str, str, str]]:
    """
    解析龙头 sheet 数据行，返回 (code, name, sheet_concept, leader_type)。

    布局与 create_leader_stocks_sheet_content 一致：
    - 名额日志之前：普通龙头
    - 名额日志之后（或无日志时空行之后）：大龙股
    """
    log_row: Optional[int] = None
    for row_idx in range(4, ws.max_row + 1):
        if _is_quota_log_row(ws, row_idx):
            log_row = row_idx
            break

    raw_entries: List[tuple[int, str, str, str]] = []
    for row_idx in range(4, ws.max_row + 1):
        parsed = _parse_stock_row(ws, row_idx)
        if parsed:
            raw_entries.append((row_idx, parsed[0], parsed[1], parsed[2]))

    results: List[tuple[str, str, str, str]] = []
    if log_row is not None:
        for row_idx, code, name, sheet_concept in raw_entries:
            leader_type = LEADER_TYPE_EXTRA if row_idx > log_row else LEADER_TYPE_NORMAL
            results.append((code, name, sheet_concept, leader_type))
        return results

    groups = _split_entries_by_row_gap(raw_entries)
    if len(groups) <= 1:
        for _, code, name, sheet_concept in raw_entries:
            results.append((code, name, sheet_concept, LEADER_TYPE_NORMAL))
        return results

    for gi, group in enumerate(groups):
        leader_type = LEADER_TYPE_NORMAL if gi == 0 else LEADER_TYPE_EXTRA
        for _, code, name, sheet_concept in group:
            results.append((code, name, sheet_concept, leader_type))
    return results


def _parse_quota_log_from_sheet(ws) -> Dict[str, str]:
    """扫描 sheet B 列，提取名额汇总日志。"""
    mapping: Dict[str, str] = {}
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=2).value
        if not val or not isinstance(val, str):
            continue
        if not re.search(r'[:：]\s*\d+\s*[;；]', val):
            continue
        mapping.update(parse_leader_quota_log(val))
    return mapping


def _build_snapshot_top_reasons(sheet_concepts: Iterable[str]) -> List[str]:
    """按当日 sheet 内概念标签估算热门概念顺序（与 ladder_chart 分组函数配合使用）。"""
    from utils.theme_color_util import extract_reasons, get_reason_colors

    all_concepts: List[str] = []
    for concept_str in sheet_concepts:
        all_concepts.extend(extract_reasons(concept_str))
    _, top_reasons = get_reason_colors(all_concepts)
    return list(top_reasons)


def resolve_concept_group(
    sheet_concept: str,
    stock_name: str,
    quota_name_map: Dict[str, str],
    top_reasons: Sequence[str],
) -> str:
    """
    解析概念组：优先名额日志（与龙头筛选名额一致），否则回退 ladder_chart 同款分组逻辑。
    """
    clean_name = _normalize_stock_display_name(stock_name)
    if clean_name in quota_name_map:
        return quota_name_map[clean_name]
    if stock_name in quota_name_map:
        return quota_name_map[stock_name]

    from analysis.ladder_chart import select_concept_group_for_stock

    return select_concept_group_for_stock(sheet_concept or '', None, list(top_reasons), None)


def _open_leader_workbooks(
    excel_path: str,
    use_leader_archive: bool,
    merge_leader_archive_splits: bool,
) -> Tuple[Workbook, List[Workbook]]:
    main_wb = load_workbook(excel_path, data_only=True)
    archive_wbs: List[Workbook] = []
    if use_leader_archive:
        for archive_path in _leader_archive_paths(excel_path, merge_leader_archive_splits):
            if os.path.exists(archive_path):
                archive_wbs.append(load_workbook(archive_path, data_only=True))
    return main_wb, archive_wbs


def _get_leader_sheet_ws(sheet_name: str, main_wb: Workbook, archive_wbs: Sequence[Workbook]):
    if sheet_name in main_wb.sheetnames:
        return main_wb[sheet_name]
    for wb in archive_wbs:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
    return None


def collect_leader_samples_from_snapshots(
    snapshots: Sequence[dict],
    main_wb: Workbook,
    archive_wbs: Sequence[Workbook],
) -> List[LeaderSample]:
    """从快照列表解析样本，并按龙头名额口径填充 concept_group。"""
    samples: List[LeaderSample] = []
    for snap in snapshots:
        sheet_name = snap['sheet_name']
        signal_date = snap['snapshot_date']
        ws = _get_leader_sheet_ws(sheet_name, main_wb, archive_wbs)
        if ws is None:
            continue

        quota_map = _parse_quota_log_from_sheet(ws)
        stock_rows = _parse_leader_sheet_stock_rows(ws)
        if not stock_rows:
            continue

        sheet_concepts = [row[2] for row in stock_rows]
        top_reasons = _build_snapshot_top_reasons(sheet_concepts)

        for code, name, sheet_concept, leader_type in stock_rows:
            concept_group = resolve_concept_group(sheet_concept, name, quota_map, top_reasons)
            samples.append(LeaderSample(
                code=str(code).strip(),
                name=name,
                sheet_concept=sheet_concept or '未分类',
                concept_group=concept_group or '其他',
                leader_type=leader_type,
                signal_date=signal_date,
                sheet_name=sheet_name,
            ))

    samples.sort(key=lambda s: (s.signal_date, s.code))
    return samples


def shift_trading_days(date_yyyymmdd: str, offset: int) -> Optional[str]:
    """自 date 起向后偏移 offset 个交易日（offset=0 返回原日）。"""
    if offset < 0:
        raise ValueError('offset 须 >= 0')
    cur = date_yyyymmdd
    for _ in range(offset):
        nxt = get_next_trading_day(cur)
        if not nxt:
            return None
        cur = nxt
    return cur


def _get_open_close(stock_code: str, date_yyyymmdd: str) -> tuple[Optional[float], Optional[float]]:
    _, row, _ = get_stock_data(stock_code, date_yyyymmdd)
    if row is None or row.empty:
        return None, None
    try:
        open_p = float(row.iloc[0]['开盘'])
        close_p = float(row.iloc[0]['收盘'])
    except (KeyError, TypeError, ValueError):
        return None, None
    if open_p <= 0 or close_p <= 0:
        return None, None
    return open_p, close_p


def evaluate_sample_at_hold_days(
    sample: LeaderSample,
    hold_days: int,
    latest_trade_date: str,
) -> Optional[TradeOutcome]:
    """
    T-1 选出，T 开盘买入，T+x 收盘观测。

    hold_days 为 x：自买入日 T 起向后 x 个交易日的收盘价。
    """
    signal_yyyymmdd = _yyyymmdd_from_iso(sample.signal_date)  # T-1
    buy_date = shift_trading_days(signal_yyyymmdd, 1)  # T
    if not buy_date:
        return None
    sell_date = shift_trading_days(buy_date, hold_days)  # T+x
    if not sell_date:
        return None
    if sell_date > latest_trade_date:
        return None

    buy_open, _ = _get_open_close(sample.code, buy_date)
    _, sell_close = _get_open_close(sample.code, sell_date)
    if buy_open is None or sell_close is None:
        return None

    return_pct = (sell_close / buy_open - 1.0) * 100.0
    return TradeOutcome(
        sample=sample,
        buy_date=buy_date,
        sell_date=sell_date,
        buy_price=buy_open,
        sell_price=sell_close,
        return_pct=return_pct,
    )


def compute_horizon_stats(
    outcomes: Sequence[Optional[TradeOutcome]],
    label: str,
    hold_days: int,
    total_samples: int,
) -> HorizonStats:
    valid = [o for o in outcomes if o is not None]
    returns = [o.return_pct for o in valid]
    wins = [r for r in returns if r > 0]
    losses = [r for r in returns if r <= 0]

    win_rate = (len(wins) / len(valid) * 100.0) if valid else 0.0
    avg_win = sum(wins) / len(wins) if wins else 0.0
    avg_loss = sum(abs(r) for r in losses) / len(losses) if losses else 0.0
    if avg_loss > 0:
        pl_ratio: Optional[float] = avg_win / avg_loss
    elif avg_win > 0:
        pl_ratio = None
    else:
        pl_ratio = 0.0
    expectancy = (win_rate / 100.0) * avg_win - (1.0 - win_rate / 100.0) * avg_loss
    median_return = float(sorted(returns)[len(returns) // 2]) if returns else 0.0
    avg_return = sum(returns) / len(returns) if returns else 0.0

    return HorizonStats(
        label=label,
        hold_days=hold_days,
        total_samples=total_samples,
        valid_samples=len(valid),
        skipped_samples=total_samples - len(valid),
        win_rate=win_rate,
        avg_win=avg_win,
        avg_loss=avg_loss,
        profit_loss_ratio=pl_ratio,
        expectancy=expectancy,
        median_return=median_return,
        avg_return=avg_return,
    )


def _format_pct(value: float, digits: int = 2) -> str:
    return f"{value:.{digits}f}%"


def _format_pl_ratio(value: Optional[float]) -> str:
    if value is None:
        return '∞'
    return f"{value:.2f}"


@dataclass
class SegmentReport:
    title: str
    samples: List[LeaderSample]
    overall_stats: List[HorizonStats]
    concept_stats: Dict[str, List[HorizonStats]]


def compute_stats_for_samples(
    samples: Sequence[LeaderSample],
    hold_days_list: Sequence[int],
    latest_trade_date: str,
) -> tuple[List[HorizonStats], Dict[str, List[HorizonStats]]]:
    overall_stats: List[HorizonStats] = []
    concept_outcomes: Dict[str, Dict[int, List[Optional[TradeOutcome]]]] = {}

    for hold_days in hold_days_list:
        label = f'T+{hold_days}'
        outcomes: List[Optional[TradeOutcome]] = []
        for sample in samples:
            outcome = evaluate_sample_at_hold_days(sample, hold_days, latest_trade_date)
            outcomes.append(outcome)
            concept_outcomes.setdefault(sample.concept_group, {}).setdefault(hold_days, []).append(outcome)
        overall_stats.append(
            compute_horizon_stats(outcomes, label, hold_days, total_samples=len(samples))
        )

    concept_stats: Dict[str, List[HorizonStats]] = {}
    for concept_group, by_hold_days in concept_outcomes.items():
        concept_sample_count = sum(1 for s in samples if s.concept_group == concept_group)
        row: List[HorizonStats] = []
        for hold_days in hold_days_list:
            outcomes = by_hold_days.get(hold_days, [])
            row.append(
                compute_horizon_stats(outcomes, f'T+{hold_days}', hold_days, total_samples=concept_sample_count)
            )
        concept_stats[concept_group] = row
    return overall_stats, concept_stats


def _append_overall_stats_table(lines: List[str], overall_stats: Sequence[HorizonStats]) -> None:
    lines.extend([
        '| 持有至 | 有效样本 | 跳过 | 胜率 | 均盈 | 均亏 | 盈亏比 | 期望值 | 平均收益 | 中位收益 |',
        '|--------|----------|------|------|------|------|--------|--------|----------|----------|',
    ])
    for st in overall_stats:
        lines.append(
            f"| {st.label} | {st.valid_samples} | {st.skipped_samples} | "
            f"{_format_pct(st.win_rate)} | {_format_pct(st.avg_win)} | {_format_pct(st.avg_loss)} | "
            f"{_format_pl_ratio(st.profit_loss_ratio)} | {_format_pct(st.expectancy)} | "
            f"{_format_pct(st.avg_return)} | {_format_pct(st.median_return)} |"
        )


def _append_concept_stats_section(
    lines: List[str],
    concept_stats: Dict[str, Sequence[HorizonStats]],
    heading: str,
) -> None:
    has_any = any(
        stats_row and stats_row[0].valid_samples >= 5
        for stats_row in concept_stats.values()
    )
    if not has_any:
        return
    lines.extend(['', heading, ''])
    for concept_group, stats_row in sorted(
        concept_stats.items(),
        key=lambda x: x[1][0].valid_samples if x[1] else 0,
        reverse=True,
    ):
        if not stats_row or stats_row[0].valid_samples < 5:
            continue
        lines.append(f'### {concept_group}')
        lines.append('')
        lines.append('| 持有至 | 有效样本 | 胜率 | 盈亏比 | 期望值 | 平均收益 |')
        lines.append('|--------|----------|------|--------|--------|----------|')
        for st in stats_row:
            lines.append(
                f"| {st.label} | {st.valid_samples} | {_format_pct(st.win_rate)} | "
                f"{_format_pl_ratio(st.profit_loss_ratio)} | {_format_pct(st.expectancy)} | "
                f"{_format_pct(st.avg_return)} |"
            )
        lines.append('')


def build_markdown_report(
    excel_path: str,
    snapshots: Sequence[dict],
    segment_reports: Sequence[SegmentReport],
    hold_days_list: Sequence[int],
    use_leader_archive: bool,
    merge_leader_archive_splits: bool,
    latest_trade_date: str,
) -> str:
    all_samples = segment_reports[0].samples if segment_reports else []
    normal_count = sum(1 for s in all_samples if s.leader_type == LEADER_TYPE_NORMAL)
    extra_count = sum(1 for s in all_samples if s.leader_type == LEADER_TYPE_EXTRA)

    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if all_samples:
        signal_dates = sorted({s.signal_date for s in all_samples})
        date_range = f"{signal_dates[0]} ~ {signal_dates[-1]}"
    else:
        date_range = '无'

    lines = [
        '# 龙头候选后续表现统计',
        '',
        f'- 生成时间：{generated_at}',
        f'- 数据文件：`{os.path.abspath(excel_path)}`',
        f'- 合并龙头归档：{"是" if use_leader_archive else "否（仅统计指定文件内龙头 sheet）"}',
        f'- 合并拆分归档：{"是" if merge_leader_archive_splits else "否"}',
        f'- 龙头 sheet 数：{len(snapshots)}',
        f'- 候选样本数：{len(all_samples)}（普通龙头 {normal_count}，大龙股 {extra_count}）',
        f'- 观测窗口：T+{"/T+".join(str(d) for d in hold_days_list)} 收盘价（T=买入日）',
        f'- 选出日范围（T-1）：{date_range}',
        f'- 数据截止交易日：{_iso_from_yyyymmdd(latest_trade_date)}',
        '',
        '## 统计口径',
        '',
        '- **T-1**：选出日（龙头 sheet 快照日，收盘后识别候选）',
        '- **T**：买入日（T-1 的次一交易日，**开盘价**买入）',
        f'- **观测**：T+x 日**收盘价**（传 int 时统计 x=1…{hold_days_list[-1] if hold_days_list else "?"}；当前窗口 T+{"/T+".join(str(d) for d in hold_days_list)}）',
        '- **普通龙头 / 大龙股**：按 sheet 布局区分（名额日志之上=普通龙头，之下=大龙股；无日志时以空行分块）',
        '- **概念组**：与龙头名额分配一致——优先读 sheet 内名额汇总日志；缺失时按 `select_concept_group_for_stock` 回退',
        '- **胜率**：收益率 > 0 的样本占比',
        '- **盈亏比**：盈利样本平均收益 / 亏损样本平均亏损（绝对值）',
        '- **期望值**：胜率 × 均盈 − (1 − 胜率) × 均亏',
        '- **跳过**：T+x 尚未到来（选出日太近）、或缺买入/卖出日行情数据的样本；不计入有效样本，但计入总样本',
    ]

    for report in segment_reports:
        lines.extend(['', f'## {report.title}', ''])
        _append_overall_stats_table(lines, report.overall_stats)
        _append_concept_stats_section(
            lines,
            report.concept_stats,
            f'### 按概念组（{report.title}，有效样本数 ≥ 5）',
        )

    lines.append('')
    return '\n'.join(lines)


def run_leader_performance_stats(
    excel_path: str = './excel/ladder_analysis.xlsx',
    output_markdown: Optional[str] = None,
    use_leader_archive: bool = False,
    merge_leader_archive_splits: bool = False,
    hold_days: HoldDaysArg = DEFAULT_HOLD_DAYS_AFTER_BUY,
) -> str:
    """
    读取龙头 sheet，统计 T 开盘买入、T+x 收盘卖出的表现，输出 Markdown。

    Args:
        excel_path: 复盘 Excel 路径（默认 ladder_analysis.xlsx，仅读该文件内龙头 sheet）
        output_markdown: 输出 md 路径；默认与 excel 同目录下 leader_performance_stats.md
        use_leader_archive: 为 True 时额外合并 {excel}_龙头归档.xlsx（及可选拆分归档）
        merge_leader_archive_splits: 是否额外合并 leader_archives 下拆分文件
        hold_days: 最大持有观测日 x：传 int 时统计 T+1…T+x（默认 3 → 1/2/3）；传序列则只统计指定 x

    Returns:
        写入的 Markdown 文件绝对路径
    """
    excel_path = os.path.abspath(excel_path)
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f'Excel 文件不存在: {excel_path}')

    if output_markdown is None:
        output_markdown = os.path.join(
            os.path.dirname(excel_path),
            'leader_performance_stats.md',
        )
    output_markdown = os.path.abspath(output_markdown)

    _, snapshots = _load_leader_snapshots_main_and_all(
        excel_path,
        use_leader_archive=use_leader_archive,
        merge_leader_archive_splits=merge_leader_archive_splits,
    )
    main_wb, archive_wbs = _open_leader_workbooks(
        excel_path, use_leader_archive, merge_leader_archive_splits
    )
    try:
        samples = collect_leader_samples_from_snapshots(snapshots, main_wb, archive_wbs)
    finally:
        main_wb.close()
        for wb in archive_wbs:
            wb.close()

    hold_days_list = normalize_hold_days(hold_days)

    latest_trade_date = get_latest_trade_date()
    if not latest_trade_date:
        raise RuntimeError('无法获取最近交易日')

    segment_defs = [
        ('整体（全部）', lambda s: True),
        ('普通龙头', lambda s: s.leader_type == LEADER_TYPE_NORMAL),
        ('大龙股', lambda s: s.leader_type == LEADER_TYPE_EXTRA),
    ]
    segment_reports: List[SegmentReport] = []
    for title, pred in segment_defs:
        seg_samples = [s for s in samples if pred(s)]
        overall, concept = compute_stats_for_samples(seg_samples, hold_days_list, latest_trade_date)
        segment_reports.append(SegmentReport(title=title, samples=seg_samples, overall_stats=overall, concept_stats=concept))

    markdown = build_markdown_report(
        excel_path=excel_path,
        snapshots=snapshots,
        segment_reports=segment_reports,
        hold_days_list=hold_days_list,
        use_leader_archive=use_leader_archive,
        merge_leader_archive_splits=merge_leader_archive_splits,
        latest_trade_date=latest_trade_date,
    )

    os.makedirs(os.path.dirname(output_markdown) or '.', exist_ok=True)
    with open(output_markdown, 'w', encoding='utf-8') as f:
        f.write(markdown)

    normal_n = sum(1 for s in samples if s.leader_type == LEADER_TYPE_NORMAL)
    extra_n = sum(1 for s in samples if s.leader_type == LEADER_TYPE_EXTRA)
    print(f'龙头表现统计已写入: {output_markdown}')
    print(
        f'样本数: {len(samples)}（普通龙头 {normal_n}，大龙股 {extra_n}）, '
        f'龙头 sheet: {len(snapshots)}, 合并归档: {"是" if use_leader_archive else "否"}'
    )
    for report in segment_reports:
        if not report.samples:
            print(f'  [{report.title}] 无样本')
            continue
        st0 = report.overall_stats[0] if report.overall_stats else None
        if st0:
            print(f"  [{report.title}] 样本 {len(report.samples)}")
        for st in report.overall_stats:
            print(
                f"    {st.label}: 有效 {st.valid_samples}, 胜率 {_format_pct(st.win_rate)}, "
                f"盈亏比 {_format_pl_ratio(st.profit_loss_ratio)}, 期望 {_format_pct(st.expectancy)}"
            )
    return output_markdown
