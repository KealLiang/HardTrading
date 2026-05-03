"""
成交量涨跌幅分析模块
用于创建基于成交量涨跌幅的涨停梯队分析表
"""

from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from analysis.ladder_chart import (
    get_stock_data, BORDER_STYLE, format_stock_code_cell,
    format_concept_cell, format_stock_name_cell, extract_pure_stock_code,
    get_stock_concept, calculate_max_board_level, get_market_marker,
    get_new_high_markers_cached, NEW_HIGH_MARKER
)
from utils.theme_color_util import add_market_indicators

# 成交量涨跌幅颜色映射 - 基于百分位数动态上色
VOLUME_CHANGE_COLORS = {
    "EXTREME_POSITIVE": "FFB000",  # 深黄色 - 极度放量
    "STRONG_POSITIVE": "FFCC00",  # 中黄色 - 大幅放量
    "MODERATE_POSITIVE": "FFE066",  # 浅黄色 - 明显放量
    "MILD_POSITIVE": "FFF2CC",  # 极浅黄色 - 适度放量
    "MILD_NEGATIVE": "CCFFCC",  # 浅绿色 - 适度缩量
    "MODERATE_NEGATIVE": "99FF99",  # 中绿色 - 明显缩量
    "STRONG_NEGATIVE": "66FF66",  # 深绿色 - 大幅缩量
    "EXTREME_NEGATIVE": "33FF33",  # 极深绿色 - 极度缩量
}

# 成交量上色配置
VOLUME_COLOR_CONFIG = {
    "COLOR_PERCENTAGE": 0.30,  # 上色比例：20%的数据会被上色
    "POSITIVE_RATIO": 0.6,  # 正值(放量)在上色数据中的比例：60%
    "NEGATIVE_RATIO": 0.4,  # 负值(缩量)在上色数据中的比例：40%
}

# 成交量趋势分析相关参数
VOLUME_MA_DAYS = 7  # 成交量均线天数
VOLUME_TREND_DAYS = 3  # 判断趋势的连续天数
VOLUME_RATIO_HIGH_THRESHOLD = 1.5  # 高活跃阈值（成交量/均线）
VOLUME_RATIO_LOW_THRESHOLD = 0.6  # 低活跃阈值
VOLUME_MA_SLOPE_THRESHOLD = 0.1  # 均线斜率阈值（日变化率）


def get_stock_daily_volume_change(stock_code, date_str_yyyymmdd):
    """
    获取指定股票在特定日期的成交量涨跌幅

    Args:
        stock_code: 股票代码
        date_str_yyyymmdd: 目标日期 (YYYYMMDD)

    Returns:
        float: 成交量涨跌幅百分比，如果数据不存在则返回None
    """
    try:
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)

        if df is None or target_row is None or target_row.empty or target_idx == 0:
            return None

        # 获取当天成交量
        current_volume = target_row['成交量'].values[0]

        # 获取前一天成交量
        prev_volume = df.iloc[target_idx - 1]['成交量']

        # 检查数据有效性
        if pd.isna(current_volume) or pd.isna(prev_volume) or prev_volume <= 0:
            return None

        # 计算成交量涨跌幅（返回倍数，不是百分比）
        volume_change = (current_volume / prev_volume) - 1

        return volume_change

    except Exception as e:
        print(f"获取股票 {stock_code} 在 {date_str_yyyymmdd} 的成交量涨跌幅时出错: {e}")
        return None


def calculate_volume_ma(stock_code, date_str_yyyymmdd, days=VOLUME_MA_DAYS):
    """
    计算指定股票在特定日期的成交量均线

    Args:
        stock_code: 股票代码
        date_str_yyyymmdd: 目标日期 (YYYYMMDD)
        days: 均线天数

    Returns:
        float: 成交量均线，如果数据不存在则返回None
    """
    try:
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)

        if df is None or target_row is None or target_row.empty:
            return None

        # 确保有足够的历史数据来计算均线
        if target_idx < days - 1:
            return None

        # 获取包含当天在内的前days天的成交量数据
        volume_data = df.iloc[target_idx - days + 1:target_idx + 1]['成交量'].values

        # 检查数据有效性
        if len(volume_data) != days or pd.isna(volume_data).any():
            return None

        # 计算均线
        volume_ma = volume_data.mean()

        return volume_ma

    except Exception as e:
        print(f"计算股票 {stock_code} 在 {date_str_yyyymmdd} 的成交量均线时出错: {e}")
        return None


def get_volume_trend_indicator(stock_code, date_str_yyyymmdd, formatted_trading_days, date_mapping):
    """
    获取成交量趋势指标

    Args:
        stock_code: 股票代码
        date_str_yyyymmdd: 目标日期 (YYYYMMDD)
        formatted_trading_days: 格式化的交易日列表
        date_mapping: 日期映射

    Returns:
        str: 趋势指标文字，如果无明显趋势则返回空字符串
    """
    try:
        # 获取当前日期的成交量和均线
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)
        if df is None or target_row is None or target_row.empty:
            return ""

        current_volume = target_row['成交量'].values[0]
        current_ma = calculate_volume_ma(stock_code, date_str_yyyymmdd)

        if pd.isna(current_volume) or current_ma is None or current_ma <= 0:
            return ""

        # 计算当前成交量相对均线的比值
        volume_ratio = current_volume / current_ma

        # 计算均线趋势（获取最近几个有数据的交易日的均线）
        ma_values = []
        current_date_idx = None

        # 找到当前日期在交易日列表中的位置
        for i, formatted_day in enumerate(formatted_trading_days):
            if date_mapping.get(formatted_day) == date_str_yyyymmdd:
                current_date_idx = i
                break

        if current_date_idx is None:
            return ""

        # 从当前日期往前收集有效的均线数据，跳过周末等无数据的日期
        collected_days = 0
        for i in range(current_date_idx + 1):  # 从当前日期往前遍历
            day_idx = current_date_idx - i
            if day_idx >= 0 and day_idx < len(formatted_trading_days):
                formatted_day = formatted_trading_days[day_idx]
                day_date = date_mapping.get(formatted_day)
                if day_date:
                    ma_value = calculate_volume_ma(stock_code, day_date)
                    if ma_value is not None:
                        ma_values.insert(0, ma_value)  # 插入到开头，保持时间顺序
                        collected_days += 1
                        if collected_days >= VOLUME_TREND_DAYS:
                            break

        # 如果收集到的数据不足，则不计算趋势
        if len(ma_values) < 2:
            # 数据不足时，只基于当前成交量和均线的比值判断
            if volume_ratio >= VOLUME_RATIO_HIGH_THRESHOLD:
                return "高量🔥"
            elif volume_ratio <= VOLUME_RATIO_LOW_THRESHOLD:
                return "低量💤"
            else:
                return ""

        # 计算均线斜率（最后一天相对第一天的变化率）
        ma_slope = (ma_values[-1] - ma_values[0]) / ma_values[0] if ma_values[0] > 0 else 0

        # 判断趋势
        if volume_ratio >= VOLUME_RATIO_HIGH_THRESHOLD:
            if ma_slope > VOLUME_MA_SLOPE_THRESHOLD:
                return "放量↗"
            else:
                return "高量🔥"
        elif volume_ratio <= VOLUME_RATIO_LOW_THRESHOLD:
            if ma_slope < -VOLUME_MA_SLOPE_THRESHOLD:
                return "缩量↘"
            else:
                return "低量💤"
        elif ma_slope > VOLUME_MA_SLOPE_THRESHOLD:
            return "量增↗"
        elif ma_slope < -VOLUME_MA_SLOPE_THRESHOLD:
            return "量减↘"

        return ""

    except Exception as e:
        print(f"获取股票 {stock_code} 在 {date_str_yyyymmdd} 的成交量趋势指标时出错: {e}")
        return ""


def get_color_for_volume_change(volume_change, thresholds=None):
    """
    根据成交量涨跌幅获取背景色

    Args:
        volume_change: 成交量涨跌幅（倍数，不是百分比）
        thresholds: 动态阈值字典，包含各级别的阈值

    Returns:
        str: 颜色代码，变化不大时返回None
    """
    if volume_change is None:
        return None

    # 如果没有提供动态阈值，使用固定阈值（向后兼容）
    if thresholds is None:
        if volume_change >= 3.0:
            return VOLUME_CHANGE_COLORS["EXTREME_POSITIVE"]
        elif volume_change >= 2.0:
            return VOLUME_CHANGE_COLORS["STRONG_POSITIVE"]
        elif volume_change >= 1.0:
            return VOLUME_CHANGE_COLORS["MODERATE_POSITIVE"]
        elif volume_change >= 0.5:
            return VOLUME_CHANGE_COLORS["MILD_POSITIVE"]
        elif volume_change <= -0.8:
            return VOLUME_CHANGE_COLORS["EXTREME_NEGATIVE"]
        elif volume_change <= -0.7:
            return VOLUME_CHANGE_COLORS["STRONG_NEGATIVE"]
        elif volume_change <= -0.5:
            return VOLUME_CHANGE_COLORS["MILD_NEGATIVE"]
        else:
            return None

    # 使用动态阈值
    if volume_change >= thresholds.get("extreme_positive", 3.0):
        return VOLUME_CHANGE_COLORS["EXTREME_POSITIVE"]
    elif volume_change >= thresholds.get("strong_positive", 2.0):
        return VOLUME_CHANGE_COLORS["STRONG_POSITIVE"]
    elif volume_change >= thresholds.get("moderate_positive", 1.0):
        return VOLUME_CHANGE_COLORS["MODERATE_POSITIVE"]
    elif volume_change >= thresholds.get("mild_positive", 0.5):
        return VOLUME_CHANGE_COLORS["MILD_POSITIVE"]
    elif volume_change <= thresholds.get("extreme_negative", -0.8):
        return VOLUME_CHANGE_COLORS["EXTREME_NEGATIVE"]
    elif volume_change <= thresholds.get("strong_negative", -0.7):
        return VOLUME_CHANGE_COLORS["STRONG_NEGATIVE"]
    elif volume_change <= thresholds.get("moderate_negative", -0.5):
        return VOLUME_CHANGE_COLORS["MODERATE_NEGATIVE"]
    elif volume_change <= thresholds.get("mild_negative", -0.3):
        return VOLUME_CHANGE_COLORS["MILD_NEGATIVE"]
    else:
        return None


def calculate_volume_change_thresholds(volume_changes):
    """
    根据成交量变化数据计算动态阈值

    Args:
        volume_changes: 成交量变化数据列表

    Returns:
        dict: 包含各级别阈值的字典
    """
    import numpy as np

    # 过滤掉None值
    valid_changes = [v for v in volume_changes if v is not None and not pd.isna(v)]

    if len(valid_changes) < 10:  # 数据太少时使用固定阈值
        return None

    valid_changes = np.array(valid_changes)

    # 分离正值和负值
    positive_changes = valid_changes[valid_changes > 0]
    negative_changes = valid_changes[valid_changes < 0]

    # 计算配置参数
    color_percentage = VOLUME_COLOR_CONFIG["COLOR_PERCENTAGE"]
    positive_ratio = VOLUME_COLOR_CONFIG["POSITIVE_RATIO"]
    negative_ratio = VOLUME_COLOR_CONFIG["NEGATIVE_RATIO"]

    thresholds = {}

    # 计算正值阈值（放量）
    if len(positive_changes) > 0:
        positive_count = int(len(valid_changes) * color_percentage * positive_ratio)
        if positive_count > 0:
            positive_sorted = np.sort(positive_changes)[::-1]  # 降序排列

            # 根据数量分配阈值
            if positive_count >= 4:
                thresholds["extreme_positive"] = positive_sorted[positive_count // 4 - 1]
                thresholds["strong_positive"] = positive_sorted[positive_count // 2 - 1]
                thresholds["moderate_positive"] = positive_sorted[positive_count * 3 // 4 - 1]
                thresholds["mild_positive"] = positive_sorted[positive_count - 1]
            elif positive_count >= 2:
                thresholds["strong_positive"] = positive_sorted[0]
                thresholds["mild_positive"] = positive_sorted[positive_count - 1]
            else:
                thresholds["mild_positive"] = positive_sorted[0]

    # 计算负值阈值（缩量）
    if len(negative_changes) > 0:
        negative_count = int(len(valid_changes) * color_percentage * negative_ratio)
        if negative_count > 0:
            negative_sorted = np.sort(negative_changes)  # 升序排列（负值越小越极端）

            # 根据数量分配阈值
            if negative_count >= 4:
                thresholds["extreme_negative"] = negative_sorted[negative_count // 4 - 1]
                thresholds["strong_negative"] = negative_sorted[negative_count // 2 - 1]
                thresholds["moderate_negative"] = negative_sorted[negative_count * 3 // 4 - 1]
                thresholds["mild_negative"] = negative_sorted[negative_count - 1]
            elif negative_count >= 2:
                thresholds["strong_negative"] = negative_sorted[0]
                thresholds["mild_negative"] = negative_sorted[negative_count - 1]
            else:
                thresholds["mild_negative"] = negative_sorted[0]

    return thresholds


def format_volume_change_cell(ws, row, col, volume_change, stock_code, current_date_obj, thresholds=None):
    """
    格式化成交量涨跌幅单元格

    Args:
        ws: Excel工作表
        row: 行索引
        col: 列索引
        volume_change: 成交量涨跌幅
        stock_code: 股票代码
        current_date_obj: 当前日期对象
        thresholds: 动态阈值字典

    Returns:
        单元格对象
    """
    cell = ws.cell(row=row, column=col)

    if pd.notna(volume_change):
        # 显示成交量涨跌幅，保留2位小数
        cell_value = f"{volume_change:.2f}"
        cell.value = cell_value

        # 设置背景色，使用动态阈值
        color = get_color_for_volume_change(volume_change, thresholds)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    else:
        cell.value = "停牌"

    # 设置单元格格式
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER_STYLE

    return cell


def process_volume_daily_cell(ws, row_idx, col_idx, stock, current_date_obj, formatted_day,
                              stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                              zaban_df, new_high_markers, thresholds=None):
    """
    处理成交量版本的每日单元格数据

    Args:
        ws: Excel工作表
        row_idx: 行索引
        col_idx: 列索引
        stock: 股票数据
        current_date_obj: 当前日期对象
        formatted_day: 格式化的日期字符串
        stock_details: 股票详细信息
        date_mapping: 日期映射
        max_tracking_days: 断板后跟踪的最大天数
        max_tracking_days_before: 入选前跟踪的最大天数
        zaban_df: 炸板数据
        new_high_markers: 新高标记

    Returns:
        最后连板日期（如果有的话）
    """
    stock_code = stock['stock_code']
    stock_name = stock['stock_name']
    all_board_data = stock['all_board_data']
    entry_date = stock['first_significant_date']

    pure_stock_code = extract_pure_stock_code(stock_code)
    date_yyyymmdd = date_mapping.get(formatted_day)

    # 检查是否有连板数据
    board_days = all_board_data.get(formatted_day)

    # 检查是否为炸板股票（优先使用缓存，如果没有缓存则重新计算）
    from analysis.helper.ladder_chart_helpers import get_cached_zaban_format, check_stock_in_zaban
    is_zaban = get_cached_zaban_format(pure_stock_code, formatted_day)
    if is_zaban is None:
        # 如果缓存中没有，重新计算
        is_zaban = check_stock_in_zaban(zaban_df, pure_stock_code, formatted_day)

    if pd.notna(board_days) and board_days > 0:
        # 有连板数据，显示连板信息（保持原有逻辑）
        from analysis.ladder_chart import format_board_cell
        cell, last_board_date = format_board_cell(
            ws, row_idx, col_idx, board_days, pure_stock_code,
            f"{stock_code}_{formatted_day}", stock_details, current_date_obj
        )

        # 如果是炸板股票，添加炸板格式
        if is_zaban:
            from analysis.ladder_chart import add_zaban_underline
            add_zaban_underline(cell)

        return last_board_date
    else:
        # 没有连板数据，显示成交量涨跌幅
        if date_yyyymmdd:
            volume_change = get_stock_daily_volume_change(pure_stock_code, date_yyyymmdd)
            cell = format_volume_change_cell(ws, row_idx, col_idx, volume_change,
                                             pure_stock_code, current_date_obj, thresholds)
        else:
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.border = BORDER_STYLE

        # 如果是炸板股票，添加炸板格式
        if is_zaban:
            from analysis.ladder_chart import add_zaban_underline
            add_zaban_underline(cell)

        return None


def fill_volume_data_rows_with_concept_groups(ws, result_df, shouban_df, stock_reason_group, reason_colors,
                                              stock_entry_count, formatted_trading_days, date_column_start,
                                              show_period_change, period_column, period_days, period_days_long,
                                              stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                                              zaban_df, show_warning_column=True, thresholds=None):
    """
    填充成交量版本的数据行，按概念分组并在组间添加分隔行
    
    Args:
        ws: Excel工作表
        result_df: 按概念分组排序的显著连板股票DataFrame
        show_warning_column: 是否显示异动预警列
        其他参数与原版本相同
    """
    # 计算所有股票的新高标记（使用缓存版本）
    new_high_markers = get_new_high_markers_cached(result_df, formatted_trading_days, date_mapping)

    current_concept_group = None
    row_idx = 4  # 从第4行开始（前3行是标题和大盘指标）

    for i, (_, stock) in enumerate(result_df.iterrows()):
        concept_group = stock.get('concept_group', '其他')

        # 如果概念组发生变化，插入概念组标题行
        if concept_group != current_concept_group:
            # 如果不是第一个概念组，先插入空行分隔
            if current_concept_group is not None:
                row_idx += 1  # 添加空行

            current_concept_group = concept_group

            # 插入概念组标题行
            concept_title_cell = ws.cell(row=row_idx, column=1, value=f"【{concept_group}】")
            concept_title_cell.font = Font(bold=True, size=12)
            concept_title_cell.alignment = Alignment(horizontal='left')

            # 设置概念组标题行的背景色
            if concept_group in reason_colors:
                bg_color = reason_colors[concept_group]
                concept_title_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
                # 如果背景色较深，使用白色字体
                if bg_color in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
                    concept_title_cell.font = Font(color="FFFFFF", bold=True, size=12)

            # 合并概念组标题行的所有列
            end_col = date_column_start + len(formatted_trading_days) - 1
            if show_period_change:
                end_col += 1
            if show_warning_column:
                end_col += 1

            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
            row_idx += 1

        # 填充股票数据行
        fill_single_volume_stock_row(ws, row_idx, stock, stock_reason_group, reason_colors, stock_entry_count,
                                     formatted_trading_days, date_column_start, show_period_change, period_column,
                                     period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                     max_tracking_days_before, zaban_df, new_high_markers, show_warning_column,
                                     thresholds)

        row_idx += 1


def fill_single_volume_stock_row(ws, row_idx, stock, stock_reason_group, reason_colors, stock_entry_count,
                                 formatted_trading_days, date_column_start, show_period_change, period_column,
                                 period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                 max_tracking_days_before, zaban_df, new_high_markers, show_warning_column=True,
                                 thresholds=None):
    """
    填充单只股票的成交量版本数据行
    
    Args:
        ws: Excel工作表
        row_idx: 行索引
        stock: 股票数据
        其他参数与原版本相同
    """
    # 提取基本股票信息
    stock_code = stock['stock_code']
    stock_name = stock['stock_name']
    all_board_data = stock['all_board_data']

    # 提取纯代码
    pure_stock_code = extract_pure_stock_code(stock_code)

    # 设置股票代码列（第一列）
    format_stock_code_cell(ws, row_idx, 1, pure_stock_code)

    # 获取概念
    concept = get_stock_concept(stock)

    # 设置概念列（第二列）
    stock_key = f"{stock_code}_{stock_name}"
    format_concept_cell(ws, row_idx, 2, concept, stock_key, stock_reason_group, reason_colors)

    # 计算股票的最高板数
    max_board_level = calculate_max_board_level(all_board_data)

    # 根据股票代码确定市场类型
    market_type = get_market_marker(pure_stock_code)

    # 设置股票名称列（第三列）
    end_date_yyyymmdd = max(date_mapping.values()) if date_mapping else None
    name_cell, apply_high_board_color = format_stock_name_cell(
        ws, row_idx, 3, stock_name, market_type, max_board_level, False,
        stock_code, stock_entry_count, end_date_yyyymmdd
    )

    # 处理周期涨跌幅列（如果启用）
    if show_period_change and period_column:
        from analysis.ladder_chart import format_period_change_cell
        entry_date = stock['first_significant_date']
        format_period_change_cell(ws, row_idx, period_column, pure_stock_code, stock_name,
                                  entry_date, period_days, period_days_long, end_date_yyyymmdd)

    # 填充日期列数据
    last_board_date = None
    for j, formatted_day in enumerate(formatted_trading_days):
        col_idx = date_column_start + j
        current_date_obj = datetime.strptime(formatted_day, '%Y年%m月%d日')

        # 处理每日单元格（成交量版本）
        cell_last_board_date = process_volume_daily_cell(
            ws, row_idx, col_idx, stock, current_date_obj, formatted_day,
            stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
            zaban_df, new_high_markers, thresholds
        )

        if cell_last_board_date:
            last_board_date = cell_last_board_date

        # 添加新高标记
        if stock_code in new_high_markers and new_high_markers[stock_code] == formatted_day:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.value = f"{cell.value}{NEW_HIGH_MARKER}"

    # 处理异动预警列（如果启用）
    if show_warning_column:
        warning_col = date_column_start + len(formatted_trading_days)
        if show_period_change:
            warning_col += 1

        # 在成交量版本中显示成交量趋势信息
        # 找到该股票最后一个有数据的交易日
        last_data_date = None
        for formatted_day_reverse in reversed(formatted_trading_days):
            date_yyyymmdd_check = date_mapping.get(formatted_day_reverse)
            if date_yyyymmdd_check:
                # 检查是否有连板数据或成交量数据
                if (all_board_data.get(formatted_day_reverse) is not None or
                        get_stock_daily_volume_change(pure_stock_code, date_yyyymmdd_check) is not None):
                    last_data_date = date_yyyymmdd_check
                    break

        volume_trend = ""
        if last_data_date:
            volume_trend = get_volume_trend_indicator(pure_stock_code, last_data_date,
                                                      formatted_trading_days, date_mapping)

        warning_cell = ws.cell(row=row_idx, column=warning_col, value=volume_trend)
        warning_cell.border = BORDER_STYLE
        warning_cell.alignment = Alignment(horizontal='center')
        warning_cell.font = Font(size=9)  # 设置小一号字体

        # 根据趋势类型设置颜色
        if volume_trend:
            if "放量" in volume_trend:
                # 放量↗ - 红色系
                warning_cell.fill = PatternFill(start_color="FFE6E6", fill_type="solid")  # 浅红色背景
                warning_cell.font = Font(color="CC0000", size=9, bold=True)  # 红色字体
            elif "缩量" in volume_trend:
                # 缩量↘ - 绿色系
                warning_cell.fill = PatternFill(start_color="E6F3E6", fill_type="solid")  # 浅绿色背景
                warning_cell.font = Font(color="006600", size=9, bold=True)  # 绿色字体
            elif "高量" in volume_trend:
                # 高量🔥 - 橙色系
                warning_cell.fill = PatternFill(start_color="FFF2E6", fill_type="solid")  # 浅橙色背景
                warning_cell.font = Font(color="FF6600", size=9, bold=True)  # 橙色字体
            elif "低量" in volume_trend:
                # 低量💤 - 蓝色系
                warning_cell.fill = PatternFill(start_color="E6F0FF", fill_type="solid")  # 浅蓝色背景
                warning_cell.font = Font(color="0066CC", size=9, bold=True)  # 蓝色字体
            # 温和变化（量增↗/量减↘）不上色


def create_volume_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                                formatted_trading_days, date_column_start, show_period_change,
                                                period_column, period_days, period_days_long, stock_details,
                                                date_mapping, max_tracking_days, max_tracking_days_before, zaban_df):
    """
    创建成交量版本的按概念分组工作表内容

    Args:
        ws: Excel工作表
        result_df: 显著连板股票DataFrame
        其他参数与原版本相同
    """
    print(f"填充成交量版本的按概念分组工作表内容")

    if result_df.empty:
        print("没有数据可显示")
        return

    # 获取股票概念数据
    stock_reason_group = stock_data['stock_reason_group']
    reason_colors = stock_data['reason_colors']

    # 计算长周期涨跌幅并重新排序
    # 由于calculate_long_period_change_for_df可能不存在，我们手动计算
    concept_grouped_df = result_df.copy()

    # 手动计算长周期涨跌幅
    def calculate_long_period_change(row):
        try:
            from analysis.ladder_chart import calculate_stock_period_change, get_n_trading_days_before
            stock_code = row['stock_code']
            end_date = date_mapping.get(formatted_trading_days[-1])
            if end_date:
                prev_date = get_n_trading_days_before(end_date, period_days_long)
                if '-' in prev_date:
                    prev_date = prev_date.replace('-', '')
                return calculate_stock_period_change(stock_code.split('_')[0] if '_' in stock_code else stock_code,
                                                     prev_date, end_date)
        except Exception as e:
            print(f"计算长周期涨跌幅时出错: {e}")
        return 0.0

    concept_grouped_df['long_period_change'] = concept_grouped_df.apply(calculate_long_period_change, axis=1)

    # 为【默默上涨】分组添加特殊排序字段
    def get_momo_sort_keys(row):
        if row.get('concept_group') == '默默上涨':
            momo_data = row.get('momo_data', {})
            # 提取成交额数值（去掉"亿"字符）
            volume_str = momo_data.get('区间成交额', '0')
            try:
                volume = float(volume_str.replace('亿', '')) if '亿' in str(volume_str) else 0.0
            except:
                volume = 0.0

            # 提取涨幅数值（去掉"%"字符）
            change_str = momo_data.get('区间涨跌幅', '0')
            try:
                change = float(change_str.replace('%', '')) if '%' in str(change_str) else 0.0
            except:
                change = 0.0

            return volume, change
        else:
            return 0.0, 0.0

    # 添加【默默上涨】排序字段
    concept_grouped_df[['momo_volume', 'momo_change']] = concept_grouped_df.apply(
        lambda row: pd.Series(get_momo_sort_keys(row)), axis=1
    )

    # 分别处理【默默上涨】分组和其他分组的排序
    momo_mask = concept_grouped_df['concept_group'] == '默默上涨'
    momo_df = concept_grouped_df[momo_mask].copy()
    other_df = concept_grouped_df[~momo_mask].copy()

    # 【默默上涨】分组：按概念优先级、成交额、涨幅倒序排列
    if not momo_df.empty:
        momo_df = momo_df.sort_values(
            by=['concept_priority', 'momo_volume', 'momo_change'],
            ascending=[True, False, False]
        )

    # 其他分组：按原版逻辑排序（概念优先级、概念组、首次显著连板日期、长周期涨跌幅、首次连板天数）
    if not other_df.empty:
        other_df = other_df.sort_values(
            by=['concept_priority', 'concept_group', 'first_significant_date', 'long_period_change',
                'board_level_at_first'],
            ascending=[True, True, True, False, False]
        )

    # 合并排序结果
    concept_grouped_df = pd.concat([other_df, momo_df], ignore_index=True)

    # 删除临时排序字段
    concept_grouped_df = concept_grouped_df.drop(columns=['momo_volume', 'momo_change'])

    print(f"按概念分组排序后的股票数量: {len(concept_grouped_df)}")

    # 设置表头，显示量趋势列（成交量版本专用）
    from analysis.ladder_chart import setup_excel_header
    show_warning_column = True
    date_columns = setup_excel_header(ws, formatted_trading_days, show_period_change, period_days,
                                      date_column_start, show_warning_column, period_days_long)

    # 修改表头标题为"量趋势"（成交量版本专用）
    if show_warning_column and formatted_trading_days:
        warning_col = len(formatted_trading_days) + date_column_start
        if show_period_change:
            warning_col += 1
        warning_cell = ws.cell(row=1, column=warning_col)
        warning_cell.value = "量趋势"
        warning_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 设置与其他表头一致的格式
        from analysis.ladder_chart import BORDER_STYLE
        warning_cell.border = BORDER_STYLE
        warning_cell.font = Font(bold=True, size=10)
        # 设置量趋势列的背景色为浅蓝色，区别于异动预警
        warning_cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")

    # 修改表头标题，标明这是成交量版本
    title_cell = ws.cell(row=1, column=1, value="成交量涨跌幅分析 - 涨停梯队按概念分组")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='left')

    # 添加大盘指标行
    add_market_indicators(ws, date_columns, label_col=2)

    # 计算动态阈值
    print("计算成交量变化动态阈值...")
    all_volume_changes = []
    for _, stock in concept_grouped_df.iterrows():
        stock_code = stock['stock_code'].split('_')[0] if '_' in stock['stock_code'] else stock['stock_code']
        for formatted_day in formatted_trading_days:
            date_str = date_mapping.get(formatted_day)
            if date_str:
                volume_change = get_stock_daily_volume_change(stock_code, date_str)
                if volume_change is not None:
                    all_volume_changes.append(volume_change)

    thresholds = calculate_volume_change_thresholds(all_volume_changes)
    if thresholds:
        print(f"动态阈值计算完成，将对{VOLUME_COLOR_CONFIG['COLOR_PERCENTAGE'] * 100:.0f}%的数据上色")
        print(f"  放量阈值: {thresholds}")
    else:
        print("使用固定阈值")

    # 填充数据行，使用成交量版本的填充函数
    fill_volume_data_rows_with_concept_groups(ws, concept_grouped_df, shouban_df, stock_reason_group,
                                              reason_colors, stock_entry_count, formatted_trading_days,
                                              date_column_start, show_period_change, period_column, period_days,
                                              period_days_long, stock_details, date_mapping, max_tracking_days,
                                              max_tracking_days_before, zaban_df, show_warning_column, thresholds)

    # 调整列宽
    from analysis.ladder_chart import adjust_column_widths
    adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, show_warning_column)

    # 冻结前三列和前三行
    freeze_cell = f"{get_column_letter(date_column_start)}4"
    ws.freeze_panes = freeze_cell


def create_volume_concept_grouped_sheet(wb, sheet_name, result_df, shouban_df, stock_data, stock_entry_count,
                                        formatted_trading_days, date_column_start, show_period_change, period_column,
                                        period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                        max_tracking_days_before, zaban_df):
    """
    创建成交量版本的按概念分组工作表

    Args:
        wb: Excel工作簿
        sheet_name: 工作表名称
        其他参数与原版本相同
    """
    print(f"创建成交量版本的按概念分组工作表: {sheet_name}")

    # 创建新的工作表
    ws = wb.create_sheet(title=sheet_name)

    # 填充内容
    create_volume_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                                formatted_trading_days, date_column_start, show_period_change,
                                                period_column, period_days, period_days_long, stock_details,
                                                date_mapping, max_tracking_days, max_tracking_days_before, zaban_df)
