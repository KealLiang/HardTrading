import glob
import logging
import os
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm  # 导入tqdm进度条库

from utils.date_util import get_next_trading_day, get_trading_days, format_date
from utils.stock_util import stock_limit_ratio

# 配置logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('limit_up_analysis')

# 配置
FUPAN_FILE_PATH = './excel/fupan_stocks.xlsx'
ASTOCKS_DIR = './data/astocks/'
RESULT_FILE_PATH = './excel/limit_up_history.xlsx'


def is_limit_up(row, stock_code):
    """判断股票是否达到涨停价格。"""
    pct_change = row.get('涨跌幅', 0)  # 涨跌幅百分比
    limit_ratio = stock_limit_ratio(stock_code)
    return pct_change >= limit_ratio * 100 - 0.1  # 考虑0.1%的误差


def get_stock_status_next_day(stock_code, current_date, stock_files_dict):
    """
    获取股票在下一个交易日的状态和高开情况。
    返回: 包含状态和是否高开的元组，状态为'promoted' (晋级), 'survived' (存活), 'died' (死亡)，或数据不可用时返回None
    
    参数:
        current_date: 格式为'YYYY-MM-DD'的日期字符串
    """
    # 将日期格式转换为YYYYMMDD以适配date_util函数
    current_date_compact = current_date.replace('-', '')
    next_date_compact = get_next_trading_day(current_date_compact)

    if not next_date_compact:
        logger.debug(f"日期 {current_date} 的下一个交易日未找到")
        return None

    # 将日期格式转回YYYY-MM-DD
    next_date = f"{next_date_compact[:4]}-{next_date_compact[4:6]}-{next_date_compact[6:8]}"

    if stock_code not in stock_files_dict:
        logger.debug(f"股票 {stock_code} 没有对应的数据文件")
        return None

    stock_file = stock_files_dict[stock_code]
    try:
        if os.path.getsize(stock_file) == 0:
            logger.debug(f"股票 {stock_code} 的数据文件为空")
            return None

        # 参考find_longtou.py中的数据读取方式
        stock_df = pd.read_csv(stock_file, header=None,
                               names=['日期', '股票代码', '开盘', '收盘', '最高', '最低', '成交量', '成交额',
                                      '振幅', '涨跌幅', '涨跌额', '换手率'])
        stock_df['日期'] = pd.to_datetime(stock_df['日期']).dt.strftime('%Y-%m-%d')

        # 获取当前日期(T日)数据
        current_day_data = stock_df[stock_df['日期'] == current_date]
        if current_day_data.empty:
            logger.debug(f"股票 {stock_code} 在日期 {current_date} 没有数据")
            return None

        # 获取下一日(T+1日)数据
        next_day_data = stock_df[stock_df['日期'] == next_date]
        if next_day_data.empty:
            logger.debug(f"股票 {stock_code} 在日期 {next_date} 没有数据")
            return None

        # 获取T日收盘价和T+1日开盘价
        current_close = current_day_data.iloc[0]['收盘']
        next_open = next_day_data.iloc[0]['开盘']
        
        # 判断是否高开（开盘价大于等于前一日收盘价）
        is_high_open = next_open >= current_close

        if is_limit_up(next_day_data.iloc[0], stock_code):
            return ('promoted', is_high_open)  # 晋级 (T+1日涨停)
        elif next_day_data.iloc[0]['涨跌幅'] > 0:
            return ('survived', is_high_open)  # 存活 (T+1日上涨)
        else:
            return ('died', is_high_open)  # 死亡 (T+1日下跌)
    except Exception as e:
        logger.error(f"处理股票 {stock_code} 日期 {current_date} 时出错: {e}")
        return None


def get_date_range(start_date, end_date=None):
    """
    获取从start_date到end_date之间的所有交易日期
    
    参数:
        start_date: 开始日期，格式可以是'YYYY-MM-DD'、'YYYYMMDD'或'YYYY年MM月DD日'
        end_date: 结束日期，格式同start_date，如果为None则等于start_date
        
    返回:
        list: 包含所有交易日期的列表，格式为'YYYY-MM-DD'
    """
    # 格式化日期
    formatted_start = format_date(start_date)
    formatted_end = format_date(end_date) if end_date else formatted_start

    if not formatted_start or not formatted_end:
        logger.error(f"日期格式无效: start_date={start_date}, end_date={end_date}")
        return []

    # 转换为YYYYMMDD格式以适配date_util函数
    start_compact = formatted_start.replace('-', '')
    end_compact = formatted_end.replace('-', '')

    try:
        # 获取交易日列表
        trading_days = get_trading_days(start_compact, end_compact)

        # 转换回YYYY-MM-DD格式
        formatted_days = [f"{d[:4]}-{d[4:6]}-{d[6:8]}" for d in trading_days]

        if not formatted_days:
            logger.warning(f"在 {formatted_start} 至 {formatted_end} 期间没有找到交易日")

        return formatted_days
    except Exception as e:
        logger.error(f"获取交易日期范围时出错: {e}")
        # 如果get_trading_days函数失败，则至少返回开始日期（假设它是交易日）
        logger.info(f"将仅使用单个日期: {formatted_start}")
        return [formatted_start]


def analyze_limit_up_progression(start_date, end_date=None):
    """
    分析涨停股票的晋级率。
    
    参数:
        start_date: 开始日期，格式可以是'YYYY-MM-DD'、'YYYYMMDD'或'YYYY年MM月DD日'
        end_date: 结束日期，格式同start_date，如果为None则等于start_date
    
    返回:
        dict: 按日期分组的连板晋级统计数据字典
    """
    try:
        # 获取要分析的日期范围
        date_list = get_date_range(start_date, end_date)
        if not date_list:
            logger.warning("没有找到符合条件的交易日")
            return {}

        logger.info(f"将分析以下日期: {date_list}")

        # 参考fupan_plot.py中的数据读取方式
        logger.info(f"正在读取Excel文件: {FUPAN_FILE_PATH}")
        lianban_data = pd.read_excel(FUPAN_FILE_PATH, sheet_name="连板数据", index_col=0)
        shouban_data = pd.read_excel(FUPAN_FILE_PATH, sheet_name="首板数据", index_col=0)

        all_limit_ups = []

        # 处理日期列表
        dates = lianban_data.columns
        filtered_dates = []
        for date in dates:
            try:
                # 尝试解析Excel文件中的日期列名
                date_obj = datetime.strptime(date, "%Y年%m月%d日")
                formatted_date = date_obj.strftime('%Y-%m-%d')

                # 如果不在指定日期范围内，则跳过
                if formatted_date not in date_list:
                    continue

                filtered_dates.append(date)
            except (ValueError, TypeError):
                logger.warning(f"无法解析日期列名: {date}")
                continue

        if not filtered_dates:
            logger.warning(f"在Excel文件中没有找到匹配的日期: {date_list}")
            return {}

        # 用于跟踪每种类型的记录数量
        record_counts = {"连板": 0, "首板": 0}
        stock_codes_by_type = {"连板": set(), "首板": set()}
        stock_details = []  # 存储详细的股票信息，便于调试

        # 按日期分组收集涨停数据
        date_grouped_limit_ups = {date: [] for date in date_list}

        # 处理连板数据
        for date in filtered_dates:
            # 解析标准日期格式
            date_obj = datetime.strptime(date, "%Y年%m月%d日")
            formatted_date = date_obj.strftime('%Y-%m-%d')

            # 连板数据处理
            lianban_col = lianban_data[date].dropna()  # 去除空单元格
            lianban_stocks = lianban_col.str.split(';').apply(lambda x: [item.strip() for item in x])  # 分列处理

            # 为每个股票创建处理好的数据
            for stock_info in lianban_stocks:
                if len(stock_info) >= 5:  # 确保有足够的数据项
                    stock_code = stock_info[0].split('.')[0]  # 移除交易所后缀
                    stock_name = stock_info[1]
                    stock_codes_by_type["连板"].add(stock_code)
                    record_counts["连板"] += 1

                    # 解析"几天几板"
                    board_info = stock_info[4]
                    board_count = 1
                    match = board_info.strip().split('天')
                    if len(match) > 1 and '板' in match[1]:
                        try:
                            board_count = int(match[1].replace('板', ''))
                        except:
                            logger.warning(f"无法解析连板数 '{board_info}'，使用默认值1")

                    stock_record = {
                        'date': formatted_date,
                        'stock_code': stock_code,
                        'stock_name': stock_name,
                        'board_count': board_count,
                        'type': '连板'
                    }
                    all_limit_ups.append(stock_record)
                    stock_details.append(stock_record)

                    # 按日期分组
                    date_grouped_limit_ups[formatted_date].append(stock_record)

            # 首板数据处理
            shouban_col = shouban_data[date].dropna()  # 去除空单元格

            shouban_stocks = shouban_col.str.split(';').apply(lambda x: [item.strip() for item in x])  # 分列处理

            for stock_info in shouban_stocks:
                if len(stock_info) >= 2:  # 确保有足够的数据项
                    stock_code = stock_info[0].split('.')[0]  # 移除交易所后缀
                    stock_name = stock_info[1]
                    stock_codes_by_type["首板"].add(stock_code)
                    record_counts["首板"] += 1

                    stock_record = {
                        'date': formatted_date,
                        'stock_code': stock_code,
                        'stock_name': stock_name,
                        'board_count': 1,  # 首板
                        'type': '首板'
                    }
                    all_limit_ups.append(stock_record)
                    stock_details.append(stock_record)

                    # 按日期分组
                    date_grouped_limit_ups[formatted_date].append(stock_record)

        # 打印处理统计
        logger.info(f"成功提取记录数: 连板 {record_counts['连板']} 条，首板 {record_counts['首板']} 条")
        logger.info(
            f"唯一股票数: 连板 {len(stock_codes_by_type['连板'])} 只，首板 {len(stock_codes_by_type['首板'])} 只")

        # 检查是否有数据
        if not all_limit_ups:
            logger.warning("没有找到符合条件的涨停数据。")
            return {}

        # 转换为DataFrame
        limit_up_df = pd.DataFrame(all_limit_ups)
        logger.info(f"处理完成，共找到 {len(limit_up_df)} 条涨停记录。")

        # 创建股票文件字典以便快速查找
        stock_files_dict = {}
        available_files = glob.glob(os.path.join(ASTOCKS_DIR, '*.csv'))
        logger.info(f"在 {ASTOCKS_DIR} 目录下找到 {len(available_files)} 个CSV文件")

        for file_path in available_files:
            stock_code = os.path.basename(file_path).split('_')[0]
            stock_files_dict[stock_code] = file_path

        # 检查股票数据文件的匹配率
        all_stock_codes = set(limit_up_df['stock_code'])
        matched_stock_codes = all_stock_codes.intersection(set(stock_files_dict.keys()))
        unmatched_stock_codes = all_stock_codes - matched_stock_codes

        logger.info(f"需要匹配的股票数: {len(all_stock_codes)}")
        logger.info(
            f"成功匹配的股票数: {len(matched_stock_codes)} ({len(matched_stock_codes) / len(all_stock_codes) * 100:.2f}%)")

        if unmatched_stock_codes:
            logger.warning(f"未找到数据文件的股票: {unmatched_stock_codes}")

        # 初始化按日期分组的结果字典
        date_results = {}

        # 添加进度条，为每个日期单独计算统计数据
        for date, date_records in tqdm(date_grouped_limit_ups.items(), desc="分析连板晋级率"):
            if not date_records:
                continue

            logger.info(f"正在处理日期 {date} 的 {len(date_records)} 条记录")

            # 将当前日期的记录转换为DataFrame
            date_df = pd.DataFrame(date_records)

            # 为当前日期初始化统计计数器，增加高开相关的统计
            stats = defaultdict(lambda: {
                'promoted': 0, 'survived': 0, 'died': 0, 'total': 0,
                'high_open_promoted': 0, 'high_open_survived': 0, 'high_open_died': 0,
                'high_open_total': 0
            })
            status_count = {"处理成功": 0, "缺少数据": 0}
            status_by_board = defaultdict(lambda: {"处理成功": 0, "缺少数据": 0})
            failed_stocks = []  # 记录处理失败的股票

            # 处理当前日期的每条涨停记录
            for _, row in date_df.iterrows():
                date_str = row['date']
                stock_code = row['stock_code']
                stock_name = row['stock_name']
                board_count = row['board_count']

                # 获取次日状态和高开情况
                next_day_status = get_stock_status_next_day(stock_code, date_str, stock_files_dict)

                if next_day_status:
                    status, is_high_open = next_day_status
                    # 使用"X进Y"的格式表示板块，例如"1进2"表示1板晋级到2板
                    level_key = f"{board_count}进{board_count + 1}"
                    stats[level_key][status] += 1
                    stats[level_key]['total'] += 1
                    
                    # 记录高开情况
                    if is_high_open:
                        stats[level_key][f'high_open_{status}'] += 1
                        stats[level_key]['high_open_total'] += 1
                    
                    status_count["处理成功"] += 1
                    status_by_board[board_count]["处理成功"] += 1
                else:
                    status_count["缺少数据"] += 1
                    status_by_board[board_count]["缺少数据"] += 1
                    failed_stocks.append({
                        'date': date_str,
                        'stock_code': stock_code,
                        'stock_name': stock_name,
                        'board_count': board_count
                    })
                    logger.warning(f"处理失败: {stock_code} {stock_name} 板数:{board_count} 日期:{date_str}")

            logger.info(
                f"日期 {date} 处理结果统计: 处理成功 {status_count['处理成功']}，缺少数据 {status_count['缺少数据']}")

            # 计算当前日期的各种比率
            date_result = {}
            for level, counts in stats.items():
                total = counts['total']
                high_open_total = counts['high_open_total']
                if total > 0:
                    date_result[level] = {
                        'promoted_rate': (counts['promoted'] / total) * 100,
                        'survived_rate': (counts['survived'] / total) * 100,
                        'died_rate': (counts['died'] / total) * 100,
                        'promoted_count': counts['promoted'],
                        'survived_count': counts['survived'],
                        'died_count': counts['died'],
                        'total_count': total,
                        # 高开相关的统计指标
                        'high_open_total': high_open_total,
                        'high_open_promoted_count': counts['high_open_promoted'],
                        'high_open_survived_count': counts['high_open_survived'],
                        'high_open_died_count': counts['high_open_died'],
                        'high_open_promoted_rate': (counts['high_open_promoted'] / high_open_total) * 100 if high_open_total > 0 else 0,
                        'high_open_survived_rate': (counts['high_open_survived'] / high_open_total) * 100 if high_open_total > 0 else 0,
                        'high_open_died_rate': (counts['high_open_died'] / high_open_total) * 100 if high_open_total > 0 else 0
                    }

            # 将当前日期的结果保存到按日期分组的结果字典中
            date_results[date] = date_result

        return date_results

    except Exception as e:
        logger.error(f"分析涨停晋级率时出错: {e}", exc_info=True)
        return {}


def format_excel_sheet(worksheet, column_order):
    """
    设置Excel工作表的格式：调整列宽并为不同日期设置交替背景色
    
    参数:
        worksheet: openpyxl的worksheet对象
        column_order: 列名顺序的列表
    """
    # 设置统一列宽为12个字符
    column_width = 12
    
    # 定义背景颜色
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 设置所有列宽相同
    for i in range(len(column_order)):
        col_letter = get_column_letter(i + 1)  # 列索引从1开始
        worksheet.column_dimensions[col_letter].width = column_width
    
    # 应用交替背景色（按日期分组）
    current_date = None
    use_gray = True
    
    # 从第2行开始（跳过标题行）
    for row_idx in range(2, worksheet.max_row + 1):
        date_value = worksheet.cell(row=row_idx, column=1).value
        if date_value != current_date:
            current_date = date_value
            use_gray = not use_gray  # 切换颜色
        
        fill = light_gray_fill if use_gray else white_fill
        
        # 为该行的所有单元格设置背景色
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = fill


def save_results_to_excel(results, date_list):
    """
    将分析结果保存到Excel文件中。
    - 如果文件已存在，则检查日期是否已存在，已存在的日期会跳过
    - 使用sheet名"晋级率"
    - 晋级目标列为"x进y"格式
    - 数据按日期和板次升序排序
    - 为不同日期设置交替的背景色
    - 自动调整列宽
    
    参数:
        results: 连板晋级率分析结果（按日期分组的嵌套字典）
        date_list: 分析的日期列表，格式为YYYY-MM-DD
    """
    if not results:
        logger.warning("没有结果可保存")
        return

    # 确保excel目录存在
    os.makedirs(os.path.dirname(RESULT_FILE_PATH), exist_ok=True)

    # 为每个分析日期创建记录
    data = []
    for date in date_list:
        # 检查日期是否有分析结果
        if date not in results or not results[date]:
            logger.warning(f"日期 {date} 没有分析结果")
            continue

        # 获取当前日期的结果
        date_result = results[date]

        # 为每个板次创建记录
        for level, stats in date_result.items():
            data.append({
                '数据日期': date,
                '总数': stats['total_count'],  # 总数放在数据日期后面
                '高开总数': stats['high_open_total'], # 添加高开总数
                '晋级目标': level,  # 直接使用"x进y"格式
                '晋级率': f"{stats['promoted_rate']:.2f}%",
                '存活率': f"{stats['survived_rate']:.2f}%",
                '死亡率': f"{stats['died_rate']:.2f}%",
                '晋级数': stats['promoted_count'],
                '存活数': stats['survived_count'],
                '死亡数': stats['died_count'],
                # 添加高开相关的列
                '高开晋级率': f"{stats['high_open_promoted_rate']:.2f}%",
                '高开存活率': f"{stats['high_open_survived_rate']:.2f}%",
                '高开死亡率': f"{stats['high_open_died_rate']:.2f}%",
                '高开晋级数': stats['high_open_promoted_count'],
                '高开存活数': stats['high_open_survived_count'],
                '高开死亡数': stats['high_open_died_count']
            })

    # 创建DataFrame并排序
    new_df = pd.DataFrame(data)

    if new_df.empty:
        logger.warning("没有有效数据可保存")
        return None

    # 提取晋级目标中的板次数字用于排序
    def extract_board_level(level_str):
        try:
            return int(level_str.split('进')[0])
        except:
            return 999  # 对于异常情况给一个大数，排在最后

    # 添加辅助列用于排序，然后进行排序
    new_df['sort_key'] = new_df['晋级目标'].apply(extract_board_level)
    new_df = new_df.sort_values(by=['数据日期', 'sort_key'])
    new_df = new_df.drop(columns=['sort_key'])  # 删除辅助列

    # 重新安排列顺序，确保"总数"和"高开总数"在"晋级目标"前面
    column_order = ['数据日期', '总数', '高开总数', '晋级目标', 
                     '晋级率', '存活率', '死亡率', 
                     '高开晋级率', '高开存活率', '高开死亡率',
                     '晋级数', '存活数', '死亡数', 
                     '高开晋级数', '高开存活数', '高开死亡数']
    new_df = new_df[column_order]

    # 准备新数据
    existing_df = None
    sheet_exists = False
    
    # 检查文件和sheet是否存在
    if os.path.exists(RESULT_FILE_PATH):
        try:
            with pd.ExcelFile(RESULT_FILE_PATH) as xls:
                sheet_exists = '晋级率' in xls.sheet_names
                if sheet_exists:
                    # 读取现有数据
                    existing_df = pd.read_excel(RESULT_FILE_PATH, sheet_name='晋级率')
        except Exception as e:
            logger.error(f"读取现有Excel文件出错: {e}")
    
    # 合并数据（如有必要）
    final_df = new_df
    if existing_df is not None and not existing_df.empty:
        # 过滤掉已存在的日期数据
        existing_dates = set(existing_df['数据日期'].astype(str))
        new_records = new_df[~new_df['数据日期'].astype(str).isin(existing_dates)]
        
        if new_records.empty:
            logger.info("所有日期数据已存在，不需要追加")
            return RESULT_FILE_PATH
            
        # 合并数据
        final_df = pd.concat([existing_df, new_records], ignore_index=True)
        
        # 重新排序合并后的数据
        final_df['sort_key'] = final_df['晋级目标'].apply(extract_board_level)
        final_df = final_df.sort_values(by=['数据日期', 'sort_key'])
        final_df = final_df.drop(columns=['sort_key'])
    
    # 保存数据
    excel_mode = 'w'  # 默认模式：写入新文件
    if os.path.exists(RESULT_FILE_PATH):
        if sheet_exists:
            # 先创建临时文件保存其他sheet的数据
            temp_writer = pd.ExcelWriter(RESULT_FILE_PATH + '.temp', engine='openpyxl')
            with pd.ExcelFile(RESULT_FILE_PATH) as xls:
                for sheet_name in xls.sheet_names:
                    if sheet_name != '晋级率':
                        pd.read_excel(RESULT_FILE_PATH, sheet_name=sheet_name).to_excel(
                            temp_writer, sheet_name=sheet_name, index=False)
            temp_writer.close()
            
            # 复制回原文件
            os.replace(RESULT_FILE_PATH + '.temp', RESULT_FILE_PATH)
            excel_mode = 'a'  # 追加模式
        else:
            excel_mode = 'a'  # 追加模式

    # 写入数据
    with pd.ExcelWriter(RESULT_FILE_PATH, engine='openpyxl', mode=excel_mode) as writer:
        final_df.to_excel(writer, sheet_name='晋级率', index=False)
        
        # 应用格式设置
        if excel_mode == 'w' or (excel_mode == 'a' and '晋级率' not in writer.book.sheetnames):
            format_excel_sheet(writer.sheets['晋级率'], column_order)
        else:
            format_excel_sheet(writer.book['晋级率'], column_order)
    
    logger.info(f"已保存结果到: {RESULT_FILE_PATH}")
    return RESULT_FILE_PATH


def analyze_rate(start_date, end_date=None):
    """
    分析连板晋级率并输出结果。
    
    参数:
        start_date: 开始日期，格式可以是'YYYY-MM-DD'、'YYYYMMDD'或'YYYY年MM月DD日'
        end_date: 结束日期，格式同start_date，如果为None则等于start_date
    """
    logger.info("开始分析连板晋级率...")

    # 获取分析的日期列表
    date_list = get_date_range(start_date, end_date)
    if not date_list:
        logger.warning("没有找到符合条件的交易日。")
        return

    # 检查是否有今天的日期，若有则需要移除（因为无法获取T+1日数据）
    today_str = datetime.today().strftime('%Y-%m-%d')
    if today_str in date_list:
        logger.info(f"检测到今日({today_str})在分析范围内，由于无法获取T+1日数据，将跳过今日分析")
        date_list.remove(today_str)
        if not date_list:
            logger.warning("移除今日后没有可分析的交易日。")
            return

    # 检查哪些日期已经存在于Excel文件中
    existing_dates = set()
    if os.path.exists(RESULT_FILE_PATH):
        try:
            with pd.ExcelFile(RESULT_FILE_PATH) as xls:
                if '晋级率' in xls.sheet_names:
                    existing_df = pd.read_excel(RESULT_FILE_PATH, sheet_name='晋级率')
                    existing_dates = set(existing_df['数据日期'].astype(str))
                    logger.info(f"检测到已存在的数据日期: {sorted(list(existing_dates))}")
        except Exception as e:
            logger.error(f"读取现有结果文件时出错: {e}")

    # 过滤掉已存在的日期，只分析新的日期
    new_dates = [date for date in date_list if date not in existing_dates]
    if not new_dates:
        logger.info("所有请求的日期数据已存在，无需重新分析。")
        return

    logger.info(f"将分析以下新日期: {new_dates}")
    print(f"将分析以下新日期: {new_dates}")

    # 只分析新的日期
    results = analyze_limit_up_progression(new_dates[0], new_dates[-1] if len(new_dates) > 1 else None)

    if not results:
        logger.warning("没有找到符合条件的数据或分析过程中出现错误。")
        return

    # 保存结果到Excel
    result_file = save_results_to_excel(results, new_dates)
    if result_file:
        logger.info(f"分析结果已保存到: {result_file}")

    # 打印分析结果概览
    print("=" * 70)
    print("连板晋级率分析结果概览")
    print("=" * 70)

    # 按日期循环打印结果
    for date in new_dates:
        if date not in results or not results[date]:
            print(f"日期 {date} 没有分析结果")
            continue

        print(f"\n--- 日期: {date} ---")
        print(
            f"{'板块':<7} {'晋级率':<8} {'存活率':<8} {'死亡率':<8} {'晋级/总数':<10} {'存活/总数':<10} {'死亡/总数':<10}")
        print("-" * 70)

        # 获取当前日期的结果
        date_result = results[date]

        # 按板次排序 (从"X进Y"提取X进行排序)
        sorted_levels = sorted(date_result.keys(), key=lambda x: int(x.split('进')[0]))

        for level in sorted_levels:
            stats = date_result[level]
            print(f"{level:<7} {stats['promoted_rate']:.2f}% {stats['survived_rate']:.2f}% {stats['died_rate']:.2f}% "
                  f"{stats['promoted_count']}/{stats['total_count']} {stats['survived_count']}/{stats['total_count']} "
                  f"{stats['died_count']}/{stats['total_count']}")

    return results

# 用法示例
# analyze_rate('2025-04-21')
# 分析一个日期范围
# analyze_rate('2025-04-01', '2025-04-30')
