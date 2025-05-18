import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from utils.date_util import get_trading_days

# 输入和输出文件路径
FUPAN_FILE = "./excel/fupan_stocks.xlsx"
OUTPUT_FILE = "./excel/dejavu_analysis.xlsx"


def process_dejavu_data(start_date, end_date, clean_output=False):
    """
    处理连板数据，转换为dejavu格式

    :param start_date: 开始日期，格式为'YYYYMMDD'
    :param end_date: 结束日期，格式为'YYYYMMDD'
    :param clean_output: 是否清空现有Excel并重新创建，默认为False
    """
    # 获取交易日列表
    trading_days = get_trading_days(start_date, end_date)

    # 读取原始Excel数据 - 只需要连板数据
    try:
        df = pd.read_excel(FUPAN_FILE, sheet_name='连板数据', index_col=0)
    except Exception as e:
        print(f"读取连板数据失败: {e}")
        return

    # 创建或加载工作簿
    if clean_output or not os.path.exists(OUTPUT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "连板跟踪"
        is_new_file = True
    else:
        try:
            wb = load_workbook(OUTPUT_FILE)
            ws = wb.active
            is_new_file = False
        except Exception as e:
            print(f"加载现有文件失败: {e}，将创建新文件")
            wb = Workbook()
            ws = wb.active
            ws.title = "连板跟踪"
            is_new_file = True

    # 收集所有股票数据
    stock_data = {}  # {股票代码: {日期: {板数, 原因等}}}
    date_formatted_list = [datetime.strptime(trade_date, '%Y%m%d').strftime('%Y年%m月%d日') for trade_date in
                           trading_days]

    # 解析Excel数据
    for date_col in df.columns:
        # 只处理在分析时间范围内的日期
        if date_col in date_formatted_list:
            column_data = df[date_col].dropna()

            for data_str in column_data:
                items = data_str.split('; ')

                # 提取需要的数据
                stock_code = ""
                stock_name = ""
                board_level = 0
                reason = ""

                for item in items:
                    if '.S' in item and len(item.split('.')[0]) == 6:
                        # 这是股票代码
                        stock_code = item.strip()
                    elif not stock_name and len(items) > 1 and not any(x in item for x in ['.', ':', '%']):
                        # 这可能是股票简称
                        stock_name = item.strip()
                    elif '天' in item and '板' in item:
                        # 连板股票，如 "2天2板"
                        try:
                            # 提取板数
                            board_text = item.strip()
                            board_level = int(board_text.split('天')[1].split('板')[0].strip())
                        except Exception:
                            pass
                    elif '+' in item and len(item.split('+')) > 1:
                        # 涨停原因
                        reason = item.strip()

                if not stock_code and len(items) >= 2:
                    # 尝试直接使用前两项作为代码和名称
                    stock_code = items[0].strip()
                    stock_name = items[1].strip()

                if stock_code and stock_name:
                    stock_key = f"{stock_code}_{stock_name}"
                    if stock_key not in stock_data:
                        stock_data[stock_key] = {
                            'code': stock_code,
                            'name': stock_name,
                            'reason': reason,
                            'dates': {}
                        }
                    elif not stock_data[stock_key]['reason'] and reason:
                        # 如果之前没有记录原因，但现在有，则更新
                        stock_data[stock_key]['reason'] = reason

                    # 保存该日期的数据
                    stock_data[stock_key]['dates'][date_col] = {
                        'board_level': board_level
                    }

    # 创建或更新Excel
    if is_new_file:
        # 设置表头
        ws.cell(row=1, column=1, value="涨停原因").font = Font(bold=True)
        ws.cell(row=1, column=2, value="股票简称").font = Font(bold=True)

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10

    # 获取现有数据的最后一列
    max_col = ws.max_column
    date_start_col = 3  # 日期从第3列开始

    # 如果是新文件，或者现有文件列数小于3，从第3列开始
    if is_new_file or max_col < 3:
        current_col = date_start_col
    else:
        current_col = max_col + 1

    # 添加新的日期列
    for trade_date in trading_days:
        date_formatted = datetime.strptime(trade_date, '%Y%m%d').strftime('%Y年%m月%d日')

        # 检查该日期列是否已存在
        date_exists = False
        for col in range(date_start_col, current_col):
            if ws.cell(row=1, column=col).value == date_formatted:
                date_exists = True
                break

        if not date_exists:
            # 添加新的日期列
            cell = ws.cell(row=1, column=current_col, value=date_formatted)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置列宽
            column_letter = get_column_letter(current_col)
            ws.column_dimensions[column_letter].width = 10

            current_col += 1

    # 获取所有日期列
    date_columns = {}
    for col in range(date_start_col, ws.max_column + 1):
        date_val = ws.cell(row=1, column=col).value
        if date_val and '年' in date_val and '月' in date_val and '日' in date_val:
            date_columns[date_val] = col

    # 如果是现有文件，读取现有数据
    existing_stocks = {}
    if not is_new_file:
        for row in range(2, ws.max_row + 1):
            stock_name = ws.cell(row=row, column=2).value
            reason = ws.cell(row=row, column=1).value
            if stock_name and stock_name != '':
                # 使用股票名称作为键
                existing_stocks[stock_name] = row

    # 按首次出现的日期和板数对股票进行排序
    sorted_stocks = []
    for stock_key, data in stock_data.items():
        # 找出首次出现的日期和最高板数
        first_date = min(data['dates'].keys())
        max_board = max(info['board_level'] for info in data['dates'].values())
        sorted_stocks.append((stock_key, data, first_date, max_board))

    # 按最高板数(从高到低)和首次出现日期(从早到晚)排序
    sorted_stocks.sort(key=lambda x: (-x[3], list(date_columns.keys()).index(x[2]) if x[2] in date_columns else 999))

    # 写入数据
    current_row = 2

    # 处理所有股票数据
    for stock_key, data, first_date, _ in sorted_stocks:
        stock_name = data['name']
        reason = data['reason']

        # 检查该股票是否已存在
        if stock_name in existing_stocks:
            row = existing_stocks[stock_name]
        else:
            row = current_row
            # 写入原因和股票名称
            ws.cell(row=row, column=1, value=reason)
            ws.cell(row=row, column=2, value=stock_name)
            current_row += 1

        # 对每个日期列进行处理
        for date, col_idx in date_columns.items():
            if date in data['dates']:
                board_level = data['dates'][date]['board_level']
                if board_level >= 2:  # 只处理2板及以上
                    ws.cell(row=row, column=col_idx, value=f"{board_level}板")
            elif date == first_date:
                # 如果是首次出现但没有板数数据(异常情况)，填充股票名称
                ws.cell(row=row, column=col_idx, value=stock_name)

    # 保存工作簿
    wb.save(OUTPUT_FILE)
    print(f"数据处理完成，已保存到 {OUTPUT_FILE}")


if __name__ == "__main__":
    # 设置当前工作目录为脚本所在目录
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    # 示例用法
    start_date = "20250421"
    end_date = "20250516"

    # 处理连板数据
    process_dejavu_data(start_date, end_date)
