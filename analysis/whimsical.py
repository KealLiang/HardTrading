import math
import os
import re
from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from utils.date_util import get_trading_days
from utils.excel_vba_util import add_vba_to_sheet

# 导入NLP工具模块 (如果可用)
try:
    from utils.nlp_utils import check_nlp_ready, calc_similarity, find_semantic_clusters, FORCE_CHAR_SIMILARITY

    NLP_UTILS_AVAILABLE = True
except ImportError:
    NLP_UTILS_AVAILABLE = False
    FORCE_CHAR_SIMILARITY = True

# 定义常见的同义词转换，顺序重要
synonym_groups = {
    "医药": ["%医药%", "创新药", "%疫苗%", "%医疗器械%", "%养老%", "%合成生物%", "%重组蛋白%", "%原料药%", "中成药"],
    "大消费": ["消费", "白酒", "食品", "饮料", "零售", "商超", "免税", "化妆品", "麦角硫因", "电商", "电子商务",
               "消费电子", "%宠物%", "家电", "%益生菌%"],
    "重组": ["%重组%", "%控制权%"],
    "新传媒": ["%IP经济%", "IP", "小红书"],
    "半导体": ["%半导体%", "%芯片%", "%存储芯片%", "%集成电路%", "%光刻%", "%集成电路%"],
    "电子元件": ["%铜缆%", "%电子元件%", "%PCB%", "%连接器%", "%卫星通信%", "%雷达%", "%电线电缆%"],
    "AI": ["%AI%", "%人工智能%", "%DeepSeek%", "%大模型%", "%GPT%", "%AIGC%", "%MCP%", "%脑机接口%"],
    "算力": ["%算力%", "%液冷%", "%数据中心电源%", "%服务器测试%", "数据中心"],
    "新能源": ["%新能源%", "%电动车%", "%动力电池%", "%光伏%", "%锂电池%", "%氢%", "%可控核聚变%", "%节能环保%",
               "特斯拉", "固态电池", "%核电%", "%核能%"],
    "电力": ["%风电%", "%电力%", "%核电%", "%电网设备%", "电机"],
    "化工": ["%化工%", "%环氧丙烷%", "%氯碱%", "%化纤%", "%涂料%", "%季戊四醇%", "%聚酯%", "%钛白粉%", "%造纸%", "%纺织%"],
    "新材料": ["%PEEK材料%", "%碳纤维%", "%高温合金%", "%稀土永磁%"],
    "军工": ["%军工%", "%国防%", "%航空%", "%战斗机%", "%大飞机%", "%军贸%", "%无人机%", "%成飞%"],
    "航天": ["%航天%", "%低空经济%", "%飞行汽车%"],
    "机器人": ["%机器人%", "%减速器%"],
    "跨境": ["%跨境%", "%外销%", "%港口%", "%一带一路%", "%航运%", "%出海%", "统一大市场"],
    "房地产": ["%房地产%", "%城市更新%", "%建筑%", "%室内设计%"],
    "汽车": ["%汽车%", "%压缩机%"],
    "消费电子": ["%消费电子%", "%苹果%", "%AR/VR%", "%智能穿戴%", "%光学元件%", "%汽车电子%", "%智能座舱%",
                 "%智能家居%"],
    "旅游": ["旅游", "酒店", "民航", "免税", "出行"],
    "金融": ["%金融%", "%保险%", "%银行%", "%信托%", "%AH%", "腾讯"],
    "华为": ["%华为%", "%鸿蒙%"],
    "国企": ["%国企%", "%国资%", "%央企%", "%中字头%", "%兵装%"],
    "业绩增长": ["%业绩%", "%报增%", "%净利%", "%预增%"],
    "扭亏为盈": ["%扭亏%", "%减亏%", "%摘帽%"]
}

# 排除列表 - 这些原因不会被选为热门原因
EXCLUDED_REASONS = [
    "业绩增长",
    "扭亏为盈",
    "国企",
]

# 选取top n的原因着色
TOP_N = 9

# 颜色列表 - 彩虹色系(深色)
COLORS = [
    "FF5A5A",  # 红色
    "FF8C42",  # 橙色
    "FFCE30",  # 黄色
    "6AD15A",  # 绿色
    "45B5FF",  # 蓝色
    "9966FF",  # 紫色
    "FF66B3",  # 粉色
    "5ACDCD",  # 青色
    "DDA0DD",  # 浅紫红
    "FFB366",  # 浅橙色
    "FFF066",  # 浅黄色
    "90EE90",  # 浅绿色
    "87CEFA",  # 浅蓝色
    "B19CD9",  # 浅紫色
    "FFB6C1",  # 浅粉色
    "E1FFFF",  # 浅青色
]

# 未分类原因的最小打印阈值
UNCLASSIFIED_PRINT_THRESHOLD = 5

# 多次上榜但无热门原因的颜色
MULTI_COLOR = "E0FFFF"

# 输入和输出文件路径
FUPAN_FILE = "./excel/fupan_stocks.xlsx"
OUTPUT_FILE = "./excel/fupan_analysis.xlsx"

# 指数数据文件路径
INDEX_FILE = "./data/indexes/sz399006_创业板指.csv"

# 表头颜色
HEADER_COLOR = "E0E0E0"  # 浅灰色

# 红绿色系定义，用于涨跌幅颜色表示
RED_COLORS = [
    "FFCCCC",  # 浅红色 (0-1%)
    "FF9999",  # 淡红色 (1-2%)
    "FF6666",  # 红色 (2-3%)
    "FF3333",  # 深红色 (3-5%)
    "FF0000",  # 大红色 (>5%)
]

GREEN_COLORS = [
    "CCFFCC",  # 浅绿色 (0-1%)
    "99FF99",  # 淡绿色 (1-2%)
    "66FF66",  # 绿色 (2-3%)
    "33FF33",  # 深绿色 (3-5%)
    "00CC00",  # 大绿色 (>5%)
]

# 涨跌幅颜色阈值
PCT_CHANGE_THRESHOLDS = [1.0, 2.0, 3.0, 5.0]


def normalize_reason(reason):
    """
    将原因标准化，处理同一类型的不同表述，优先匹配最具体的模式
    """
    # 移除所有空格
    reason = re.sub(r'\s+', '', reason)
    original_reason = reason

    # 存储所有匹配结果及其匹配信息
    matches = []

    # 检查原因属于哪个组
    for main_reason, synonyms in synonym_groups.items():
        for synonym in synonyms:
            # 处理通配符匹配
            if '%' in synonym:
                # 转换SQL风格通配符为正则表达式
                pattern = synonym.replace('%', '(.*)')
                regex = re.compile(f"^{pattern}$")
                match = regex.search(reason)
                
                if match:
                    # 计算匹配的具体部分
                    matched_text = reason
                    for group in match.groups():
                        matched_text = matched_text.replace(group, '')
                    
                    # 匹配强度 = 匹配文本长度 / 原文长度
                    match_strength = len(matched_text) / len(reason)
                    matches.append((main_reason, match_strength, synonym))
            
            # 保留原有的包含匹配
            elif synonym in reason:
                match_strength = len(synonym) / len(reason)
                matches.append((main_reason, match_strength, synonym))

    # 如果有匹配，选择匹配强度最高的
    if matches:
        matches.sort(key=lambda x: x[1], reverse=True)
        return matches[0][0]

    return f"未分类_{original_reason}"


def extract_reasons(reason_text):
    """
    从原因文本中提取所有原因
    """
    if not reason_text or pd.isna(reason_text):
        return []

    # 以"+"分割不同原因
    reasons = reason_text.split('+')
    return [normalize_reason(r.strip()) for r in reasons if r.strip()]


def get_color_by_pct_change(pct_change):
    """
    根据涨跌幅返回对应的颜色代码
    
    :param pct_change: 涨跌幅百分比
    :return: 颜色代码
    """
    if pd.isna(pct_change):
        return "FFFFFF"  # 白色

    # 初始化颜色索引为0 (最浅)
    color_idx = 0
    abs_change = abs(pct_change)

    # 根据涨跌幅绝对值确定颜色深浅
    for i, threshold in enumerate(PCT_CHANGE_THRESHOLDS):
        if abs_change >= threshold:
            color_idx = i + 1

    # 如果超过最后一个阈值，使用最深的颜色
    color_idx = min(color_idx, len(RED_COLORS) - 1)

    # 根据正负选择红色或绿色
    if pct_change >= 0:
        return RED_COLORS[color_idx]
    else:
        return GREEN_COLORS[color_idx]


def load_index_data():
    """
    加载创业板指数数据
    
    :return: 以日期为键的字典，包含涨跌幅和成交量信息
    """
    try:
        # 检查指数数据文件是否存在
        if not os.path.exists(INDEX_FILE):
            print(f"创业板指数文件不存在: {INDEX_FILE}")
            return {}

        # 读取CSV文件 (无表头)
        df = pd.read_csv(INDEX_FILE, header=None,
                         names=['date', 'open', 'high', 'low', 'close', 'volume'])

        # 计算每日涨跌幅
        df['pct_change'] = df['close'].pct_change() * 100

        # 将成交量转换为亿元 (原始数据单位为成交量，而非手数)
        # 注意：volume可能已经是以元为单位，需根据实际数据调整
        df['volume_100m'] = df['volume'] / 100000000  # 转换为亿元

        # 以日期为键创建字典
        index_data = {}
        for _, row in df.iterrows():
            # 确保日期格式与date_columns中的键完全一致
            date_obj = pd.to_datetime(row['date'])
            date_str = date_obj.strftime('%Y年%m月%d日')
            index_data[date_str] = {
                'pct_change': row['pct_change'],
                'volume_100m': row['volume_100m']
            }

        return index_data
    except Exception as e:
        print(f"读取创业板指数数据失败: {e}")
        return {}


def process_zt_data(start_date, end_date, clean_output=False):
    """
    处理涨停数据，转换为更易于分析的格式

    :param start_date: 开始日期，格式为'YYYYMMDD'
    :param end_date: 结束日期，格式为'YYYYMMDD'
    :param clean_output: 是否清空现有Excel并重新创建，默认为False
    """
    # 获取交易日列表
    trading_days = get_trading_days(start_date, end_date)

    # 读取指数数据
    index_data = load_index_data()
    if not index_data:
        print(f"警告: 未能加载创业板指数数据，请确保文件存在: {INDEX_FILE}")
        print("将继续处理涨停数据，但不会显示指数信息。")

    # 读取原始Excel数据
    sheets_to_process = ['连板数据', '首板数据']
    sheet_data = {}

    for sheet_name in sheets_to_process:
        try:
            df = pd.read_excel(FUPAN_FILE, sheet_name=sheet_name, index_col=0)
            sheet_data[sheet_name] = df
        except Exception as e:
            print(f"读取{sheet_name}失败: {e}")
            continue

    if not sheet_data:
        print("没有找到需要处理的数据")
        return

    # 创建新的工作簿或清空现有工作簿
    if clean_output and os.path.exists(OUTPUT_FILE):
        try:
            # 尝试删除现有文件
            os.remove(OUTPUT_FILE)
            print(f"已清空并将重新创建文件: {OUTPUT_FILE}")
        except Exception as e:
            print(f"清空文件失败: {e}")
            print("将尝试使用新的工作簿覆盖现有文件")

    wb = Workbook()
    ws = wb.active
    ws.title = "涨停分析"

    # 收集所有涨停原因和每日涨停股票数据
    all_reasons = []
    daily_stocks = {}

    # 找出当前分析时间范围内的交易日对应的列
    date_formatted_list = [datetime.strptime(trade_date, '%Y%m%d').strftime('%Y年%m月%d日') for trade_date in
                           trading_days]

    for sheet_name, df in sheet_data.items():
        for date_col in df.columns:
            # 只处理在分析时间范围内的日期
            if date_col in date_formatted_list:
                if date_col not in daily_stocks:
                    daily_stocks[date_col] = []

                column_data = df[date_col].dropna()

                for idx, data_str in enumerate(column_data):
                    items = data_str.split('; ')
                    stock_info = {}

                    # 提取需要的数据
                    stock_code = ""
                    stock_name = ""
                    board_level = 0
                    first_time = "23:59:59"

                    for item in items:
                        if '.S' in item and len(item.split('.')[0]) == 6:
                            # 这是股票代码
                            stock_code = item.strip()
                        elif not stock_name and len(items) > 1 and not any(x in item for x in ['.', ':', '%']):
                            # 这可能是股票简称
                            stock_name = item.strip()
                        elif '天' in item and '板' in item:
                            # 连板股票，如 "2天2板"
                            stock_info['几天几板'] = item.strip()
                            try:
                                # 提取板数
                                board_text = item.strip()
                                board_level = int(board_text.split('天')[1].split('板')[0].strip())
                            except Exception:
                                pass
                        elif '首板涨停' in item:
                            # 首板股票
                            stock_info['几天几板'] = item.strip()
                            board_level = 1
                        elif ':' in item and '涨停时间' not in item and '开板时间' not in item:
                            # 可能是首次涨停时间
                            stock_info['首次涨停时间'] = item.strip()
                            first_time = item.strip()
                        elif item.isdigit() or (item.replace('.', '', 1).isdigit() and item.count('.') <= 1):
                            # 可能是开板次数或价格
                            if '涨停开板次数' not in stock_info:
                                stock_info['涨停开板次数'] = item.strip()
                        elif '+' in item and len(item.split('+')) > 1:
                            # 可能是涨停原因
                            stock_info['涨停原因类别'] = item.strip()
                            reasons = extract_reasons(item.strip())
                            all_reasons.extend(reasons)

                    if not stock_code and len(items) >= 2:
                        # 尝试直接使用前两项作为代码和名称
                        stock_code = items[0].strip()
                        stock_name = items[1].strip()

                    if stock_code and stock_name:
                        # 确保板数正确
                        if board_level == 0 and '首板涨停' in str(stock_info.get('几天几板', '')):
                            board_level = 1

                        # 添加到每日数据
                        daily_stocks[date_col].append({
                            'code': stock_code,
                            'name': stock_name,
                            'info': stock_info,
                            'board_level': board_level,
                            'first_time': first_time,
                            'sheet_name': sheet_name  # 记录来源sheet，用于区分连板和首板
                        })

    # 统计所有原因并找出未分类的原因
    reason_counter = Counter(all_reasons)
    unclassified_reasons = [reason for reason in reason_counter.keys() if reason.startswith('未分类_')]

    # 打印出现次数超过阈值的未分类原因
    if unclassified_reasons:
        print("\n======== 未分类原因统计（出现次数 >= {}）========".format(UNCLASSIFIED_PRINT_THRESHOLD))
        for reason in sorted(unclassified_reasons, key=lambda x: reason_counter[x], reverse=True):
            count = reason_counter[reason]
            if count >= UNCLASSIFIED_PRINT_THRESHOLD:
                original_reason = reason.replace('未分类_', '')
                print(f"  {original_reason}: {count}次")
        print("================================================\n")

    # 创建未分类原因的sheet
    if unclassified_reasons:
        unclassified_ws = wb.create_sheet(title="未分类原因")

        # 设置表头
        unclassified_ws.cell(row=1, column=1, value="原因").font = Font(bold=True)
        unclassified_ws.cell(row=1, column=2, value="出现次数").font = Font(bold=True)

        # 按出现次数排序
        sorted_reasons = sorted(unclassified_reasons, key=lambda x: reason_counter[x], reverse=True)

        # 计算每列应该显示的行数
        total_reasons = len(sorted_reasons)
        columns_count = 7  # 每页显示7列
        rows_per_column = math.ceil(total_reasons / columns_count)

        # 写入未分类原因和次数
        for i, reason in enumerate(sorted_reasons):
            # 计算当前行应该在哪一列
            col_idx = (i // rows_per_column) * 2 + 1
            row_idx = i % rows_per_column + 2

            # 写入原因和次数
            original_reason = reason.replace('未分类_', '')
            unclassified_ws.cell(row=row_idx, column=col_idx, value=original_reason)
            unclassified_ws.cell(row=row_idx, column=col_idx + 1, value=reason_counter[reason])

        # 设置列宽
        for i in range(1, columns_count * 2 + 1):
            col_letter = get_column_letter(i)
            if i % 2 == 1:  # 原因列
                unclassified_ws.column_dimensions[col_letter].width = 25
            else:  # 次数列
                unclassified_ws.column_dimensions[col_letter].width = 10

        print(f"未分类的涨停原因已保存到工作簿的 '未分类原因' 页")

    # 过滤掉未分类的原因，获取热门原因
    classified_reasons = [reason for reason in reason_counter.keys()
                          if not reason.startswith('未分类_') and reason not in EXCLUDED_REASONS]

    # 确保所有原因的出现次数都被正确计算
    # 首先创建包含所有可能原因的计数字典
    all_reason_counts = {reason: 0 for reason in synonym_groups.keys()}

    # 然后用实际统计的次数更新
    for reason, count in reason_counter.items():
        if not reason.startswith('未分类_'):
            all_reason_counts[reason] = count

    # 选择热门原因 (排除指定的原因)
    # 按出现次数倒序选择TOP_N个原因
    top_reasons = [reason for reason, count in sorted(all_reason_counts.items(),
                                                      key=lambda x: x[1],
                                                      reverse=True)
                   if count > 0 and reason not in EXCLUDED_REASONS][:TOP_N]

    # 为每个原因分配颜色
    reason_colors = {reason: COLORS[i % len(COLORS)] for i, reason in enumerate(top_reasons)}

    # 计算每个股票在分析时间范围内的出现次数和主要原因
    all_stocks = {}
    for date, stocks in daily_stocks.items():
        for stock in stocks:
            stock_key = f"{stock['code']}_{stock['name']}"

            if stock_key not in all_stocks:
                all_stocks[stock_key] = {
                    'name': stock['name'],
                    'appearances': [],
                    'reasons': []
                }

            all_stocks[stock_key]['appearances'].append(date)

            if '涨停原因类别' in stock['info']:
                # 只使用已分类的原因进行分组
                reasons = [r for r in extract_reasons(stock['info']['涨停原因类别'])
                           if not r.startswith('未分类_')]
                all_stocks[stock_key]['reasons'].extend(reasons)

    # 确定每支股票主要属于哪个原因组
    stock_reason_group = {}
    for stock_key, data in all_stocks.items():
        if not data['reasons']:
            continue

        # 统计该股票的原因
        stock_reason_counter = Counter(data['reasons'])

        # 先检查哪些原因是热门原因
        top_reasons_found = [reason for reason in top_reasons if reason in stock_reason_counter]

        if top_reasons_found:
            # 如果有多个热门原因，选择出现次数最多的
            top_reason_counts = [(reason, stock_reason_counter[reason]) for reason in top_reasons_found]
            top_reason_counts.sort(key=lambda x: x[1], reverse=True)
            stock_reason_group[stock_key] = top_reason_counts[0][0]
        elif stock_reason_counter:
            # 如果没有热门原因，使用该股票最常见的原因
            most_common_reason = stock_reason_counter.most_common(1)[0][0]
            stock_reason_group[stock_key] = most_common_reason

    # 创建图例作为第一列
    ws.column_dimensions['A'].width = 15
    ws.cell(row=1, column=1, value="热门概念图例").font = Font(bold=True)

    # 添加各个热门原因图例（带次数）- 图例从第4行开始，因为第2、3行要留给涨跌幅和成交量的标签
    current_row = 4
    for i, reason in enumerate(top_reasons, start=0):
        count = all_reason_counts.get(reason, 0)
        cell = ws.cell(row=current_row, column=1, value=f"{reason}({count})")
        cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")
        current_row += 1

    # 添加未入选着色的原因（按出现次数倒序）
    # 所有synonym_groups中的原因（除了TOP_N和EXCLUDED_REASONS）
    non_top_reasons = [r for r in synonym_groups.keys()
                       if r not in top_reasons and r not in EXCLUDED_REASONS]
    sorted_non_top = sorted(non_top_reasons, key=lambda x: all_reason_counts[x], reverse=True)

    for reason in sorted_non_top:
        count = all_reason_counts.get(reason, 0)
        ws.cell(row=current_row, column=1, value=f"{reason}({count})")
        current_row += 1

    # 在EXCLUDED_REASONS原因前添加分隔线
    if EXCLUDED_REASONS:
        separator_cell = ws.cell(row=current_row, column=1, value="以下是排除的概念")
        separator_cell.border = Border(top=Side(style='hair', color='000000'))
        separator_cell.font = Font(italic=True, size=9)
        current_row += 1

    # 添加EXCLUDED_REASONS原因（放在最后）
    for reason in EXCLUDED_REASONS:
        count = all_reason_counts.get(reason, 0)
        ws.cell(row=current_row, column=1, value=f"{reason}({count})")
        current_row += 1

    # 添加多次上榜图例
    cell = ws.cell(row=current_row, column=1, value="多次上榜")
    cell.fill = PatternFill(start_color=MULTI_COLOR, fill_type="solid")

    # 添加首板/连板分隔cell说明
    separator_row = current_row + 1
    ws.cell(row=separator_row, column=1, value="分隔cell = 首板")
    separator_cell = ws.cell(row=separator_row + 1, column=1)
    separator_cell.border = Border(bottom=Side(style='dashed', color='000000'))

    # 设置日期表头(从第二列开始)
    date_columns = {}
    for idx, trade_date in enumerate(trading_days, start=2):
        date_obj = datetime.strptime(trade_date, '%Y%m%d')
        date_formatted = date_obj.strftime('%Y年%m月%d日')

        # 获取星期几的中文名称
        weekday_name = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日'][date_obj.weekday()]
        cell = ws.cell(row=1, column=idx, value=f"{date_obj.strftime('%Y-%m-%d')}\n{weekday_name}")

        # 设置表头样式
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 增强表头样式
        cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

        # 设置列宽
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = 15

        date_columns[date_formatted] = idx

        # 添加指数涨跌幅数据 (第二行)
        if date_formatted in index_data:
            # 涨跌幅数据 (第二行)
            pct_change = index_data[date_formatted]['pct_change']
            pct_cell = ws.cell(row=2, column=idx, value=f"{pct_change:.2f}%")
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.fill = PatternFill(start_color=get_color_by_pct_change(pct_change), fill_type="solid")

            # 成交量数据 (第三行)
            volume = index_data[date_formatted]['volume_100m']
            volume_cell = ws.cell(row=3, column=idx, value=f"{volume:.2f}")
            volume_cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # 如果没有指数数据，添加空单元格
            ws.cell(row=2, column=idx, value="--")
            ws.cell(row=3, column=idx, value="--")

    # 在第一列添加涨跌幅和成交量的标签，并设置样式
    label_cell_1 = ws.cell(row=2, column=1, value="创业指")
    label_cell_1.alignment = Alignment(horizontal="center", vertical="center")
    label_cell_1.font = Font(bold=True)
    label_cell_1.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

    label_cell_2 = ws.cell(row=3, column=1, value="成交量(亿)")
    label_cell_2.alignment = Alignment(horizontal="center", vertical="center")
    label_cell_2.font = Font(bold=True)
    label_cell_2.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")

    # 美化标题行
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill(start_color=HEADER_COLOR, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加指数行和个股行之间的分隔线
    for col_idx in range(1, len(trading_days) + 2):
        cell = ws.cell(row=3, column=col_idx)
        cell.border = Border(
            bottom=Side(style='double', color='000000')
        )

    # 冻结首行和首列
    ws.freeze_panes = ws.cell(row=4, column=2)  # 冻结A1:A3和B1:X3

    # 对每日数据进行排序并写入工作表
    for date, col_idx in date_columns.items():
        if date in daily_stocks:
            # 按照板块数倒序，首次涨停时间正序排序
            sorted_stocks = sorted(daily_stocks[date],
                                   key=lambda x: (-x['board_level'], x['first_time']))

            # 先找出每列的连板和首板分界行号
            first_connection_idx = None
            for idx, stock in enumerate(sorted_stocks):
                if stock['board_level'] == 1 and stock['sheet_name'] == '首板数据':
                    first_connection_idx = idx
                    break

            # 写入连板股票数据 (从第4行开始，因为前3行是日期、涨跌幅和成交量)
            row_idx = 4
            for idx, stock in enumerate(sorted_stocks):
                # 达到分界点，插入空行作为分隔
                if first_connection_idx is not None and idx == first_connection_idx:
                    # 添加空行作为分隔
                    row_idx += 1
                    separator_cell = ws.cell(row=row_idx - 1, column=col_idx)
                    # 设置底色为黑色
                    separator_cell.fill = PatternFill(start_color="333333", fill_type="solid")
                    # separator_cell.border = Border(
                    #     bottom=Side(style='double', color='FFFFFF')
                    # )

                stock_key = f"{stock['code']}_{stock['name']}"

                # 计算该股票在当前分析区间内的上榜次数
                appearances_count = len(all_stocks[stock_key]['appearances'])

                # 获取当前板数，用于显示在名称后
                current_board_level = stock['board_level'] if stock['board_level'] != 1 else ''

                # 在股票名称后添加当天涨停板数（y板数）
                display_name = f"{stock['name']}{current_board_level}"

                cell = ws.cell(row=row_idx, column=col_idx, value=display_name)

                # 设置单元格样式
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 创建备注
                comment_text = ""
                if '几天几板' in stock['info']:
                    comment_text += f"{stock['info']['几天几板']} "
                if '首次涨停时间' in stock['info']:
                    comment_text += f"{stock['info']['首次涨停时间']} "
                if '涨停开板次数' in stock['info']:
                    comment_text += f"{stock['info']['涨停开板次数']}\n"
                if '涨停原因类别' in stock['info']:
                    comment_text += f"{stock['info']['涨停原因类别']}"

                cell.comment = Comment(comment_text, "分析系统")

                # 根据股票所属原因组设置颜色
                if stock_key in stock_reason_group:
                    reason = stock_reason_group[stock_key]
                    if reason in reason_colors:
                        cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")

                # 如果股票在多个日期出现，但没有热门原因，使用浅灰色
                if appearances_count > 1:
                    if stock_key not in stock_reason_group or stock_reason_group[stock_key] not in reason_colors:
                        cell.fill = PatternFill(start_color=MULTI_COLOR, fill_type="solid")

                row_idx += 1

    # 保存工作簿
    wb.save(OUTPUT_FILE)
    print(f"数据处理完成，已保存到 {OUTPUT_FILE}")


def consolidate_unclassified_reasons():
    """
    读取未分类原因sheet，分析可以归类到现有synonym_groups的原因
    并将结果打印到控制台
    """
    # 检查NLP环境
    use_semantic = False
    word_vectors = None

    if NLP_UTILS_AVAILABLE:
        if FORCE_CHAR_SIMILARITY:
            print("当前使用字符相似度模式，不使用gensim")
            use_semantic = False
        else:
            try:
                use_semantic, word_vectors, message = check_nlp_ready()
                if use_semantic:
                    print(f"NLP环境就绪: {message}")
                else:
                    print(f"NLP环境未就绪: {message}")
                    print("将使用基本字符匹配进行分析")
            except Exception as e:
                print(f"NLP模块加载失败: {e}")
                print("将使用基本字符匹配进行分析")
    else:
        print("未找到NLP工具模块，将使用基本字符匹配")
        print("如需启用语义分析功能，请使用: pip install --only-binary=:all: gensim numpy")

    try:
        # 读取Excel文件中的未分类原因sheet
        df_unclassified = pd.read_excel(OUTPUT_FILE, sheet_name="未分类原因", header=0)

        # 收集所有未分类原因及其出现次数
        unclassified_reasons = {}

        # 处理可能的多列布局
        for i in range(0, df_unclassified.shape[1], 2):
            if i + 1 < df_unclassified.shape[1]:
                reason_col = df_unclassified.iloc[:, i]
                count_col = df_unclassified.iloc[:, i + 1]

                for j in range(len(reason_col)):
                    reason = reason_col.iloc[j]
                    count = count_col.iloc[j]

                    if pd.notna(reason) and pd.notna(count):
                        unclassified_reasons[reason] = count

        # 获取当前的同义词组
        current_groups = {
            group_name: set(synonyms)
            for group_name, synonyms in synonym_groups.items()
        }

        # 分析每个未分类原因是否可以归入现有分组
        suggestions = {}
        remaining_unclassified = {}

        # 基础匹配逻辑，无需NLP库支持
        for reason, count in unclassified_reasons.items():
            matched = False
            best_match = None
            best_match_score = 0
            match_type = ""

            # 1. 基于字符匹配的方法
            for group_name, synonym_set in current_groups.items():
                # 直接包含匹配
                for synonym in synonym_set:
                    if synonym in reason or reason in synonym:
                        match_score = len(synonym) / max(len(reason), 1)
                        if match_score > best_match_score:
                            best_match = group_name
                            best_match_score = match_score
                            matched = True
                            match_type = "字符匹配"

                # 组名匹配
                if group_name in reason or reason in group_name:
                    match_score = len(group_name) / max(len(reason), 1)
                    if match_score > best_match_score:
                        best_match = group_name
                        best_match_score = match_score
                        matched = True
                        match_type = "组名匹配"

            # 2. 进行语义相似度匹配 (如果NLP环境可用)
            if not matched and use_semantic and word_vectors is not None:
                for group_name in current_groups.keys():
                    # 计算与组名的语义相似度
                    group_sim = calc_similarity(reason, group_name, word_vectors)

                    # 计算与组内词汇的最大相似度
                    max_synonym_sim = 0
                    for synonym in current_groups[group_name]:
                        syn_sim = calc_similarity(reason, synonym, word_vectors)
                        max_synonym_sim = max(max_synonym_sim, syn_sim)

                    # 取与组名和组内词汇相似度的最大值
                    final_sim = max(group_sim, max_synonym_sim)

                    # 使用相对保守的阈值
                    if final_sim > 0.6 and final_sim > best_match_score:
                        best_match = group_name
                        best_match_score = final_sim
                        matched = True
                        match_type = f"语义相似度({final_sim:.2f})"

            if matched:
                if best_match not in suggestions:
                    suggestions[best_match] = []
                suggestions[best_match].append((reason, count, match_type))
            else:
                remaining_unclassified[reason] = count

        # 先收集所有推荐结果再显示

        # 现有分组的更新建议
        updated_groups = {}
        for group_name in synonym_groups.keys():
            current_synonyms = list(synonym_groups[group_name])
            new_synonyms = []

            if group_name in suggestions:
                for reason, _, _ in suggestions[group_name]:
                    if reason not in current_synonyms:
                        new_synonyms.append(reason)

            # 只保存有更新的组
            if new_synonyms:
                updated_groups[group_name] = current_synonyms + new_synonyms

        # 打印可以归类的原因
        print("=" * 60)
        print("可以归类到现有分组的原因:")
        print("=" * 60)

        for group_name, reasons in suggestions.items():
            print(f"\n【{group_name}】分组可添加以下同义词:")
            for reason, count, match_type in sorted(reasons, key=lambda x: x[1], reverse=True):
                print(f"  - \"{reason}\" (出现{count}次) - 匹配方式: {match_type}")

        # 打印剩余未分类的原因
        if remaining_unclassified:
            print("\n" + "=" * 60)
            print("仍然无法归类的原因:")
            print("=" * 60)

            for reason, count in sorted(remaining_unclassified.items(), key=lambda x: x[1], reverse=True):
                if count >= 3:  # 仅显示出现次数较多的原因
                    print(f"  - \"{reason}\" (出现{count}次)")

        # 发现新的潜在分组
        print("\n" + "=" * 60)
        print("潜在的新分组分析:")
        print("=" * 60)

        # 基于常见关键词分析
        common_keywords = [
            "ChatGPT", "风电", "稀土", "碳中和", "元宇宙", "区块链",
            "锂电池", "生成式AI", "储能", "数据中心", "虚拟现实", "航天",
            "电池", "生物科技", "云计算", "大数据"
        ]

        potential_new_groups = {}
        for keyword in common_keywords:
            matching_reasons = []
            for reason, count in remaining_unclassified.items():
                if keyword.lower() in reason.lower():
                    matching_reasons.append((reason, count))

            if matching_reasons:
                potential_new_groups[keyword] = matching_reasons

        # 收集潜在新分组
        new_group_suggestions = {}
        for group_name, reasons in potential_new_groups.items():
            total_count = sum(count for _, count in reasons)
            if total_count >= 5:  # 累计出现5次以上才创建新分组
                # 按出现次数降序排序的原因列表
                sorted_reasons = [reason for reason, _ in sorted(reasons, key=lambda x: x[1], reverse=True)]
                new_group_suggestions[group_name] = sorted_reasons

                print(f"\n【{group_name}】新分组可包含 (总计{total_count}次):")
                for reason, count in sorted(reasons, key=lambda x: x[1], reverse=True):
                    print(f"  - \"{reason}\" (出现{count}次)")

        # 仅在NLP环境可用时执行语义聚类
        semantic_clusters = []
        if use_semantic and word_vectors is not None and len(remaining_unclassified) > 5:
            print("\n" + "=" * 60)
            print("基于语义相似度的聚类分析:")
            print("=" * 60)

            # 使用NLP工具进行语义聚类
            clusters = find_semantic_clusters(remaining_unclassified, word_vectors, threshold=0.7)

            # 打印发现的语义簇
            for i, cluster in enumerate(clusters, 1):
                total = sum(count for _, count in cluster)
                if total >= 5:  # 显示总频次达到阈值的簇
                    # 找出簇中出现频率最高的原因作为组名
                    most_common_reason = max(cluster, key=lambda x: x[1])[0]
                    # 按出现次数降序排序的原因列表
                    sorted_reasons = [reason for reason, _ in sorted(cluster, key=lambda x: x[1], reverse=True)]
                    # 避免重复的组名
                    if most_common_reason not in new_group_suggestions:
                        semantic_clusters.append((most_common_reason, sorted_reasons))

                    print(f"\n语义簇 #{i} (总计{total}次):")
                    for reason, count in sorted(cluster, key=lambda x: x[1], reverse=True):
                        print(f"  - \"{reason}\" (出现{count}次)")

                    # 推荐可能的分组名称
                    print(f"  推荐分组名: \"{most_common_reason}\"")

        # 如果没有NLP环境，使用基础的相似度分析
        char_similarity_clusters = []
        if not use_semantic and len(remaining_unclassified) > 5:
            print("\n" + "=" * 60)
            print("基于字符相似度的分析:")
            print("=" * 60)
            print("提示: 安装NLP工具可获得更准确的语义聚类分析")

            # 简单字符相似度方法
            def get_common_chars(str1, str2):
                # 计算两个字符串的公共字符数量
                common = 0
                for c in str1:
                    if c in str2:
                        common += 1
                return common

            # 使用简单字符相似度进行聚类
            clustered = set()
            clusters = []

            reasons_list = list(remaining_unclassified.keys())
            for i, reason1 in enumerate(reasons_list):
                if reason1 in clustered:
                    continue

                cluster = [(reason1, remaining_unclassified[reason1])]
                clustered.add(reason1)

                for reason2 in reasons_list[i + 1:]:
                    if reason2 in clustered:
                        continue

                    # 计算基本字符相似度
                    common_chars = get_common_chars(reason1, reason2)
                    sim_score = common_chars / max(len(reason1), len(reason2))

                    if sim_score > 0.5:  # 阈值比NLP版本更保守
                        cluster.append((reason2, remaining_unclassified[reason2]))
                        clustered.add(reason2)

                if len(cluster) > 1:  # 至少形成了一个簇
                    clusters.append(cluster)

            # 打印发现的字符簇
            for i, cluster in enumerate(clusters, 1):
                total = sum(count for _, count in cluster)
                if total >= 5:  # 显示总频次达到阈值的簇
                    # 找出簇中出现频率最高的原因作为组名
                    most_common_reason = max(cluster, key=lambda x: x[1])[0]
                    # 按出现次数降序排序的原因列表
                    sorted_reasons = [reason for reason, _ in sorted(cluster, key=lambda x: x[1], reverse=True)]
                    # 避免重复的组名
                    if most_common_reason not in new_group_suggestions:
                        char_similarity_clusters.append((most_common_reason, sorted_reasons))

                    print(f"\n字符簇 #{i} (总计{total}次):")
                    for reason, count in sorted(cluster, key=lambda x: x[1], reverse=True):
                        print(f"  - \"{reason}\" (出现{count}次)")

                    # 推荐可能的分组名称
                    print(f"  推荐分组名: \"{most_common_reason}\"")

        # 先打印现有分组的建议更新
        print("\n" + "=" * 60)
        print("建议更新的synonym_groups代码:")
        print("=" * 60)

        for group_name, synonyms in updated_groups.items():
            print(f'    "{group_name}": {synonyms},')

        # 打印所有建议的新分组（包括潜在分组、语义簇和字符簇）的代码
        if new_group_suggestions or semantic_clusters or char_similarity_clusters:
            print("\n" + "=" * 60)
            print("建议添加的新分组代码:")
            print("=" * 60)

            # 创建已使用组名的集合，避免重复
            used_group_names = set()

            # 输出潜在新分组
            for group_name, reasons in new_group_suggestions.items():
                if group_name in used_group_names:
                    continue
                used_group_names.add(group_name)
                reasons_str = ", ".join([f'"{r}"' for r in reasons])
                print(f'    "{group_name}": [{reasons_str}],')

            # 输出语义聚类分组
            for group_name, reasons in semantic_clusters:
                if group_name in used_group_names:
                    continue
                used_group_names.add(group_name)
                reasons_str = ", ".join([f'"{r}"' for r in reasons])
                print(f'    "{group_name}": [{reasons_str}],')

            # 输出字符相似度聚类分组
            for group_name, reasons in char_similarity_clusters:
                if group_name in used_group_names:
                    continue
                used_group_names.add(group_name)
                reasons_str = ", ".join([f'"{r}"' for r in reasons])
                print(f'    "{group_name}": [{reasons_str}],')

    except Exception as e:
        import traceback
        print(f"处理未分类原因时出错: {e}")
        print(traceback.format_exc())


def add_vba_for_excel():
    vba_file = "./scripts/highlight_same_cell.vba"
    output_xlsm = add_vba_to_sheet(OUTPUT_FILE, vba_file, "涨停分析")
    print(f"已添加高亮相同股票的VBA脚本，保存为: {output_xlsm}")


if __name__ == "__main__":
    # 设置当前工作目录为脚本所在目录
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    # 示例用法
    start_date = "20240101"
    end_date = "20240601"

    # 处理涨停数据，设置clean_output=True可清空现有Excel重新生成
    # process_zt_data(start_date, end_date, clean_output=True)
    process_zt_data(start_date, end_date)

    # 为【未分类原因】归类
    consolidate_unclassified_reasons()
