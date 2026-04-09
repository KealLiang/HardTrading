"""
龙头sheet全量HTML交互图表生成器

数据来源：
- excel/ladder_analysis.xlsx 中所有“龙头xxxx”工作表
 - 可选：默认仅合并同目录 ladder_analysis_龙头归档.xlsx；若开启 merge_leader_archive_splits，再合并 excel/leader_archives 下拆分归档（与主天梯合并去重，主天梯同名 sheet 优先）

功能：
- 读取全部龙头sheet，提取所有曾入选的股票
- 基于每日快照计算“入选日”和“移除日”事件（支持多次进出）
- 生成与策略扫描一致风格的K线+成交量HTML图表
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Sequence, Tuple, Union

import pandas as pd
from openpyxl import load_workbook

from analysis.html_gen.strategy_scan_html_chart import (
    _create_combined_html,
    _create_single_chart_figure,
)
from analysis.loader.fupan_data_loader import load_zaban_data
from fetch.stock_concept_map import get_stock_concepts, is_map_available
from utils.backtrade.visualizer import read_stock_data
from utils.date_util import get_n_trading_days_before, get_next_trading_day

logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

VirtualBarInput = Union[Dict[str, float], Tuple[float, float], List[float]]


def _normalize_virtual_bars(virtual_bars: Optional[Sequence[VirtualBarInput]]) -> List[Dict[str, float]]:
    """标准化虚拟K线参数。"""
    if not virtual_bars:
        return []
    normalized: List[Dict[str, float]] = []
    for idx, bar in enumerate(virtual_bars):
        if isinstance(bar, dict):
            if 'open_pct' not in bar or 'close_pct' not in bar:
                raise ValueError(f"第{idx + 1}根虚拟K线缺少 open_pct/close_pct")
            open_pct = float(bar['open_pct'])
            close_pct = float(bar['close_pct'])
        elif isinstance(bar, (tuple, list)) and len(bar) == 2:
            open_pct = float(bar[0])
            close_pct = float(bar[1])
        else:
            raise ValueError(f"第{idx + 1}根虚拟K线格式无效: {bar}")
        normalized.append({'open_pct': open_pct, 'close_pct': close_pct})
    return normalized


def _append_virtual_bars_to_chart_df(chart_df: pd.DataFrame, virtual_bars: Sequence[Dict[str, float]]) -> pd.DataFrame:
    """
    对单只股票图表数据追加虚拟K线。

    规则：
    - 每根虚拟K线的开收盘相对前一日收盘价计算
    - high/low 取 open/close 的 max/min
    - 虚拟量柱固定为0
    """
    if chart_df.empty or not virtual_bars:
        return chart_df

    df = chart_df.copy()
    last_close = float(df.iloc[-1]['Close'])
    current_day = df.index.max().strftime('%Y%m%d')
    virtual_rows = []

    for i, bar in enumerate(virtual_bars):
        next_day = get_next_trading_day(current_day)
        if not next_day:
            logging.warning(f"第{i + 1}根虚拟K线未找到下一个交易日，提前结束")
            break

        open_price = round(last_close * (1 + bar['open_pct'] / 100.0), 4)
        close_price = round(last_close * (1 + bar['close_pct'] / 100.0), 4)
        high_price = max(open_price, close_price)
        low_price = min(open_price, close_price)

        row_dt = datetime.strptime(next_day, '%Y%m%d')
        virtual_rows.append((row_dt, open_price, high_price, low_price, close_price, 0.0))

        last_close = close_price
        current_day = next_day

    if not virtual_rows:
        return df

    virtual_df = pd.DataFrame(
        [(r[1], r[2], r[3], r[4], r[5]) for r in virtual_rows],
        index=[r[0] for r in virtual_rows],
        columns=['Open', 'High', 'Low', 'Close', 'Volume'],
    )
    return pd.concat([df, virtual_df], axis=0)


def _extract_date_from_header(header_val: object) -> Optional[str]:
    """从表头单元格提取 YYYY-MM-DD 日期字符串。"""
    if not header_val:
        return None
    text = str(header_val).strip()
    if not text:
        return None

    first_line = text.split('\n')[0].strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', first_line):
        return first_line
    return None


def _infer_snapshot_date_from_sheet_name(sheet_name: str, header_dates: List[str]) -> Optional[str]:
    """
    从工作表名“龙头MMDD”推断快照日期（YYYY-MM-DD）。

    说明：
    - 历史龙头sheet会被回填最新日期列，不能用“最大表头日期”作为该sheet快照日。
    - 优先根据sheet名中的MMDD，在表头日期中寻找同月同日并选择最早年份（通常是该sheet创建日）。
    """
    m = re.search(r'^龙头(\d{2})(\d{2})$', str(sheet_name).strip())
    if not m:
        return None

    mm = int(m.group(1))
    dd = int(m.group(2))
    if not header_dates:
        return None

    matched = []
    for d in header_dates:
        try:
            dt = datetime.strptime(d, '%Y-%m-%d')
        except Exception:
            continue
        if dt.month == mm and dt.day == dd:
            matched.append(dt)

    if matched:
        matched.sort()
        return matched[0].strftime('%Y-%m-%d')

    # 兜底：若表头中不存在同MMDD（极少数异常sheet），按当前年推一个日期
    year = datetime.now().year
    try:
        dt = datetime(year, mm, dd)
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return None


def _parse_leader_sheet_snapshot(sheet_name: str, ws) -> Tuple[Optional[str], Dict[str, Dict]]:
    """
    解析单个龙头sheet快照。

    Returns:
        (snapshot_date, stock_map)
        snapshot_date: 该sheet所代表的日期(YYYY-MM-DD)，取表头中最大日期
        stock_map: {code: {'code','name','sheet_concept'}}
    """
    # 识别日期列（第5列开始通常是交易日列）
    header_dates: List[str] = []
    for col_idx in range(5, ws.max_column + 1):
        date_str = _extract_date_from_header(ws.cell(row=1, column=col_idx).value)
        if date_str:
            header_dates.append(date_str)

    if not header_dates:
        return None, {}

    # 快照日期优先取“龙头MMDD”的推断结果，避免被历史回填列污染
    snapshot_date = _infer_snapshot_date_from_sheet_name(sheet_name, header_dates)
    if not snapshot_date:
        # 兼容异常命名sheet，最后兜底才使用最大表头日期
        snapshot_date = max(header_dates)
    stock_map: Dict[str, Dict] = {}

    for row_idx in range(4, ws.max_row + 1):
        code_val = ws.cell(row=row_idx, column=1).value
        if code_val is None:
            continue
        code = str(code_val).strip()
        if not re.match(r'^\d{6}$', code):
            continue

        name = str(ws.cell(row=row_idx, column=3).value or '').strip()
        sheet_concept = str(ws.cell(row=row_idx, column=2).value or '').strip()
        stock_map[code] = {
            'code': code,
            'name': name,
            'sheet_concept': sheet_concept,
        }

    return snapshot_date, stock_map


def _leader_archive_path(excel_path: str) -> str:
    """与 ladder_chart.archive_leader_sheets 一致：{主文件名}_龙头归档.xlsx，与主文件同目录。"""
    p = os.path.abspath(excel_path)
    d = os.path.dirname(p)
    base = os.path.splitext(os.path.basename(p))[0]
    return os.path.join(d, f"{base}_龙头归档.xlsx")


def _leader_archive_paths(excel_path: str, merge_splits: bool = False) -> List[str]:
    """
    返回需要合并读取的归档路径列表（只读）。

    merge_splits=False（默认）：仅 {base}_龙头归档.xlsx（控制读取量，适合 HTML 场景）。
    merge_splits=True：再包含 leader_archives 下拆分文件（按拆分日期升序），最后追加主归档
    （与拆分中同名 sheet 冲突时以主归档为准）。
    """
    p = os.path.abspath(excel_path)
    d = os.path.dirname(p)
    base = os.path.splitext(os.path.basename(p))[0]
    archive_prefix = f"{base}_龙头归档"
    main_path = os.path.join(d, f"{archive_prefix}.xlsx")

    if not merge_splits:
        return [main_path] if os.path.exists(main_path) else []

    split_dir = os.path.join(d, "leader_archives")

    pattern = re.compile(rf"^{re.escape(archive_prefix)}_(\d{{8}})(?:_dup\d+)?\.xlsx$")
    items: List[Tuple[str, float, str]] = []
    if os.path.exists(split_dir):
        for fn in os.listdir(split_dir):
            if not fn.endswith(".xlsx"):
                continue
            m = pattern.match(fn)
            if not m:
                continue
            date_str = m.group(1)
            path = os.path.join(split_dir, fn)
            try:
                mtime = os.path.getmtime(path)
            except Exception:
                mtime = 0.0
            items.append((date_str, mtime, path))
    items.sort(key=lambda x: (x[0], x[1]))

    paths = [t[2] for t in items]
    if os.path.exists(main_path):
        paths.append(main_path)
    return paths


def _leader_snapshots_from_workbook(wb) -> List[Dict]:
    """从已打开的工作簿解析全部龙头 sheet 快照（未排序）。"""
    snapshots: List[Dict] = []
    for sheet_name in wb.sheetnames:
        if not str(sheet_name).startswith("龙头"):
            continue

        ws = wb[sheet_name]
        snapshot_date, stock_map = _parse_leader_sheet_snapshot(sheet_name, ws)
        if not snapshot_date:
            logging.warning(f"跳过sheet {sheet_name}：未识别到有效日期")
            continue

        snapshots.append({
            'sheet_name': sheet_name,
            'snapshot_date': snapshot_date,
            'stock_map': stock_map,
        })
    return snapshots


def _load_leader_snapshots_main_and_all(
    excel_path: str,
    use_leader_archive: bool = True,
    merge_leader_archive_splits: bool = False,
) -> Tuple[List[Dict], List[Dict]]:
    """
    返回 (main_snaps, all_snaps)：
    - main_snaps：只解析主文件；用于决定“最终 HTML 渲染哪些股票”（保证数量不变）
    - all_snaps：主文件 + 可选归档合并；用于计算更准确的入选/移除标记

    merge_leader_archive_splits：是否在合并归档时包含 leader_archives 下拆分文件（默认否，仅主归档）。
    """
    wb = load_workbook(excel_path, data_only=True)
    main_snaps = _leader_snapshots_from_workbook(wb)
    main_snaps.sort(key=lambda x: x['snapshot_date'])

    if not use_leader_archive:
        return main_snaps, main_snaps

    archive_paths = _leader_archive_paths(excel_path, merge_splits=merge_leader_archive_splits)
    if not archive_paths:
        logging.info(f"龙头归档文件不存在，仅使用主文件: {_leader_archive_path(excel_path)}")
        return main_snaps, main_snaps

    by_name: Dict[str, Dict] = {}
    all_arch_count = 0
    for archive_path in archive_paths:
        archive_wb = load_workbook(archive_path, data_only=True)
        arch_snaps = _leader_snapshots_from_workbook(archive_wb)
        all_arch_count += len(arch_snaps)
        for s in arch_snaps:
            by_name[s['sheet_name']] = s

    # 主文件覆盖归档同名 sheet（与 archive_leader_sheets 行为保持一致）
    for s in main_snaps:
        by_name[s['sheet_name']] = s

    all_snaps = list(by_name.values())
    all_snaps.sort(key=lambda x: x['snapshot_date'])
    split_note = "含拆分" if merge_leader_archive_splits else "仅主归档"
    logging.info(
        f"龙头快照已合并归档（{split_note}）: 主 {len(main_snaps)} 张, 归档读入 {all_arch_count} 张, 去重后 {len(all_snaps)} 张（归档文件 {len(archive_paths)} 个）"
    )
    return main_snaps, all_snaps


def _build_stock_signals_from_snapshots(snapshots: List[Dict]) -> Dict[str, List[Dict]]:
    """
    从每日龙头快照构建股票事件序列。

    事件定义：
    - 入选日：当日存在，前一日不存在
    - 移除日：当日不存在，前一日存在
    """
    if not snapshots:
        return {}

    all_codes = set()
    stock_meta: Dict[str, Dict] = {}
    for snap in snapshots:
        for code, info in snap['stock_map'].items():
            all_codes.add(code)
            if code not in stock_meta:
                stock_meta[code] = {
                    'name': info.get('name', ''),
                    'sheet_concept': info.get('sheet_concept', ''),
                }
            else:
                # 如果后续sheet名称更完整，则更新
                if not stock_meta[code].get('name') and info.get('name'):
                    stock_meta[code]['name'] = info.get('name')
                if not stock_meta[code].get('sheet_concept') and info.get('sheet_concept'):
                    stock_meta[code]['sheet_concept'] = info.get('sheet_concept')

    stock_signals_map: Dict[str, List[Dict]] = {}

    for code in all_codes:
        name = stock_meta.get(code, {}).get('name', '')
        sheet_concept = stock_meta.get(code, {}).get('sheet_concept', '')
        prev_in = False
        signals: List[Dict] = []

        for snap in snapshots:
            current_in = code in snap['stock_map']
            signal_date = snap['snapshot_date']

            if current_in and not prev_in:
                signals.append({
                    'code': code,
                    'name': name,
                    'signal_date': signal_date,
                    'signal_type': '龙头入选',
                    'price': None,
                    'details': f"来源: {snap['sheet_name']}",
                    'sheet_concept': sheet_concept,
                })
            elif (not current_in) and prev_in:
                signals.append({
                    'code': code,
                    'name': name,
                    'signal_date': signal_date,
                    'signal_type': '龙头移除',
                    'price': None,
                    'details': f"来源: {snap['sheet_name']}",
                    'sheet_concept': sheet_concept,
                })

            prev_in = current_in

        if signals:
            stock_signals_map[code] = signals

    return stock_signals_map


def _latest_leader_entry_date(signals: List[Dict]) -> str:
    """最近一次龙头入选日；无入选事件时退回全部信号中的最大日期。"""
    dates = [s['signal_date'] for s in signals if s.get('signal_type') == '龙头入选']
    if dates:
        return max(dates)
    return max(s['signal_date'] for s in signals)


def _sheet_concept_for_latest_entry(signals: List[Dict]) -> str:
    """以最近一次龙头入选为基准取 sheet 题材概念列。"""
    entries = [s for s in signals if s.get('signal_type') == '龙头入选']
    if not entries:
        return (signals[0].get('sheet_concept', '') if signals else '') or ''
    latest = max(entries, key=lambda x: x['signal_date'])
    return str(latest.get('sheet_concept', '') or '')


def _prepare_zaban_lookup(
    snapshots: List[Dict],
    before_days: int,
    after_days: int,
) -> Dict[str, set]:
    """
    基于快照日期范围加载炸板数据，返回 {YYYYMMDD: {code,...}} 结构。
    """
    if not snapshots:
        return {}
    snap_dates = [s.get('snapshot_date') for s in snapshots if s.get('snapshot_date')]
    if not snap_dates:
        return {}

    start_ymd = min(snap_dates).replace('-', '')
    end_ymd = max(snap_dates).replace('-', '')
    try:
        start_ymd = get_n_trading_days_before(start_ymd, max(0, before_days))
        if '-' in start_ymd:
            start_ymd = start_ymd.replace('-', '')
    except Exception:
        pass

    try:
        for _ in range(max(0, after_days)):
            next_day = get_next_trading_day(end_ymd)
            if not next_day:
                break
            end_ymd = next_day
    except Exception:
        pass

    zaban_df = load_zaban_data(start_ymd, end_ymd)
    if zaban_df is None or zaban_df.empty:
        logging.info("炸板数据为空，龙头HTML将不添加炸板标记")
        return {}

    lookup = zaban_df.attrs.get('zaban_lookup')
    if isinstance(lookup, dict):
        logging.info(f"炸板查找表已加载：{len(lookup)} 个交易日")
        return {k: set(v) for k, v in lookup.items()}

    fallback: Dict[str, set] = {}
    try:
        for _, row in zaban_df.iterrows():
            d = str(row.get('date', '')).strip()
            code = str(row.get('stock_code', '')).strip()
            if len(d) != 8:
                continue
            m = re.search(r'(\d{6})', code)
            if not m:
                continue
            fallback.setdefault(d, set()).add(m.group(1))
    except Exception:
        pass
    logging.info(f"炸板查找表已回退构建：{len(fallback)} 个交易日")
    return fallback


def _build_zaban_signals_for_chart(
    stock_code: str,
    chart_df: pd.DataFrame,
    zaban_lookup: Dict[str, set],
) -> List[Dict]:
    """为单只股票在当前图窗内构建炸板信号。"""
    if chart_df is None or chart_df.empty or not zaban_lookup:
        return []
    code = str(stock_code).zfill(6)
    out: List[Dict] = []
    for dt in chart_df.index:
        try:
            ymd = dt.strftime('%Y%m%d')
            if code in zaban_lookup.get(ymd, set()):
                out.append({
                    'code': code,
                    'signal_date': dt.strftime('%Y-%m-%d'),
                    'signal_type': '炸板',
                    'price': None,
                    'details': '炸板日',
                })
        except Exception:
            continue
    return out


def generate_leader_sheet_html_charts(
        excel_path: str = './excel/ladder_analysis.xlsx',
        columns: int = 2,
        before_days: int = 60,
        after_days: int = 30,
        output_dir: str = './excel/html_charts',
        data_dir: str = './data/astocks',
        virtual_bars: Optional[Sequence[VirtualBarInput]] = None,
        use_leader_archive: bool = True,
        merge_leader_archive_splits: bool = False,
) -> Optional[str]:
    """
    从全部龙头sheet生成全量HTML图表（包含入选/移除标记）。

    use_leader_archive: 是否合并「{主文件名}_龙头归档.xlsx」中的历史龙头 sheet（默认开启）。
    merge_leader_archive_splits: 是否在上述基础上再合并 leader_archives 下拆分归档（默认关闭，避免读入过久远/过大）。
    """
    if columns not in [1, 2, 3]:
        logging.warning(f"列数 {columns} 无效，使用默认值2")
        columns = 2

    if not os.path.exists(excel_path):
        logging.error(f"Excel文件不存在: {excel_path}")
        return None

    main_snaps, all_snaps = _load_leader_snapshots_main_and_all(
        excel_path,
        use_leader_archive=use_leader_archive,
        merge_leader_archive_splits=merge_leader_archive_splits,
    )
    if not main_snaps:
        logging.error("未找到可用的龙头sheet")
        return None

    normalized_virtual_bars = _normalize_virtual_bars(virtual_bars)

    # stock_signals_map_main：决定最终渲染哪些股票（保证数量不变）
    stock_signals_map_main = _build_stock_signals_from_snapshots(main_snaps)
    if not stock_signals_map_main:
        logging.warning("未提取到有效的入选/移除信号")
        return None

    # stock_signals_map_all：用于给上述股票计算更准确的入选/移除标记
    # 说明：
    # - 当开启归档合并时，all_snaps 可能包含比 main_snaps 更完整的历史，信号应基于 all_snaps 计算；
    # - 当未合并归档（或无可用归档）时，all_snaps 与 main_snaps 同源，此时直接复用，避免重复计算。
    if all_snaps is main_snaps:
        stock_signals_map_all = stock_signals_map_main
    else:
        stock_signals_map_all = _build_stock_signals_from_snapshots(all_snaps) if all_snaps else stock_signals_map_main

    # 炸板查找表（用于追加“炸板”标记）
    zaban_lookup = _prepare_zaban_lookup(all_snaps or main_snaps, before_days=before_days, after_days=after_days)

    os.makedirs(output_dir, exist_ok=True)

    # 标准概念映射（可选）
    concept_lookup: Dict[str, List[str]] = {}
    try:
        if is_map_available():
            # 只为最终渲染范围取概念，避免归档新增股票把概念拉取放大
            concept_lookup = {code: get_stock_concepts(code) for code in stock_signals_map_main}
    except Exception as e:
        logging.warning(f"加载概念映射失败，跳过标准概念标签: {e}")

    chart_figures = []
    chart_titles = []

    # 排序：按最近一次「龙头入选」日期倒序；无入选则按最后一条信号日期
    sorted_codes = sorted(
        list(stock_signals_map_main.keys()),
        key=lambda c: _latest_leader_entry_date(stock_signals_map_all.get(c) or stock_signals_map_main[c]),
        reverse=True,
    )

    for stock_code in sorted_codes:
        signals_info = stock_signals_map_all.get(stock_code) or stock_signals_map_main[stock_code]
        try:
            stock_name = signals_info[0].get('name', '')

            # 同股去重：同一天同类型只保留一次
            uniq = {}
            for sig in signals_info:
                k = (sig['signal_date'], sig['signal_type'])
                if k not in uniq:
                    uniq[k] = sig
            dedup_signals = sorted(uniq.values(), key=lambda x: x['signal_date'])

            signal_dates = [sig['signal_date'] for sig in dedup_signals]
            earliest_date = min(signal_dates)
            latest_date = max(signal_dates)

            earliest_date_yyyymmdd = earliest_date.replace('-', '')
            latest_date_yyyymmdd = latest_date.replace('-', '')

            chart_start = get_n_trading_days_before(earliest_date_yyyymmdd, before_days)
            chart_end = latest_date_yyyymmdd
            for _ in range(after_days):
                next_day = get_next_trading_day(chart_end)
                if not next_day:
                    break
                chart_end = next_day

            stock_data = read_stock_data(stock_code, data_dir)
            if stock_data is None or stock_data.empty:
                logging.warning(f"未找到股票数据: {stock_code} {stock_name}")
                continue

            start_dt = datetime.strptime(chart_start.replace('-', ''), '%Y%m%d')
            end_dt = datetime.strptime(chart_end, '%Y%m%d')
            chart_df = stock_data.loc[start_dt:end_dt].copy()
            chart_df = chart_df.dropna(subset=['Open', 'High', 'Low', 'Close'])
            if chart_df.empty:
                continue

            overlay_segment_start = None
            if normalized_virtual_bars:
                real_len = len(chart_df)
                chart_df = _append_virtual_bars_to_chart_df(chart_df, normalized_virtual_bars)
                if len(chart_df) > real_len:
                    # 叠加浅色起点 = 第一根成功追加的虚拟K线日期
                    overlay_segment_start = chart_df.index[real_len].strftime('%Y-%m-%d')

            # 追加炸板信号（独立标记，不与龙头入选/移除合并）
            zaban_signals = _build_zaban_signals_for_chart(stock_code, chart_df, zaban_lookup)
            merged_signals = dedup_signals + zaban_signals
            uniq2 = {}
            for sig in merged_signals:
                k = (sig['signal_date'], sig['signal_type'])
                if k not in uniq2:
                    uniq2[k] = sig
            merged_signals = sorted(uniq2.values(), key=lambda x: x['signal_date'])

            concepts = concept_lookup.get(stock_code) or []
            fig = _create_single_chart_figure(
                stock_code=stock_code,
                stock_name=stock_name,
                chart_df=chart_df,
                signal_dates_info=merged_signals,
                before_days=before_days,
                after_days=after_days,
                data_dir=data_dir,
                concepts=concepts,
                overlay_segment_start=overlay_segment_start,
                overlay_up_color='#ff8a8a',
                overlay_down_color='#66cc66',
                # 龙头图：建仓区间只锚定最新“龙头入选”信号
                entry_range_anchor_signal_types=['龙头入选'],
            )
            if fig is None:
                continue

            chart_figures.append(fig)
            # 卡片标题：以最近一次龙头入选为基准的题材概念
            sheet_concept = _sheet_concept_for_latest_entry(merged_signals)
            if sheet_concept:
                title = f"{stock_code} {stock_name} {sheet_concept}"
            else:
                title = f"{stock_code} {stock_name}"
            chart_titles.append(title)

        except Exception as e:
            logging.error(f"生成股票图失败 {stock_code}: {e}")
            continue

    if not chart_figures:
        logging.warning("没有可用图表，未生成HTML")
        return None

    rows = (len(chart_figures) + columns - 1) // columns
    html_content = _create_combined_html(chart_figures, chart_titles, columns, rows)

    html_filename = f"leader_sheet_all_{columns}cols.html"
    html_path = os.path.join(output_dir, html_filename)
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    logging.info(f"龙头sheet HTML生成完成: {html_path}（共 {len(chart_figures)} 只）")
    return html_path

