"""
【默默上涨】（概念分组口径）HTML交互图表生成器

数据来源：
- analysis.momo_shangzhang_processor.load_momo_shangzhang_data(..., keep_latest_per_stock=False)
  保留近3个月内全部入选记录（不做每股 keep='last' 去重）

功能：
- 为每只股票绘制K线+成交量图
- 标记全部【默默上涨入选】日期
- 叠加【炸板】与【关注度入榜】标记
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd

from analysis.html_gen.strategy_scan_html_chart import (
    _create_combined_html,
    _create_single_chart_figure,
)
from analysis.loader.fupan_data_loader import load_attention_data, load_zaban_data
from analysis.momo_shangzhang_processor import load_momo_shangzhang_data
from fetch.stock_concept_map import get_stock_concepts, is_map_available
from utils.backtrade.visualizer import read_stock_data
from utils.date_util import get_n_trading_days_before, get_next_trading_day

logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

_MOMO_SIGNAL_TYPE = '默默上涨入选'
MOMO_HTML_DRAW_TOP_N = 35  # 默认绘制“默默上涨”前N只（按 ladder_analysis.xlsx 的排序口径）


def _normalize_stock_code(raw_code: object) -> Optional[str]:
    code = str(raw_code or '').strip()
    if not code:
        return None
    if code.startswith(('sh', 'sz', 'bj')) and len(code) > 2:
        code = code[2:]
    if '.' in code:
        base, _, _ = code.partition('.')
        if base.isdigit() and len(base) == 6:
            code = base
    m = re.search(r'(\d{6})', code)
    return m.group(1) if m else None


def _parse_sheet_date(raw_date: object) -> Optional[str]:
    text = str(raw_date or '').strip()
    if not text:
        return None
    # 表头里很多日期单元格是类似：'2026-01-05\n周一'，只取第一行日期部分
    if '\n' in text:
        text = text.split('\n', 1)[0].strip()
    try:
        if re.match(r'^\d{4}年\d{1,2}月\d{1,2}日$', text):
            return datetime.strptime(text, '%Y年%m月%d日').strftime('%Y-%m-%d')
        if re.match(r'^\d{4}/\d{1,2}/\d{1,2}$', text):
            return pd.to_datetime(text).strftime('%Y-%m-%d')
        if re.match(r'^\d{4}-\d{2}-\d{2}$', text):
            return text
        if re.match(r'^\d{8}$', text):
            return datetime.strptime(text, '%Y%m%d').strftime('%Y-%m-%d')
        return pd.to_datetime(text).strftime('%Y-%m-%d')
    except Exception:
        return None


def _build_momo_signals_map(momo_df: pd.DataFrame) -> Dict[str, List[Dict]]:
    stock_signals_map: Dict[str, List[Dict]] = {}
    if momo_df is None or momo_df.empty:
        return stock_signals_map

    for _, row in momo_df.iterrows():
        code = _normalize_stock_code(row.get('纯代码') or row.get('股票代码'))
        signal_date = _parse_sheet_date(row.get('日期'))
        if not code or not signal_date:
            continue

        stock_name = str(row.get('股票名称', '') or '').strip()
        interval_change = str(row.get('区间涨跌幅', '') or '').strip()
        interval_volume = str(row.get('区间成交额', '') or '').strip()
        details = f"区间涨跌幅: {interval_change} | 区间成交额: {interval_volume}".strip()

        stock_signals_map.setdefault(code, []).append({
            'code': code,
            'name': stock_name,
            'signal_date': signal_date,
            'signal_type': _MOMO_SIGNAL_TYPE,
            'price': None,
            'details': details,
        })

    # 同股同日去重并按日期排序
    for code, signals in list(stock_signals_map.items()):
        uniq = {}
        for sig in signals:
            uniq[(sig['signal_date'], sig['signal_type'])] = sig
        merged = sorted(uniq.values(), key=lambda x: x['signal_date'])
        if merged:
            stock_signals_map[code] = merged
        else:
            del stock_signals_map[code]

    return stock_signals_map


def _extract_concept_group_max_date_ymd(ws) -> Optional[str]:
    """
    从概念分组sheet表头解析最大交易日（YYYYMMDD），用于将fupan近3个月窗口与ladder_analysis对齐。
    """
    max_ymd: Optional[str] = None
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        parsed = _parse_sheet_date(header_val)
        if not parsed:
            continue
        ymd = parsed.replace('-', '')
        if len(ymd) == 8 and ymd.isdigit():
            if max_ymd is None or ymd > max_ymd:
                max_ymd = ymd
    return max_ymd


def _extract_momo_codes_in_order_from_concept_group(ws, top_n: int) -> Tuple[List[str], Dict[str, str]]:
    """
    从 ladder_analysis.xlsx 的“*概念分组”sheet中提取【默默上涨】分组的股票列表（按excel已排序顺序），并限制前top_n只。

    返回：
    - momo_codes: [code1, code2, ...]（去重保序）
    - code_to_name: {code: stock_name}
    """
    momo_codes: List[str] = []
    code_to_name: Dict[str, str] = {}
    momo_seen: set = set()

    current_group: Optional[str] = None
    group_header_pat = re.compile(r'^【(.+)】$')

    # 从行号4附近开始更稳（前几行是日期/大盘指标）
    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if not isinstance(cell_val, str):
            continue
        s = cell_val.strip()
        if not s:
            continue

        m = group_header_pat.match(s)
        if m:
            current_group = m.group(1)
            continue

        if current_group != '默默上涨':
            continue

        # 股票代码可能直接就是 6位数字，也可能带符号/空格，这里抓取6位数字
        code_m = re.search(r'(\d{6})', s)
        if not code_m:
            continue
        code = code_m.group(1)

        if code in momo_seen:
            continue

        stock_name = ws.cell(row=row_idx, column=3).value
        if stock_name is None:
            stock_name = ''

        momo_seen.add(code)
        momo_codes.append(code)
        code_to_name[code] = str(stock_name).strip()

        if len(momo_codes) >= top_n:
            break

    return momo_codes, code_to_name


def _calc_signal_window(signals_map: Dict[str, List[Dict]], before_days: int, after_days: int) -> Tuple[Optional[str], Optional[str]]:
    all_dates = []
    for sigs in signals_map.values():
        for s in sigs:
            d = s.get('signal_date')
            if d:
                all_dates.append(d.replace('-', ''))
    if not all_dates:
        return None, None

    start_ymd = min(all_dates)
    end_ymd = max(all_dates)
    try:
        start_ymd = get_n_trading_days_before(start_ymd, max(0, before_days))
        if '-' in start_ymd:
            start_ymd = start_ymd.replace('-', '')
    except Exception:
        pass

    try:
        for _ in range(max(0, after_days)):
            next_day = get_next_trading_day(end_ymd)
            if not next_day:
                break
            end_ymd = next_day
    except Exception:
        pass
    return start_ymd, end_ymd


def _normalize_date_code_lookup(raw_lookup: Dict) -> Dict[str, set]:
    out: Dict[str, set] = {}
    if not isinstance(raw_lookup, dict):
        return out
    for raw_date, codes in raw_lookup.items():
        ds = str(raw_date).strip().replace('-', '')
        if len(ds) != 8 or not ds.isdigit():
            continue
        norm_codes = set()
        try:
            iterable = list(codes) if codes is not None else []
        except Exception:
            iterable = []
        for code in iterable:
            m = re.search(r'(\d{6})', str(code))
            if m:
                norm_codes.add(m.group(1))
        if norm_codes:
            out[ds] = norm_codes
    return out


def _prepare_zaban_lookup(start_ymd: str, end_ymd: str) -> Dict[str, set]:
    zaban_df = load_zaban_data(start_ymd, end_ymd)
    if zaban_df is None or zaban_df.empty:
        logging.info("炸板数据为空，默默上涨HTML将不添加炸板标记")
        return {}

    lookup = zaban_df.attrs.get('zaban_lookup')
    if isinstance(lookup, dict):
        return _normalize_date_code_lookup(lookup)

    fallback: Dict[str, set] = {}
    try:
        for _, row in zaban_df.iterrows():
            d = str(row.get('date', '')).strip()
            code = _normalize_stock_code(row.get('stock_code'))
            if len(d) == 8 and code:
                fallback.setdefault(d, set()).add(code)
    except Exception:
        pass
    return _normalize_date_code_lookup(fallback)


def _prepare_attention_lookup(start_ymd: str, end_ymd: str) -> Dict[str, set]:
    lookup: Dict[str, set] = {}
    for is_main_board in (True, False):
        df = load_attention_data(start_ymd, end_ymd, is_main_board=is_main_board)
        if df is None or df.empty:
            continue
        try:
            for _, row in df.iterrows():
                raw_date = str(row.get('日期', '')).strip()
                code = _normalize_stock_code(row.get('股票代码'))
                if not raw_date or not code:
                    continue
                if '年' in raw_date:
                    d_obj = datetime.strptime(raw_date, '%Y年%m月%d日')
                else:
                    d_obj = pd.to_datetime(raw_date)
                d = d_obj.strftime('%Y%m%d')
                lookup.setdefault(d, set()).add(code)
        except Exception:
            continue
    return lookup


def _build_signals_from_lookup_for_chart(
    stock_code: str,
    chart_df: pd.DataFrame,
    lookup: Dict[str, set],
    signal_type: str,
    details: str,
) -> List[Dict]:
    if chart_df is None or chart_df.empty or not lookup:
        return []
    code = str(stock_code).zfill(6)
    out: List[Dict] = []
    for dt in chart_df.index:
        ymd = dt.strftime('%Y%m%d')
        if code in lookup.get(ymd, set()):
            out.append({
                'code': code,
                'signal_date': dt.strftime('%Y-%m-%d'),
                'signal_type': signal_type,
                'price': None,
                'details': details,
            })
    return out


def generate_momo_concept_group_html_charts(
    columns: int = 2,
    before_days: int = 60,
    after_days: int = 30,
    output_dir: str = './excel/html_charts',
    data_dir: str = './data/astocks',
    ladder_excel_path: str = './excel/ladder_analysis.xlsx',
    top_n: int = MOMO_HTML_DRAW_TOP_N,
) -> Optional[str]:
    """
    生成【默默上涨】（概念分组口径）HTML图表。

    说明：
    - 事件源来自【默默上涨】sheet近3个月内全部入选记录（不做 keep='last'）
    - 额外叠加炸板与关注度入榜标记
    - 但绘图只取 ladder_analysis.xlsx 概念分组里【默默上涨】分组排序的前 top_n 只
    """
    if columns not in [1, 2, 3]:
        logging.warning(f"列数 {columns} 无效，使用默认值2")
        columns = 2

    if not os.path.exists(ladder_excel_path):
        logging.error(f"ladder分析Excel不存在: {ladder_excel_path}")
        return None

    # 1) 从 ladder_analysis.xlsx 中提取“默默上涨”分组的前top_n只（按excel排序）
    from openpyxl import load_workbook

    ladder_wb = load_workbook(ladder_excel_path, data_only=True)
    concept_sheets = [sn for sn in ladder_wb.sheetnames if '概念分组' in str(sn)]
    if not concept_sheets:
        logging.error(f"在 {ladder_excel_path} 未找到“概念分组”类sheet")
        return None

    # 优先选择“涨停梯队xxx_概念分组”，否则取第一个概念分组sheet
    preferred = [sn for sn in concept_sheets if '涨停梯队' in str(sn)]
    concept_sheet_name = preferred[0] if preferred else concept_sheets[0]

    concept_ws = ladder_wb[concept_sheet_name]

    momo_codes, code_to_name = _extract_momo_codes_in_order_from_concept_group(concept_ws, top_n=top_n)
    if not momo_codes:
        logging.error("在ladder_analysis.xlsx中未提取到【默默上涨】股票列表")
        return None

    # 2) 仍用 momo_shangzhang_processor 拿“这些股票的全部入选日”（keep_latest_per_stock=False）
    #    通过sheet最大日期对齐fupan近3个月窗口，减少与ladder_analysis口径偏差。
    sheet_end_ymd = _extract_concept_group_max_date_ymd(concept_ws)
    if not sheet_end_ymd:
        sheet_end_ymd = datetime.now().strftime('%Y%m%d')

    momo_df_all = load_momo_shangzhang_data(
        start_date=sheet_end_ymd,
        end_date=sheet_end_ymd,
        keep_latest_per_stock=False,
    )
    if momo_df_all is None or momo_df_all.empty:
        logging.warning("fupan_stocks.xlsx中未能加载到【默默上涨】数据")
        return None

    if '纯代码' not in momo_df_all.columns:
        logging.error("【默默上涨】数据缺少“纯代码”列，无法匹配ladder_analysis选股集合")
        return None

    momo_set = set(momo_codes)
    # processor 的“纯代码”可能是 000338.SZ / 600000.SH，这里统一归一化后再匹配
    momo_df_all = momo_df_all.copy()
    momo_df_all['_norm_code'] = momo_df_all['纯代码'].apply(_normalize_stock_code)
    momo_df = momo_df_all[momo_df_all['_norm_code'].isin(momo_set)].copy()
    if '_norm_code' in momo_df.columns:
        momo_df.drop(columns=['_norm_code'], inplace=True)
    signals_map = _build_momo_signals_map(momo_df)
    if not signals_map:
        logging.warning("未提取到有效的【默默上涨】入选事件")
        return None

    # 标准概念映射（与原有实现一致）：用于在图表标题显示 [概念A+概念B+...]
    concept_lookup: Dict[str, List[str]] = {}
    try:
        if is_map_available():
            # 只拉取当前绘图范围内股票，避免不必要的映射读取
            concept_lookup = {code: get_stock_concepts(code) for code in momo_codes}
    except Exception as e:
        logging.warning(f"加载标准概念映射失败，标题将不显示概念标签: {e}")

    window_start_ymd, window_end_ymd = _calc_signal_window(signals_map, before_days=before_days, after_days=after_days)
    zaban_lookup: Dict[str, set] = {}
    attention_lookup: Dict[str, set] = {}
    if window_start_ymd and window_end_ymd:
        zaban_lookup = _prepare_zaban_lookup(window_start_ymd, window_end_ymd)
        attention_lookup = _prepare_attention_lookup(window_start_ymd, window_end_ymd)

    os.makedirs(output_dir, exist_ok=True)

    chart_figures = []
    chart_titles = []
    # 按ladder_analysis.xlsx提取的顺序绘图（与“市值排序/成交额排序”口径一致）
    for stock_code in momo_codes:
        if stock_code not in signals_map:
            continue
        signals = signals_map[stock_code]
        stock_name = code_to_name.get(stock_code) or str(next((s.get('name') for s in signals if s.get('name')), '') or '')
        signal_dates = [s['signal_date'] for s in signals]
        earliest_ymd = min(signal_dates).replace('-', '')
        latest_ymd = max(signal_dates).replace('-', '')

        chart_start = get_n_trading_days_before(earliest_ymd, before_days)
        if '-' in chart_start:
            chart_start = chart_start.replace('-', '')
        chart_end = latest_ymd
        for _ in range(after_days):
            next_day = get_next_trading_day(chart_end)
            if not next_day:
                break
            chart_end = next_day

        stock_data = read_stock_data(stock_code, data_dir)
        if stock_data is None or stock_data.empty:
            logging.warning(f"未找到股票数据: {stock_code} {stock_name}")
            continue

        start_dt = datetime.strptime(chart_start, '%Y%m%d')
        end_dt = datetime.strptime(chart_end, '%Y%m%d')
        chart_df = stock_data.loc[start_dt:end_dt].copy()
        chart_df = chart_df.dropna(subset=['Open', 'High', 'Low', 'Close'])
        if chart_df.empty:
            continue

        zaban_signals = _build_signals_from_lookup_for_chart(
            stock_code, chart_df, zaban_lookup, signal_type='炸板', details='炸板日'
        )
        attention_signals = _build_signals_from_lookup_for_chart(
            stock_code, chart_df, attention_lookup, signal_type='关注度入榜', details='关注度入榜'
        )

        merged_signals = signals + zaban_signals + attention_signals
        uniq = {}
        for sig in merged_signals:
            uniq[(sig['signal_date'], sig['signal_type'])] = sig
        merged_signals = sorted(uniq.values(), key=lambda x: x['signal_date'])

        fig = _create_single_chart_figure(
            stock_code=stock_code,
            stock_name=stock_name,
            chart_df=chart_df,
            signal_dates_info=merged_signals,
            before_days=before_days,
            after_days=after_days,
            data_dir=data_dir,
            concepts=concept_lookup.get(stock_code) or [],
            entry_range_anchor_signal_types=[_MOMO_SIGNAL_TYPE],
        )
        if fig is None:
            continue

        chart_figures.append(fig)
        chart_titles.append(f"{stock_code} {stock_name}".strip())

    if not chart_figures:
        logging.warning("没有可用图表，未生成HTML")
        return None

    rows = (len(chart_figures) + columns - 1) // columns
    html_content = _create_combined_html(chart_figures, chart_titles, columns, rows)

    html_filename = f"momo_concept_group_all_{columns}cols.html"
    html_path = os.path.join(output_dir, html_filename)
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    logging.info(f"默默上涨（概念分组）HTML生成完成: {html_path}（共 {len(chart_figures)} 只）")
    return html_path

