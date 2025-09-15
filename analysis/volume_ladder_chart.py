"""
æˆäº¤é‡æ¶¨è·Œå¹…åˆ†ææ¨¡å—
ç”¨äºåˆ›å»ºåŸºäºæˆäº¤é‡æ¶¨è·Œå¹…çš„æ¶¨åœæ¢¯é˜Ÿåˆ†æè¡¨
"""

from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from analysis.ladder_chart import (
    get_stock_data, BORDER_STYLE, format_stock_code_cell,
    format_concept_cell, format_stock_name_cell, extract_pure_stock_code,
    get_stock_concept, calculate_max_board_level, get_market_marker,
    get_new_high_markers_cached, NEW_HIGH_MARKER
)
from utils.theme_color_util import add_market_indicators

# æˆäº¤é‡æ¶¨è·Œå¹…é¢œè‰²æ˜ å°„ - åŸºäºç™¾åˆ†ä½æ•°åŠ¨æ€ä¸Šè‰²
VOLUME_CHANGE_COLORS = {
    "EXTREME_POSITIVE": "FFB000",  # æ·±é»„è‰² - æåº¦æ”¾é‡
    "STRONG_POSITIVE": "FFCC00",  # ä¸­é»„è‰² - å¤§å¹…æ”¾é‡
    "MODERATE_POSITIVE": "FFE066",  # æµ…é»„è‰² - æ˜æ˜¾æ”¾é‡
    "MILD_POSITIVE": "FFF2CC",  # ææµ…é»„è‰² - é€‚åº¦æ”¾é‡
    "MILD_NEGATIVE": "CCFFCC",  # æµ…ç»¿è‰² - é€‚åº¦ç¼©é‡
    "MODERATE_NEGATIVE": "99FF99",  # ä¸­ç»¿è‰² - æ˜æ˜¾ç¼©é‡
    "STRONG_NEGATIVE": "66FF66",  # æ·±ç»¿è‰² - å¤§å¹…ç¼©é‡
    "EXTREME_NEGATIVE": "33FF33",  # ææ·±ç»¿è‰² - æåº¦ç¼©é‡
}

# æˆäº¤é‡ä¸Šè‰²é…ç½®
VOLUME_COLOR_CONFIG = {
    "COLOR_PERCENTAGE": 0.30,  # ä¸Šè‰²æ¯”ä¾‹ï¼š20%çš„æ•°æ®ä¼šè¢«ä¸Šè‰²
    "POSITIVE_RATIO": 0.6,  # æ­£å€¼(æ”¾é‡)åœ¨ä¸Šè‰²æ•°æ®ä¸­çš„æ¯”ä¾‹ï¼š60%
    "NEGATIVE_RATIO": 0.4,  # è´Ÿå€¼(ç¼©é‡)åœ¨ä¸Šè‰²æ•°æ®ä¸­çš„æ¯”ä¾‹ï¼š40%
}

# æˆäº¤é‡è¶‹åŠ¿åˆ†æç›¸å…³å‚æ•°
VOLUME_MA_DAYS = 7  # æˆäº¤é‡å‡çº¿å¤©æ•°
VOLUME_TREND_DAYS = 3  # åˆ¤æ–­è¶‹åŠ¿çš„è¿ç»­å¤©æ•°
VOLUME_RATIO_HIGH_THRESHOLD = 1.5  # é«˜æ´»è·ƒé˜ˆå€¼ï¼ˆæˆäº¤é‡/å‡çº¿ï¼‰
VOLUME_RATIO_LOW_THRESHOLD = 0.6  # ä½æ´»è·ƒé˜ˆå€¼
VOLUME_MA_SLOPE_THRESHOLD = 0.1  # å‡çº¿æ–œç‡é˜ˆå€¼ï¼ˆæ—¥å˜åŒ–ç‡ï¼‰


def get_stock_daily_volume_change(stock_code, date_str_yyyymmdd):
    """
    è·å–æŒ‡å®šè‚¡ç¥¨åœ¨ç‰¹å®šæ—¥æœŸçš„æˆäº¤é‡æ¶¨è·Œå¹…

    Args:
        stock_code: è‚¡ç¥¨ä»£ç 
        date_str_yyyymmdd: ç›®æ ‡æ—¥æœŸ (YYYYMMDD)

    Returns:
        float: æˆäº¤é‡æ¶¨è·Œå¹…ç™¾åˆ†æ¯”ï¼Œå¦‚æœæ•°æ®ä¸å­˜åœ¨åˆ™è¿”å›None
    """
    try:
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)

        if df is None or target_row is None or target_row.empty or target_idx == 0:
            return None

        # è·å–å½“å¤©æˆäº¤é‡
        current_volume = target_row['æˆäº¤é‡'].values[0]

        # è·å–å‰ä¸€å¤©æˆäº¤é‡
        prev_volume = df.iloc[target_idx - 1]['æˆäº¤é‡']

        # æ£€æŸ¥æ•°æ®æœ‰æ•ˆæ€§
        if pd.isna(current_volume) or pd.isna(prev_volume) or prev_volume <= 0:
            return None

        # è®¡ç®—æˆäº¤é‡æ¶¨è·Œå¹…ï¼ˆè¿”å›å€æ•°ï¼Œä¸æ˜¯ç™¾åˆ†æ¯”ï¼‰
        volume_change = (current_volume / prev_volume) - 1

        return volume_change

    except Exception as e:
        print(f"è·å–è‚¡ç¥¨ {stock_code} åœ¨ {date_str_yyyymmdd} çš„æˆäº¤é‡æ¶¨è·Œå¹…æ—¶å‡ºé”™: {e}")
        return None


def calculate_volume_ma(stock_code, date_str_yyyymmdd, days=VOLUME_MA_DAYS):
    """
    è®¡ç®—æŒ‡å®šè‚¡ç¥¨åœ¨ç‰¹å®šæ—¥æœŸçš„æˆäº¤é‡å‡çº¿

    Args:
        stock_code: è‚¡ç¥¨ä»£ç 
        date_str_yyyymmdd: ç›®æ ‡æ—¥æœŸ (YYYYMMDD)
        days: å‡çº¿å¤©æ•°

    Returns:
        float: æˆäº¤é‡å‡çº¿ï¼Œå¦‚æœæ•°æ®ä¸å­˜åœ¨åˆ™è¿”å›None
    """
    try:
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)

        if df is None or target_row is None or target_row.empty:
            return None

        # ç¡®ä¿æœ‰è¶³å¤Ÿçš„å†å²æ•°æ®æ¥è®¡ç®—å‡çº¿
        if target_idx < days - 1:
            return None

        # è·å–åŒ…å«å½“å¤©åœ¨å†…çš„å‰dayså¤©çš„æˆäº¤é‡æ•°æ®
        volume_data = df.iloc[target_idx - days + 1:target_idx + 1]['æˆäº¤é‡'].values

        # æ£€æŸ¥æ•°æ®æœ‰æ•ˆæ€§
        if len(volume_data) != days or pd.isna(volume_data).any():
            return None

        # è®¡ç®—å‡çº¿
        volume_ma = volume_data.mean()

        return volume_ma

    except Exception as e:
        print(f"è®¡ç®—è‚¡ç¥¨ {stock_code} åœ¨ {date_str_yyyymmdd} çš„æˆäº¤é‡å‡çº¿æ—¶å‡ºé”™: {e}")
        return None


def get_volume_trend_indicator(stock_code, date_str_yyyymmdd, formatted_trading_days, date_mapping):
    """
    è·å–æˆäº¤é‡è¶‹åŠ¿æŒ‡æ ‡

    Args:
        stock_code: è‚¡ç¥¨ä»£ç 
        date_str_yyyymmdd: ç›®æ ‡æ—¥æœŸ (YYYYMMDD)
        formatted_trading_days: æ ¼å¼åŒ–çš„äº¤æ˜“æ—¥åˆ—è¡¨
        date_mapping: æ—¥æœŸæ˜ å°„

    Returns:
        str: è¶‹åŠ¿æŒ‡æ ‡æ–‡å­—ï¼Œå¦‚æœæ— æ˜æ˜¾è¶‹åŠ¿åˆ™è¿”å›ç©ºå­—ç¬¦ä¸²
    """
    try:
        # è·å–å½“å‰æ—¥æœŸçš„æˆäº¤é‡å’Œå‡çº¿
        df, target_row, target_idx = get_stock_data(stock_code, date_str_yyyymmdd)
        if df is None or target_row is None or target_row.empty:
            return ""

        current_volume = target_row['æˆäº¤é‡'].values[0]
        current_ma = calculate_volume_ma(stock_code, date_str_yyyymmdd)

        if pd.isna(current_volume) or current_ma is None or current_ma <= 0:
            return ""

        # è®¡ç®—å½“å‰æˆäº¤é‡ç›¸å¯¹å‡çº¿çš„æ¯”å€¼
        volume_ratio = current_volume / current_ma

        # è®¡ç®—å‡çº¿è¶‹åŠ¿ï¼ˆè·å–æœ€è¿‘å‡ ä¸ªæœ‰æ•°æ®çš„äº¤æ˜“æ—¥çš„å‡çº¿ï¼‰
        ma_values = []
        current_date_idx = None

        # æ‰¾åˆ°å½“å‰æ—¥æœŸåœ¨äº¤æ˜“æ—¥åˆ—è¡¨ä¸­çš„ä½ç½®
        for i, formatted_day in enumerate(formatted_trading_days):
            if date_mapping.get(formatted_day) == date_str_yyyymmdd:
                current_date_idx = i
                break

        if current_date_idx is None:
            return ""

        # ä»å½“å‰æ—¥æœŸå¾€å‰æ”¶é›†æœ‰æ•ˆçš„å‡çº¿æ•°æ®ï¼Œè·³è¿‡å‘¨æœ«ç­‰æ— æ•°æ®çš„æ—¥æœŸ
        collected_days = 0
        for i in range(current_date_idx + 1):  # ä»å½“å‰æ—¥æœŸå¾€å‰éå†
            day_idx = current_date_idx - i
            if day_idx >= 0 and day_idx < len(formatted_trading_days):
                formatted_day = formatted_trading_days[day_idx]
                day_date = date_mapping.get(formatted_day)
                if day_date:
                    ma_value = calculate_volume_ma(stock_code, day_date)
                    if ma_value is not None:
                        ma_values.insert(0, ma_value)  # æ’å…¥åˆ°å¼€å¤´ï¼Œä¿æŒæ—¶é—´é¡ºåº
                        collected_days += 1
                        if collected_days >= VOLUME_TREND_DAYS:
                            break

        # å¦‚æœæ”¶é›†åˆ°çš„æ•°æ®ä¸è¶³ï¼Œåˆ™ä¸è®¡ç®—è¶‹åŠ¿
        if len(ma_values) < 2:
            # æ•°æ®ä¸è¶³æ—¶ï¼ŒåªåŸºäºå½“å‰æˆäº¤é‡å’Œå‡çº¿çš„æ¯”å€¼åˆ¤æ–­
            if volume_ratio >= VOLUME_RATIO_HIGH_THRESHOLD:
                return "é«˜é‡ğŸ”¥"
            elif volume_ratio <= VOLUME_RATIO_LOW_THRESHOLD:
                return "ä½é‡ğŸ’¤"
            else:
                return ""

        # è®¡ç®—å‡çº¿æ–œç‡ï¼ˆæœ€åä¸€å¤©ç›¸å¯¹ç¬¬ä¸€å¤©çš„å˜åŒ–ç‡ï¼‰
        ma_slope = (ma_values[-1] - ma_values[0]) / ma_values[0] if ma_values[0] > 0 else 0

        # åˆ¤æ–­è¶‹åŠ¿
        if volume_ratio >= VOLUME_RATIO_HIGH_THRESHOLD:
            if ma_slope > VOLUME_MA_SLOPE_THRESHOLD:
                return "æ”¾é‡â†—"
            else:
                return "é«˜é‡ğŸ”¥"
        elif volume_ratio <= VOLUME_RATIO_LOW_THRESHOLD:
            if ma_slope < -VOLUME_MA_SLOPE_THRESHOLD:
                return "ç¼©é‡â†˜"
            else:
                return "ä½é‡ğŸ’¤"
        elif ma_slope > VOLUME_MA_SLOPE_THRESHOLD:
            return "é‡å¢â†—"
        elif ma_slope < -VOLUME_MA_SLOPE_THRESHOLD:
            return "é‡å‡â†˜"

        return ""

    except Exception as e:
        print(f"è·å–è‚¡ç¥¨ {stock_code} åœ¨ {date_str_yyyymmdd} çš„æˆäº¤é‡è¶‹åŠ¿æŒ‡æ ‡æ—¶å‡ºé”™: {e}")
        return ""


def get_color_for_volume_change(volume_change, thresholds=None):
    """
    æ ¹æ®æˆäº¤é‡æ¶¨è·Œå¹…è·å–èƒŒæ™¯è‰²

    Args:
        volume_change: æˆäº¤é‡æ¶¨è·Œå¹…ï¼ˆå€æ•°ï¼Œä¸æ˜¯ç™¾åˆ†æ¯”ï¼‰
        thresholds: åŠ¨æ€é˜ˆå€¼å­—å…¸ï¼ŒåŒ…å«å„çº§åˆ«çš„é˜ˆå€¼

    Returns:
        str: é¢œè‰²ä»£ç ï¼Œå˜åŒ–ä¸å¤§æ—¶è¿”å›None
    """
    if volume_change is None:
        return None

    # å¦‚æœæ²¡æœ‰æä¾›åŠ¨æ€é˜ˆå€¼ï¼Œä½¿ç”¨å›ºå®šé˜ˆå€¼ï¼ˆå‘åå…¼å®¹ï¼‰
    if thresholds is None:
        if volume_change >= 3.0:
            return VOLUME_CHANGE_COLORS["EXTREME_POSITIVE"]
        elif volume_change >= 2.0:
            return VOLUME_CHANGE_COLORS["STRONG_POSITIVE"]
        elif volume_change >= 1.0:
            return VOLUME_CHANGE_COLORS["MODERATE_POSITIVE"]
        elif volume_change >= 0.5:
            return VOLUME_CHANGE_COLORS["MILD_POSITIVE"]
        elif volume_change <= -0.8:
            return VOLUME_CHANGE_COLORS["EXTREME_NEGATIVE"]
        elif volume_change <= -0.7:
            return VOLUME_CHANGE_COLORS["STRONG_NEGATIVE"]
        elif volume_change <= -0.5:
            return VOLUME_CHANGE_COLORS["MILD_NEGATIVE"]
        else:
            return None

    # ä½¿ç”¨åŠ¨æ€é˜ˆå€¼
    if volume_change >= thresholds.get("extreme_positive", 3.0):
        return VOLUME_CHANGE_COLORS["EXTREME_POSITIVE"]
    elif volume_change >= thresholds.get("strong_positive", 2.0):
        return VOLUME_CHANGE_COLORS["STRONG_POSITIVE"]
    elif volume_change >= thresholds.get("moderate_positive", 1.0):
        return VOLUME_CHANGE_COLORS["MODERATE_POSITIVE"]
    elif volume_change >= thresholds.get("mild_positive", 0.5):
        return VOLUME_CHANGE_COLORS["MILD_POSITIVE"]
    elif volume_change <= thresholds.get("extreme_negative", -0.8):
        return VOLUME_CHANGE_COLORS["EXTREME_NEGATIVE"]
    elif volume_change <= thresholds.get("strong_negative", -0.7):
        return VOLUME_CHANGE_COLORS["STRONG_NEGATIVE"]
    elif volume_change <= thresholds.get("moderate_negative", -0.5):
        return VOLUME_CHANGE_COLORS["MODERATE_NEGATIVE"]
    elif volume_change <= thresholds.get("mild_negative", -0.3):
        return VOLUME_CHANGE_COLORS["MILD_NEGATIVE"]
    else:
        return None


def calculate_volume_change_thresholds(volume_changes):
    """
    æ ¹æ®æˆäº¤é‡å˜åŒ–æ•°æ®è®¡ç®—åŠ¨æ€é˜ˆå€¼

    Args:
        volume_changes: æˆäº¤é‡å˜åŒ–æ•°æ®åˆ—è¡¨

    Returns:
        dict: åŒ…å«å„çº§åˆ«é˜ˆå€¼çš„å­—å…¸
    """
    import numpy as np

    # è¿‡æ»¤æ‰Noneå€¼
    valid_changes = [v for v in volume_changes if v is not None and not pd.isna(v)]

    if len(valid_changes) < 10:  # æ•°æ®å¤ªå°‘æ—¶ä½¿ç”¨å›ºå®šé˜ˆå€¼
        return None

    valid_changes = np.array(valid_changes)

    # åˆ†ç¦»æ­£å€¼å’Œè´Ÿå€¼
    positive_changes = valid_changes[valid_changes > 0]
    negative_changes = valid_changes[valid_changes < 0]

    # è®¡ç®—é…ç½®å‚æ•°
    color_percentage = VOLUME_COLOR_CONFIG["COLOR_PERCENTAGE"]
    positive_ratio = VOLUME_COLOR_CONFIG["POSITIVE_RATIO"]
    negative_ratio = VOLUME_COLOR_CONFIG["NEGATIVE_RATIO"]

    thresholds = {}

    # è®¡ç®—æ­£å€¼é˜ˆå€¼ï¼ˆæ”¾é‡ï¼‰
    if len(positive_changes) > 0:
        positive_count = int(len(valid_changes) * color_percentage * positive_ratio)
        if positive_count > 0:
            positive_sorted = np.sort(positive_changes)[::-1]  # é™åºæ’åˆ—

            # æ ¹æ®æ•°é‡åˆ†é…é˜ˆå€¼
            if positive_count >= 4:
                thresholds["extreme_positive"] = positive_sorted[positive_count // 4 - 1]
                thresholds["strong_positive"] = positive_sorted[positive_count // 2 - 1]
                thresholds["moderate_positive"] = positive_sorted[positive_count * 3 // 4 - 1]
                thresholds["mild_positive"] = positive_sorted[positive_count - 1]
            elif positive_count >= 2:
                thresholds["strong_positive"] = positive_sorted[0]
                thresholds["mild_positive"] = positive_sorted[positive_count - 1]
            else:
                thresholds["mild_positive"] = positive_sorted[0]

    # è®¡ç®—è´Ÿå€¼é˜ˆå€¼ï¼ˆç¼©é‡ï¼‰
    if len(negative_changes) > 0:
        negative_count = int(len(valid_changes) * color_percentage * negative_ratio)
        if negative_count > 0:
            negative_sorted = np.sort(negative_changes)  # å‡åºæ’åˆ—ï¼ˆè´Ÿå€¼è¶Šå°è¶Šæç«¯ï¼‰

            # æ ¹æ®æ•°é‡åˆ†é…é˜ˆå€¼
            if negative_count >= 4:
                thresholds["extreme_negative"] = negative_sorted[negative_count // 4 - 1]
                thresholds["strong_negative"] = negative_sorted[negative_count // 2 - 1]
                thresholds["moderate_negative"] = negative_sorted[negative_count * 3 // 4 - 1]
                thresholds["mild_negative"] = negative_sorted[negative_count - 1]
            elif negative_count >= 2:
                thresholds["strong_negative"] = negative_sorted[0]
                thresholds["mild_negative"] = negative_sorted[negative_count - 1]
            else:
                thresholds["mild_negative"] = negative_sorted[0]

    return thresholds


def format_volume_change_cell(ws, row, col, volume_change, stock_code, current_date_obj, thresholds=None):
    """
    æ ¼å¼åŒ–æˆäº¤é‡æ¶¨è·Œå¹…å•å…ƒæ ¼

    Args:
        ws: Excelå·¥ä½œè¡¨
        row: è¡Œç´¢å¼•
        col: åˆ—ç´¢å¼•
        volume_change: æˆäº¤é‡æ¶¨è·Œå¹…
        stock_code: è‚¡ç¥¨ä»£ç 
        current_date_obj: å½“å‰æ—¥æœŸå¯¹è±¡
        thresholds: åŠ¨æ€é˜ˆå€¼å­—å…¸

    Returns:
        å•å…ƒæ ¼å¯¹è±¡
    """
    cell = ws.cell(row=row, column=col)

    if pd.notna(volume_change):
        # æ˜¾ç¤ºæˆäº¤é‡æ¶¨è·Œå¹…ï¼Œä¿ç•™2ä½å°æ•°
        cell_value = f"{volume_change:.2f}"
        cell.value = cell_value

        # è®¾ç½®èƒŒæ™¯è‰²ï¼Œä½¿ç”¨åŠ¨æ€é˜ˆå€¼
        color = get_color_for_volume_change(volume_change, thresholds)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    else:
        cell.value = "åœç‰Œ"

    # è®¾ç½®å•å…ƒæ ¼æ ¼å¼
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER_STYLE

    return cell


def process_volume_daily_cell(ws, row_idx, col_idx, stock, current_date_obj, formatted_day,
                              stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                              zaban_df, new_high_markers, thresholds=None):
    """
    å¤„ç†æˆäº¤é‡ç‰ˆæœ¬çš„æ¯æ—¥å•å…ƒæ ¼æ•°æ®

    Args:
        ws: Excelå·¥ä½œè¡¨
        row_idx: è¡Œç´¢å¼•
        col_idx: åˆ—ç´¢å¼•
        stock: è‚¡ç¥¨æ•°æ®
        current_date_obj: å½“å‰æ—¥æœŸå¯¹è±¡
        formatted_day: æ ¼å¼åŒ–çš„æ—¥æœŸå­—ç¬¦ä¸²
        stock_details: è‚¡ç¥¨è¯¦ç»†ä¿¡æ¯
        date_mapping: æ—¥æœŸæ˜ å°„
        max_tracking_days: æ–­æ¿åè·Ÿè¸ªçš„æœ€å¤§å¤©æ•°
        max_tracking_days_before: å…¥é€‰å‰è·Ÿè¸ªçš„æœ€å¤§å¤©æ•°
        zaban_df: ç‚¸æ¿æ•°æ®
        new_high_markers: æ–°é«˜æ ‡è®°

    Returns:
        æœ€åè¿æ¿æ—¥æœŸï¼ˆå¦‚æœæœ‰çš„è¯ï¼‰
    """
    stock_code = stock['stock_code']
    stock_name = stock['stock_name']
    all_board_data = stock['all_board_data']
    entry_date = stock['first_significant_date']

    pure_stock_code = extract_pure_stock_code(stock_code)
    date_yyyymmdd = date_mapping.get(formatted_day)

    # æ£€æŸ¥æ˜¯å¦æœ‰è¿æ¿æ•°æ®
    board_days = all_board_data.get(formatted_day)

    # æ£€æŸ¥æ˜¯å¦ä¸ºç‚¸æ¿è‚¡ç¥¨ï¼ˆä¼˜å…ˆä½¿ç”¨ç¼“å­˜ï¼Œå¦‚æœæ²¡æœ‰ç¼“å­˜åˆ™é‡æ–°è®¡ç®—ï¼‰
    from analysis.ladder_chart import get_cached_zaban_format, check_stock_in_zaban
    is_zaban = get_cached_zaban_format(pure_stock_code, formatted_day)
    if is_zaban is None:
        # å¦‚æœç¼“å­˜ä¸­æ²¡æœ‰ï¼Œé‡æ–°è®¡ç®—
        is_zaban = check_stock_in_zaban(zaban_df, pure_stock_code, formatted_day)

    if pd.notna(board_days) and board_days > 0:
        # æœ‰è¿æ¿æ•°æ®ï¼Œæ˜¾ç¤ºè¿æ¿ä¿¡æ¯ï¼ˆä¿æŒåŸæœ‰é€»è¾‘ï¼‰
        from analysis.ladder_chart import format_board_cell
        cell, last_board_date = format_board_cell(
            ws, row_idx, col_idx, board_days, pure_stock_code,
            f"{stock_code}_{formatted_day}", stock_details, current_date_obj
        )

        # å¦‚æœæ˜¯ç‚¸æ¿è‚¡ç¥¨ï¼Œæ·»åŠ ç‚¸æ¿æ ¼å¼
        if is_zaban:
            from analysis.ladder_chart import add_zaban_underline
            add_zaban_underline(cell)

        return last_board_date
    else:
        # æ²¡æœ‰è¿æ¿æ•°æ®ï¼Œæ˜¾ç¤ºæˆäº¤é‡æ¶¨è·Œå¹…
        if date_yyyymmdd:
            volume_change = get_stock_daily_volume_change(pure_stock_code, date_yyyymmdd)
            cell = format_volume_change_cell(ws, row_idx, col_idx, volume_change,
                                             pure_stock_code, current_date_obj, thresholds)
        else:
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.border = BORDER_STYLE

        # å¦‚æœæ˜¯ç‚¸æ¿è‚¡ç¥¨ï¼Œæ·»åŠ ç‚¸æ¿æ ¼å¼
        if is_zaban:
            from analysis.ladder_chart import add_zaban_underline
            add_zaban_underline(cell)

        return None


def fill_volume_data_rows_with_concept_groups(ws, result_df, shouban_df, stock_reason_group, reason_colors,
                                              stock_entry_count, formatted_trading_days, date_column_start,
                                              show_period_change, period_column, period_days, period_days_long,
                                              stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
                                              zaban_df, show_warning_column=True, thresholds=None):
    """
    å¡«å……æˆäº¤é‡ç‰ˆæœ¬çš„æ•°æ®è¡Œï¼ŒæŒ‰æ¦‚å¿µåˆ†ç»„å¹¶åœ¨ç»„é—´æ·»åŠ åˆ†éš”è¡Œ
    
    Args:
        ws: Excelå·¥ä½œè¡¨
        result_df: æŒ‰æ¦‚å¿µåˆ†ç»„æ’åºçš„æ˜¾è‘—è¿æ¿è‚¡ç¥¨DataFrame
        show_warning_column: æ˜¯å¦æ˜¾ç¤ºå¼‚åŠ¨é¢„è­¦åˆ—
        å…¶ä»–å‚æ•°ä¸åŸç‰ˆæœ¬ç›¸åŒ
    """
    # è®¡ç®—æ‰€æœ‰è‚¡ç¥¨çš„æ–°é«˜æ ‡è®°ï¼ˆä½¿ç”¨ç¼“å­˜ç‰ˆæœ¬ï¼‰
    new_high_markers = get_new_high_markers_cached(result_df, formatted_trading_days, date_mapping)

    current_concept_group = None
    row_idx = 4  # ä»ç¬¬4è¡Œå¼€å§‹ï¼ˆå‰3è¡Œæ˜¯æ ‡é¢˜å’Œå¤§ç›˜æŒ‡æ ‡ï¼‰

    for i, (_, stock) in enumerate(result_df.iterrows()):
        concept_group = stock.get('concept_group', 'å…¶ä»–')

        # å¦‚æœæ¦‚å¿µç»„å‘ç”Ÿå˜åŒ–ï¼Œæ’å…¥æ¦‚å¿µç»„æ ‡é¢˜è¡Œ
        if concept_group != current_concept_group:
            # å¦‚æœä¸æ˜¯ç¬¬ä¸€ä¸ªæ¦‚å¿µç»„ï¼Œå…ˆæ’å…¥ç©ºè¡Œåˆ†éš”
            if current_concept_group is not None:
                row_idx += 1  # æ·»åŠ ç©ºè¡Œ

            current_concept_group = concept_group

            # æ’å…¥æ¦‚å¿µç»„æ ‡é¢˜è¡Œ
            concept_title_cell = ws.cell(row=row_idx, column=1, value=f"ã€{concept_group}ã€‘")
            concept_title_cell.font = Font(bold=True, size=12)
            concept_title_cell.alignment = Alignment(horizontal='left')

            # è®¾ç½®æ¦‚å¿µç»„æ ‡é¢˜è¡Œçš„èƒŒæ™¯è‰²
            if concept_group in reason_colors:
                bg_color = reason_colors[concept_group]
                concept_title_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
                # å¦‚æœèƒŒæ™¯è‰²è¾ƒæ·±ï¼Œä½¿ç”¨ç™½è‰²å­—ä½“
                if bg_color in ["FF5A5A", "FF8C42", "9966FF", "45B5FF"]:
                    concept_title_cell.font = Font(color="FFFFFF", bold=True, size=12)

            # åˆå¹¶æ¦‚å¿µç»„æ ‡é¢˜è¡Œçš„æ‰€æœ‰åˆ—
            end_col = date_column_start + len(formatted_trading_days) - 1
            if show_period_change:
                end_col += 1
            if show_warning_column:
                end_col += 1

            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
            row_idx += 1

        # å¡«å……è‚¡ç¥¨æ•°æ®è¡Œ
        fill_single_volume_stock_row(ws, row_idx, stock, stock_reason_group, reason_colors, stock_entry_count,
                                     formatted_trading_days, date_column_start, show_period_change, period_column,
                                     period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                     max_tracking_days_before, zaban_df, new_high_markers, show_warning_column,
                                     thresholds)

        row_idx += 1


def fill_single_volume_stock_row(ws, row_idx, stock, stock_reason_group, reason_colors, stock_entry_count,
                                 formatted_trading_days, date_column_start, show_period_change, period_column,
                                 period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                 max_tracking_days_before, zaban_df, new_high_markers, show_warning_column=True,
                                 thresholds=None):
    """
    å¡«å……å•åªè‚¡ç¥¨çš„æˆäº¤é‡ç‰ˆæœ¬æ•°æ®è¡Œ
    
    Args:
        ws: Excelå·¥ä½œè¡¨
        row_idx: è¡Œç´¢å¼•
        stock: è‚¡ç¥¨æ•°æ®
        å…¶ä»–å‚æ•°ä¸åŸç‰ˆæœ¬ç›¸åŒ
    """
    # æå–åŸºæœ¬è‚¡ç¥¨ä¿¡æ¯
    stock_code = stock['stock_code']
    stock_name = stock['stock_name']
    all_board_data = stock['all_board_data']

    # æå–çº¯ä»£ç 
    pure_stock_code = extract_pure_stock_code(stock_code)

    # è®¾ç½®è‚¡ç¥¨ä»£ç åˆ—ï¼ˆç¬¬ä¸€åˆ—ï¼‰
    format_stock_code_cell(ws, row_idx, 1, pure_stock_code)

    # è·å–æ¦‚å¿µ
    concept = get_stock_concept(stock)

    # è®¾ç½®æ¦‚å¿µåˆ—ï¼ˆç¬¬äºŒåˆ—ï¼‰
    stock_key = f"{stock_code}_{stock_name}"
    format_concept_cell(ws, row_idx, 2, concept, stock_key, stock_reason_group, reason_colors)

    # è®¡ç®—è‚¡ç¥¨çš„æœ€é«˜æ¿æ•°
    max_board_level = calculate_max_board_level(all_board_data)

    # æ ¹æ®è‚¡ç¥¨ä»£ç ç¡®å®šå¸‚åœºç±»å‹
    market_type = get_market_marker(pure_stock_code)

    # è®¾ç½®è‚¡ç¥¨åç§°åˆ—ï¼ˆç¬¬ä¸‰åˆ—ï¼‰
    end_date_yyyymmdd = max(date_mapping.values()) if date_mapping else None
    name_cell, apply_high_board_color = format_stock_name_cell(
        ws, row_idx, 3, stock_name, market_type, max_board_level, False,
        stock_code, stock_entry_count, end_date_yyyymmdd
    )

    # å¤„ç†å‘¨æœŸæ¶¨è·Œå¹…åˆ—ï¼ˆå¦‚æœå¯ç”¨ï¼‰
    if show_period_change and period_column:
        from analysis.ladder_chart import format_period_change_cell
        entry_date = stock['first_significant_date']
        format_period_change_cell(ws, row_idx, period_column, pure_stock_code, stock_name,
                                  entry_date, period_days, period_days_long, end_date_yyyymmdd)

    # å¡«å……æ—¥æœŸåˆ—æ•°æ®
    last_board_date = None
    for j, formatted_day in enumerate(formatted_trading_days):
        col_idx = date_column_start + j
        current_date_obj = datetime.strptime(formatted_day, '%Yå¹´%mæœˆ%dæ—¥')

        # å¤„ç†æ¯æ—¥å•å…ƒæ ¼ï¼ˆæˆäº¤é‡ç‰ˆæœ¬ï¼‰
        cell_last_board_date = process_volume_daily_cell(
            ws, row_idx, col_idx, stock, current_date_obj, formatted_day,
            stock_details, date_mapping, max_tracking_days, max_tracking_days_before,
            zaban_df, new_high_markers, thresholds
        )

        if cell_last_board_date:
            last_board_date = cell_last_board_date

        # æ·»åŠ æ–°é«˜æ ‡è®°
        if stock_code in new_high_markers and new_high_markers[stock_code] == formatted_day:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.value = f"{cell.value}{NEW_HIGH_MARKER}"

    # å¤„ç†å¼‚åŠ¨é¢„è­¦åˆ—ï¼ˆå¦‚æœå¯ç”¨ï¼‰
    if show_warning_column:
        warning_col = date_column_start + len(formatted_trading_days)
        if show_period_change:
            warning_col += 1

        # åœ¨æˆäº¤é‡ç‰ˆæœ¬ä¸­æ˜¾ç¤ºæˆäº¤é‡è¶‹åŠ¿ä¿¡æ¯
        # æ‰¾åˆ°è¯¥è‚¡ç¥¨æœ€åä¸€ä¸ªæœ‰æ•°æ®çš„äº¤æ˜“æ—¥
        last_data_date = None
        for formatted_day_reverse in reversed(formatted_trading_days):
            date_yyyymmdd_check = date_mapping.get(formatted_day_reverse)
            if date_yyyymmdd_check:
                # æ£€æŸ¥æ˜¯å¦æœ‰è¿æ¿æ•°æ®æˆ–æˆäº¤é‡æ•°æ®
                if (all_board_data.get(formatted_day_reverse) is not None or
                    get_stock_daily_volume_change(pure_stock_code, date_yyyymmdd_check) is not None):
                    last_data_date = date_yyyymmdd_check
                    break

        volume_trend = ""
        if last_data_date:
            volume_trend = get_volume_trend_indicator(pure_stock_code, last_data_date,
                                                    formatted_trading_days, date_mapping)


        warning_cell = ws.cell(row=row_idx, column=warning_col, value=volume_trend)
        warning_cell.border = BORDER_STYLE
        warning_cell.alignment = Alignment(horizontal='center')
        warning_cell.font = Font(size=9)  # è®¾ç½®å°ä¸€å·å­—ä½“

        # æ ¹æ®è¶‹åŠ¿ç±»å‹è®¾ç½®é¢œè‰²
        if volume_trend:
            if "æ”¾é‡" in volume_trend:
                # æ”¾é‡â†— - çº¢è‰²ç³»
                warning_cell.fill = PatternFill(start_color="FFE6E6", fill_type="solid")  # æµ…çº¢è‰²èƒŒæ™¯
                warning_cell.font = Font(color="CC0000", size=9, bold=True)  # çº¢è‰²å­—ä½“
            elif "ç¼©é‡" in volume_trend:
                # ç¼©é‡â†˜ - ç»¿è‰²ç³»
                warning_cell.fill = PatternFill(start_color="E6F3E6", fill_type="solid")  # æµ…ç»¿è‰²èƒŒæ™¯
                warning_cell.font = Font(color="006600", size=9, bold=True)  # ç»¿è‰²å­—ä½“
            elif "é«˜é‡" in volume_trend:
                # é«˜é‡ğŸ”¥ - æ©™è‰²ç³»
                warning_cell.fill = PatternFill(start_color="FFF2E6", fill_type="solid")  # æµ…æ©™è‰²èƒŒæ™¯
                warning_cell.font = Font(color="FF6600", size=9, bold=True)  # æ©™è‰²å­—ä½“
            elif "ä½é‡" in volume_trend:
                # ä½é‡ğŸ’¤ - è“è‰²ç³»
                warning_cell.fill = PatternFill(start_color="E6F0FF", fill_type="solid")  # æµ…è“è‰²èƒŒæ™¯
                warning_cell.font = Font(color="0066CC", size=9, bold=True)  # è“è‰²å­—ä½“
            # æ¸©å’Œå˜åŒ–ï¼ˆé‡å¢â†—/é‡å‡â†˜ï¼‰ä¸ä¸Šè‰²


def create_volume_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                                formatted_trading_days, date_column_start, show_period_change,
                                                period_column, period_days, period_days_long, stock_details,
                                                date_mapping, max_tracking_days, max_tracking_days_before, zaban_df):
    """
    åˆ›å»ºæˆäº¤é‡ç‰ˆæœ¬çš„æŒ‰æ¦‚å¿µåˆ†ç»„å·¥ä½œè¡¨å†…å®¹

    Args:
        ws: Excelå·¥ä½œè¡¨
        result_df: æ˜¾è‘—è¿æ¿è‚¡ç¥¨DataFrame
        å…¶ä»–å‚æ•°ä¸åŸç‰ˆæœ¬ç›¸åŒ
    """
    print(f"å¡«å……æˆäº¤é‡ç‰ˆæœ¬çš„æŒ‰æ¦‚å¿µåˆ†ç»„å·¥ä½œè¡¨å†…å®¹")

    if result_df.empty:
        print("æ²¡æœ‰æ•°æ®å¯æ˜¾ç¤º")
        return

    # è·å–è‚¡ç¥¨æ¦‚å¿µæ•°æ®
    stock_reason_group = stock_data['stock_reason_group']
    reason_colors = stock_data['reason_colors']

    # è®¡ç®—é•¿å‘¨æœŸæ¶¨è·Œå¹…å¹¶é‡æ–°æ’åº
    # ç”±äºcalculate_long_period_change_for_dfå¯èƒ½ä¸å­˜åœ¨ï¼Œæˆ‘ä»¬æ‰‹åŠ¨è®¡ç®—
    concept_grouped_df = result_df.copy()

    # æ‰‹åŠ¨è®¡ç®—é•¿å‘¨æœŸæ¶¨è·Œå¹…
    def calculate_long_period_change(row):
        try:
            from analysis.ladder_chart import calculate_stock_period_change, get_n_trading_days_before
            stock_code = row['stock_code']
            end_date = date_mapping.get(formatted_trading_days[-1])
            if end_date:
                prev_date = get_n_trading_days_before(end_date, period_days_long)
                if '-' in prev_date:
                    prev_date = prev_date.replace('-', '')
                return calculate_stock_period_change(stock_code.split('_')[0] if '_' in stock_code else stock_code,
                                                     prev_date, end_date)
        except Exception as e:
            print(f"è®¡ç®—é•¿å‘¨æœŸæ¶¨è·Œå¹…æ—¶å‡ºé”™: {e}")
        return 0.0

    concept_grouped_df['long_period_change'] = concept_grouped_df.apply(calculate_long_period_change, axis=1)

    # ä¸ºã€é»˜é»˜ä¸Šæ¶¨ã€‘åˆ†ç»„æ·»åŠ ç‰¹æ®Šæ’åºå­—æ®µ
    def get_momo_sort_keys(row):
        if row.get('concept_group') == 'é»˜é»˜ä¸Šæ¶¨':
            momo_data = row.get('momo_data', {})
            # æå–æˆäº¤é¢æ•°å€¼ï¼ˆå»æ‰"äº¿"å­—ç¬¦ï¼‰
            volume_str = momo_data.get('åŒºé—´æˆäº¤é¢', '0')
            try:
                volume = float(volume_str.replace('äº¿', '')) if 'äº¿' in str(volume_str) else 0.0
            except:
                volume = 0.0

            # æå–æ¶¨å¹…æ•°å€¼ï¼ˆå»æ‰"%"å­—ç¬¦ï¼‰
            change_str = momo_data.get('åŒºé—´æ¶¨è·Œå¹…', '0')
            try:
                change = float(change_str.replace('%', '')) if '%' in str(change_str) else 0.0
            except:
                change = 0.0

            return volume, change
        else:
            return 0.0, 0.0

    # æ·»åŠ ã€é»˜é»˜ä¸Šæ¶¨ã€‘æ’åºå­—æ®µ
    concept_grouped_df[['momo_volume', 'momo_change']] = concept_grouped_df.apply(
        lambda row: pd.Series(get_momo_sort_keys(row)), axis=1
    )

    # åˆ†åˆ«å¤„ç†ã€é»˜é»˜ä¸Šæ¶¨ã€‘åˆ†ç»„å’Œå…¶ä»–åˆ†ç»„çš„æ’åº
    momo_mask = concept_grouped_df['concept_group'] == 'é»˜é»˜ä¸Šæ¶¨'
    momo_df = concept_grouped_df[momo_mask].copy()
    other_df = concept_grouped_df[~momo_mask].copy()

    # ã€é»˜é»˜ä¸Šæ¶¨ã€‘åˆ†ç»„ï¼šæŒ‰æ¦‚å¿µä¼˜å…ˆçº§ã€æˆäº¤é¢ã€æ¶¨å¹…å€’åºæ’åˆ—
    if not momo_df.empty:
        momo_df = momo_df.sort_values(
            by=['concept_priority', 'momo_volume', 'momo_change'],
            ascending=[True, False, False]
        )

    # å…¶ä»–åˆ†ç»„ï¼šæŒ‰åŸç‰ˆé€»è¾‘æ’åºï¼ˆæ¦‚å¿µä¼˜å…ˆçº§ã€æ¦‚å¿µç»„ã€é¦–æ¬¡æ˜¾è‘—è¿æ¿æ—¥æœŸã€é•¿å‘¨æœŸæ¶¨è·Œå¹…ã€é¦–æ¬¡è¿æ¿å¤©æ•°ï¼‰
    if not other_df.empty:
        other_df = other_df.sort_values(
            by=['concept_priority', 'concept_group', 'first_significant_date', 'long_period_change',
                'board_level_at_first'],
            ascending=[True, True, True, False, False]
        )

    # åˆå¹¶æ’åºç»“æœ
    concept_grouped_df = pd.concat([other_df, momo_df], ignore_index=True)

    # åˆ é™¤ä¸´æ—¶æ’åºå­—æ®µ
    concept_grouped_df = concept_grouped_df.drop(columns=['momo_volume', 'momo_change'])

    print(f"æŒ‰æ¦‚å¿µåˆ†ç»„æ’åºåçš„è‚¡ç¥¨æ•°é‡: {len(concept_grouped_df)}")

    # è®¾ç½®è¡¨å¤´ï¼Œæ˜¾ç¤ºé‡è¶‹åŠ¿åˆ—ï¼ˆæˆäº¤é‡ç‰ˆæœ¬ä¸“ç”¨ï¼‰
    from analysis.ladder_chart import setup_excel_header
    show_warning_column = True
    date_columns = setup_excel_header(ws, formatted_trading_days, show_period_change, period_days,
                                      date_column_start, show_warning_column)

    # ä¿®æ”¹è¡¨å¤´æ ‡é¢˜ä¸º"é‡è¶‹åŠ¿"ï¼ˆæˆäº¤é‡ç‰ˆæœ¬ä¸“ç”¨ï¼‰
    if show_warning_column and formatted_trading_days:
        warning_col = len(formatted_trading_days) + date_column_start
        if show_period_change:
            warning_col += 1
        warning_cell = ws.cell(row=1, column=warning_col)
        warning_cell.value = "é‡è¶‹åŠ¿"
        warning_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # è®¾ç½®ä¸å…¶ä»–è¡¨å¤´ä¸€è‡´çš„æ ¼å¼
        from analysis.ladder_chart import BORDER_STYLE
        warning_cell.border = BORDER_STYLE
        warning_cell.font = Font(bold=True, size=10)
        # è®¾ç½®é‡è¶‹åŠ¿åˆ—çš„èƒŒæ™¯è‰²ä¸ºæµ…è“è‰²ï¼ŒåŒºåˆ«äºå¼‚åŠ¨é¢„è­¦
        warning_cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")

    # ä¿®æ”¹è¡¨å¤´æ ‡é¢˜ï¼Œæ ‡æ˜è¿™æ˜¯æˆäº¤é‡ç‰ˆæœ¬
    title_cell = ws.cell(row=1, column=1, value="æˆäº¤é‡æ¶¨è·Œå¹…åˆ†æ - æ¶¨åœæ¢¯é˜ŸæŒ‰æ¦‚å¿µåˆ†ç»„")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='left')

    # æ·»åŠ å¤§ç›˜æŒ‡æ ‡è¡Œ
    add_market_indicators(ws, date_columns, label_col=2)

    # è®¡ç®—åŠ¨æ€é˜ˆå€¼
    print("è®¡ç®—æˆäº¤é‡å˜åŒ–åŠ¨æ€é˜ˆå€¼...")
    all_volume_changes = []
    for _, stock in concept_grouped_df.iterrows():
        stock_code = stock['stock_code'].split('_')[0] if '_' in stock['stock_code'] else stock['stock_code']
        for formatted_day in formatted_trading_days:
            date_str = date_mapping.get(formatted_day)
            if date_str:
                volume_change = get_stock_daily_volume_change(stock_code, date_str)
                if volume_change is not None:
                    all_volume_changes.append(volume_change)

    thresholds = calculate_volume_change_thresholds(all_volume_changes)
    if thresholds:
        print(f"åŠ¨æ€é˜ˆå€¼è®¡ç®—å®Œæˆï¼Œå°†å¯¹{VOLUME_COLOR_CONFIG['COLOR_PERCENTAGE'] * 100:.0f}%çš„æ•°æ®ä¸Šè‰²")
        print(f"  æ”¾é‡é˜ˆå€¼: {thresholds}")
    else:
        print("ä½¿ç”¨å›ºå®šé˜ˆå€¼")

    # å¡«å……æ•°æ®è¡Œï¼Œä½¿ç”¨æˆäº¤é‡ç‰ˆæœ¬çš„å¡«å……å‡½æ•°
    fill_volume_data_rows_with_concept_groups(ws, concept_grouped_df, shouban_df, stock_reason_group,
                                              reason_colors, stock_entry_count, formatted_trading_days,
                                              date_column_start, show_period_change, period_column, period_days,
                                              period_days_long, stock_details, date_mapping, max_tracking_days,
                                              max_tracking_days_before, zaban_df, show_warning_column, thresholds)

    # è°ƒæ•´åˆ—å®½
    from analysis.ladder_chart import adjust_column_widths
    adjust_column_widths(ws, formatted_trading_days, date_column_start, show_period_change, show_warning_column)

    # å†»ç»“å‰ä¸‰åˆ—å’Œå‰ä¸‰è¡Œ
    freeze_cell = f"{get_column_letter(date_column_start)}4"
    ws.freeze_panes = freeze_cell


def create_volume_concept_grouped_sheet(wb, sheet_name, result_df, shouban_df, stock_data, stock_entry_count,
                                        formatted_trading_days, date_column_start, show_period_change, period_column,
                                        period_days, period_days_long, stock_details, date_mapping, max_tracking_days,
                                        max_tracking_days_before, zaban_df):
    """
    åˆ›å»ºæˆäº¤é‡ç‰ˆæœ¬çš„æŒ‰æ¦‚å¿µåˆ†ç»„å·¥ä½œè¡¨

    Args:
        wb: Excelå·¥ä½œç°¿
        sheet_name: å·¥ä½œè¡¨åç§°
        å…¶ä»–å‚æ•°ä¸åŸç‰ˆæœ¬ç›¸åŒ
    """
    print(f"åˆ›å»ºæˆäº¤é‡ç‰ˆæœ¬çš„æŒ‰æ¦‚å¿µåˆ†ç»„å·¥ä½œè¡¨: {sheet_name}")

    # åˆ›å»ºæ–°çš„å·¥ä½œè¡¨
    ws = wb.create_sheet(title=sheet_name)

    # å¡«å……å†…å®¹
    create_volume_concept_grouped_sheet_content(ws, result_df, shouban_df, stock_data, stock_entry_count,
                                                formatted_trading_days, date_column_start, show_period_change,
                                                period_column, period_days, period_days_long, stock_details,
                                                date_mapping, max_tracking_days, max_tracking_days_before, zaban_df)
