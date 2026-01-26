"""
fupan_stocks.xlsx 历史数据清理工具

功能：
- 保留最近 N 天的数据
- 删除更早的历史数据
- 删除前自动备份原文件
"""

import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def parse_date_from_column(col_name):
    """
    从列名解析日期，格式为 '2025年01月20日' -> '20250120'
    
    :param col_name: 列名
    :return: 日期字符串 'YYYYMMDD' 或 None
    """
    try:
        dt = datetime.strptime(str(col_name), '%Y年%m月%d日')
        return dt.strftime('%Y%m%d')
    except (ValueError, TypeError):
        return None


def get_date_range_from_sheet(excel_path, sheet_name='首板数据'):
    """
    从指定 sheet 获取数据的起止日期
    
    :param excel_path: Excel 文件路径
    :param sheet_name: sheet 名称
    :return: (earliest_date, latest_date) 元组，格式为 'YYYYMMDD'
    """
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, index_col=0)
        dates = []
        for col in df.columns:
            date_str = parse_date_from_column(col)
            if date_str:
                dates.append(date_str)

        if not dates:
            return None, None

        dates.sort()
        return dates[0], dates[-1]
    except Exception as e:
        print(f"读取 sheet '{sheet_name}' 失败: {e}")
        return None, None


def backup_excel_file(excel_path):
    """
    备份 Excel 文件，备份文件名格式：fupan_stocks_YYYYMMDD-YYYYMMDD.xlsx
    
    :param excel_path: 原文件路径
    :return: 备份文件路径，失败返回 None
    """
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return None

    # 获取起止日期
    earliest, latest = get_date_range_from_sheet(excel_path, '首板数据')
    if not earliest or not latest:
        print("无法获取数据日期范围，使用当前时间作为备份文件名")
        earliest = latest = datetime.now().strftime('%Y%m%d')

    # 构造备份文件名
    dir_name = os.path.dirname(excel_path)
    base_name = os.path.basename(excel_path)
    name_without_ext = os.path.splitext(base_name)[0]
    backup_name = f"{name_without_ext}_{earliest}-{latest}.xlsx"
    backup_path = os.path.join(dir_name, backup_name)

    # 如果备份文件已存在，添加序号
    counter = 1
    original_backup_path = backup_path
    while os.path.exists(backup_path):
        backup_name = f"{name_without_ext}_{earliest}-{latest}_{counter}.xlsx"
        backup_path = os.path.join(dir_name, backup_name)
        counter += 1

    try:
        shutil.copy2(excel_path, backup_path)
        print(f"✅ 已备份到: {backup_path}")
        return backup_path
    except Exception as e:
        print(f"❌ 备份失败: {e}")
        return None


def clean_fupan_excel(excel_path, keep_days=60, dry_run=False):
    """
    清理 fupan_stocks.xlsx 中的历史数据
    
    :param excel_path: Excel 文件路径
    :param keep_days: 保留最近 N 天的数据（以交易日期列为准）
    :param dry_run: 如果为 True，只打印要删除的列，不实际删除
    :return: 清理是否成功
    """
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return False

    print(f"\n{'=' * 60}")
    print(f"清理文件: {excel_path}")
    print(f"保留天数: {keep_days}")
    print(f"模式: {'模拟运行 (dry_run)' if dry_run else '实际执行'}")
    print(f"{'=' * 60}")

    # 获取当前所有日期列
    _, latest_date = get_date_range_from_sheet(excel_path, '首板数据')
    if not latest_date:
        print("无法获取最新日期，跳过清理")
        return False

    print(f"最新数据日期: {latest_date}")

    # 计算保留日期的阈值（简单方式：按列数保留，而非实际交易日）
    # 这里直接读取首板数据 sheet，按列的日期排序后保留最近 keep_days 列

    try:
        # 读取所有 sheet 名称
        xls = pd.ExcelFile(excel_path)
        sheet_names = xls.sheet_names
        print(f"共 {len(sheet_names)} 个 sheet: {sheet_names}")

        # 收集所有 sheet 中的日期列
        all_dates = set()
        for sheet in sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet, index_col=0)
            for col in df.columns:
                date_str = parse_date_from_column(col)
                if date_str:
                    all_dates.add(date_str)

        if not all_dates:
            print("未找到任何日期列，跳过清理")
            return False

        # 排序日期，保留最近 keep_days 天
        sorted_dates = sorted(all_dates, reverse=True)
        dates_to_keep = set(sorted_dates[:keep_days])
        dates_to_remove = set(sorted_dates[keep_days:])

        print(f"\n总日期数: {len(sorted_dates)}")
        print(f"保留日期数: {len(dates_to_keep)}")
        print(f"删除日期数: {len(dates_to_remove)}")

        if not dates_to_remove:
            print("\n没有需要删除的数据，当前数据量未超过保留天数")
            return True

        print(f"\n将删除的日期范围: {min(dates_to_remove)} ~ {max(dates_to_remove)}")

        if dry_run:
            print("\n[模拟运行] 以下列将被删除:")
            for date in sorted(dates_to_remove):
                col_name = datetime.strptime(date, '%Y%m%d').strftime('%Y年%m月%d日')
                print(f"  - {col_name}")
            return True

        # 备份原文件
        print("\n正在备份原文件...")
        backup_path = backup_excel_file(excel_path)
        if not backup_path:
            print("备份失败，终止清理操作")
            return False

        # 开始清理
        print("\n正在清理数据...")

        # 使用 openpyxl 直接操作，避免格式丢失
        wb = load_workbook(excel_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  处理 sheet: {sheet_name}")

            # 找出需要删除的列索引（从后往前删，避免索引变化）
            cols_to_delete = []
            for col_idx in range(1, ws.max_column + 1):
                col_header = ws.cell(row=1, column=col_idx).value
                date_str = parse_date_from_column(col_header)
                if date_str and date_str in dates_to_remove:
                    cols_to_delete.append(col_idx)

            # 从后往前删除列
            for col_idx in sorted(cols_to_delete, reverse=True):
                ws.delete_cols(col_idx)

            print(f"    删除了 {len(cols_to_delete)} 列")

        # 保存
        wb.save(excel_path)
        wb.close()

        # 显示清理后的文件大小
        new_size = os.path.getsize(excel_path) / (1024 * 1024)
        print(f"\n✅ 清理完成！")
        print(f"清理后文件大小: {new_size:.2f} MB")

        return True

    except Exception as e:
        print(f"❌ 清理过程出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def clean_all_fupan_files(keep_days=60, dry_run=False):
    """
    清理所有 fupan_stocks 相关文件
    
    :param keep_days: 保留最近 N 天的数据
    :param dry_run: 如果为 True，只打印要删除的列，不实际删除
    """
    files = [
        "./excel/fupan_stocks.xlsx",
        "./excel/fupan_stocks_non_main.xlsx"
    ]

    for file_path in files:
        if os.path.exists(file_path):
            clean_fupan_excel(file_path, keep_days=keep_days, dry_run=dry_run)
        else:
            print(f"文件不存在，跳过: {file_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='清理 fupan_stocks.xlsx 历史数据')
    parser.add_argument('--keep-days', type=int, default=60, help='保留最近 N 天的数据（默认60天）')
    parser.add_argument('--dry-run', action='store_true', help='模拟运行，不实际删除')
    parser.add_argument('--file', type=str, default=None, help='指定清理的文件路径（默认清理所有）')

    args = parser.parse_args()

    if args.file:
        clean_fupan_excel(args.file, keep_days=args.keep_days, dry_run=args.dry_run)
    else:
        clean_all_fupan_files(keep_days=args.keep_days, dry_run=args.dry_run)
