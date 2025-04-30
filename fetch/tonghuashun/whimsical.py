import os
import sys
import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from collections import Counter
import math

# 添加项目根目录到sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
from utils.date_util import get_trading_days

# 输入和输出文件路径
FUPAN_FILE = "./excel/fupan_stocks.xlsx"
OUTPUT_FILE = "./excel/fupan_analysis.xlsx"

# 颜色列表 - 彩虹色系(深色)
COLORS = [
    "FF5A5A",  # 红色
    "FF8C42",  # 橙色
    "FFCE30",  # 黄色
    "6AD15A",  # 绿色
    "45B5FF",  # 蓝色
    "9966FF",  # 紫色
    "FF66B3",  # 粉色
    "5ACDCD",  # 青色
    "FF8A8A",  # 浅红色
    "FFAA33"   # 金色
]

# 多次上榜但无热门原因的颜色
MULTI_COLOR = "E0E0E0"  # 浅灰色

def normalize_reason(reason):
    """
    将原因标准化，处理同一类型的不同表述
    """
    # 移除所有空格
    reason = re.sub(r'\s+', '', reason)
    
    # 定义常见的同义词转换
    synonym_groups = {
        "机器人": ["机器人", "人形机器人", "服务机器人", "工业机器人"],
        "新能源": ["新能源", "新能源汽车", "新能源车", "电动车", "动力电池", "光伏"],
        "AI": ["AI", "人工智能", "算力", "大模型", "GPT", "AIGC"],
        "数字经济": ["数字经济", "数字化", "数字技术", "数字转型"],
        "半导体": ["半导体", "芯片", "存储芯片", "集成电路"],
        "国企改革": ["国企改革", "国资改革", "国资国企改革", "国企整合"],
        "华为": ["华为", "华为产业链", "鸿蒙", "昇腾"],
        "电子": ["电子", "消费电子", "苹果概念"],
        "券商": ["券商", "证券", "参股券商"],
        "医药": ["医药", "创新药", "疫苗", "生物医药", "医疗器械"],
        "军工": ["军工", "国防军工", "航空航天", "战斗机"],
        "大消费": ["消费", "白酒", "食品", "饮料", "零售", "商超", "免税"],
        "汽车": ["汽车", "整车", "汽配", "车载"],
        "旅游": ["旅游", "酒店", "民航", "免税", "出行"],
        "互联网": ["互联网", "电商", "社交", "游戏"],
        "金融": ["金融", "保险", "银行", "信托", "支付"]
    }
    
    # 保存原始原因，用于后续分析未分类的原因
    original_reason = reason
    
    # 检查原因属于哪个组
    for main_reason, synonyms in synonym_groups.items():
        for synonym in synonyms:
            if synonym in reason:
                return main_reason
    
    # 如果没有匹配到组，返回原始原因，并标记为未分类
    return f"未分类_{original_reason}"

def extract_reasons(reason_text):
    """
    从原因文本中提取所有原因
    """
    if not reason_text or pd.isna(reason_text):
        return []
    
    # 以"+"分割不同原因
    reasons = reason_text.split('+')
    return [normalize_reason(r.strip()) for r in reasons if r.strip()]

def process_zt_data(start_date, end_date):
    """
    处理涨停数据，转换为更易于分析的格式
    
    :param start_date: 开始日期，格式为'YYYYMMDD'
    :param end_date: 结束日期，格式为'YYYYMMDD'
    """
    # 获取交易日列表
    trading_days = get_trading_days(start_date, end_date)
    
    # 读取原始Excel数据
    sheets_to_process = ['连板数据', '首板数据']
    sheet_data = {}
    
    for sheet_name in sheets_to_process:
        try:
            df = pd.read_excel(FUPAN_FILE, sheet_name=sheet_name, index_col=0)
            sheet_data[sheet_name] = df
        except Exception as e:
            print(f"读取{sheet_name}失败: {e}")
            continue
    
    if not sheet_data:
        print("没有找到需要处理的数据")
        return
    
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "涨停分析"
    
    # 收集所有涨停原因和每日涨停股票数据
    all_reasons = []
    daily_stocks = {}
    
    # 找出当前分析时间范围内的交易日对应的列
    date_formatted_list = [datetime.strptime(trade_date, '%Y%m%d').strftime('%Y年%m月%d日') for trade_date in trading_days]
    
    for sheet_name, df in sheet_data.items():
        for date_col in df.columns:
            # 只处理在分析时间范围内的日期
            if date_col in date_formatted_list:
                if date_col not in daily_stocks:
                    daily_stocks[date_col] = []
                    
                column_data = df[date_col].dropna()
                
                for idx, data_str in enumerate(column_data):
                    items = data_str.split('; ')
                    stock_info = {}
                    
                    # 提取需要的数据
                    stock_code = ""
                    stock_name = ""
                    board_level = 0
                    first_time = "23:59:59"
                    
                    for item in items:
                        if '.S' in item and len(item.split('.')[0]) == 6:
                            # 这是股票代码
                            stock_code = item.strip()
                        elif not stock_name and len(items) > 1 and not any(x in item for x in ['.', ':', '%']):
                            # 这可能是股票简称
                            stock_name = item.strip()
                        elif '天' in item and '板' in item:
                            # 连板股票，如 "2天2板"
                            stock_info['几天几板'] = item.strip()
                            try:
                                # 提取板数
                                board_text = item.strip()
                                board_level = int(board_text.split('天')[1].split('板')[0].strip())
                            except Exception:
                                pass
                        elif '首板涨停' in item:
                            # 首板股票
                            stock_info['几天几板'] = item.strip()
                            board_level = 1
                        elif ':' in item and '涨停时间' not in item and '开板时间' not in item:
                            # 可能是首次涨停时间
                            stock_info['首次涨停时间'] = item.strip()
                            first_time = item.strip()
                        elif item.isdigit() or (item.replace('.', '', 1).isdigit() and item.count('.') <= 1):
                            # 可能是开板次数或价格
                            if '涨停开板次数' not in stock_info:
                                stock_info['涨停开板次数'] = item.strip()
                        elif '+' in item and len(item.split('+')) > 1:
                            # 可能是涨停原因
                            stock_info['涨停原因类别'] = item.strip()
                            reasons = extract_reasons(item.strip())
                            all_reasons.extend(reasons)
                    
                    if not stock_code and len(items) >= 2:
                        # 尝试直接使用前两项作为代码和名称
                        stock_code = items[0].strip()
                        stock_name = items[1].strip()
                    
                    if stock_code and stock_name:
                        # 确保板数正确
                        if board_level == 0 and '首板涨停' in str(stock_info.get('几天几板', '')):
                            board_level = 1
                        
                        # 添加到每日数据
                        daily_stocks[date_col].append({
                            'code': stock_code,
                            'name': stock_name,
                            'info': stock_info,
                            'board_level': board_level,
                            'first_time': first_time,
                            'sheet_name': sheet_name  # 记录来源sheet，用于区分连板和首板
                        })
    
    # 统计所有原因并找出未分类的原因
    reason_counter = Counter(all_reasons)
    unclassified_reasons = [reason for reason in reason_counter.keys() if reason.startswith('未分类_')]
    
    # 创建未分类原因的sheet
    if unclassified_reasons:
        unclassified_ws = wb.create_sheet(title="未分类原因")
        
        # 设置表头
        unclassified_ws.cell(row=1, column=1, value="原因").font = Font(bold=True)
        unclassified_ws.cell(row=1, column=2, value="出现次数").font = Font(bold=True)
        
        # 按出现次数排序
        sorted_reasons = sorted(unclassified_reasons, key=lambda x: reason_counter[x], reverse=True)
        
        # 计算每列应该显示的行数
        total_reasons = len(sorted_reasons)
        columns_count = 7  # 每页显示7列
        rows_per_column = math.ceil(total_reasons / columns_count)
        
        # 写入未分类原因和次数
        for i, reason in enumerate(sorted_reasons):
            # 计算当前行应该在哪一列
            col_idx = (i // rows_per_column) * 2 + 1
            row_idx = i % rows_per_column + 2
            
            # 写入原因和次数
            original_reason = reason.replace('未分类_', '')
            unclassified_ws.cell(row=row_idx, column=col_idx, value=original_reason)
            unclassified_ws.cell(row=row_idx, column=col_idx+1, value=reason_counter[reason])
        
        # 设置列宽
        for i in range(1, columns_count*2+1):
            col_letter = get_column_letter(i)
            if i % 2 == 1:  # 原因列
                unclassified_ws.column_dimensions[col_letter].width = 25
            else:  # 次数列
                unclassified_ws.column_dimensions[col_letter].width = 10
        
        print(f"未分类的涨停原因已保存到工作簿的 '未分类原因' 页")
    
    # 过滤掉未分类的原因，获取热门原因
    classified_reasons = [reason for reason in reason_counter.keys() if not reason.startswith('未分类_')]
    top_reasons = [reason for reason, count in Counter(classified_reasons).most_common(8) if count > 0]
    
    # 如果没有足够的热门原因，使用默认分类
    if len(top_reasons) < 5:
        default_reasons = ["新能源", "AI", "医药", "半导体", "军工", "大消费", "汽车", "旅游"]
        for reason in default_reasons:
            if reason not in top_reasons:
                top_reasons.append(reason)
            if len(top_reasons) >= 8:
                break
    
    # 为每个原因分配颜色
    reason_colors = {reason: COLORS[i % len(COLORS)] for i, reason in enumerate(top_reasons)}
    
    # 计算每个股票在分析时间范围内的出现次数和主要原因
    all_stocks = {}
    for date, stocks in daily_stocks.items():
        for stock in stocks:
            stock_key = f"{stock['code']}_{stock['name']}"
            
            if stock_key not in all_stocks:
                all_stocks[stock_key] = {
                    'name': stock['name'],
                    'appearances': [],
                    'reasons': []
                }
            
            all_stocks[stock_key]['appearances'].append(date)
            
            if '涨停原因类别' in stock['info']:
                # 只使用已分类的原因进行分组
                reasons = [r for r in extract_reasons(stock['info']['涨停原因类别']) 
                           if not r.startswith('未分类_')]
                all_stocks[stock_key]['reasons'].extend(reasons)
    
    # 确定每支股票主要属于哪个原因组
    stock_reason_group = {}
    for stock_key, data in all_stocks.items():
        if not data['reasons']:
            continue
            
        # 统计该股票的原因
        stock_reason_counter = Counter(data['reasons'])
        
        # 检查是否有热门原因
        found_top_reason = False
        for top_reason in top_reasons:
            if top_reason in stock_reason_counter:
                stock_reason_group[stock_key] = top_reason
                found_top_reason = True
                break
                
        # 如果没有热门原因，使用该股票最常见的原因
        if not found_top_reason and stock_reason_counter:
            most_common_reason = stock_reason_counter.most_common(1)[0][0]
            stock_reason_group[stock_key] = most_common_reason
    
    # 创建图例作为第一列
    ws.column_dimensions['A'].width = 15
    ws.cell(row=1, column=1, value="热门概念图例").font = Font(bold=True)
    
    # 添加各个原因图例
    for i, reason in enumerate(top_reasons, start=2):
        cell = ws.cell(row=i, column=1, value=reason)
        cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")
    
    # 添加多次上榜图例
    multi_row = len(top_reasons) + 2
    cell = ws.cell(row=multi_row, column=1, value="多次上榜")
    cell.fill = PatternFill(start_color=MULTI_COLOR, fill_type="solid")
    
    # 添加首板/连板分割线说明
    separator_row = multi_row + 1
    ws.cell(row=separator_row, column=1, value="分隔线 = 首板")
    separator_cell = ws.cell(row=separator_row+1, column=1)
    separator_cell.border = Border(bottom=Side(style='double', color='000000'))
    
    # 设置日期表头(从第二列开始)
    date_columns = {}
    for idx, trade_date in enumerate(trading_days, start=2):
        date_formatted = datetime.strptime(trade_date, '%Y%m%d').strftime('%Y年%m月%d日')
        cell = ws.cell(row=1, column=idx, value=date_formatted)
        
        # 设置表头样式
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 设置列宽
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = 15
        
        date_columns[date_formatted] = idx
    
    # 对每日数据进行排序并写入工作表
    for date, col_idx in date_columns.items():
        if date in daily_stocks:
            # 按照板块数倒序，首次涨停时间正序排序
            sorted_stocks = sorted(daily_stocks[date], 
                                  key=lambda x: (-x['board_level'], x['first_time']))
            
            # 先找出每列的连板和首板分界行号
            first_connection_idx = None
            
            # 写入排序后的数据
            for row_idx, stock in enumerate(sorted_stocks, start=2):
                stock_key = f"{stock['code']}_{stock['name']}"
                
                # 记录首个首板的位置
                if first_connection_idx is None and stock['board_level'] == 1 and stock['sheet_name'] == '首板数据':
                    first_connection_idx = row_idx
                
                # 计算该股票在当前分析区间内的上榜次数
                appearances_count = len(all_stocks[stock_key]['appearances'])
                
                # 在股票名称后添加上榜次数
                display_name = stock['name']
                if appearances_count > 1:
                    display_name = f"{stock['name']}{appearances_count}"
                
                cell = ws.cell(row=row_idx, column=col_idx, value=display_name)
                
                # 设置单元格样式
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 创建备注
                comment_text = ""
                if '几天几板' in stock['info']:
                    comment_text += f"几天几板: {stock['info']['几天几板']}\n"
                if '首次涨停时间' in stock['info']:
                    comment_text += f"首次涨停时间: {stock['info']['首次涨停时间']}\n"
                if '涨停开板次数' in stock['info']:
                    comment_text += f"涨停开板次数: {stock['info']['涨停开板次数']}\n"
                if '涨停原因类别' in stock['info']:
                    comment_text += f"涨停原因类别: {stock['info']['涨停原因类别']}"
                
                cell.comment = Comment(comment_text, "分析系统")
                
                # 根据股票所属原因组设置颜色
                if stock_key in stock_reason_group:
                    reason = stock_reason_group[stock_key]
                    if reason in reason_colors:
                        cell.fill = PatternFill(start_color=reason_colors[reason], fill_type="solid")
                
                # 如果股票在多个日期出现，但没有热门原因，使用浅灰色
                if appearances_count > 1:
                    if stock_key not in stock_reason_group or stock_reason_group[stock_key] not in reason_colors:
                        cell.fill = PatternFill(start_color=MULTI_COLOR, fill_type="solid")
            
            # 添加首板和连板的分隔线
            if first_connection_idx is not None:
                # 第一个首板的前一行添加下边框
                separator_cell = ws.cell(row=first_connection_idx-1, column=col_idx)
                separator_cell.border = Border(
                    left=separator_cell.border.left if separator_cell.border else None,
                    right=separator_cell.border.right if separator_cell.border else None,
                    top=separator_cell.border.top if separator_cell.border else None,
                    bottom=Side(style='double', color='000000')
                )
    
    # 保存工作簿
    wb.save(OUTPUT_FILE)
    print(f"数据处理完成，已保存到 {OUTPUT_FILE}")

if __name__ == "__main__":
    # 设置当前工作目录为脚本所在目录
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    # 示例用法
    start_date = "20240101"
    end_date = "20240601"
    process_zt_data(start_date, end_date)
