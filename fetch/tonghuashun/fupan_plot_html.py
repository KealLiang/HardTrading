"""
HTML交互式复盘图生成器 - 使用Plotly

优势：
1. 鼠标悬停显示详细信息，完全解决标签重叠问题
2. 支持缩放、平移、保存图片
3. 可添加更多交互功能
4. 生成单个HTML文件，方便分享
"""

import os
import re
import sys
from datetime import datetime
from functools import lru_cache

import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# 导入工具函数
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from utils.stock_util import stock_limit_ratio
from utils.following_profit_calculator import calculate_daily_profit_with_stock_details_from_momo_results

# 配置悬浮窗换行阈值
LIANBAN_STOCKS_PER_LINE = 5  # 连板天梯图层：每5只股票换行
MOMO_STOCKS_PER_LINE = 3  # 默默上涨图层：每3只股票换行
ZHANGTING_OPEN_THRESHOLD = 10  # 涨停开板次数阈值（超过此值加下划线标记）
JI_BAN_TIERS = 2  # 次高几板显示阶数（2表示显示第2高和第3高）
ATTENTION_TOP_N = 20  # 关注度榜取前N名（用于加粗股票名称）

# 高门槛股代码前缀配置（用于区分不同交易门槛的股票）
HIGH_THRESHOLD_STOCK_PREFIXES = ['8', '688']  # 以这些前缀开头的股票为高门槛股


def format_stock_name_with_indicators(stock_code: str, stock_name: str,
                                      zhangting_open_times: str = None,
                                      first_zhangting_time: str = None,
                                      final_zhangting_time: str = None) -> str:
    """
    格式化股票名称，添加涨跌幅标识和一字板标识
    
    标识说明：
    - | = 一字板涨停
    - * = 20%涨跌幅限制
    - ** = 30%涨跌幅限制
    - 下划线 = 涨停开板次数超过阈值
    """
    try:
        clean_code = stock_code.split('.')[0] if '.' in stock_code else stock_code
        limit_ratio = stock_limit_ratio(clean_code)
        formatted_name = stock_name

        # 判断是否为一字板
        is_yizi_ban = is_yizi_board_zhangting(zhangting_open_times, first_zhangting_time, final_zhangting_time)
        if is_yizi_ban:
            formatted_name = f"{formatted_name}|"

        # 根据涨跌幅比例添加星号
        if limit_ratio == 0.2:
            formatted_name = f"{formatted_name}*"
        elif limit_ratio == 0.3:
            formatted_name = f"{formatted_name}**"

        # 判断是否涨停开板次数超过阈值（加下划线）
        if zhangting_open_times is not None and str(zhangting_open_times).strip() != '':
            try:
                open_times = int(str(zhangting_open_times).strip())
                if open_times > ZHANGTING_OPEN_THRESHOLD:
                    formatted_name = f"<u>{formatted_name}</u>"
            except:
                pass

        return formatted_name
    except:
        return stock_name


def is_yizi_board_zhangting(zhangting_open_times: str, first_zhangting_time: str, final_zhangting_time: str) -> bool:
    """判断是否为一字板涨停"""
    try:
        if zhangting_open_times is None or str(zhangting_open_times).strip() == '':
            return False
        open_times = int(str(zhangting_open_times).strip())
        if open_times != 0:
            return False

        if (first_zhangting_time is None or final_zhangting_time is None or
                str(first_zhangting_time).strip() == '' or str(final_zhangting_time).strip() == ''):
            return False

        first_time = str(first_zhangting_time).strip()
        final_time = str(final_zhangting_time).strip()

        if first_time != final_time:
            return False

        if not is_market_open_time(first_time):
            return False

        return True
    except:
        return False


def is_market_open_time(time_str: str) -> bool:
    """判断是否为开盘时间"""
    try:
        time_str = time_str.strip()
        if time_str == "09:30:00" or time_str == "09:25:00":
            return True
        if time_str.startswith("09:30") or time_str.startswith("09:25"):
            return True
        return False
    except:
        return False


def format_stock_name_with_lianban_count(stock_code: str, stock_name: str,
                                         lianban_days: int,
                                         zhangting_open_times: str = None,
                                         first_zhangting_time: str = None,
                                         final_zhangting_time: str = None) -> str:
    """
    格式化股票名称，添加涨跌幅标识、一字板标识和连板数
    
    标识说明：
    - | = 一字板涨停
    - * = 20%涨跌幅限制
    - ** = 30%涨跌幅限制
    - 末尾数字 = 连续涨停天数
    """
    try:
        # 先使用原有函数格式化名称
        base_name = format_stock_name_with_indicators(
            stock_code, stock_name, zhangting_open_times, first_zhangting_time, final_zhangting_time
        )
        # 在末尾加上连板数
        return f"{base_name}{lianban_days}"
    except:
        return f"{stock_name}{lianban_days}"


@lru_cache(maxsize=1)
def load_attention_stocks_by_date(start_date_yyyymmdd: str, end_date_yyyymmdd: str, top_n: int = ATTENTION_TOP_N):
    """
    加载整个分析周期内每个日期的关注度榜前N名股票
    
    Args:
        start_date_yyyymmdd: 开始日期 (YYYYMMDD格式)
        end_date_yyyymmdd: 结束日期 (YYYYMMDD格式)
        top_n: 取前N名，默认为ATTENTION_TOP_N
        
    Returns:
        dict: {日期字符串(YYYY年MM月DD日): [股票代码列表]}
    """
    try:
        from openpyxl import load_workbook

        # 使用相对路径找到 fupan_stocks.xlsx
        current_dir = os.path.dirname(os.path.abspath(__file__))
        fupan_file = os.path.join(current_dir, '..', '..', 'excel', 'fupan_stocks.xlsx')

        if not os.path.exists(fupan_file):
            print(f"关注度榜数据文件不存在: {fupan_file}")
            return {}

        wb = load_workbook(fupan_file, data_only=True)

        # 将日期范围转换为日期对象
        start_date_obj = datetime.strptime(start_date_yyyymmdd, '%Y%m%d')
        end_date_obj = datetime.strptime(end_date_yyyymmdd, '%Y%m%d')

        attention_by_date = {}

        # 处理两个sheet：【关注度榜】和【非主关注度榜】
        for sheet_name in ['关注度榜', '非主关注度榜']:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # 遍历所有列，查找分析周期内的数据
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                if not header_cell.value:
                    continue

                # 解析日期（格式：2025年11月18日）
                col_date_str = str(header_cell.value).strip()
                try:
                    col_date_obj = datetime.strptime(col_date_str, '%Y年%m月%d日')
                except:
                    continue

                # 检查是否在分析周期内
                if not (start_date_obj <= col_date_obj <= end_date_obj):
                    continue

                # 初始化该日期的股票列表
                if col_date_str not in attention_by_date:
                    attention_by_date[col_date_str] = set()

                # 读取该列的前top_n行数据（从第2行开始）
                for row_idx in range(2, min(2 + top_n, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if not cell_value:
                        continue

                    # 解析数据：600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1
                    stock_code = extract_stock_code_from_attention_data(cell_value)
                    if stock_code:
                        attention_by_date[col_date_str].add(stock_code)

        # 将 set 转换为 list
        attention_by_date = {date: list(codes) for date, codes in attention_by_date.items()}

        print(f"✓ 加载关注度榜数据：分析周期内共{len(attention_by_date)}个交易日有关注度榜数据")
        return attention_by_date

    except Exception as e:
        print(f"✗ 加载关注度榜数据失败: {e}")
        import traceback
        traceback.print_exc()
        return {}


def extract_stock_code_from_attention_data(cell_value: str) -> str:
    """
    从关注度榜数据中提取股票代码
    
    输入: "600340.SH; 华夏幸福; 3.31; 10.0%; 998637.5; 1"
    输出: "600340"（标准化后的纯代码）
    
    Args:
        cell_value: 单元格值
        
    Returns:
        str: 标准化后的股票代码，解析失败返回None
    """
    try:
        parts = str(cell_value).split(';')
        if len(parts) >= 1:
            stock_code = parts[0].strip()  # "600340.SH"
            # 去除市场后缀 .SH/.SZ
            if '.' in stock_code:
                stock_code = stock_code.split('.')[0]
            return stock_code
    except:
        pass
    return None


def apply_bold_for_attention_stocks(stock_name: str, stock_code: str, date_str: str, attention_by_date: dict) -> str:
    """
    如果股票在指定日期入选关注度榜前N名，则对股票名称应用加粗
    
    Args:
        stock_name: 股票名称
        stock_code: 股票代码（可能包含市场后缀或下划线）
        date_str: 日期字符串（格式：YYYY年MM月DD日）
        attention_by_date: 关注度榜数据 {日期: [股票代码列表]}
        
    Returns:
        str: 加粗后的股票名称（如果入选），或原始名称
    """
    if not attention_by_date:
        return stock_name

    # 提取纯股票代码（去除市场后缀和下划线）
    clean_code = stock_code.split('.')[0] if '.' in stock_code else stock_code
    clean_code = clean_code.split('_')[0] if '_' in clean_code else clean_code

    # 去掉可能的市场前缀（sh/sz/bj）
    if clean_code.startswith(('sh', 'sz', 'bj')):
        clean_code = clean_code[2:]

    # 检查该股票在该日期是否入选关注度榜
    attention_codes = attention_by_date.get(date_str, [])
    if clean_code in attention_codes:
        return f"<b>{stock_name}</b>"

    return stock_name


# ========== 工具函数：避免重复代码 ==========

def _inject_click_copy_script(html_path, copyable_trace_indices):
    """
    向HTML文件注入JavaScript代码，实现点击节点复制股票代码功能（通用版）
    
    Args:
        html_path: HTML文件路径
        copyable_trace_indices: 支持复制的trace索引列表
    """
    # 读取HTML文件
    with open(html_path, 'r', encoding='utf-8') as f:
        html_content = f.read()

    # 准备JavaScript代码
    js_code = f"""
<script>
// 点击复制股票代码功能（通用版）
(function() {{
    const copyableTraceIndices = {copyable_trace_indices};  // 支持复制的trace索引列表
    const plotDiv = document.querySelector('.plotly-graph-div');
    
    if (!plotDiv) {{
        console.error('未找到Plotly图表元素');
        return;
    }}
    
    // 监听点击事件
    plotDiv.on('plotly_click', function(data) {{
        // 检查是否点击了支持复制的trace
        if (data.points && data.points.length > 0) {{
            const point = data.points[0];
            
            // 判断是否点击的是支持复制的图层
            // 注意：默默上涨和盈亏折线可能重叠，如果点击的是盈亏折线（通常是默默上涨索引+1），
            // 需要从默默上涨的trace中获取数据
            let targetTraceIndex = point.curveNumber;
            let stockCodes = null;
            
            // 首先检查点击的trace是否在可复制列表中
            if (copyableTraceIndices.includes(targetTraceIndex)) {{
                // 正常处理：从当前trace的customdata获取数据
                if (point.customdata) {{
                    if (Array.isArray(point.customdata)) {{
                        // 如果是数组，直接取第二个元素
                        if (point.customdata.length > 1) {{
                            stockCodes = point.customdata[1];
                        }}
                        // 如果是嵌套数组（二维数组），取第二个子数组
                        else if (point.customdata.length > 0 && Array.isArray(point.customdata[0])) {{
                            if (point.customdata[0].length > 1) {{
                                stockCodes = point.customdata[0][1];
                            }}
                        }}
                    }} else if (typeof point.customdata === 'object') {{
                        // 如果是对象，尝试访问第二个属性
                        const keys = Object.keys(point.customdata);
                        if (keys.length > 1) {{
                            stockCodes = point.customdata[keys[1]];
                        }}
                    }}
                }}
            }} else {{
                // 如果不在可复制列表中，检查是否是盈亏折线（默默上涨索引+1）
                // 尝试从默默上涨trace获取数据
                const momoTraceIndex = targetTraceIndex - 1;
                if (copyableTraceIndices.includes(momoTraceIndex)) {{
                    // 从Plotly的图形数据中获取默默上涨trace的customdata
                    const gd = plotDiv;
                    if (gd && gd.data && gd.data[momoTraceIndex]) {{
                        const momoTrace = gd.data[momoTraceIndex];
                        if (momoTrace.customdata) {{
                            // 根据点击的x坐标找到对应的数据点
                            const xValues = momoTrace.x;
                            let xIndex = -1;
                            for (let i = 0; i < xValues.length; i++) {{
                                if (xValues[i] === point.x) {{
                                    xIndex = i;
                                    break;
                                }}
                            }}
                            if (xIndex >= 0 && momoTrace.customdata[xIndex]) {{
                                const customdataPoint = momoTrace.customdata[xIndex];
                                if (Array.isArray(customdataPoint) && customdataPoint.length > 1) {{
                                    stockCodes = customdataPoint[1];
                                }}
                            }}
                        }}
                    }}
                }}
            }}
            
            // 处理股票代码字符串
            if (stockCodes !== null && stockCodes !== undefined) {{
                const codesStr = String(stockCodes).trim();
                if (codesStr !== '') {{
                    // 复制到剪贴板
                    copyToClipboard(codesStr);
                    
                    // 显示提示信息
                    const codeCount = codesStr.split('\\n').filter(line => line.trim() !== '').length;
                    showCopyNotification(point.x, '已复制 ' + codeCount + ' 只股票代码！');
                }} else {{
                    showCopyNotification(point.x, '该日期无股票代码');
                }}
            }}
        }}
    }});
    
    // 复制到剪贴板函数
    function copyToClipboard(text) {{
        if (navigator.clipboard && navigator.clipboard.writeText) {{
            // 现代浏览器
            navigator.clipboard.writeText(text).then(function() {{
                console.log('复制成功');
            }}).catch(function(err) {{
                console.error('复制失败:', err);
                fallbackCopy(text);
            }});
        }} else {{
            // 兼容旧浏览器
            fallbackCopy(text);
        }}
    }}
    
    // 兼容旧浏览器的复制方法
    function fallbackCopy(text) {{
        const textarea = document.createElement('textarea');
        textarea.value = text;
        textarea.style.position = 'fixed';
        textarea.style.opacity = '0';
        document.body.appendChild(textarea);
        textarea.select();
        try {{
            document.execCommand('copy');
            console.log('复制成功（兼容模式）');
        }} catch (err) {{
            console.error('复制失败（兼容模式）:', err);
        }}
        document.body.removeChild(textarea);
    }}
    
    // 显示复制成功提示
    function showCopyNotification(date, message) {{
        // 创建提示元素
        const notification = document.createElement('div');
        notification.textContent = message;
        notification.style.position = 'fixed';
        notification.style.top = '20px';
        notification.style.left = '50%';
        notification.style.transform = 'translateX(-50%)';
        notification.style.backgroundColor = 'rgba(76, 175, 80, 0.95)';
        notification.style.color = 'white';
        notification.style.padding = '12px 24px';
        notification.style.borderRadius = '6px';
        notification.style.fontSize = '14px';
        notification.style.fontWeight = 'bold';
        notification.style.zIndex = '10000';
        notification.style.boxShadow = '0 4px 12px rgba(0,0,0,0.3)';
        notification.style.fontFamily = 'SimHei, Arial, sans-serif';
        
        document.body.appendChild(notification);
        
        // 2秒后自动消失
        setTimeout(function() {{
            notification.style.transition = 'opacity 0.5s';
            notification.style.opacity = '0';
            setTimeout(function() {{
                document.body.removeChild(notification);
            }}, 500);
        }}, 2000);
    }}
}})();
</script>
"""

    # 在</body>之前插入JavaScript代码
    html_content = html_content.replace('</body>', js_code + '\n</body>')

    # 写回HTML文件
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print("✅ 已注入点击复制功能")


def format_stock_list_for_hover(stock_list, stocks_per_line=5):
    """
    格式化股票列表用于悬浮窗显示（每N只换一行）
    
    Args:
        stock_list: 股票列表
        stocks_per_line: 每行显示的股票数，默认5
        
    Returns:
        格式化后的字符串，用<br>分隔
    """
    if len(stock_list) > stocks_per_line:
        stock_lines = [', '.join(stock_list[i:i + stocks_per_line]) for i in range(0, len(stock_list), stocks_per_line)]
        return '<br>'.join(stock_lines)
    else:
        return ', '.join(stock_list)


def remove_bold_tags(text: str) -> str:
    """
    去除文本中的HTML加粗标签
    
    Args:
        text: 可能包含<b>和</b>标签的文本
        
    Returns:
        去除加粗标签后的文本
    """
    return text.replace('<b>', '').replace('</b>', '')


def create_display_labels(stock_list, max_display=3):
    """
    创建图表上显示的标签（超过max_display个时添加省略号）
    去除加粗标签，因为图表标签不支持HTML
    
    Args:
        stock_list: 股票列表（可能包含<b>标签）
        max_display: 最大显示数量，默认3
        
    Returns:
        格式化后的标签文本（不含加粗标签）
    """
    # 去除加粗标签
    clean_list = [remove_bold_tags(stock) for stock in stock_list]

    if len(clean_list) > max_display:
        return '<br>'.join(clean_list[:max_display]) + '<br>……'
    else:
        return '<br>'.join(clean_list) if clean_list else ''


def extract_stock_codes_from_df(df, code_column='股票代码'):
    """
    从DataFrame中提取股票代码列表（去掉交易所后缀）
    
    Args:
        df: 包含股票代码的DataFrame
        code_column: 股票代码列名，默认'股票代码'
        
    Returns:
        股票代码列表（list）
    """
    if df.empty:
        return []

    codes = []
    for code in df[code_column]:
        clean_code = str(code).split('.')[0] if '.' in str(code) else str(code)
        codes.append(clean_code)

    return codes


def _generate_profit_explanation(buy_days_before: int) -> str:
    """
    生成盈亏说明文字（根据buy_days_before动态生成）
    
    Args:
        buy_days_before: 选股日相对于当前日的前N个交易日
        
    Returns:
        str: 说明文字
    """
    # 当前逻辑：在日期t，计算的是t-(N+1)日选出的股票，t-N日买入，t日卖出的盈亏
    # 对于buy_days_before=1：t-2日选出，t-1日买入，t日卖出
    select_offset = buy_days_before + 1
    buy_offset = buy_days_before
    sell_offset = 0  # 卖出日是当前日期t

    select_day = f"t-{select_offset}" if select_offset > 0 else "t"
    buy_day = f"t-{buy_offset}" if buy_offset > 0 else "t"
    sell_day = "t"

    return f"{select_day}日选出，{buy_day}日开盘买入，{sell_day}日收盘卖出"


def _generate_stock_info_explanation(buy_days_before: int) -> str:
    """
    生成股票信息说明文字（说明括号内数据的含义）
    
    Args:
        buy_days_before: 选股日相对于当前日的前N个交易日
        
    Returns:
        str: 说明文字
    """
    # 对于buy_days_before=1，显示"未来2日的盈亏"（选出的第二天买入，第三天卖出）
    days_after = buy_days_before + 1
    profit_label = f"未来{days_after}日的盈亏" if buy_days_before == 1 else f"未来{days_after}日的盈亏"
    return f"说明：括号内为(区间涨幅, 区间成交额, {profit_label}%)"


def _add_profit_to_stock_info(stock_info: str, profit: float) -> str:
    """
    在股票信息字符串的括号中添加盈亏信息
    
    Args:
        stock_info: 股票信息字符串，格式为"股票名称(涨幅, 成交额)"
        profit: 盈亏百分比
        
    Returns:
        str: 添加盈亏后的股票信息，格式为"股票名称(涨幅, 成交额, 盈亏%)"
    """
    if not stock_info.endswith(')'):
        return stock_info

    last_open = stock_info.rfind('(')
    if last_open <= 0:
        return stock_info

    base_info = stock_info[:last_open]
    inside_paren = stock_info[last_open + 1:-1]
    return f"{base_info}({inside_paren}, {profit:.2f}%)"


def _is_high_threshold_stock(stock_code: str) -> bool:
    """
    判断股票是否属于高门槛股
    
    Args:
        stock_code: 股票代码（可能包含市场后缀，如 "000001.SZ" 或 "688001.SH"）
        
    Returns:
        bool: True表示高门槛股，False表示正常股
    """
    # 清理股票代码（去除市场后缀）
    clean_code = stock_code.split('.')[0] if '.' in stock_code else stock_code
    clean_code = clean_code.split('_')[0] if '_' in clean_code else clean_code

    # 检查是否以配置的前缀开头
    for prefix in HIGH_THRESHOLD_STOCK_PREFIXES:
        if clean_code.startswith(prefix):
            return True
    return False


def read_and_plot_html(fupan_file, start_date=None, end_date=None, output_path=None, buy_days_before=1):
    """
    读取数据并生成HTML交互式图表
    
    Args:
        fupan_file: Excel文件路径
        start_date: 开始日期（格式: YYYYMMDD）
        end_date: 结束日期（格式: YYYYMMDD）
        output_path: 输出HTML文件路径（可选）
        buy_days_before: 选股日相对于当前日的前N个交易日，默认为1（表示t-1日选出）
            - 1: t-1日选出，t-1+1=t日开盘买入，t-1+2=t+1日收盘卖出（隔日盈亏）
            - 2: t-2日选出，t-2+1=t-1日开盘买入，t-2+2=t日收盘卖出
            - 3: t-3日选出，t-3+1=t-2日开盘买入，t-3+2=t-1日收盘卖出
    
    Returns:
        str: 生成的HTML文件路径
    """
    # 读取Excel数据
    lianban_data = pd.read_excel(fupan_file, sheet_name="连板数据", index_col=0)
    dieting_data = pd.read_excel(fupan_file, sheet_name="跌停数据", index_col=0)
    shouban_data = pd.read_excel(fupan_file, sheet_name="首板数据", index_col=0)

    # 读取默默上涨数据（可能不存在）
    try:
        momo_data = pd.read_excel(fupan_file, sheet_name="默默上涨", index_col=0)
        has_momo_data = True
    except:
        momo_data = None
        has_momo_data = False
        print("未找到【默默上涨】数据sheet，将跳过该数据")

    # 提取日期列
    dates = lianban_data.columns

    # 筛选时间范围
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y%m%d")
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y%m%d")

    # 先筛选日期范围
    filtered_dates = []
    for date in dates:
        date_obj = datetime.strptime(date, "%Y年%m月%d日")
        if (not start_date_obj or date_obj >= start_date_obj) and (not end_date_obj or date_obj <= end_date_obj):
            filtered_dates.append(date)

    dates = filtered_dates

    # 加载关注度榜数据（用于股票名称加粗）
    attention_by_date = {}
    if start_date and dates:
        # 如果没有提供 end_date，使用筛选后的最后一个交易日
        actual_end_date = end_date
        if not actual_end_date and dates:
            # 从最后一个日期字符串（YYYY年MM月DD日）转换为YYYYMMDD格式
            last_date_obj = datetime.strptime(dates[-1], "%Y年%m月%d日")
            actual_end_date = last_date_obj.strftime("%Y%m%d")

        if actual_end_date:
            try:
                attention_by_date = load_attention_stocks_by_date(start_date, actual_end_date)
            except Exception as e:
                print(f"⚠ 加载关注度榜数据失败: {e}")
                attention_by_date = {}

    # 初始化结果存储
    lianban_results = []
    lianban_second_results = []
    dieting_results = []
    shouban_counts = []
    max_ji_ban_results = []
    second_ji_ban_results = []  # 存储次高几板数据
    momo_results = []  # 默默上涨数据
    all_codes_by_date = {}  # 存储每个日期的所有股票代码（用于点击复制）
    lianban_4plus_results = []  # 存储4连板及以上股票

    # 逐列提取数据
    for date in dates:
        # 初始化当日代码列表
        date_codes = []
        # 连板数据处理
        lianban_col = lianban_data[date].dropna()
        lianban_stocks = lianban_col.str.split(';').apply(lambda x: [item.strip() for item in x])
        lianban_df = pd.DataFrame(lianban_stocks.tolist(), columns=[
            '股票代码', '股票简称', '涨停开板次数', '最终涨停时间',
            '几天几板', '最新价', '首次涨停时间', '最新涨跌幅',
            '连续涨停天数', '涨停原因类别'
        ])

        # 清理数据
        lianban_df['连续涨停天数'] = lianban_df['连续涨停天数'].fillna(0)
        lianban_df['连续涨停天数'] = lianban_df['连续涨停天数'].replace('', 0)
        lianban_df['连续涨停天数'] = pd.to_numeric(lianban_df['连续涨停天数'], errors='coerce').fillna(0).astype(int)

        # 提取几板数值
        def extract_ji_ban(ji_tian_ji_ban):
            match = re.search(r'(\d+)天(\d+)板', str(ji_tian_ji_ban))
            if match:
                return int(match.group(2))
            return 0

        lianban_df['几板'] = lianban_df['几天几板'].apply(extract_ji_ban)

        # 提取最高几板（确保即使为0也显示）
        max_ji_ban = lianban_df['几板'].max() if not lianban_df.empty else 0
        if pd.isna(max_ji_ban):
            max_ji_ban = 0
        max_ji_ban_filtered = lianban_df[lianban_df['几板'] == max_ji_ban]
        max_ji_ban_stocks = []
        if not max_ji_ban_filtered.empty:
            max_ji_ban_stocks = [apply_bold_for_attention_stocks(
                format_stock_name_with_indicators(
                    row['股票代码'], row['股票简称'],
                    row['涨停开板次数'], row['首次涨停时间'], row['最终涨停时间']
                ),
                row['股票代码'], date, attention_by_date
            ) for _, row in max_ji_ban_filtered.iterrows()]
            # 提取最高几板的股票代码
            date_codes.extend(extract_stock_codes_from_df(max_ji_ban_filtered))
        max_ji_ban_results.append((date, max_ji_ban, max_ji_ban_stocks))

        # 提取最高连板（确保即使为0也显示）
        max_lianban = lianban_df['连续涨停天数'].max() if not lianban_df.empty else 0
        if pd.isna(max_lianban):
            max_lianban = 0
        max_lianban_filtered = lianban_df[lianban_df['连续涨停天数'] == max_lianban]
        max_lianban_stocks = []
        max_lianban_codes = set()  # 记录最高连板的股票代码（用于去重）
        if not max_lianban_filtered.empty:
            max_lianban_stocks = [apply_bold_for_attention_stocks(
                format_stock_name_with_indicators(
                    row['股票代码'], row['股票简称'],
                    row['涨停开板次数'], row['首次涨停时间'], row['最终涨停时间']
                ),
                row['股票代码'], date, attention_by_date
            ) for _, row in max_lianban_filtered.iterrows()]
            # 提取最高连板的股票代码
            date_codes.extend(extract_stock_codes_from_df(max_lianban_filtered))
            # 记录股票代码用于去重
            max_lianban_codes = set(extract_stock_codes_from_df(max_lianban_filtered))

        # 提取次高连板（确保即使为0也显示）
        second_lianban = lianban_df[lianban_df['连续涨停天数'] < max_lianban][
            '连续涨停天数'].max() if not lianban_df.empty and max_lianban > 0 else 0
        if pd.isna(second_lianban):
            second_lianban = 0
        second_lianban_filtered = lianban_df[lianban_df['连续涨停天数'] == second_lianban]
        second_lianban_stocks = []
        second_lianban_codes = set()  # 记录次高连板的股票代码（用于去重）
        if not second_lianban_filtered.empty and second_lianban > 0:
            second_lianban_stocks = [apply_bold_for_attention_stocks(
                format_stock_name_with_indicators(
                    row['股票代码'], row['股票简称'],
                    row['涨停开板次数'], row['首次涨停时间'], row['最终涨停时间']
                ),
                row['股票代码'], date, attention_by_date
            ) for _, row in second_lianban_filtered.iterrows()]
            # 提取次高连板的股票代码
            date_codes.extend(extract_stock_codes_from_df(second_lianban_filtered))
            # 记录股票代码用于去重
            second_lianban_codes = set(extract_stock_codes_from_df(second_lianban_filtered))

        # 提取次高几板（多阶，根据 JI_BAN_TIERS 配置）
        # 需要排除已经在最高连板和次高连板中出现的股票（避免重复显示）
        tier_ji_ban_stocks = []  # 存储所有阶次的股票（带板数标记）
        lianban_codes_to_exclude = max_lianban_codes | second_lianban_codes  # 合并两个集合
        if not lianban_df.empty and max_ji_ban > 0:
            # 获取所有不同的几板数（降序）
            unique_ji_bans = sorted(lianban_df['几板'].unique(), reverse=True)
            # 去掉最高几板，取接下来的 JI_BAN_TIERS 个阶次
            tier_ji_bans = [jb for jb in unique_ji_bans if jb < max_ji_ban][:JI_BAN_TIERS]

            for tier_ji_ban in tier_ji_bans:
                if tier_ji_ban > 0:
                    tier_filtered = lianban_df[lianban_df['几板'] == tier_ji_ban]
                    if not tier_filtered.empty:
                        for _, row in tier_filtered.iterrows():
                            # 提取股票代码（去掉交易所后缀）
                            clean_code = str(row['股票代码']).split('.')[0] if '.' in str(row['股票代码']) else str(
                                row['股票代码'])
                            # 如果股票代码已经在连板股票中，跳过（避免重复）
                            if clean_code in lianban_codes_to_exclude:
                                continue
                            # 使用类似 format_stock_name_with_lianban_count 的格式，在股票名后加板数
                            base_name = apply_bold_for_attention_stocks(
                                format_stock_name_with_indicators(
                                    row['股票代码'], row['股票简称'],
                                    row['涨停开板次数'], row['首次涨停时间'], row['最终涨停时间']
                                ),
                                row['股票代码'], date, attention_by_date
                            )
                            stock_with_count = f"{base_name}{tier_ji_ban}"
                            tier_ji_ban_stocks.append(stock_with_count)
                        # 提取股票代码（仅限未排除的股票）
                        tier_codes = [code for code in extract_stock_codes_from_df(tier_filtered)
                                      if code not in lianban_codes_to_exclude]
                        date_codes.extend(tier_codes)

        second_ji_ban_results.append((date, tier_ji_ban_stocks))

        # 筛选4连板及以上股票（仅在次高连板>4时）
        # 只显示未入选次高连板的部分，即4连板及以上但没达到次高连板数的股票
        lianban_4plus_stocks = []
        if second_lianban > 4:
            # 筛选所有连续涨停天数>=4且<次高连板数的股票（排除次高连板的股票）
            lianban_4plus_filtered = lianban_df[
                (lianban_df['连续涨停天数'] >= 4) &
                (lianban_df['连续涨停天数'] < second_lianban)
                ]
            if not lianban_4plus_filtered.empty:
                lianban_4plus_stocks = []
                for _, row in lianban_4plus_filtered.iterrows():
                    # 先格式化名称并应用加粗
                    base_name = apply_bold_for_attention_stocks(
                        format_stock_name_with_indicators(
                            row['股票代码'], row['股票简称'],
                            row['涨停开板次数'], row['首次涨停时间'], row['最终涨停时间']
                        ),
                        row['股票代码'], date, attention_by_date
                    )
                    # 然后加上连板数
                    stock_with_count = f"{base_name}{int(row['连续涨停天数'])}"
                    lianban_4plus_stocks.append(stock_with_count)
                # 提取4连板及以上股票的代码（用于点击复制）
                date_codes.extend(extract_stock_codes_from_df(lianban_4plus_filtered))
        lianban_4plus_results.append((date, lianban_4plus_stocks))

        lianban_results.append((date, max_lianban, max_lianban_stocks))
        lianban_second_results.append((date, second_lianban, second_lianban_stocks))

        # 跌停数据处理
        dieting_col = dieting_data[date].dropna()
        dieting_col = dieting_col.fillna('').astype(str)
        dieting_stocks = dieting_col.str.split(';').apply(lambda x: [item.strip() for item in x])
        dieting_df = pd.DataFrame(dieting_stocks.tolist(), columns=[
            '股票代码', '股票简称', '跌停开板次数', '首次跌停时间',
            '跌停类型', '最新价', '最新涨跌幅',
            '连续跌停天数', '跌停原因类型'
        ])

        if not dieting_df.empty:
            dieting_df['连续跌停天数'] = dieting_df['连续跌停天数'].fillna(0)
            dieting_df['连续跌停天数'] = dieting_df['连续跌停天数'].replace('', 0)
            dieting_df['连续跌停天数'] = pd.to_numeric(dieting_df['连续跌停天数'], errors='coerce').fillna(0).astype(
                int)

            max_dieting = dieting_df['连续跌停天数'].max()
            max_dieting_filtered = dieting_df[dieting_df['连续跌停天数'] == max_dieting]
            max_dieting_stocks = []
            if not max_dieting_filtered.empty:
                max_dieting_stocks = [apply_bold_for_attention_stocks(
                    format_stock_name_with_indicators(row['股票代码'], row['股票简称']),
                    row['股票代码'], date, attention_by_date
                ) for _, row in max_dieting_filtered.iterrows()]
                # 提取最大连续跌停的股票代码
                date_codes.extend(extract_stock_codes_from_df(max_dieting_filtered))
        else:
            max_dieting = 0
            max_dieting_stocks = []
        dieting_results.append((date, -max_dieting, max_dieting_stocks))

        # 首板数据
        shouban_col = shouban_data[date].dropna()
        shouban_counts.append(len(shouban_col))

        # 默默上涨数据处理
        if has_momo_data and date in momo_data.columns:
            momo_col = momo_data[date].dropna()
            momo_stocks_data = []  # 完整信息（涨幅+成交额）用于悬浮窗
            momo_stocks_simple = []  # 简化信息（仅涨幅）用于节点标签
            momo_stock_codes = []  # 股票代码列表，用于点击复制
            momo_zhangfus = []

            for cell in momo_col:
                if pd.isna(cell) or str(cell).strip() == '':
                    continue
                parts = str(cell).split(';')
                if len(parts) >= 6:
                    # 格式：股票代码; 股票简称; 最新价; 最新涨跌幅; 区间涨跌幅; 区间成交额; 区间振幅; 上市交易日天数
                    stock_code = parts[0].strip()
                    stock_name = parts[1].strip()
                    qujian_zhangfu = parts[4].strip()  # 区间涨跌幅（第5个字段）
                    qujian_chengjiao = parts[5].strip() if len(parts) > 5 else ''  # 区间成交额（第6个字段）

                    try:
                        # 去掉百分号，转换为浮点数
                        zhangfu_value = float(qujian_zhangfu.rstrip('%'))
                        momo_zhangfus.append(zhangfu_value)
                        # 提取股票代码（去掉交易所后缀，如 .SH .SZ）
                        clean_code = stock_code.split('.')[0] if '.' in stock_code else stock_code
                        momo_stock_codes.append(clean_code)
                        # 应用加粗逻辑
                        formatted_name = apply_bold_for_attention_stocks(stock_name, stock_code, date,
                                                                         attention_by_date)
                        # 完整信息：股票名称(涨幅, 成交额) - 用于悬浮窗
                        momo_stocks_data.append(f"{formatted_name}({qujian_zhangfu}, {qujian_chengjiao})")
                        # 简化信息：股票名称(涨幅) - 用于节点标签
                        momo_stocks_simple.append(f"{formatted_name}({qujian_zhangfu})")
                    except:
                        pass

            # 计算平均涨幅或最大涨幅
            if momo_zhangfus:
                avg_zhangfu = sum(momo_zhangfus) / len(momo_zhangfus)
                max_zhangfu = max(momo_zhangfus)
                sample_count = len(momo_zhangfus)  # 样本数量
                # 找出涨幅最高的前3只股票（用简化信息）
                top_3_indices = sorted(range(len(momo_zhangfus)), key=lambda i: momo_zhangfus[i], reverse=True)[:3]
                top_3_stocks = [momo_stocks_simple[i] for i in top_3_indices if i < len(momo_stocks_simple)]
                # 将股票代码列表转换为换行符分隔的字符串（便于复制）
                codes_str = '\n'.join(momo_stock_codes)
                # 添加样本数量和股票代码到结果中
                momo_results.append((date, avg_zhangfu, momo_stocks_data, top_3_stocks, sample_count, codes_str))
            else:
                # 没有数据时用None，不影响Y轴范围
                momo_results.append((date, None, [], [], 0, ''))
        elif has_momo_data:
            # 该日期没有默默上涨数据，用None
            momo_results.append((date, None, [], [], 0, ''))

        # 去重并存储当日所有股票代码（用于点击复制）
        unique_codes = list(dict.fromkeys(date_codes))  # 保持顺序的去重
        all_codes_by_date[date] = '\n'.join(unique_codes)

    # === 开始绘制Plotly图表 ===

    # 提取日期和数据
    lianban_dates = [datetime.strptime(item[0], "%Y年%m月%d日") for item in lianban_results]
    date_labels = [d.strftime('%Y-%m-%d') for d in lianban_dates]  # 修改日期格式为 yyyy-MM-dd

    # 创建代码列表（与date_labels对应，用于customdata）
    all_codes_list = [all_codes_by_date.get(item[0], '') for item in lianban_results]

    # 创建多Y轴图表（需要为默默上涨单独创建一个Y轴）
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # 用于记录支持复制的trace索引
    copyable_trace_indices = []

    # 最高几板线（副Y轴）
    max_ji_ban_days = [item[1] for item in max_ji_ban_results]
    max_ji_ban_stocks = [format_stock_list_for_hover(item[2], LIANBAN_STOCKS_PER_LINE) for item in max_ji_ban_results]
    max_ji_ban_labels = [create_display_labels(item[2]) for item in max_ji_ban_results]

    # 格式化次高几板股票（用于悬浮窗显示）
    second_ji_ban_stocks_formatted = []
    for idx, item in enumerate(second_ji_ban_results):
        if item[1]:  # 如果有次高几板股票（item[1]是股票列表）
            formatted = format_stock_list_for_hover(item[1], LIANBAN_STOCKS_PER_LINE)
            second_ji_ban_stocks_formatted.append(f'<br>---<br>次高几板:<br>{formatted}')
        else:
            second_ji_ban_stocks_formatted.append('')

    # 组合customdata：[股票列表, 该日所有代码, 次高几板股票文本]
    max_ji_ban_customdata = list(zip(max_ji_ban_stocks, all_codes_list, second_ji_ban_stocks_formatted))

    copyable_trace_indices.append(len(fig.data))
    fig.add_trace(
        go.Scatter(
            x=date_labels,
            y=max_ji_ban_days,
            name='最高几板',
            mode='lines+markers+text',
            line=dict(color='purple', width=2),
            marker=dict(symbol='star', size=10),
            text=max_ji_ban_labels,
            textposition='top center',
            textfont=dict(size=9, color='purple'),
            customdata=max_ji_ban_customdata,
            hovertemplate='几板: %{y}板<br>股票: %{customdata[0]}%{customdata[2]}<br><extra></extra>',
        ),
        secondary_y=True,
    )

    # 最高连板线（副Y轴）
    lianban_days = [item[1] for item in lianban_results]
    lianban_stocks = [format_stock_list_for_hover(item[2], LIANBAN_STOCKS_PER_LINE) for item in lianban_results]
    lianban_labels = [create_display_labels(item[2]) for item in lianban_results]
    # 组合customdata：[股票列表, 该日所有代码]
    lianban_customdata = list(zip(lianban_stocks, all_codes_list))

    copyable_trace_indices.append(len(fig.data))
    fig.add_trace(
        go.Scatter(
            x=date_labels,
            y=lianban_days,
            name='最高连续涨停天数',
            mode='lines+markers+text',  # 添加text模式
            line=dict(color='red', width=2),
            marker=dict(symbol='circle', size=10),
            text=lianban_labels,  # 永久显示的标签
            textposition='top center',
            textfont=dict(size=9, color='red'),
            customdata=lianban_customdata,
            hovertemplate='连板: %{y}板<br>股票: %{customdata[0]}<br><extra></extra>',
        ),
        secondary_y=True,
    )

    # 次高连板线（副Y轴）
    lianban_second_days = [item[1] for item in lianban_second_results]
    lianban_second_stocks = [format_stock_list_for_hover(item[2], LIANBAN_STOCKS_PER_LINE) for item in
                             lianban_second_results]
    lianban_second_labels = [create_display_labels(item[2]) for item in lianban_second_results]
    # 格式化4连板及以上股票（用于悬浮窗显示）
    lianban_4plus_stocks_formatted = []
    for item in lianban_4plus_results:
        if item[1]:  # 如果有4连板及以上股票
            formatted = format_stock_list_for_hover(item[1], LIANBAN_STOCKS_PER_LINE)
            lianban_4plus_stocks_formatted.append(f'<br>---<br>4连板及以上:<br>{formatted}')
        else:
            lianban_4plus_stocks_formatted.append('')
    # 组合customdata：[次高连板股票列表, 该日所有代码, 4连板及以上股票文本]
    lianban_second_customdata = list(zip(lianban_second_stocks, all_codes_list, lianban_4plus_stocks_formatted))

    copyable_trace_indices.append(len(fig.data))
    fig.add_trace(
        go.Scatter(
            x=date_labels,
            y=lianban_second_days,
            name='次高连续涨停天数',
            mode='lines+markers+text',
            line=dict(color='orange', width=2),
            marker=dict(symbol='square', size=8),
            text=lianban_second_labels,
            textposition='bottom center',
            textfont=dict(size=9, color='orange'),
            customdata=lianban_second_customdata,
            hovertemplate='次高连板: %{y}板<br>股票: %{customdata[0]}%{customdata[2]}<br><extra></extra>',
        ),
        secondary_y=True,
    )

    # 首板数量线（主Y轴）
    fig.add_trace(
        go.Scatter(
            x=date_labels,
            y=shouban_counts,
            name='首板数量',
            mode='lines+markers+text',  # 添加text模式，永久显示标签
            line=dict(color='blue', width=2, dash='dash'),
            marker=dict(symbol='diamond', size=8),
            text=[f'{count}' for count in shouban_counts],  # 显示数量
            textposition='top center',
            textfont=dict(size=10, color='blue'),
            opacity=0.3,
            hovertemplate='首板数量: %{y}<extra></extra>',  # 去掉日期，顶部统一显示
        ),
        secondary_y=False,
    )

    # 跌停线（副Y轴）
    dieting_days = [item[1] for item in dieting_results]
    dieting_stocks = [format_stock_list_for_hover(item[2], LIANBAN_STOCKS_PER_LINE) for item in dieting_results]
    dieting_labels = [create_display_labels(item[2]) for item in dieting_results]
    # 组合customdata：[股票列表, 该日所有代码]
    dieting_customdata = list(zip(dieting_stocks, all_codes_list))

    copyable_trace_indices.append(len(fig.data))
    fig.add_trace(
        go.Scatter(
            x=date_labels,
            y=dieting_days,
            name='最大连续跌停天数',
            mode='lines+markers+text',
            line=dict(color='green', width=2),
            marker=dict(symbol='triangle-down', size=8),
            text=dieting_labels,
            textposition='bottom center',
            textfont=dict(size=9, color='green'),
            customdata=dieting_customdata,
            hovertemplate='<br>跌停: %{y}天<br>股票: %{customdata[0]}<br><i>💡 点击节点复制当日所有股票代码</i><br><extra></extra>',
        ),
        secondary_y=True,
    )

    # 默默上涨线（独立Y轴）- 显示平均涨幅
    momo_trace_index = None
    profit_trace_index = None  # 盈亏折线索引
    stock_profits_detail = {}  # 每只股票的盈亏详情
    momo_annotations = []  # 用于存储样本数量的annotations
    if has_momo_data and momo_results:
        # 先计算每日盈亏并获取每只股票的盈亏详情（在创建trace之前）
        try:
            profit_results, stock_profits_detail = calculate_daily_profit_with_stock_details_from_momo_results(
                momo_results, buy_days_before=buy_days_before
            )

            # 获取所有日期中的最大日期，用于判断t+2日是否已经过去
            max_date_yyyymmdd = None
            if momo_results:
                max_date_str = max([item[0] for item in momo_results])
                max_date_obj = datetime.strptime(max_date_str, "%Y年%m月%d日")
                max_date_yyyymmdd = max_date_obj.strftime("%Y%m%d")

            # 重新格式化股票名称，加入"未来2日的盈亏"信息
            # 重要说明：
            # - 在日期t，显示的是t日选出的股票的"未来2日盈亏"（t+1日买入，t+2日卖出）
            # - 只有当t+2日已经过去（即最大日期 >= t+2）时，才能显示这个盈亏
            # - 如果t+2日还没到，则不显示盈亏（保持原格式）
            # momo_results格式: (date, avg_zhangfu, momo_stocks_data, top_3_stocks, sample_count, codes_str)
            for idx, item in enumerate(momo_results):
                date_str = item[0]
                original_stocks_data = item[2]  # 原始格式：股票名称(涨幅, 成交额)
                codes_str = item[5]  # t日显示的股票代码列表

                # 重新格式化股票名称，加入盈亏
                new_stocks_data = []
                stock_codes_list = codes_str.split('\n') if codes_str else []

                # 将当前日期转换为日期对象，用于比较
                current_date_obj = datetime.strptime(date_str, "%Y年%m月%d日")
                current_date_yyyymmdd = current_date_obj.strftime("%Y%m%d")

                for stock_idx, stock_info in enumerate(original_stocks_data):
                    if stock_idx < len(stock_codes_list):
                        clean_code = stock_codes_list[stock_idx].strip()

                        # 查找这只股票的"未来2日的盈亏"
                        # 重要：在日期t，显示的是t日选出的股票的"未来2日盈亏"（t+1日买入，t+2日卖出）
                        # 只有当t+2日已经过去（即当前日期 >= t+2）时，才能显示这个盈亏
                        # 如果t+2日还没到，则不显示盈亏
                        profit = None
                        from utils.date_util import get_next_trading_day
                        from utils.file_util import read_stock_data

                        try:
                            # 当前日期t就是选股日
                            select_date_yyyymmdd = current_date_yyyymmdd

                            # 买入日 = t+1日
                            buy_date_yyyymmdd = get_next_trading_day(select_date_yyyymmdd)
                            if buy_date_yyyymmdd:
                                # 卖出日 = t+2日
                                sell_date_yyyymmdd = get_next_trading_day(buy_date_yyyymmdd)

                                # 只有当卖出日已经过去（即最大日期 >= t+2）时，才能计算盈亏
                                if sell_date_yyyymmdd and max_date_yyyymmdd and sell_date_yyyymmdd <= max_date_yyyymmdd:
                                    # 读取股票数据，计算盈亏
                                    stock_data = read_stock_data(clean_code, './data/astocks')
                                    if stock_data is not None and not stock_data.empty:
                                        stock_data = stock_data.sort_values('日期').reset_index(drop=True)
                                        stock_data['日期_str'] = stock_data['日期'].dt.strftime('%Y%m%d')

                                        # 查找买入日和卖出日的数据
                                        buy_data = stock_data[stock_data['日期_str'] == buy_date_yyyymmdd]
                                        sell_data = stock_data[stock_data['日期_str'] == sell_date_yyyymmdd]

                                        if not buy_data.empty and not sell_data.empty:
                                            buy_row = buy_data.iloc[0]
                                            sell_row = sell_data.iloc[0]

                                            buy_open = buy_row['开盘']
                                            sell_close = sell_row['收盘']

                                            if buy_open is not None and sell_close is not None and buy_open != 0:
                                                # 计算盈亏：(t+2日收盘价 - t+1日开盘价) / t+1日开盘价
                                                profit = ((sell_close - buy_open) / buy_open) * 100
                        except Exception as e:
                            # 如果计算失败，跳过
                            pass

                        if profit is not None:
                            # 格式：股票名称(涨幅, 成交额, 盈亏%)
                            # 含义：在选股日选出，选股日+1日开盘买入，选股日+2日收盘卖出的盈亏为profit%
                            new_stock_info = _add_profit_to_stock_info(stock_info, profit)
                            new_stocks_data.append(new_stock_info)
                        else:
                            # 如果这只股票从未被选出过，或者卖出日还没到，则不显示盈亏
                            new_stocks_data.append(stock_info)
                    else:
                        new_stocks_data.append(stock_info)

                # 更新momo_results中的股票信息
                date_str, avg_zhangfu, _, top_3_stocks, sample_count, codes_str = item
                momo_results[idx] = (date_str, avg_zhangfu, new_stocks_data,
                                     top_3_stocks, sample_count, codes_str)
        except Exception as e:
            print(f"⚠ 计算默默上涨每日盈亏时出错: {e}")
            import traceback
            traceback.print_exc()
            stock_profits_detail = {}

        momo_zhangfus = [item[1] for item in momo_results]  # 平均涨幅
        # 悬浮窗显示所有股票（包含成交额和盈亏），每3只换行
        momo_all_stocks = [format_stock_list_for_hover(item[2], MOMO_STOCKS_PER_LINE) for item in momo_results]
        # 提取股票代码字符串（用于点击复制）
        momo_stock_codes = [item[5] for item in momo_results]  # item[5] 是代码字符串

        # 记录默默上涨trace的索引（当前是最后一个）
        momo_trace_index = len(fig.data)

        # 创建标签：显示前3只涨幅最高的股票（注意：是item[3]而不是item[2]）
        momo_labels = []
        for item in momo_results:
            if item[1] is None:  # 没有数据
                momo_labels.append('')
            else:
                momo_labels.append(create_display_labels(item[3]))

        # 创建样本数量的annotations（显示在节点下方）
        for i, item in enumerate(momo_results):
            if item[1] is not None:  # 有数据时才显示
                sample_count = item[4]  # 样本数量
                momo_annotations.append(
                    dict(
                        x=date_labels[i],
                        y=item[1],  # Y坐标为平均涨幅
                        xref='x',
                        yref='y3',  # 使用y3轴
                        text=f'{sample_count}只',
                        showarrow=False,
                        font=dict(size=8, color='brown'),
                        xanchor='center',
                        yanchor='top',
                        yshift=-10,  # 向下偏移10像素
                        visible=False,  # 默认隐藏（跟随图层切换）
                    )
                )

        # 准备 customdata（二维数组：[股票列表用于hover, 股票代码用于复制]）
        # 使用嵌套列表格式，与旧版本保持一致（旧版本此格式工作正常）
        momo_customdata = [[stocks, codes] for stocks, codes in zip(momo_all_stocks, momo_stock_codes)]

        # 先记录当前trace索引（添加trace之前的索引，添加后就是新trace的索引）
        momo_trace_index_before = len(fig.data)
        fig.add_trace(
            go.Scatter(
                x=date_labels,
                y=momo_zhangfus,
                name='默默上涨(平均涨幅%)',
                mode='lines+markers+text',  # 添加text显示标签
                line=dict(color='brown', width=2, dash='dot'),
                marker=dict(symbol='diamond-open', size=8),
                text=momo_labels,  # 显示TOP3股票
                textposition='top center',
                textfont=dict(size=9, color='brown'),
                visible=False,  # 默认隐藏，不显示
                showlegend=True,  # 显示图例
                legendgroup='momo',  # 图例分组
                customdata=momo_customdata,
                # 独立悬浮窗，去掉日期（顶部已有），添加点击提示和说明
                hovertemplate=f'平均涨幅: %{{y:.1f}}%<br>股票: %{{customdata[0]}}<br><i>💡 点击节点复制股票代码</i><br><i>{_generate_stock_info_explanation(buy_days_before)}</i><extra></extra>',
                hoverinfo='all',
                hoverlabel=dict(
                    bgcolor='rgba(139, 69, 19, 0.9)',  # 棕色背景
                    font=dict(color='white', size=12, family='SimHei')
                ),
                yaxis='y3',  # 使用第三个Y轴
            )
        )
        # 添加trace后，记录其索引（新添加的trace索引就是添加前的len(fig.data)）
        copyable_trace_indices.append(momo_trace_index_before)

        # 初始化y4_title（如果没有盈亏数据，使用默认值）
        y4_title = '隔日平均盈亏(%)' if buy_days_before == 1 else f't-{buy_days_before}日选出平均盈亏(%)'

        # 添加盈亏折线（使用y4轴，默认隐藏）
        if stock_profits_detail:
            try:
                # 按照date_labels的顺序，分别计算高门槛股和正常股的平均盈亏
                profit_values_high = []  # 高门槛股盈亏值
                profit_counts_high = []  # 高门槛股样本数
                profit_values_normal = []  # 正常股盈亏值
                profit_counts_normal = []  # 正常股样本数

                for item in momo_results:
                    date_str = item[0]
                    codes_str = item[5]  # 股票代码字符串（换行符分隔）

                    # 获取该日期的所有股票盈亏数据
                    if date_str in stock_profits_detail:
                        date_profits = stock_profits_detail[date_str]

                        # 将股票分为两组
                        high_threshold_profits = []
                        normal_profits = []

                        # 从codes_str中提取股票代码列表
                        stock_codes_list = codes_str.split('\n') if codes_str else []

                        for stock_code in stock_codes_list:
                            # 清理股票代码（去除市场后缀，与stock_profits_detail中的格式一致）
                            clean_code = stock_code.strip()
                            if not clean_code:
                                continue
                            clean_code = clean_code.split('.')[0] if '.' in clean_code else clean_code
                            clean_code = clean_code.split('_')[0] if '_' in clean_code else clean_code

                            # 查找该股票的盈亏
                            if clean_code in date_profits:
                                profit = date_profits[clean_code]
                                if profit is not None:
                                    # 判断是否为高门槛股
                                    if _is_high_threshold_stock(clean_code):
                                        high_threshold_profits.append(profit)
                                    else:
                                        normal_profits.append(profit)

                        # 计算平均盈亏
                        if high_threshold_profits:
                            avg_high = sum(high_threshold_profits) / len(high_threshold_profits)
                            profit_values_high.append(avg_high)
                            profit_counts_high.append(len(high_threshold_profits))
                        else:
                            profit_values_high.append(None)
                            profit_counts_high.append(0)

                        if normal_profits:
                            avg_normal = sum(normal_profits) / len(normal_profits)
                            profit_values_normal.append(avg_normal)
                            profit_counts_normal.append(len(normal_profits))
                        else:
                            profit_values_normal.append(None)
                            profit_counts_normal.append(0)
                    else:
                        # 该日期没有盈亏数据
                        profit_values_high.append(None)
                        profit_counts_high.append(0)
                        profit_values_normal.append(None)
                        profit_counts_normal.append(0)

                # 添加高门槛股盈亏折线（使用y4轴，默认隐藏）
                profit_trace_index_high = len(fig.data)
                fig.add_trace(
                    go.Scatter(
                        x=date_labels,
                        y=profit_values_high,
                        name=f'{y4_title}(高门槛)',
                        mode='lines+markers',
                        line=dict(color='darkorange', width=2, dash='dash'),
                        marker=dict(symbol='circle', size=6),
                        visible=False,  # 默认隐藏，跟随默默上涨图层
                        showlegend=True,
                        legendgroup='momo',  # 与默默上涨同一图例组
                        hovertemplate=f'平均盈亏: %{{y:.2f}}%<br>有效样本: %{{customdata}}只<br><i>说明：{_generate_profit_explanation(buy_days_before)}</i><br><extra></extra>',
                        customdata=profit_counts_high,
                        yaxis='y4',  # 使用第四个Y轴（独立Y轴）
                    )
                )

                # 添加正常股盈亏折线（使用y4轴，默认隐藏）
                profit_trace_index_normal = len(fig.data)
                fig.add_trace(
                    go.Scatter(
                        x=date_labels,
                        y=profit_values_normal,
                        name=f'{y4_title}(正常)',
                        mode='lines+markers',
                        line=dict(color='darkgreen', width=2, dash='dot'),
                        marker=dict(symbol='square', size=6),
                        visible=False,  # 默认隐藏，跟随默默上涨图层
                        showlegend=True,
                        legendgroup='momo',  # 与默默上涨同一图例组
                        hovertemplate=f'平均盈亏: %{{y:.2f}}%<br>有效样本: %{{customdata}}只<br><i>说明：{_generate_profit_explanation(buy_days_before)}</i><br><extra></extra>',
                        customdata=profit_counts_normal,
                        yaxis='y4',  # 使用第四个Y轴（独立Y轴）
                    )
                )

                # 记录两条折线的索引（用于图层切换）
                profit_trace_index = [profit_trace_index_high, profit_trace_index_normal]
            except Exception as e:
                print(f"⚠ 添加盈亏折线时出错: {e}")
                import traceback
                traceback.print_exc()
                profit_trace_index = None

    # 创建图层切换按钮（如果有默默上涨数据）
    updatemenus = []
    if momo_trace_index is not None:
        total_traces = len(fig.data)
        # 为每个annotation设置visible属性（跟随图层切换）
        annotations_count = len(momo_annotations)

        # 确定盈亏折线的索引（如果存在）
        profit_trace_indices = []
        if profit_trace_index is not None:
            # profit_trace_index可能是单个索引或索引列表
            if isinstance(profit_trace_index, list):
                profit_trace_indices = profit_trace_index
            else:
                profit_trace_indices = [profit_trace_index]

        updatemenus = [
            dict(
                type="buttons",
                direction="left",
                buttons=[
                    dict(
                        args=[
                            {"visible": [True if (i != momo_trace_index and i not in profit_trace_indices) else False
                                         for i in range(total_traces)]},
                            {
                                "yaxis.visible": True,
                                "yaxis2.visible": True,
                                "yaxis3.visible": False,
                                "yaxis4.visible": False,
                                # 隐藏所有样本数量annotations
                                "annotations": [dict(ann, visible=False) for ann in momo_annotations],
                            }
                        ],
                        label="📊 连板天梯",
                        method="update"
                    ),
                    dict(
                        args=[
                            {"visible": [False if (i != momo_trace_index and i not in profit_trace_indices) else True
                                         for i in range(total_traces)]},
                            {
                                "yaxis.visible": False,
                                "yaxis2.visible": False,
                                "yaxis3.visible": True,
                                "yaxis4.visible": True,
                                # 显示所有样本数量annotations
                                "annotations": [dict(ann, visible=True) for ann in momo_annotations],
                            }
                        ],
                        label="📈 默默上涨",
                        method="update"
                    ),
                ],
                pad={"r": 10, "t": 10},
                showactive=True,
                active=0,  # 默认选中"连板天梯"
                x=0.15,
                xanchor="left",
                y=1.09,
                yanchor="top",
                bgcolor='rgba(255, 255, 255, 0.95)',
                bordercolor='#2196F3',
                borderwidth=2,
                font=dict(size=13, family='SimHei', color='#333'),
            ),
        ]

    # 更新布局
    fig.update_xaxes(
        title_text="日期",
        tickangle=-45,
        tickfont=dict(size=10),
        type='category',  # 确保日期按分类显示，不会自动格式化
    )

    fig.update_yaxes(
        title_text="数量",
        secondary_y=False,
        tickformat=',d',
        zeroline=True,
        zerolinewidth=2,
        zerolinecolor='gray',
    )

    fig.update_yaxes(
        title_text="连板/跌停/几板天数",
        secondary_y=True,
        tickformat=',d',
        zeroline=True,
        zerolinewidth=2,
        zerolinecolor='gray',
    )

    fig.update_layout(
        title=dict(
            text="连板/跌停/首板/默默上涨个股走势",
            x=0.5,
            xanchor='center',
            font=dict(size=20, family='SimHei'),
        ),
        hovermode='x unified',
        legend=dict(
            x=0.01,
            y=0.99,
            xanchor='left',
            yanchor='top',
            bgcolor='rgba(255, 255, 255, 0.8)',
            bordercolor='gray',
            borderwidth=1,
        ),
        updatemenus=updatemenus,  # 添加切换按钮
        annotations=momo_annotations,  # 添加样本数量标注（默认隐藏）
        width=None,  # 改为自适应宽度，避免横向滚动条
        height=900,
        font=dict(family='SimHei'),
        plot_bgcolor='white',
        paper_bgcolor='white',
        autosize=True,  # 启用自适应大小
        # 配置第三个Y轴（默默上涨专用）
        yaxis3=dict(
            title=dict(
                text="默默上涨涨幅(%)",
                font=dict(color='brown', size=12, family='SimHei')
            ),
            overlaying='y',  # 覆盖在主Y轴上
            side='right',  # 显示在右侧
            # 不设置position，让它自然靠近图表右侧
            showgrid=True,  # 显示网格线
            gridwidth=1,
            gridcolor='lightgray',
            zeroline=True,
            zerolinewidth=2,
            zerolinecolor='gray',
            tickfont=dict(color='brown', size=10),
            tickformat='.1f',
            ticksuffix='%',
            visible=False,  # 默认隐藏（连板天梯图层不显示）
        ),
        # 配置第四个Y轴（盈亏专用，显示在左侧，标题根据buy_days_before动态生成）
        yaxis4=dict(
            title=dict(
                text=y4_title,
                font=dict(color='darkorange', size=12, family='SimHei')
            ),
            overlaying='y',  # 覆盖在主Y轴上
            side='left',  # 显示在左侧
            showgrid=False,  # 不显示网格线，避免与其他轴混淆
            zeroline=True,
            zerolinewidth=2,
            zerolinecolor='gray',
            tickfont=dict(color='darkorange', size=10),
            tickformat='.2f',
            ticksuffix='%',
            visible=False,  # 默认隐藏（连板天梯图层不显示）
        ),
    )

    # 添加网格线
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray', secondary_y=False)
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray', secondary_y=True)

    # 生成文件名
    if output_path is None:
        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date}_to_{end_date}"
        elif start_date:
            date_range = f"from_{start_date}"
        elif end_date:
            date_range = f"to_{end_date}"
        else:
            date_range = datetime.now().strftime('%Y%m%d')

        output_path = f"images/fupan_lb_{date_range}.html"

    # 确保目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    # 保存HTML文件
    fig.write_html(
        output_path,
        config={
            'displayModeBar': True,
            'displaylogo': False,
            'modeBarButtonsToRemove': ['lasso2d', 'select2d'],
            'toImageButtonOptions': {
                'format': 'png',
                'filename': 'fupan_lb',
                'height': 900,
                'width': 1800,
                'scale': 2
            }
        }
    )

    # 注入JavaScript实现点击复制股票代码功能（通用版）
    if copyable_trace_indices:
        _inject_click_copy_script(output_path, copyable_trace_indices)

    print(f"HTML图表已保存到: {output_path}")
    return output_path


def draw_fupan_lb_html(start_date=None, end_date=None, output_path=None, buy_days_before=1):
    """
    生成HTML交互式复盘图的便捷函数
    
    Args:
        start_date: 开始日期（格式: YYYYMMDD）
        end_date: 结束日期（格式: YYYYMMDD）
        output_path: 输出HTML文件路径（可选）
        buy_days_before: 选股日相对于当前日的前N个交易日，默认为1（表示t-1日选出）
            - 1: t-1日选出，t-1+1=t日开盘买入，t-1+2=t+1日收盘卖出（隔日盈亏）
            - 2: t-2日选出，t-2+1=t-1日开盘买入，t-2+2=t日收盘卖出
            - 3: t-3日选出，t-3+1=t-2日开盘买入，t-3+2=t-1日收盘卖出
    
    Returns:
        str: 生成的HTML文件路径
    """
    fupan_file = "./excel/fupan_stocks.xlsx"
    return read_and_plot_html(fupan_file, start_date, end_date, output_path, buy_days_before)


if __name__ == '__main__':
    # 测试
    start_date = '20250830'
    end_date = None
    draw_fupan_lb_html(start_date, end_date)
